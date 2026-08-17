"""
Unit tests for the researched repair-cost table.

What matters here is not the arithmetic. It is that a number nobody
researched can never appear as though somebody did:

  * every figure names its sources and the date they were read;
  * a figure that is not in the table cannot be produced at all -- no
    fallback, no nearest-neighbour, no interpolation;
  * a national average never overwrites an inspector who stood in the
    room;
  * "unpriced" and "not a cost item" stay distinct, so measurements do
    not sit on the needs-pricing list forever.
"""

import unittest

from tools import site_dd_bank as bank
from tools import site_dd_checklist as cl
from tools import site_dd_costs as costs
from tools import site_dd_reference_costs as ref
from tools import site_dd_unit_checklist as uc


def all_item_keys():
    keys = []
    for room_type, _ in uc.ROOM_TYPES:
        keys += [i["key"] for i in uc.items_for_room(room_type)]
    keys += [i["key"] for i in uc.items_for_unit()]
    keys += list(cl.ITEM_LABELS)
    keys += [b["key"] for b in bank.BANK_ITEMS]
    return list(dict.fromkeys(keys))


class ProvenanceTests(unittest.TestCase):
    def test_every_figure_names_its_sources(self):
        for key, entry in ref.REFERENCE_COSTS.items():
            with self.subTest(key):
                self.assertTrue(entry.sources, "no source cited")

    def test_every_figure_says_how_it_was_derived(self):
        for key, entry in ref.REFERENCE_COSTS.items():
            with self.subTest(key):
                self.assertTrue(entry.note.strip())

    def test_every_provenance_carries_the_required_phrase(self):
        """So a later edit cannot present an average as a quote."""
        for key, entry in ref.REFERENCE_COSTS.items():
            with self.subTest(key):
                self.assertIn(ref.REQUIRED_PROVENANCE_PHRASE,
                              entry.provenance.lower())

    def test_every_provenance_says_it_is_not_a_quote(self):
        for key, entry in ref.REFERENCE_COSTS.items():
            with self.subTest(key):
                self.assertIn("not a quote", entry.provenance)

    def test_every_provenance_carries_the_research_date(self):
        for key, entry in ref.REFERENCE_COSTS.items():
            with self.subTest(key):
                self.assertIn(ref.RESEARCHED_ON, entry.provenance)

    def test_the_visible_label_warns_it_is_national(self):
        self.assertIn("not a quote", costs.SOURCE_LABELS[costs.SOURCE_REFERENCE])

    def test_nothing_here_can_reach_a_network(self):
        """Michelle asked explicitly for no scraping. The module is made
        incapable of it rather than merely instructed not to."""
        from pathlib import Path
        src = Path("tools/site_dd_reference_costs.py").read_text(encoding="utf-8")
        for forbidden in ("requests", "urllib", "http", "socket", "webbrowser"):
            with self.subTest(forbidden):
                self.assertNotIn(f"import {forbidden}", src)


class TableShapeTests(unittest.TestCase):
    def test_every_cost_is_positive(self):
        for key, entry in ref.REFERENCE_COSTS.items():
            with self.subTest(key):
                self.assertGreater(entry.unit_cost, 0)

    def test_every_unit_is_a_known_unit(self):
        for key, entry in ref.REFERENCE_COSTS.items():
            with self.subTest(key):
                self.assertIn(entry.unit, ref.UNITS)

    def test_per_sqft_figures_are_plausibly_per_sqft(self):
        """A per-sqft figure in the hundreds is a units mistake, and one
        that would multiply into a six-figure error on a real area."""
        for key, entry in ref.REFERENCE_COSTS.items():
            if entry.unit in (ref.UNIT_SQFT, ref.UNIT_LF):
                with self.subTest(key):
                    self.assertLess(entry.unit_cost, 100.0)

    def test_every_key_maps_to_a_real_item(self):
        known = set(all_item_keys())
        for key in ref.REFERENCE_COSTS:
            with self.subTest(key):
                self.assertIn(key, known, "prices an item that does not exist")

    def test_no_key_is_both_priced_and_unpriced(self):
        overlap = set(ref.REFERENCE_COSTS) & set(ref.UNPRICED)
        self.assertEqual(overlap, set())

    def test_no_key_is_both_unpriced_and_not_a_cost_item(self):
        overlap = set(ref.UNPRICED) & set(ref.NOT_A_COST_ITEM)
        self.assertEqual(overlap, set())


class CoverageTests(unittest.TestCase):
    def test_every_checklist_item_is_accounted_for(self):
        """No item may be silently absent. It is priced, explicitly
        unpriced, or explicitly not a cost item -- never unknown."""
        unknown = [k for k in all_item_keys() if ref.status(k) == "unknown"]
        self.assertEqual(unknown, [], f"unaccounted items: {unknown}")

    def test_measurements_are_not_on_the_needs_pricing_list(self):
        for key in ("water_heater_age", "hvac_age", "water_heater_gal"):
            with self.subTest(key):
                self.assertEqual(ref.status(key), "not_a_cost_item")
                self.assertNotIn(key, ref.UNPRICED)

    def test_every_unpriced_item_explains_itself(self):
        for key, why in ref.UNPRICED.items():
            with self.subTest(key):
                self.assertGreater(len(why.strip()), 20,
                                   "reason too thin to be useful")

    def test_the_unpriced_report_is_readable(self):
        rows = ref.unpriced_report({"foundation": "Foundation"})
        self.assertTrue(rows)
        self.assertTrue(all(r["reason"] for r in rows))
        labels = [r["label"] for r in rows]
        self.assertEqual(labels, sorted(labels, key=str.lower))


class LookupTests(unittest.TestCase):
    def test_a_priced_item_returns_its_figure(self):
        entry = ref.for_item("water_heater")
        self.assertEqual(entry.unit_cost, 1725.00)
        self.assertEqual(entry.unit, ref.UNIT_EACH)

    def test_an_unpriced_item_returns_nothing(self):
        self.assertIsNone(ref.for_item("foundation"))

    def test_a_measurement_returns_nothing(self):
        self.assertIsNone(ref.for_item("hvac_age"))

    def test_an_unknown_key_returns_nothing(self):
        self.assertIsNone(ref.for_item("not_a_real_item"))

    def test_flooring_is_priced_by_material(self):
        self.assertEqual(ref.for_item("flooring", "carpet").unit_cost, 3.50)
        self.assertEqual(ref.for_item("flooring", "hardwood").unit_cost, 12.75)
        self.assertNotEqual(ref.for_item("flooring", "carpet").unit_cost,
                            ref.for_item("flooring", "tile").unit_cost)

    def test_flooring_with_no_type_falls_back_to_the_common_case(self):
        self.assertEqual(ref.for_item("flooring").unit_cost, 6.50)

    def test_flooring_types_match_the_checklist_options(self):
        options = {v for v, _ in uc.FLOORING_TYPES}
        priced = set(ref.FLOORING_BY_TYPE)
        missing = options - priced - {"other"}
        self.assertEqual(missing, set(), f"unpriced flooring types: {missing}")


class ApplicationTests(unittest.TestCase):
    def test_a_reference_cost_is_applied_and_labelled(self):
        out = costs.apply_reference({"item_key": "co_alarm"})
        self.assertEqual(out["est_unit_cost"], 195.00)
        self.assertEqual(out["est_cost_source"], costs.SOURCE_REFERENCE)

    def test_it_never_overwrites_an_inspector(self):
        out = costs.apply_reference({"item_key": "co_alarm",
                                     "est_unit_cost": 80.0,
                                     "est_cost_source": costs.SOURCE_MANUAL})
        self.assertEqual(out["est_unit_cost"], 80.0)
        self.assertEqual(out["est_cost_source"], costs.SOURCE_MANUAL)

    def test_an_unpriced_item_is_left_alone(self):
        out = costs.apply_reference({"item_key": "foundation"})
        self.assertIsNone(out.get("est_unit_cost"))

    def test_a_bank_item_is_priced_through_its_bank_key(self):
        out = costs.apply_reference({"item_key": "custom_x",
                                     "bank_item_key": "tankless_water_heater"})
        self.assertEqual(out["est_unit_cost"], 2375.00)

    def test_describe_reports_it_as_a_reference(self):
        out = costs.apply_reference({"item_key": "toilet"})
        d = costs.describe(out)
        self.assertEqual(d["source"], costs.SOURCE_REFERENCE)
        self.assertFalse(d["is_estimate"], "a reference is not an inspector estimate")
        self.assertIn("not a quote", d["label"])


class ExportTests(unittest.TestCase):
    LABELS = {"water_heater": "Water heater", "toilet": "Toilet",
              "foundation": "Foundation"}

    def _lines(self):
        from tools import site_dd_capex_export as capex
        findings = [
            costs.apply_reference({"item_key": "water_heater", "scope": "unit",
                                   "condition": "replace", "category_key": "mep"}),
            {"item_key": "toilet", "scope": "room", "condition": "replace",
             "category_key": "mep", "est_unit_cost": 450.0,
             "est_cost_source": costs.SOURCE_MANUAL},
            {"item_key": "foundation", "scope": "property", "condition": "replace",
             "category_key": "structural_envelope"},
        ]
        return capex.build_lines(findings, self.LABELS), capex

    def test_the_three_sources_render_distinctly(self):
        lines, capex = self._lines()
        self.assertEqual([l["source_label"] for l in lines],
                         ["Researched average", "Inspector estimate", "No estimate"])

    def test_the_totals_are_split_by_source(self):
        lines, capex = self._lines()
        s = capex.summarize(lines)
        self.assertEqual(s["by_source"][costs.SOURCE_REFERENCE], 1725.00)
        self.assertEqual(s["by_source"][costs.SOURCE_MANUAL], 450.00)
        self.assertEqual(s["total"], 2175.00)

    def test_the_researched_share_is_reported(self):
        lines, capex = self._lines()
        s = capex.summarize(lines)
        self.assertAlmostEqual(s["researched_pct"], 1725 / 2175 * 100, places=6)

    def test_an_unpriced_line_is_listed_but_adds_nothing(self):
        lines, capex = self._lines()
        s = capex.summarize(lines)
        self.assertEqual(s["unpriced_count"], 1)
        found = next(l for l in lines if l["item_key"] == "foundation")
        # None rather than 0.0: a zero would sum into the total as though
        # the work were free.
        self.assertIsNone(found["total"])
        self.assertTrue(found["reason"])

    def test_a_budget_with_nothing_in_it_does_not_divide_by_zero(self):
        from tools import site_dd_capex_export as capex
        s = capex.summarize([])
        self.assertEqual(s["total"], 0.0)
        self.assertEqual(s["researched_pct"], 0.0)


if __name__ == "__main__":
    unittest.main()


class ReadabilityTests(unittest.TestCase):
    """The 'Not priced' sheet is the thing that goes to Michelle. A list
    meant to be read by somebody who does not work in this codebase must
    not be written in its identifiers."""

    def _labels(self):
        labels = dict(cl.ITEM_LABELS)
        for room_type, _ in uc.ROOM_TYPES:
            labels.update({i["key"]: i["label"]
                           for i in uc.items_for_room(room_type)})
        labels.update({i["key"]: i["label"] for i in uc.items_for_unit()})
        labels.update({b["key"]: b["label"] for b in bank.BANK_ITEMS})
        return labels

    def test_the_unpriced_report_uses_human_labels_when_given_them(self):
        rows = ref.unpriced_report(self._labels())
        by_key = {r["key"]: r["label"] for r in rows}
        self.assertEqual(by_key["ada_parking_path"], "ADA parking & path of travel")
        self.assertEqual(by_key["appliance_dishwasher"], "Dishwasher")
        self.assertEqual(by_key["wd_hookups"], "W/D hookups only")

    def test_almost_every_unpriced_item_has_a_real_label(self):
        """A raw key surviving into the list means an item nobody named."""
        labels = self._labels()
        raw = [r["key"] for r in ref.unpriced_report(labels)
               if r["label"] == r["key"]]
        # concrete_flooring is a flooring MATERIAL, not a checklist item,
        # so it legitimately has no item label of its own.
        self.assertEqual(raw, ["concrete_flooring"], f"unlabelled: {raw}")

    def test_the_export_sheets_carry_labels(self):
        import io
        import tempfile
        from openpyxl import load_workbook
        from tools import site_dd_capex_export as capex

        out = Path(tempfile.mkdtemp()) / "b.xlsx"
        capex.build_xlsx(out, {"property_label": "T"}, [], capex.summarize([]),
                         self._labels())
        wb = load_workbook(out)
        un = wb["Not priced"]
        self.assertEqual([c.value for c in un[1]][:2], ["Item", "Key"])
        first_col = {un.cell(row=r, column=1).value
                     for r in range(2, un.max_row + 1)}
        self.assertIn("ADA parking & path of travel", first_col)
        ref_sheet = wb["Reference costs"]
        self.assertEqual([c.value for c in ref_sheet[1]][:2], ["Item", "Key"])


from pathlib import Path  # noqa: E402
