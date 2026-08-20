"""Ctrl-P on the downloaded summary prints the summary, not 33 pages.

WHAT WENT WRONG

The download is the source MMR with a Summary sheet prepended. The source
workbook arrives with whatever tabs were selected when somebody last saved
it, and setting Summary active did not deselect those. Excel's default
print option is "Print Active Sheets" -- plural -- so with two tabs
selected it printed both, and the second was a raw export sheet with no
print area. General Ledger is 1,168 rows on OXPT.

IT WAS REPORTED AS PROPERTY-SPECIFIC AND IT IS NOT

Measured across four real MMRs before the fix, the stray selected tab
differed per SOURCE FILE, not per property:

    OXPT          Summary + Prospect Source Summary
    Maple Valley  Summary + Cash Flow
    Canyon        Summary + Work Order Summary
    ERA           Summary only

ERA was clean by luck of how its source workbook was saved, which is why
it looked like the odd one out. Fixing "the Oxford bug" would have fixed
nothing. ERA is kept here as a control: it was already correct and must
stay correct.

WHY SOURCE SHEETS GET A PRINT AREA TOO

A different need, not belt-and-braces. Someone who deliberately selects
Rent Roll to print should get its used range rather than 685 rows
sprawling across however many sheets of paper Excel decides on.

Nothing else is touched. No fit-to-page, no column widths, no
reformatting, no reordering -- these sheets are somebody else's export,
and a print area is the one setting that is additive and reversible.
"""

import unittest

import openpyxl

from tools.mmr_report.builders import scope_workbook_for_print


def workbook(selected=(), rows=None):
    """A stand-in for the prepended-Summary shape, with stray selections."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    rows = rows or {"Summary": 40, "General Ledger": 1168, "Rent Roll": 685,
                    "Cash Flow": 105, "Empty Sheet": 0}
    for title, n in rows.items():
        ws = wb.create_sheet(title)
        for r in range(1, n + 1):
            ws.cell(row=r, column=1, value=r)
            ws.cell(row=r, column=3, value="x")
        ws.sheet_view.tabSelected = title in selected
    return wb


class OnlyTheSummaryIsSelectedTests(unittest.TestCase):
    def selected(self, wb):
        return [ws.title for ws in wb.worksheets if ws.sheet_view.tabSelected]

    def test_a_stray_selected_tab_is_deselected(self):
        """The reported bug, in one assertion."""
        wb = workbook(selected=("Summary", "General Ledger"))
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(self.selected(wb), ["Summary"])

    def test_every_observed_real_world_selection_collapses_to_summary(self):
        """The four shapes measured on real MMRs before the fix."""
        for stray in ("Prospect Source Summary", "Cash Flow",
                      "Work Order Summary", None):
            with self.subTest(stray=stray):
                sel = ("Summary",) + ((stray,) if stray else ())
                rows = {"Summary": 40, "Cash Flow": 105,
                        "Prospect Source Summary": 20,
                        "Work Order Summary": 24}
                wb = workbook(selected=sel, rows=rows)
                scope_workbook_for_print(wb, wb["Summary"])
                self.assertEqual(self.selected(wb), ["Summary"])

    def test_an_already_clean_workbook_stays_clean(self):
        """ERA was correct before the fix. It is the control."""
        wb = workbook(selected=("Summary",))
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(self.selected(wb), ["Summary"])

    def test_summary_is_selected_even_if_nothing_was(self):
        wb = workbook(selected=())
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(self.selected(wb), ["Summary"])

    def test_no_future_source_workbook_can_reintroduce_it(self):
        """Every sheet selected -- the worst case a source file could carry."""
        wb = workbook(selected=tuple(
            t for t in ("Summary", "General Ledger", "Rent Roll", "Cash Flow")))
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(self.selected(wb), ["Summary"])


class SourceSheetsGetAPrintAreaTests(unittest.TestCase):
    def test_a_long_source_sheet_is_scoped_to_its_used_range(self):
        wb = workbook()
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(wb["General Ledger"].print_area,
                         "'General Ledger'!$A$1:$C$1168")

    def test_every_source_sheet_with_content_ends_up_scoped(self):
        """Summary is excluded on purpose -- setup_summary_print owns it,
        and this helper must not fight that."""
        wb = workbook()
        scope_workbook_for_print(wb, wb["Summary"])
        unscoped = [ws.title for ws in wb.worksheets
                    if ws.title != "Summary" and not ws.print_area
                    and ws.max_row > 1]
        self.assertEqual(unscoped, [])

    def test_an_empty_sheet_is_left_alone(self):
        """Pinning A1:A1 would disguise an empty sheet as a one-cell one."""
        wb = workbook()
        scope_workbook_for_print(wb, wb["Summary"])
        # openpyxl reports an unset print area as '' rather than None.
        self.assertFalse(wb["Empty Sheet"].print_area)

    def test_an_existing_print_area_is_not_overwritten(self):
        wb = workbook()
        wb["Rent Roll"].print_area = "A1:B10"
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(wb["Rent Roll"].print_area, "'Rent Roll'!$A$1:$B$10")

    def test_the_summary_keeps_the_print_area_it_was_given(self):
        """setup_summary_print owns the Summary; this must not fight it."""
        wb = workbook()
        wb["Summary"].print_area = "A1:P54"
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(wb["Summary"].print_area, "'Summary'!$A$1:$P$54")


class NothingElseWasTouchedTests(unittest.TestCase):
    """These are somebody else's export sheets."""

    def test_column_widths_are_untouched(self):
        wb = workbook()
        wb["Rent Roll"].column_dimensions["A"].width = 42
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(wb["Rent Roll"].column_dimensions["A"].width, 42)

    def test_sheet_order_is_untouched(self):
        wb = workbook()
        before = wb.sheetnames[:]
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(wb.sheetnames, before)

    def test_source_sheets_are_not_forced_to_fit_to_page(self):
        wb = workbook()
        scope_workbook_for_print(wb, wb["Summary"])
        pr = wb["General Ledger"].sheet_properties.pageSetUpPr
        self.assertFalse(pr and pr.fitToPage)

    def test_cell_values_are_untouched(self):
        wb = workbook()
        before = wb["Cash Flow"]["A105"].value
        scope_workbook_for_print(wb, wb["Summary"])
        self.assertEqual(wb["Cash Flow"]["A105"].value, before)


if __name__ == "__main__":
    unittest.main()
