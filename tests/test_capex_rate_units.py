"""A rate is not a price, and it may never be multiplied by a headcount.

WHAT WENT WRONG

Assessment 11 -- a real walked unit, Michelle's, 23 findings -- exported
a capital budget totalling $5.75. One line: interior repaint, condition
"repair". The researched figure for walls_ceiling is $5.75 PER SQUARE
FOOT. The export multiplied it by a quantity of 1, where 1 was the
instance count the grouping produced ("forty toilets are one line of
quantity 40"), and printed the result as a total.

Then it said "No estimate: 0 item(s), not costed" and "100% of the total
is researched". A number roughly two orders of magnitude low, presented
as complete and fully sourced.

Seven of the thirty-six researched figures are rates, and they are the
expensive ones: flooring, interior repaint, roof covering, facade,
paving, roof drainage. On a full property walk those dominate a real
budget, and every one of them would have come out in single or double
digits.

THE RULE THIS FILE ENFORCES

    per-item figure   total = unit cost x instance count      unchanged
    rate, measured    total = unit cost x measured quantity
    rate, unmeasured  NO TOTAL. The line keeps its rate, states the
                      measurement it needs, and is excluded from every
                      figure in the summary.

None, never 0.0, for the unpriced case. Zero is a claim that the work is
free and it sums into the total silently; None cannot be added up by
accident.

The last test in EachItemsAreUnaffectedTests is the regression guard on
the other direction: this change must not make per-item pricing stop
working, which is most of the table.
"""

import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

from tools import site_dd_capex_export as capex
from tools import site_dd_reference_costs as refcosts


LABELS = {"walls_ceiling": "Walls & ceiling", "flooring": "Flooring",
          "appliance_range": "Range / oven", "smoke_alarm_unit": "Smoke alarm"}


def finding(item_key, cost, source="reference", unit=None, quantity=None,
            area_id=1, room_id=1, ident=1):
    """One priced finding, shaped as apply_reference() leaves it."""
    row = {"id": ident, "item_key": item_key, "area_id": area_id,
           "room_id": room_id, "scope": "room", "condition": "repair",
           "est_unit_cost": cost, "est_cost_source": source,
           "quantity": quantity, "measure": None, "instance_label": ""}
    if unit is not None:
        row["_reference"] = refcosts.ReferenceCost(
            key=item_key, unit_cost=cost, unit=unit, sources=("test",))
    return row


class TheTableKnowsWhichFiguresAreRatesTests(unittest.TestCase):
    def test_sqft_and_lf_are_rates(self):
        self.assertTrue(refcosts.is_rate(refcosts.UNIT_SQFT))
        self.assertTrue(refcosts.is_rate(refcosts.UNIT_LF))

    def test_each_is_not_a_rate(self):
        self.assertFalse(refcosts.is_rate(refcosts.UNIT_EACH))

    def test_an_unknown_unit_is_not_treated_as_a_rate(self):
        self.assertFalse(refcosts.is_rate(None))
        self.assertFalse(refcosts.is_rate("furlong"))

    def test_every_rate_names_what_must_be_measured(self):
        for unit in refcosts.RATE_UNITS:
            with self.subTest(unit=unit):
                self.assertTrue(refcosts.measurement_needed(unit))

    def test_walls_ceiling_is_the_rate_that_caused_this(self):
        ref = refcosts.REFERENCE_COSTS["walls_ceiling"]
        self.assertEqual(ref.unit, refcosts.UNIT_SQFT)
        self.assertEqual(ref.unit_cost, 5.75)


class ARateIsNeverTotalledByInstanceCountTests(unittest.TestCase):
    """The assertion this whole change exists for."""

    def line(self, **kw):
        return capex.build_lines([finding("walls_ceiling", 5.75,
                                          unit=refcosts.UNIT_SQFT, **kw)],
                                 LABELS)[0]

    def test_an_unmeasured_rate_has_no_total(self):
        self.assertIsNone(self.line()["total"])

    def test_it_is_not_zero_either(self):
        """Zero would sum in as though the work were free."""
        self.assertIsNot(self.line()["total"], 0.0)
        self.assertIsNone(self.line()["total"])

    def test_the_total_is_never_the_bare_rate(self):
        """$5.75 as a total is exactly the bug."""
        self.assertNotEqual(self.line()["total"], 5.75)

    def test_the_rate_itself_stays_visible(self):
        line = self.line()
        self.assertEqual(line["unit_cost"], 5.75)
        self.assertEqual(line["unit"], refcosts.UNIT_SQFT)
        self.assertEqual(line["unit_label"], "per sq ft")
        self.assertTrue(line["is_rate"])

    def test_it_says_what_it_needs(self):
        self.assertIn("square feet", self.line()["reason"])

    def test_three_instances_do_not_become_a_quantity_of_three(self):
        """The instance count is not an area, however many there are."""
        rows = [finding("walls_ceiling", 5.75, unit=refcosts.UNIT_SQFT,
                        ident=i) for i in (1, 2, 3)]
        line = capex.build_lines(rows, LABELS)[0]
        self.assertIsNone(line["quantity"])
        self.assertIsNone(line["total"])
        self.assertEqual(line["instances"], 3.0)

    def test_a_measured_rate_is_priced_normally(self):
        line = self.line(quantity=120.0)
        self.assertEqual(line["quantity"], 120.0)
        self.assertAlmostEqual(line["total"], 690.0)
        self.assertFalse(line["reason"])


class EachItemsAreUnaffectedTests(unittest.TestCase):
    """Most of the table is per-item and must behave exactly as before."""

    def test_a_single_each_item_totals_its_unit_cost(self):
        line = capex.build_lines(
            [finding("appliance_range", 1150.0, unit=refcosts.UNIT_EACH)],
            LABELS)[0]
        self.assertEqual(line["quantity"], 1.0)
        self.assertEqual(line["total"], 1150.0)

    def test_forty_of_them_still_multiply(self):
        rows = [finding("smoke_alarm_unit", 260.0, unit=refcosts.UNIT_EACH,
                        ident=i) for i in range(40)]
        line = capex.build_lines(rows, LABELS)[0]
        self.assertEqual(line["quantity"], 40.0)
        self.assertEqual(line["total"], 260.0 * 40)

    def test_a_manual_figure_with_no_reference_is_treated_as_per_item(self):
        """An inspector typing a number is quoting a job, not a rate."""
        line = capex.build_lines(
            [finding("cabinets", 900.0, source="manual")], LABELS)[0]
        self.assertFalse(line["is_rate"])
        self.assertEqual(line["total"], 900.0)


class TheSummaryTellsTheTruthTests(unittest.TestCase):
    """A confidently wrong number is worse than an honestly incomplete one."""

    def summary(self):
        rows = [finding("walls_ceiling", 5.75, unit=refcosts.UNIT_SQFT,
                        room_id=1),
                finding("appliance_range", 1150.0, unit=refcosts.UNIT_EACH,
                        room_id=2)]
        return capex.summarize(capex.build_lines(rows, LABELS))

    def test_the_total_excludes_the_unmeasured_rate(self):
        self.assertEqual(self.summary()["total"], 1150.0)

    def test_the_unmeasured_rate_is_counted_as_unpriced(self):
        s = self.summary()
        self.assertEqual(s["unpriced_count"], 1)
        self.assertEqual(s["priced_count"], 1)
        self.assertEqual(s["line_count"], 2)

    def test_unmeasured_rates_are_reported_separately(self):
        """Needing a tape measure is a different problem from needing
        research that may not exist."""
        s = self.summary()
        self.assertEqual(s["unmeasured_count"], 1)
        self.assertEqual(s["unmeasured"][0]["item_key"], "walls_ceiling")

    def test_an_item_with_no_figure_at_all_is_unpriced_but_not_unmeasured(self):
        rows = [finding("foundation", None, source="none")]
        s = capex.summarize(capex.build_lines(rows, LABELS))
        self.assertEqual(s["unpriced_count"], 1)
        self.assertEqual(s["unmeasured_count"], 0)

    def test_zero_priced_lines_reports_no_total_at_all(self):
        """$0.00 would read as "this costs nothing"."""
        rows = [finding("walls_ceiling", 5.75, unit=refcosts.UNIT_SQFT)]
        s = capex.summarize(capex.build_lines(rows, LABELS))
        self.assertIsNone(s["total"])
        self.assertTrue(s["total_is_partial"])

    def test_and_does_not_claim_zero_percent_researched(self):
        """We HAVE the research -- $5.75/sqft from the real table. What is
        missing is a measurement. Reporting 0% researched would say we hold
        no cost data, which is the opposite of true."""
        rows = [finding("walls_ceiling", 5.75, unit=refcosts.UNIT_SQFT)]
        s = capex.summarize(capex.build_lines(rows, LABELS))
        self.assertIsNone(s["researched_pct"])
        self.assertEqual(s["unmeasured_count"], 1)
        self.assertEqual(s["unresearched_count"], 0)

    def test_an_empty_budget_really_is_zero(self):
        """Nothing recorded as needing work is a finding, not a gap."""
        s = capex.summarize([])
        self.assertEqual(s["total"], 0.0)
        self.assertFalse(s["total_is_partial"])


class TheCoverageSentenceNamesAllThreeBucketsTests(unittest.TestCase):
    """The summary must not do, one level up, what the line refuses to do.

    A line that declines to state a total, sitting under a summary that
    says "$0.00" and "0% researched", has simply moved the false claim
    from the row to the header. $0.00 reads as free; 0% researched reads
    as "we hold no cost data" when in fact we hold $5.75/sqft and lack
    only a tape measure.
    """

    def sentence(self, priced, total, unmeasured, unresearched):
        return capex.coverage_sentence(priced, total, unmeasured, unresearched)

    def test_nothing_priced_says_there_is_no_total(self):
        s = self.sentence(0, 1, 1, 0)
        self.assertIn("NO total", s)
        self.assertIn("researched rate", s)

    def test_it_distinguishes_unmeasured_from_unresearched(self):
        s = self.sentence(1, 3, 1, 1)
        self.assertIn("researched rate but nothing measured", s)
        self.assertIn("no researched figure at all", s)

    def test_a_fully_priced_budget_says_so_plainly(self):
        s = self.sentence(4, 4, 0, 0)
        self.assertIn("whole recorded budget", s)
        self.assertNotIn("NOT", s)

    def test_a_partial_budget_says_it_is_not_the_full_one(self):
        self.assertIn("NOT", self.sentence(2, 3, 1, 0))

    def test_an_empty_budget_is_described_as_empty(self):
        self.assertIn("No items", self.sentence(0, 0, 0, 0))

    def test_the_summary_carries_the_sentence(self):
        rows = [finding("walls_ceiling", 5.75, unit=refcosts.UNIT_SQFT)]
        s = capex.summarize(capex.build_lines(rows, LABELS))
        self.assertTrue(s["coverage_sentence"])
        self.assertIn("NO total", s["coverage_sentence"])


class BothExportsCarryTheUnitTests(unittest.TestCase):
    """The PDF shares build_lines() with the XLSX, so both were wrong."""

    def rows(self):
        return capex.build_lines(
            [finding("walls_ceiling", 5.75, unit=refcosts.UNIT_SQFT),
             finding("appliance_range", 1150.0, unit=refcosts.UNIT_EACH,
                     room_id=2)], LABELS)

    def test_the_xlsx_has_a_unit_column(self):
        import tempfile
        from pathlib import Path
        from openpyxl import load_workbook
        lines = self.rows()
        out = Path(tempfile.mkdtemp()) / "b.xlsx"
        capex.build_xlsx(out, {"property_label": "Nabob Hill",
                               "assessed_on": "2026-08-16"},
                         lines, capex.summarize(lines), LABELS)
        wb = load_workbook(str(out))
        ws = wb["Capital Budget"]
        headers = [c.value for c in next(
            r for r in ws.iter_rows() if r[0].value == "Item")]
        self.assertIn("Unit", headers)
        self.assertIn("Needs measurement", wb.sheetnames)

    def test_the_pdf_renders_without_a_total(self):
        """None must not blow up the renderer that used to get 0.0."""
        import tempfile
        from pathlib import Path
        lines = self.rows()
        out = Path(tempfile.mkdtemp()) / "b.pdf"
        capex.build_pdf(out, {"property_label": "Nabob Hill",
                              "assessed_on": "2026-08-16"},
                        lines, capex.summarize(lines))
        self.assertTrue(out.exists())
        self.assertEqual(out.read_bytes()[:4], b"%PDF")


if __name__ == "__main__":
    unittest.main()


class TheManualPathIsClosedTooTests(unittest.TestCase):
    """The same bug, arriving through a different door.

    build_lines() used to read the unit off the _reference object, which
    is attached only when a RESEARCHED cost was applied. An inspector
    typing their own $5.75 on walls_ceiling produced no _reference, fell
    through to "each", and got the rate multiplied by an instance count.

    The unit belongs to the item. walls_ceiling is per-square-foot
    whoever prices it.

    FREEFORM ITEMS ARE THE NARROWER CASE, AND THEY ARE NOT ASSUMED

    An item with no reference entry has no unit to inherit. Treating it
    as "each" because nothing describes it is the same assumption one
    step further out, so a figure below FREEFORM_RATE_CEILING is read as
    a rate and refused a total.

    The threshold is grounded, not invented: every researched rate is at
    most $11.50 and the cheapest researched per-item figure is $195.
    Its failure mode is a genuine sub-$15 job price being refused, which
    is asserted below so the trade is on the record rather than implied.
    """

    def line(self, item, cost, source="manual", **kw):
        row = {"id": 1, "item_key": item, "area_id": 1, "room_id": 1,
               "scope": "room", "condition": "repair", "est_unit_cost": cost,
               "est_cost_source": source, "quantity": None, "measure": None,
               "instance_label": ""}
        row.update(kw)
        return capex.build_lines([row], LABELS)[0]

    def test_a_manual_figure_on_a_rate_item_is_a_rate(self):
        l = self.line("walls_ceiling", 5.75)
        self.assertTrue(l["is_rate"])
        self.assertEqual(l["unit"], refcosts.UNIT_SQFT)
        self.assertIsNone(l["total"])

    def test_it_is_never_the_bare_rate_as_a_total(self):
        self.assertNotEqual(self.line("walls_ceiling", 5.75)["total"], 5.75)

    def test_all_seven_rate_items_are_closed_on_the_manual_path(self):
        for key, entry in refcosts.REFERENCE_COSTS.items():
            if not refcosts.is_rate(entry.unit):
                continue
            with self.subTest(item=key):
                l = self.line(key, entry.unit_cost)
                self.assertTrue(l["is_rate"])
                self.assertIsNone(l["total"])

    def test_each_items_are_completely_unaffected(self):
        for key, entry in refcosts.REFERENCE_COSTS.items():
            if refcosts.is_rate(entry.unit):
                continue
            with self.subTest(item=key):
                l = self.line(key, entry.unit_cost)
                self.assertFalse(l["is_rate"])
                self.assertEqual(l["total"], entry.unit_cost)

    def test_a_freeform_job_price_still_totals(self):
        """The common, legitimate case must not regress."""
        self.assertEqual(self.line("replace_gate", 800.0)["total"], 800.0)

    def test_a_freeform_rate_is_refused_a_total(self):
        l = self.line("replace_gate", 5.75)
        self.assertTrue(l["is_rate"])
        self.assertIsNone(l["total"])

    def test_the_threshold_sits_above_every_researched_rate(self):
        rates = [e.unit_cost for e in refcosts.REFERENCE_COSTS.values()
                 if refcosts.is_rate(e.unit)]
        self.assertLess(max(rates), refcosts.FREEFORM_RATE_CEILING)

    def test_and_below_every_researched_per_item_figure(self):
        each = [e.unit_cost for e in refcosts.REFERENCE_COSTS.values()
                if not refcosts.is_rate(e.unit)]
        self.assertGreater(min(each), refcosts.FREEFORM_RATE_CEILING)

    def test_the_threshold_is_labelled_provisional(self):
        """It is a reasoned guess about an uncurated category.

        The seventeenfold gap is evidence about the 36 curated reference
        entries, and a freeform item is by definition not one of them. The
        number is applied to exactly the category it was not measured on,
        so the source must say so or somebody later reads $15 as derived.
        """
        src = (ROOT / "tools" / "site_dd_reference_costs.py").read_text(
            encoding="utf-8")
        self.assertIn("PROVISIONAL", src)
        self.assertIn("not a derived constant", src.lower().replace(
            "NOT A DERIVED CONSTANT", "not a derived constant"))

    def test_the_known_false_positive_is_on_the_record(self):
        """A real sub-$15 job price is refused. Stated, not hidden."""
        l = self.line("replace_gate", 12.00)
        self.assertIsNone(l["total"],
                          "a genuine $12 job price is refused; the trade is "
                          "that no capital job costs twelve dollars")
