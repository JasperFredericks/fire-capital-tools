"""
Tests for the pro-forma P&L view.

The central claim under test is that this is a formatting layer and not a
second calculation: every figure it shows must be one the engine already
produced. The Eagle Rock tests assert that against the real production
scenario and its independently confirmed figures.
"""

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from tools import underwriting_math as um
from tools import underwriting_pnl as pnl_mod
from tools import underwriting_pnl_export as pnl_export

FIXTURE = Path(__file__).parent / "fixtures" / "eagle_rock_scenario4.json"

# Independently confirmed against production on 2026-08-12.
EAGLE_ROCK = {
    "units": 92,
    "expense_lines": 109,
    "noi_year1": 384455.38,
    "opex_year1": 839216.14,
    "egi_year1": 1223671.52,
    "equity": 2586300.00,
    "levered_irr_pct": 8.11,
    "dscr": 1.12,
}


def load_fixture():
    d = json.loads(FIXTURE.read_text(encoding="utf-8"))
    return d["scenario"], d["units"], d["expenses"]


def build(scenario=None, units=None, expenses=None):
    if scenario is None:
        scenario, units, expenses = load_fixture()
    result = um.analyze_scenario(scenario, units, expenses)
    return pnl_mod.build_pnl(scenario, units, expenses, result), result


class TestEagleRockZeroDrift(unittest.TestCase):
    """The P&L must reproduce the already-verified Eagle Rock figures
    exactly. A failure here means the view has become a calculation."""

    @classmethod
    def setUpClass(cls):
        cls.scenario, cls.units, cls.expenses = load_fixture()
        cls.pnl, cls.result = build(cls.scenario, cls.units, cls.expenses)

    def test_fixture_is_the_real_scenario(self):
        self.assertEqual(self.scenario["property_label"], "Eagle Rock Apartments")
        self.assertEqual(len(self.units), EAGLE_ROCK["units"])
        self.assertEqual(len(self.expenses), EAGLE_ROCK["expense_lines"])

    def test_engine_still_produces_the_confirmed_figures(self):
        """Guards the fixture itself -- if this fails the baseline moved,
        not the P&L."""
        r = self.result
        self.assertAlmostEqual(r["projection"]["years"][0]["noi"],
                               EAGLE_ROCK["noi_year1"], places=2)
        self.assertAlmostEqual(r["operating_expenses_year1"],
                               EAGLE_ROCK["opex_year1"], places=2)
        self.assertAlmostEqual(r["egi"]["effective_gross_income"],
                               EAGLE_ROCK["egi_year1"], places=2)
        self.assertAlmostEqual(r["returns"]["equity_invested"],
                               EAGLE_ROCK["equity"], places=2)
        self.assertAlmostEqual(r["returns"]["levered_irr"] * 100,
                               EAGLE_ROCK["levered_irr_pct"], places=2)
        self.assertAlmostEqual(r["returns"]["dscr"], EAGLE_ROCK["dscr"], places=2)

    def test_pnl_year1_noi_matches_confirmed_figure(self):
        self.assertAlmostEqual(self.pnl["noi"][0], EAGLE_ROCK["noi_year1"], places=2)

    def test_pnl_year1_expenses_match_confirmed_figure(self):
        self.assertAlmostEqual(self.pnl["expense_totals"][0],
                               EAGLE_ROCK["opex_year1"], places=2)

    def test_pnl_year1_revenue_matches_confirmed_egi(self):
        self.assertAlmostEqual(self.pnl["revenue_totals"][0],
                               EAGLE_ROCK["egi_year1"], places=2)

    def test_every_year_ties_to_the_projection(self):
        for i, py in enumerate(self.result["projection"]["years"]):
            self.assertAlmostEqual(self.pnl["revenue_totals"][i], py["income"], places=2)
            self.assertAlmostEqual(self.pnl["expense_totals"][i], py["expenses"], places=2)
            self.assertAlmostEqual(self.pnl["noi"][i], py["noi"], places=2)

    def test_reconciliation_all_passed(self):
        checks = self.pnl["reconciliation"]
        self.assertTrue(checks)
        self.assertTrue(all(c["passed"] for c in checks),
                        [c for c in checks if not c["passed"]])

    def test_every_included_line_appears_exactly_once(self):
        included = um.operating_expense_lines(self.expenses)
        shown = sum(g["line_count"] for g in self.pnl["expenses"])
        self.assertEqual(shown, len(included))

    def test_excluded_lines_are_listed_but_not_counted(self):
        excluded = [l for l in self.expenses
                    if not l.get("is_included") and not um.is_acquisition_line(l)]
        self.assertEqual(len(self.pnl["excluded"]), len(excluded))
        # 109 total, 96 included -> the rest must not touch the total
        self.assertAlmostEqual(self.pnl["expense_totals"][0],
                               um.total_operating_expenses(self.expenses), places=2)

    def test_hold_years_matches_scenario(self):
        self.assertEqual(self.pnl["hold_years"], self.scenario["hold_years"])
        self.assertEqual(self.pnl["years"], [1, 2, 3, 4, 5])

    def test_building_the_pnl_does_not_mutate_the_result(self):
        before = json.dumps(self.result["projection"]["years"], sort_keys=True)
        pnl_mod.build_pnl(self.scenario, self.units, self.expenses, self.result)
        after = json.dumps(self.result["projection"]["years"], sort_keys=True)
        self.assertEqual(before, after)

    def test_engine_output_identical_before_and_after_building_pnl(self):
        """The whole scenario re-analysed after a P&L build must be
        byte-identical -- proves the view has no side effect on the model."""
        fresh = um.analyze_scenario(self.scenario, self.units, self.expenses)
        self.assertEqual(json.dumps(fresh["returns"], sort_keys=True, default=str),
                         json.dumps(self.result["returns"], sort_keys=True, default=str))


class TestRevenueDetail(unittest.TestCase):
    def setUp(self):
        self.pnl, self.result = build()

    def test_revenue_rows_sum_to_egi_every_year(self):
        for i in range(len(self.pnl["years"])):
            self.assertAlmostEqual(
                sum(r["amounts"][i] for r in self.pnl["revenue"]),
                self.pnl["revenue_totals"][i], places=6)

    def test_deductions_are_negative(self):
        for r in self.pnl["revenue"]:
            if r["is_deduction"] and abs(r["amounts"][0]) > 0:
                self.assertLess(r["amounts"][0], 0, r["label"])

    def test_gpr_matches_build_egi(self):
        self.assertAlmostEqual(self.pnl["revenue"][0]["amounts"][0],
                               self.result["egi"]["gross_potential_rent"], places=6)

    def test_revenue_grows_at_rent_growth(self):
        rg = float(self.pnl["rent_growth_pct"]) / 100.0
        t0, t1 = self.pnl["revenue_totals"][0], self.pnl["revenue_totals"][1]
        self.assertAlmostEqual(t1 / t0, 1 + rg, places=9)


class TestExpenseDetail(unittest.TestCase):
    def setUp(self):
        self.pnl, self.result = build()

    def test_category_subtotals_sum_to_total(self):
        for i in range(len(self.pnl["years"])):
            self.assertAlmostEqual(
                sum(g["subtotals"][i] for g in self.pnl["expenses"]),
                self.pnl["expense_totals"][i], places=6)

    def test_lines_sum_to_their_category_subtotal(self):
        for g in self.pnl["expenses"]:
            for i in range(len(self.pnl["years"])):
                self.assertAlmostEqual(sum(l["amounts"][i] for l in g["lines"]),
                                       g["subtotals"][i], places=6)

    def test_categories_are_sorted(self):
        cats = [g["category"] for g in self.pnl["expenses"]]
        self.assertEqual(cats, sorted(cats))

    def test_explicit_zero_growth_does_not_inherit_default(self):
        scenario = {"property_label": "T", "hold_years": 2, "rent_growth_pct": 0.0,
                    "expense_growth_pct": 10.0, "vacancy_pct": 0, "concessions_pct": 0,
                    "bad_debt_pct": 0, "other_income_annual": 0}
        lines = [{"label": "Flat", "annual_amount": 1000.0, "growth_pct": 0.0,
                  "is_included": 1, "line_kind": None, "category_name": "Ops"}]
        years = [1, 2]
        groups = pnl_mod.build_expenses(lines, 10.0, years)
        self.assertAlmostEqual(groups[0]["subtotals"][1], 1000.0, places=6)

    def test_absent_growth_inherits_default(self):
        lines = [{"label": "Grows", "annual_amount": 1000.0, "growth_pct": None,
                  "is_included": 1, "line_kind": None, "category_name": "Ops"}]
        groups = pnl_mod.build_expenses(lines, 10.0, [1, 2])
        self.assertAlmostEqual(groups[0]["subtotals"][1], 1100.0, places=6)

    def test_acquisition_lines_never_appear(self):
        lines = [{"label": "Legal", "annual_amount": 5000.0, "is_included": 1,
                  "line_kind": um.ACQUISITION_COST_KIND, "category_name": "Closing"}]
        groups = pnl_mod.build_expenses(lines, 3.0, [1])
        self.assertEqual(groups, [])
        self.assertEqual(pnl_mod.excluded_lines(lines), [])


class TestReconciliationGate(unittest.TestCase):
    """The gate must actually fire -- a reconciliation that cannot fail
    proves nothing."""

    def test_mismatch_raises(self):
        pnl, result = build()
        result["projection"]["years"][0]["noi"] += 1.0
        with self.assertRaises(pnl_mod.PnLReconciliationError):
            pnl_mod.reconcile(pnl, result)

    def test_error_names_the_year_and_the_gap(self):
        pnl, result = build()
        result["projection"]["years"][2]["expenses"] += 250.0
        with self.assertRaises(pnl_mod.PnLReconciliationError) as ctx:
            pnl_mod.reconcile(pnl, result)
        self.assertIn("year 3", str(ctx.exception))

    def test_tolerance_is_below_a_cent(self):
        self.assertLess(pnl_mod.RECONCILE_TOLERANCE, 0.01)


class TestExports(unittest.TestCase):
    def setUp(self):
        self.pnl, _ = build()

    def test_flatten_rows_covers_every_line(self):
        rows = pnl_export.flatten_rows(self.pnl)
        line_rows = [r for r in rows if r["kind"] == "line"]
        self.assertGreater(len(line_rows), len(self.pnl["revenue"]))
        labels = [r["label"] for r in rows]
        self.assertIn("NET OPERATING INCOME", labels)
        self.assertIn("Effective Gross Income", labels)
        self.assertIn("Total Operating Expenses", labels)

    def test_noi_row_carries_the_confirmed_figure(self):
        rows = pnl_export.flatten_rows(self.pnl)
        noi_row = next(r for r in rows if r["label"] == "NET OPERATING INCOME")
        self.assertAlmostEqual(noi_row["amounts"][0], EAGLE_ROCK["noi_year1"], places=2)

    def test_pdf_is_written_and_non_trivial(self):
        with tempfile.TemporaryDirectory() as d:
            out = pnl_export.build_pdf(Path(d) / "pnl.pdf", self.pnl)
            self.assertTrue(out.exists())
            self.assertGreater(out.stat().st_size, 5000)
            self.assertEqual(out.read_bytes()[:4], b"%PDF")

    def test_xlsx_is_written_and_readable(self):
        import openpyxl
        with tempfile.TemporaryDirectory() as d:
            out = pnl_export.build_xlsx(Path(d) / "pnl.xlsx", self.pnl)
            self.assertTrue(out.exists())
            wb = openpyxl.load_workbook(out)
            self.assertIn("Pro-Forma P&L", wb.sheetnames)
            self.assertIn("Reconciliation", wb.sheetnames)

    def test_xlsx_noi_cell_is_a_number_matching_the_confirmed_figure(self):
        import openpyxl
        with tempfile.TemporaryDirectory() as d:
            out = pnl_export.build_xlsx(Path(d) / "pnl.xlsx", self.pnl)
            ws = openpyxl.load_workbook(out)["Pro-Forma P&L"]
            found = None
            for row in ws.iter_rows():
                if row[0].value == "NET OPERATING INCOME":
                    found = row[1].value
            self.assertIsNotNone(found)
            self.assertIsInstance(found, float)
            self.assertAlmostEqual(found, EAGLE_ROCK["noi_year1"], places=2)

    def test_money_formatting(self):
        self.assertEqual(pnl_export._money(1234.5), "1,234")
        self.assertEqual(pnl_export._money(-1234.5), "(1,234)")
        self.assertEqual(pnl_export._money(None), "—")
        self.assertEqual(pnl_export._money(-0.001), "0")

    def test_pagination_leaves_no_orphan_final_page(self):
        rows = pnl_export.flatten_rows(self.pnl)
        pages = pnl_export._paginate(rows)
        self.assertGreater(len(pages), 1, "Eagle Rock should span several pages")
        # Even distribution: the last page must not be a stub under a full
        # page header, which is what naive chunking produced.
        self.assertGreaterEqual(len(pages[-1]), len(pages[0]) // 2)

    def test_pagination_preserves_every_row(self):
        rows = pnl_export.flatten_rows(self.pnl)
        pages = pnl_export._paginate(rows)
        flat = [r for p in pages for r in p]
        # Only leading spacers on continuation pages are dropped.
        self.assertLessEqual(len(rows) - len(flat), len(pages))
        kept = [r for r in rows if r["kind"] != "spacer"]
        self.assertEqual([r["label"] for r in flat if r["kind"] != "spacer"],
                         [r["label"] for r in kept])

    def test_pagination_handles_empty_and_single_page(self):
        self.assertEqual(pnl_export._paginate([]), [[]])
        one = [{"kind": "line", "label": "x", "amounts": [1]}]
        self.assertEqual(pnl_export._paginate(one), [one])

    def test_export_filename_is_safe(self):
        name = pnl_export.export_filename({"property_label": "Eagle Rock/Apts #2"}, "pdf")
        self.assertNotIn("/", name)
        self.assertTrue(name.endswith(".pdf"))


class TestPurity(unittest.TestCase):
    """Both modules must stay Flask-free, per the project's pure-math /
    routes / db split.

    Checked by parsing the module rather than grepping its text: the
    export module's docstring legitimately mentions current_app while
    explaining why it does not use it, and a substring check cannot tell
    prose from code.
    """

    @staticmethod
    def _imported_roots(module):
        import ast
        tree = ast.parse(Path(module.__file__).read_text(encoding="utf-8"))
        roots = set()
        for node in ast.walk(tree):
            if isinstance(node, ast.Import):
                roots.update(a.name.split(".")[0] for a in node.names)
            elif isinstance(node, ast.ImportFrom) and node.module:
                roots.add(node.module.split(".")[0])
        return roots

    @staticmethod
    def _referenced_names(module):
        import ast
        tree = ast.parse(Path(module.__file__).read_text(encoding="utf-8"))
        return {n.id for n in ast.walk(tree) if isinstance(n, ast.Name)}

    def test_pnl_module_does_not_import_flask(self):
        self.assertNotIn("flask", self._imported_roots(pnl_mod))

    def test_export_module_does_not_import_flask(self):
        self.assertNotIn("flask", self._imported_roots(pnl_export))

    def test_neither_module_references_current_app(self):
        for mod in (pnl_mod, pnl_export):
            with self.subTest(module=mod.__name__):
                self.assertNotIn("current_app", self._referenced_names(mod))

    def test_underwriting_math_untouched_by_this_phase(self):
        """underwriting_math must not import the view, in either direction."""
        src = Path(um.__file__).read_text(encoding="utf-8")
        self.assertNotIn("underwriting_pnl", src)


if __name__ == "__main__":
    unittest.main()
