#!/usr/bin/env python3
"""Add ACS median home/condo value growth metrics to Clean Cities 100k+ sheet.

Source: ACS 1-year API
Variable: B25077_001E (Median value in dollars for owner-occupied housing units)
"""

import datetime as dt
import json
import re
import shutil
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from fire_metrics_updater.config import get_secret


BASE_DIR = Path(__file__).resolve().parent
INPUT_WORKBOOK = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_CRIME_INDEX.xlsx"
OUTPUT_WORKBOOK = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_HOME_VALUE_GROWTH.xlsx"
CENSUS_API_KEY = get_secret("CENSUS_API_KEY", "data/cache/census_api_key.txt")

ACS_VAR = "B25077_001E"
ACS_MISSING = {"", "null", "None", "-666666666", "-777777777", "-888888888", "-999999999"}

STATE_TO_ABBR = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR", "CALIFORNIA": "CA",
    "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE", "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID", "ILLINOIS": "IL",
    "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA",
    "MAINE": "ME", "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
    "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK", "OREGON": "OR",
    "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA",
    "WASHINGTON": "WA", "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
}

ABBR_TO_STATE = {abbr: name for name, abbr in STATE_TO_ABBR.items()}

STATE_ABBR_TO_FIPS = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06", "CO": "08", "CT": "09", "DE": "10",
    "DC": "11", "FL": "12", "GA": "13", "HI": "15", "ID": "16", "IL": "17", "IN": "18", "IA": "19",
    "KS": "20", "KY": "21", "LA": "22", "ME": "23", "MD": "24", "MA": "25", "MI": "26", "MN": "27",
    "MS": "28", "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33", "NJ": "34", "NM": "35",
    "NY": "36", "NC": "37", "ND": "38", "OH": "39", "OK": "40", "OR": "41", "PA": "42", "RI": "44",
    "SC": "45", "SD": "46", "TN": "47", "TX": "48", "UT": "49", "VT": "50", "VA": "51", "WA": "53",
    "WV": "54", "WI": "55", "WY": "56",
}

PUNCT = re.compile(r"[^\w\s]", re.U)
MULTI = re.compile(r"\s+")
TRAILING_DIGITS = re.compile(r"\d+$")
SUFFIX = re.compile(
    r"\b(city|town|village|cdp|municipality|metropolitan government)\b$",
    re.I,
)


def normalize_city(text):
    value = "" if text is None else str(text).strip().lower()
    value = PUNCT.sub(" ", value)
    value = TRAILING_DIGITS.sub("", value).strip()
    value = SUFFIX.sub("", value).strip()
    value = MULTI.sub(" ", value)
    return value


def normalize_state_abbr(text):
    value = "" if text is None else str(text).strip().upper()
    value = TRAILING_DIGITS.sub("", value).strip()
    if len(value) == 2 and value.isalpha():
        return value
    return STATE_TO_ABBR.get(value, "")


def load_api_key():
    if CENSUS_API_KEY:
        return CENSUS_API_KEY
    raise RuntimeError(
        "Census API key not found. Set CENSUS_API_KEY or create data/cache/census_api_key.txt"
    )


def api_get_json(year, params, api_key):
    payload = dict(params)
    payload["key"] = api_key
    query = urllib.parse.urlencode(payload)
    url = f"https://api.census.gov/data/{year}/acs/acs1?{query}"
    with urllib.request.urlopen(url, timeout=90) as resp:
        body = resp.read().decode("utf-8")
    if not body.strip().startswith("["):
        raise RuntimeError(f"Census API returned non-JSON for year {year}: {body[:200]}")
    return json.loads(body)


def discover_latest_year(api_key):
    current = dt.date.today().year
    for year in range(current - 1, 2021, -1):
        try:
            rows = api_get_json(year, {"get": f"NAME,{ACS_VAR}", "for": "place:*", "in": "state:01"}, api_key)
            if len(rows) > 1:
                return year
        except Exception:
            continue
    raise RuntimeError("Could not detect latest ACS 1-year year for place-level B25077_001E")


def parse_missing_to_int(value):
    if value is None:
        return None
    s = str(value).strip()
    if s in ACS_MISSING:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def fetch_place_values(year, api_key):
    rows = api_get_json(year, {"get": f"NAME,{ACS_VAR}", "for": "place:*", "in": "state:*"}, api_key)
    head = rows[0]
    idx = {k: i for i, k in enumerate(head)}

    by_city_state = {}
    by_fips = {}
    for row in rows[1:]:
        val = parse_missing_to_int(row[idx[ACS_VAR]])
        if val is None:
            continue

        name = row[idx["NAME"]]
        state_fips = row[idx["state"]]
        place_fips = row[idx["place"]]

        # NAME format: "Place name, State Name"
        if "," in name:
            place_name, state_name = [x.strip() for x in name.rsplit(",", 1)]
            abbr = STATE_TO_ABBR.get(state_name.upper(), "")
        else:
            place_name = name.strip()
            abbr = ""

        if abbr:
            key = (abbr, normalize_city(place_name))
            by_city_state.setdefault(key, val)

        geoid = f"{state_fips}{place_fips}"
        by_fips[geoid] = val

    return {"city_state": by_city_state, "fips": by_fips}


def find_header_index(ws, names):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    norm = [str(h).strip().lower() if h is not None else "" for h in headers]
    for n in names:
        if n in norm:
            return norm.index(n) + 1
    return None


def remove_existing_home_value_columns(ws):
    patterns = (
        "median home/condo value in ",
        "median home/condo value growth ",
    )
    to_delete = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        text = str(h).strip().lower() if h is not None else ""
        if any(text.startswith(p) for p in patterns):
            to_delete.append(c)
    for c in reversed(to_delete):
        ws.delete_cols(c, 1)


def append_home_value_columns(ws, years, values_by_year):
    y2021, yprior, ylatest = years
    city_col = find_header_index(ws, ["city"])
    state_col = find_header_index(ws, ["state_abbr", "state"])
    state_name_col = find_header_index(ws, ["state"])
    geoid_col = find_header_index(ws, ["geoid", "place_geoid", "census_place_geoid"])

    if city_col is None or state_col is None:
        raise RuntimeError("Could not find city/state columns in Clean Cities 100k+ sheet")

    remove_existing_home_value_columns(ws)

    start = ws.max_column + 1
    headers = [
        f"Median home/condo value in {y2021}",
        f"Median home/condo value in {yprior}",
        f"Median home/condo value in {ylatest}",
        f"Median home/condo value growth {y2021}\u2013{ylatest} (%)",
        f"Median home/condo value growth {yprior}\u2013{ylatest} (%)",
    ]
    for i, h in enumerate(headers):
        ws.cell(1, start + i, h)

    matched = 0
    unmatched = []
    sanity = []

    for r in range(2, ws.max_row + 1):
        city = ws.cell(r, city_col).value
        raw_state = ws.cell(r, state_col).value
        state_name = ws.cell(r, state_name_col).value if state_name_col else raw_state
        geoid = ws.cell(r, geoid_col).value if geoid_col else None

        abbr = normalize_state_abbr(raw_state)
        if not abbr:
            abbr = normalize_state_abbr(state_name)
        city_norm = normalize_city(city)

        key = (abbr, city_norm)
        geoid_key = str(geoid).strip() if geoid is not None else ""

        def get_val(year):
            if geoid_key:
                v = values_by_year[year]["fips"].get(geoid_key)
                if v is not None:
                    return v
            return values_by_year[year]["city_state"].get(key)

        v2021 = get_val(y2021)
        vprior = get_val(yprior)
        vlast = get_val(ylatest)

        c1 = ws.cell(r, start)
        c2 = ws.cell(r, start + 1)
        c3 = ws.cell(r, start + 2)
        g1 = ws.cell(r, start + 3)
        g2 = ws.cell(r, start + 4)

        c1.value = v2021
        c2.value = vprior
        c3.value = vlast

        if v2021 and vlast and v2021 != 0:
            g1.value = (vlast - v2021) / v2021
        else:
            g1.value = None

        if vprior and vlast and vprior != 0:
            g2.value = (vlast - vprior) / vprior
        else:
            g2.value = None

        c1.number_format = "$#,##0"
        c2.number_format = "$#,##0"
        c3.number_format = "$#,##0"
        g1.number_format = "0.00%"
        g2.number_format = "0.00%"

        if vlast is not None:
            matched += 1
            if len(sanity) < 5:
                sanity.append((city, abbr, v2021, vprior, vlast, g1.value, g2.value))
        else:
            unmatched.append((city, abbr))

    return matched, unmatched, sanity


def append_readme_note(wb):
    if "README" not in wb.sheetnames:
        return
    ws = wb["README"]
    note = (
        "Median home/condo value comes from ACS variable B25077_001E, Median Value (Dollars), "
        "universe: owner-occupied housing units. This is an ACS survey estimate, not a Zillow market value."
    )

    col1_values = [str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1)]
    if any(note in v for v in col1_values):
        return
    ws.append([""])
    ws.append([note])


def main():
    if not INPUT_WORKBOOK.exists():
        raise FileNotFoundError(f"Input workbook not found: {INPUT_WORKBOOK}")

    api_key = load_api_key()
    latest = discover_latest_year(api_key)
    prior = latest - 1
    years = (2021, prior, latest)

    values_by_year = {
        2021: fetch_place_values(2021, api_key),
        prior: fetch_place_values(prior, api_key),
        latest: fetch_place_values(latest, api_key),
    }

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = INPUT_WORKBOOK.with_name(f"{INPUT_WORKBOOK.stem}_backup_{stamp}{INPUT_WORKBOOK.suffix}")
    shutil.copy2(INPUT_WORKBOOK, backup_path)

    shutil.copy2(INPUT_WORKBOOK, OUTPUT_WORKBOOK)
    wb = load_workbook(OUTPUT_WORKBOOK)
    if "Clean Cities 100k+" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Clean Cities 100k+' not found")

    ws = wb["Clean Cities 100k+"]
    matched, unmatched, sanity = append_home_value_columns(ws, years, values_by_year)
    append_readme_note(wb)
    wb.save(OUTPUT_WORKBOOK)

    total_rows = ws.max_row - 1

    print(f"Output file path: {OUTPUT_WORKBOOK}")
    print(f"ACS years used: {years}")
    print(f"Number of city rows: {total_rows}")
    print(f"Number matched to ACS home/condo value data: {matched}")
    print(f"Number failed to match: {len(unmatched)}")
    print("First 10 unmatched cities:")
    for city, st in unmatched[:10]:
        print(f"- {city}, {st}")
    print("5 sanity-check rows:")
    y2021, yprior, ylatest = years
    for city, st, v1, vp, vl, g1, g2 in sanity:
        print(
            f"- {city}, {st} | {y2021}: {v1} | {yprior}: {vp} | {ylatest}: {vl} "
            f"| growth {y2021}\u2013{ylatest}: {g1} | growth {yprior}\u2013{ylatest}: {g2}"
        )


if __name__ == "__main__":
    main()
