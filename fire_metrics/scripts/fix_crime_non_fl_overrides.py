#!/usr/bin/env python3
"""Targeted non-Florida crime blank-row fix using existing FBI CIUS Table 8 source.

Scope:
- Only updates listed non-Florida target cities.
- Leaves Florida rows untouched.
- Preserves workbook structure and formatting as much as possible.
"""

import datetime
import re
import shutil
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from add_crime_index import get_fbi_crime_workbook_path, normalize_table_8_header

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_WB = BASE_DIR / "output" / "us_cities_100k_population_ranked_FORMATTED_FINAL_BEFORE_CRIME_FIX.xlsx"
OUTPUT_WB = BASE_DIR / "output" / "us_cities_100k_population_ranked_CRIME_NON_FL_FIXED.xlsx"
BACKUP_WB = BASE_DIR / "output" / "us_cities_100k_population_ranked_FORMATTED_FINAL_BEFORE_CRIME_FIX.backup_before_non_fl_crime_fix.xlsx"

STATE_TO_ABBR = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR", "CALIFORNIA": "CA",
    "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE", "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID", "ILLINOIS": "IL",
    "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA",
    "MAINE": "ME", "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
    "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK", "OREGON": "OR",
    "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA",
    "WASHINGTON": "WA", "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
}

TARGET_ALIASES = {
    ("Charlotte", "NC"): [
        "Charlotte-Mecklenburg Police Department",
        "Charlotte-Mecklenburg PD",
        "Charlotte Police Department",
    ],
    ("Nashville-Davidson", "TN"): [
        "Metropolitan Nashville Police Department",
        "Nashville Metropolitan Police Department",
        "Nashville Police Department",
        "Davidson County",
    ],
    ("Las Vegas", "NV"): [
        "Las Vegas Metropolitan Police Department",
        "Las Vegas Metro Police Department",
        "Las Vegas Police Department",
    ],
    ("Atlanta", "GA"): [
        "Atlanta Police Department",
    ],
    ("New Orleans", "LA"): [
        "New Orleans Police Department",
    ],
    ("Lexington-Fayette", "KY"): [
        "Lexington Police Department",
        "Lexington-Fayette Police Department",
        "Fayette County",
    ],
    ("Baton Rouge", "LA"): [
        "Baton Rouge Police Department",
    ],
    ("Augusta-Richmond", "GA"): [
        "Richmond County Sheriff's Office",
        "Augusta Police Department",
        "Augusta-Richmond County",
    ],
    ("Macon-Bibb", "GA"): [
        "Bibb County Sheriff's Office",
        "Macon Police Department",
        "Macon-Bibb County",
    ],
    ("Jackson", "MS"): [
        "Jackson Police Department",
    ],
    ("Athens-Clarke", "GA"): [
        "Athens-Clarke County Police Department",
        "Athens Police Department",
        "Clarke County",
    ],
    ("San Buenaventura / Ventura", "CA"): [
        "Ventura Police Department",
        "San Buenaventura Police Department",
    ],
    ("Downey", "CA"): [
        "Downey Police Department",
    ],
}

TARGET_ROW_LOOKUP_ALIASES = {
    ("San Buenaventura / Ventura", "CA"): ["Ventura", "San Buenaventura", "San Buenaventura (Ventura)"],
}

KEY_CRIME_COLS = [
    "FBI Population",
    "Coverage Rate",
    "Violent Crime Rate per 100k",
    "Property Crime Rate per 100k",
]

PUNCT = re.compile(r"[^\w\s]")
SPACE = re.compile(r"\s+")


def normalize_text(value):
    txt = str(value or "").strip().lower()
    txt = txt.replace("-", " ").replace("/", " ")
    txt = PUNCT.sub(" ", txt)
    txt = SPACE.sub(" ", txt).strip()
    return txt


def normalize_for_match(value):
    txt = normalize_text(value)
    replacements = {
        "police department": "police",
        "department": "",
        "dept": "",
        "pd": "police",
        "metro": "metropolitan",
        "sheriff s": "sheriff",
        "county": "",
        "city": "",
    }
    for old, new in replacements.items():
        txt = txt.replace(old, new)
    txt = SPACE.sub(" ", txt).strip()
    return txt


def core_tokens(value):
    txt = normalize_for_match(value)
    stop = {
        "police", "department", "metropolitan", "metro", "sheriff", "office",
        "county", "city", "pd", "dept",
    }
    tokens = [t for t in txt.split() if t and t not in stop]
    return " ".join(tokens), set(tokens)


def get_rating(score):
    if score is None or pd.isna(score):
        return ""
    s = float(score)
    if s <= 20:
        return "Very Low"
    if s <= 40:
        return "Low"
    if s <= 60:
        return "Moderate"
    if s <= 75:
        return "Elevated"
    if s <= 90:
        return "High"
    return "Very High"


def is_blank_row(ws, row_num, col_idx):
    for c in KEY_CRIME_COLS:
        v = ws.cell(row_num, col_idx[c]).value
        if v not in (None, ""):
            return False
    return True


def ensure_columns(ws, col_idx, required):
    for name in required:
        if name in col_idx:
            continue
        ws.cell(1, ws.max_column + 1).value = name
        col_idx[name] = ws.max_column


def load_fbi_table_8(fbi_file=None):
    fbi_file = Path(fbi_file) if fbi_file is not None else get_fbi_crime_workbook_path()
    df = pd.read_excel(fbi_file, sheet_name=0, header=3, engine="openpyxl")
    df.columns = [normalize_table_8_header(c) for c in df.columns]
    df = df.copy()
    df["state"] = df["state"].astype(str).str.upper().map(STATE_TO_ABBR)
    df["city"] = df["city"].astype(str).str.strip()
    df["population"] = pd.to_numeric(df.get("population", 0), errors="coerce")
    df["violent crime"] = pd.to_numeric(df.get("violent crime", 0), errors="coerce")
    df["property crime"] = pd.to_numeric(df.get("property crime", 0), errors="coerce")
    df = df[(df["state"].notna()) & (df["city"].notna()) & (df["population"] > 0)]
    df["city_norm"] = df["city"].map(normalize_for_match)
    return df[["state", "city", "city_norm", "population", "violent crime", "property crime"]].reset_index(drop=True)


def pick_match(state_abbr, aliases, state_df):
    if state_df.empty:
        return None, "No FBI rows for state"

    # First pass: exact normalized alias match to FBI city/agency label
    alias_norm = [normalize_for_match(a) for a in aliases]
    exact_hits = []
    for raw_alias, norm_alias in zip(aliases, alias_norm):
        matches = state_df[state_df["city_norm"] == norm_alias]
        if len(matches) == 1:
            row = matches.iloc[0]
            return row, f"Exact alias match: {raw_alias}"
        if len(matches) > 1:
            exact_hits.append((raw_alias, matches))

    if exact_hits:
        return None, "Ambiguous exact alias match in state"

    # Second pass: conservative fuzzy with token containment/overlap.
    scored = []
    for raw_alias, norm_alias in zip(aliases, alias_norm):
        alias_core, alias_tokens = core_tokens(raw_alias)
        for _, r in state_df.iterrows():
            city_norm = r["city_norm"]
            city_core, city_tokens = core_tokens(r["city"])
            ratio = SequenceMatcher(None, norm_alias, city_norm).ratio()
            core_ratio = SequenceMatcher(None, alias_core, city_core).ratio() if alias_core and city_core else 0.0
            contains = False
            if alias_core and city_core:
                if alias_core == city_core:
                    contains = True
                elif len(alias_tokens) >= 2 and (alias_core in city_core or city_core in alias_core):
                    contains = True
            overlap = len(alias_tokens & city_tokens)

            strong_single = overlap == 1 and core_ratio >= 0.95
            if contains or overlap >= 2 or strong_single or ratio >= 0.90:
                scored.append((contains, overlap, core_ratio, ratio, raw_alias, r))

    if not scored:
        return None, "No exact/strong fuzzy alias match"

    dedup = {}
    for item in scored:
        city_key = str(item[5]["city"])
        prev = dedup.get(city_key)
        if prev is None or item[:4] > prev[:4]:
            dedup[city_key] = item
    scored = list(dedup.values())

    scored.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[5]["population"]), reverse=True)
    best_contains, best_overlap, best_core_ratio, best_ratio, best_alias, best_row = scored[0]
    if len(scored) > 1:
        _, second_overlap, second_core_ratio, second_ratio, _, _ = scored[1]
        # Require clear separation to avoid bad assignments.
        if (
            best_contains == scored[1][0]
            and abs(best_overlap - second_overlap) < 1
            and abs(best_core_ratio - second_core_ratio) < 0.03
            and abs(best_ratio - second_ratio) < 0.03
        ):
            return None, f"Ambiguous fuzzy match (top {best_core_ratio:.3f}/{best_ratio:.3f} vs {second_core_ratio:.3f}/{second_ratio:.3f})"

    return best_row, f"Fuzzy alias match: {best_alias} (core={best_core_ratio:.3f}, full={best_ratio:.3f})"


def fix_crime_non_fl_overrides(input_path=None, output_path=None, backup_path=None, fbi_file=None):
    """Fill blank non-Florida Crime Index rows via targeted agency-name overrides.

    Returns a summary dict identical in shape to what the CLI used to print.
    """
    input_wb = Path(input_path) if input_path is not None else INPUT_WB
    output_wb = Path(output_path) if output_path is not None else OUTPUT_WB
    backup_wb = Path(backup_path) if backup_path is not None else BACKUP_WB
    fbi_file = Path(fbi_file) if fbi_file is not None else get_fbi_crime_workbook_path()

    if not input_wb.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_wb}")
    if not fbi_file.exists():
        raise FileNotFoundError(f"FBI Table 8 source not found: {fbi_file}")

    shutil.copy2(input_wb, backup_wb)

    wb = load_workbook(input_wb)
    if "Crime Index" not in wb.sheetnames:
        raise RuntimeError("Crime Index sheet not found")

    ws = wb["Crime Index"]
    headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else "" for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}

    required_cols = [
        "City", "State", "Census City Population", "FBI Population", "Coverage Rate",
        "Violent Crime Rate per 100k", "Property Crime Rate per 100k", "Crime Index Score",
        "Crime Rating", "Manual Review", "Last Updated",
    ]
    for c in required_cols:
        if c not in col_idx:
            raise RuntimeError(f"Required column missing in Crime Index: {c}")

    ensure_columns(ws, col_idx, ["FBI Agency Name", "FBI ORI", "FBI Match Method", "FBI Match Notes"])

    # Snapshot Florida blank rows before for untouched validation.
    fl_blank_before = 0
    fl_key_snapshot = {}

    targets = {(city, st) for city, st in TARGET_ALIASES.keys()}
    target_rows = {}

    non_fl_blank_before = 0
    for r in range(2, ws.max_row + 1):
        city = str(ws.cell(r, col_idx["City"]).value or "").strip()
        st = str(ws.cell(r, col_idx["State"]).value or "").strip()
        if st == "FL" and is_blank_row(ws, r, col_idx):
            fl_blank_before += 1
            fl_key_snapshot[r] = tuple(ws.cell(r, col_idx[k]).value for k in KEY_CRIME_COLS + ["Manual Review"])

        key = (city, st)
        if key in targets:
            target_rows[(city, st)] = r
            if is_blank_row(ws, r, col_idx):
                non_fl_blank_before += 1

    # Fill lookup gaps where workbook city text differs from requested target label.
    for target_key, lookup_names in TARGET_ROW_LOOKUP_ALIASES.items():
        if target_key in target_rows:
            continue
        _, st = target_key
        for r in range(2, ws.max_row + 1):
            city = str(ws.cell(r, col_idx["City"]).value or "").strip()
            row_st = str(ws.cell(r, col_idx["State"]).value or "").strip()
            if row_st != st:
                continue
            if city in lookup_names:
                target_rows[target_key] = r
                if is_blank_row(ws, r, col_idx):
                    non_fl_blank_before += 1
                break

    fbi_df = load_fbi_table_8(fbi_file)

    # Existing distributions used by original methodology for percentile-based scoring.
    existing_violent = []
    existing_property = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col_idx["Violent Crime Rate per 100k"]).value
        p = ws.cell(r, col_idx["Property Crime Rate per 100k"]).value
        if isinstance(v, (int, float)) and v >= 0:
            existing_violent.append(float(v))
        if isinstance(p, (int, float)) and p >= 0:
            existing_property.append(float(p))

    populated = []
    still_blank = []

    for city_state, aliases in TARGET_ALIASES.items():
        city, st = city_state
        row = target_rows.get(city_state)
        if row is None:
            still_blank.append((city, st, "City/state row not found in Crime Index"))
            continue

        # Only target blank/non-Florida rows for this pass.
        if st == "FL":
            continue
        if not is_blank_row(ws, row, col_idx):
            still_blank.append((city, st, "Row already populated; left unchanged"))
            continue

        state_df = fbi_df[fbi_df["state"] == st].copy()
        match_row, reason = pick_match(st, aliases, state_df)

        if match_row is None:
            ws.cell(row, col_idx["Manual Review"]).value = "TRUE"
            ws.cell(row, col_idx["FBI Match Method"]).value = "non-FL override"
            ws.cell(row, col_idx["FBI Match Notes"]).value = reason
            still_blank.append((city, st, reason))
            continue

        census_pop = pd.to_numeric(ws.cell(row, col_idx["Census City Population"]).value, errors="coerce")
        fbi_pop = float(match_row["population"])
        coverage = (fbi_pop / census_pop) if pd.notna(census_pop) and census_pop > 0 else None

        ws.cell(row, col_idx["FBI Agency Name"]).value = str(match_row["city"])
        ws.cell(row, col_idx["FBI ORI"]).value = ""
        ws.cell(row, col_idx["FBI Match Method"]).value = "non-FL override"

        if coverage is None:
            ws.cell(row, col_idx["Manual Review"]).value = "TRUE"
            ws.cell(row, col_idx["FBI Match Notes"]).value = f"{reason}; missing census population"
            still_blank.append((city, st, "Missing census population"))
            continue

        ws.cell(row, col_idx["FBI Population"]).value = int(round(fbi_pop))
        ws.cell(row, col_idx["Coverage Rate"]).value = round(float(coverage), 4)

        if coverage < 0.75 or coverage > 1.25:
            ws.cell(row, col_idx["Manual Review"]).value = "TRUE"
            ws.cell(row, col_idx["FBI Match Notes"]).value = f"{reason}; coverage out of range ({coverage:.4f})"
            still_blank.append((city, st, f"Coverage out of range ({coverage:.4f})"))
            continue

        violent_count = float(match_row["violent crime"])
        property_count = float(match_row["property crime"])
        violent_rate = round((violent_count / fbi_pop) * 100000, 2)
        property_rate = round((property_count / fbi_pop) * 100000, 2)

        ws.cell(row, col_idx["Violent Crime Rate per 100k"]).value = violent_rate
        ws.cell(row, col_idx["Property Crime Rate per 100k"]).value = property_rate

        v_series = pd.Series(existing_violent + [violent_rate])
        p_series = pd.Series(existing_property + [property_rate])
        v_pct = float(v_series.rank(pct=True, method="max").iloc[-1] * 100.0)
        p_pct = float(p_series.rank(pct=True, method="max").iloc[-1] * 100.0)
        score = round(0.75 * v_pct + 0.25 * p_pct, 2)

        ws.cell(row, col_idx["Crime Index Score"]).value = score
        ws.cell(row, col_idx["Crime Rating"]).value = get_rating(score)
        ws.cell(row, col_idx["Manual Review"]).value = ""
        ws.cell(row, col_idx["Last Updated"]).value = datetime.date.today().isoformat()
        ws.cell(row, col_idx["FBI Match Notes"]).value = reason

        populated.append({
            "City": city,
            "State": st,
            "FBI Agency Name": str(match_row["city"]),
            "FBI ORI": "",
            "FBI Population": int(round(fbi_pop)),
            "Coverage Rate": round(float(coverage), 4),
            "Manual Review": "",
        })

    # Validate Florida rows unchanged for key blank fields.
    fl_unchanged = 0
    for r, snap in fl_key_snapshot.items():
        now = tuple(ws.cell(r, col_idx[k]).value for k in KEY_CRIME_COLS + ["Manual Review"])
        if now == snap:
            fl_unchanged += 1

    # Save output workbook from in-memory updated copy.
    output_wb.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_wb)

    # Sanity summary
    print(f"Backup created: {backup_wb}")
    print(f"Output saved: {output_wb}")
    print(f"Non-Florida blank rows before: {non_fl_blank_before}")
    print(f"Non-Florida rows successfully populated: {len(populated)}")

    remaining_blanks = 0
    for (city, st), row in target_rows.items():
        if st != "FL" and is_blank_row(ws, row, col_idx):
            remaining_blanks += 1
    print(f"Non-Florida rows still blank: {remaining_blanks}")
    print(f"Florida blank rows left untouched: {fl_unchanged} (of {fl_blank_before} tracked blank FL rows)")

    print("\nNewly populated rows:")
    if populated:
        for p in populated:
            print(
                f"- {p['City']}, {p['State']} | Agency: {p['FBI Agency Name']} | ORI: {p['FBI ORI']} "
                f"| FBI Pop: {p['FBI Population']} | Coverage: {p['Coverage Rate']:.4f} | Manual Review: {p['Manual Review'] or 'FALSE'}"
            )
    else:
        print("- none")

    print("\nStill-blank non-Florida target rows:")
    if still_blank:
        for city, st, reason in still_blank:
            # Only report true blanks and untouched target rows for transparency.
            row = target_rows.get((city, st))
            if row is None:
                print(f"- {city}, {st} | reason: {reason}")
                continue
            if is_blank_row(ws, row, col_idx):
                print(f"- {city}, {st} | reason: {reason}")
    else:
        print("- none")

    return {
        "output_path": str(output_wb),
        "backup_path": str(backup_wb),
        "non_fl_blank_before": non_fl_blank_before,
        "populated": populated,
        "still_blank": still_blank,
        "remaining_blanks": remaining_blanks,
        "fl_unchanged": fl_unchanged,
        "fl_blank_before": fl_blank_before,
    }


def main():
    fix_crime_non_fl_overrides()


if __name__ == "__main__":
    main()
