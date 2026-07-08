#!/usr/bin/env python3
"""Simple climate risk join: PLACE -> COUNTY -> FEMA NRI

Produces: output/us_cities_100k_population_ranked_WITH_CLIMATE_RISK_SIMPLE.xlsx

Adds only these columns to the Clean Cities 100k+ sheet:
- climate_risk_county
- climate_risk_county_fips
- climate_risk_score
- climate_risk_rating

Uses Census TIGER PLACE + COUNTY and FEMA National Risk Index Counties.
"""
import sys
from pathlib import Path
import re
import pandas as pd
import requests
try:
    import geopandas as gpd
except Exception:
    print("Missing geopandas. Install with: python3 -m pip install geopandas pyogrio shapely")
    sys.exit(1)
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_LANDLORD_AND_POP_CHANGE.xlsx"
OUTPUT_FILE = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_CLIMATE_RISK_SIMPLE.xlsx"
RAW_DIR = BASE_DIR / "data" / "raw"
TIGER_DIR = RAW_DIR / "tiger"

TIGER_YEARS = [2025, 2024]
TIGER_PLACE_URL_TEMPLATE = (
    "https://www2.census.gov/geo/tiger/TIGER{year}/PLACE/tl_{year}_us_place.zip"
)
TIGER_PLACE_STATE_URL_TEMPLATE = (
    "https://www2.census.gov/geo/tiger/TIGER{year}/PLACE/tl_{year}_{state_fips}_place.zip"
)
TIGER_COUNTY_URL_TEMPLATE = (
    "https://www2.census.gov/geo/tiger/TIGER{year}/COUNTY/tl_{year}_us_county.zip"
)

FEMA_NRI_QUERY_URL = (
    "https://services.arcgis.com/XG15cJAlne2vxtgt/ArcGIS/rest/services/"
    "National_Risk_Index_Counties/FeatureServer/0/query"
)

STATE_ABBR_TO_FIPS = {
    "AL": "01", "AK": "02", "AZ": "04", "AR": "05", "CA": "06",
    "CO": "08", "CT": "09", "DE": "10", "DC": "11", "FL": "12",
    "GA": "13", "HI": "15", "ID": "16", "IL": "17", "IN": "18",
    "IA": "19", "KS": "20", "KY": "21", "LA": "22", "ME": "23",
    "MD": "24", "MA": "25", "MI": "26", "MN": "27", "MS": "28",
    "MO": "29", "MT": "30", "NE": "31", "NV": "32", "NH": "33",
    "NJ": "34", "NM": "35", "NY": "36", "NC": "37", "ND": "38",
    "OH": "39", "OK": "40", "OR": "41", "PA": "42", "RI": "44",
    "SC": "45", "SD": "46", "TN": "47", "TX": "48", "UT": "49",
    "VT": "50", "VA": "51", "WA": "53", "WV": "54", "WI": "55",
    "WY": "56", "PR": "72",
}

PLACE_SUFFIX_PATTERN = re.compile(r"\b(city|town|village|cdp|municipality|metropolitan government)\b$", flags=re.IGNORECASE)
PUNCTUATION_PATTERN = re.compile(r"[^\w\s]", flags=re.UNICODE)
PAREN_PATTERN = re.compile(r"\s*\([^)]*\)")
MULTI_SPACE_PATTERN = re.compile(r"\s+")


def ensure_dirs():
    TIGER_DIR.mkdir(parents=True, exist_ok=True)
    (TIGER_DIR / "PLACE").mkdir(parents=True, exist_ok=True)
    (TIGER_DIR / "COUNTY").mkdir(parents=True, exist_ok=True)


def download_file(url: str, dest: Path):
    if dest.exists():
        return
    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"Downloading {url}")
    try:
        r = requests.get(url, stream=True, timeout=120)
    except requests.RequestException as e:
        raise RuntimeError(f"Failed to download {url}: {e}")
    if r.status_code != 200:
        raise RuntimeError(f"Failed to download {url} HTTP {r.status_code}")
    with open(dest, "wb") as fh:
        for chunk in r.iter_content(8192):
            if chunk:
                fh.write(chunk)


def normalize_for_matching(value):
    if pd.isna(value):
        return None
    text = str(value).strip().lower()
    text = PAREN_PATTERN.sub("", text)
    text = PUNCTUATION_PATTERN.sub(" ", text)
    text = PLACE_SUFFIX_PATTERN.sub("", text).strip()
    text = MULTI_SPACE_PATTERN.sub(" ", text)
    return text


def normalize_state_input(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    if len(text) == 2:
        return text.upper()
    return text.title() if text.title() in STATE_ABBR_TO_FIPS else text.upper()


def choose_columns(columns):
    city_col = next((c for c in ["city", "City"] if c in columns), None)
    state_col = next((c for c in ["state_abbr", "state", "State"] if c in columns), None)
    return city_col, state_col


def load_city_sheet():
    if not INPUT_FILE.exists():
        print(f"Missing input workbook: {INPUT_FILE}")
        sys.exit(1)
    df = pd.read_excel(INPUT_FILE, sheet_name="Clean Cities 100k+", engine="openpyxl", dtype=str)
    city_col, state_col = choose_columns(df.columns)
    if city_col is None or state_col is None:
        print("Unable to find city/state columns in workbook.")
        sys.exit(1)
    df = df.reset_index(drop=True)
    df['__city_original'] = df[city_col].astype(str).fillna("")
    df['__state_original'] = df[state_col].astype(str).fillna("")
    df['city_normalized'] = df['__city_original'].map(normalize_for_matching)
    df['state_abbr'] = df['__state_original'].map(normalize_state_input)
    df['statefp'] = df['state_abbr'].map(STATE_ABBR_TO_FIPS)
    df['match_key'] = df.apply(lambda r: f"{r['statefp']}|{r['city_normalized']}" if pd.notna(r['statefp']) and pd.notna(r['city_normalized']) and r['city_normalized']!='' else None, axis=1)
    df['row_id'] = df.index.astype(str)
    return df


def load_tiger_place_gdf(year, needed_states):
    place_dir = TIGER_DIR / "PLACE"
    us_zip = place_dir / f"tl_{year}_us_place.zip"
    if not us_zip.exists():
        try:
            download_file(TIGER_PLACE_URL_TEMPLATE.format(year=year), us_zip)
        except Exception:
            us_zip = None
    if us_zip and us_zip.exists():
        return gpd.read_file(us_zip)
    # fallback to state zips
    print(f"Falling back to state-level PLACE zips for TIGER {year}")
    gdfs = []
    for abbr in sorted(set(needed_states)):
        fips = STATE_ABBR_TO_FIPS.get(abbr)
        if not fips:
            continue
        state_zip = place_dir / f"tl_{year}_{fips}_place.zip"
        if not state_zip.exists():
            download_file(TIGER_PLACE_STATE_URL_TEMPLATE.format(year=year, state_fips=fips), state_zip)
        gdfs.append(gpd.read_file(state_zip))
    if not gdfs:
        raise RuntimeError(f"No place shapefiles available for {year}")
    return pd.concat(gdfs, ignore_index=True)


def prepare_place_match_table(place_gdf):
    place_gdf = place_gdf.to_crs("EPSG:4269")
    place_gdf = place_gdf[['STATEFP', 'GEOID', 'NAME', 'geometry']].copy()
    place_gdf['place_name_match'] = place_gdf['NAME'].map(normalize_for_matching)
    place_gdf['match_key'] = place_gdf.apply(lambda r: f"{r['STATEFP']}|{r['place_name_match']}" if pd.notna(r['place_name_match']) else None, axis=1)
    place_gdf = place_gdf[place_gdf['match_key'].notna()].copy()
    place_gdf = place_gdf.set_index('match_key')
    return place_gdf


def match_city_to_place(df, place_gdf):
    matched = []
    for _, row in df.iterrows():
        key = row.get('match_key')
        if key is None:
            matched.append(None)
            continue
        try:
            match = place_gdf.loc[key]
        except Exception:
            match = None
        if match is None:
            matched.append(None)
        elif isinstance(match, pd.DataFrame):
            matched.append(match.iloc[0].geometry)
        else:
            matched.append(match.geometry)
    df['place_geometry'] = matched
    return df


def load_tiger_county_gdf(year):
    county_dir = TIGER_DIR / "COUNTY"
    county_zip = county_dir / f"tl_{year}_us_county.zip"
    if not county_zip.exists():
        download_file(TIGER_COUNTY_URL_TEMPLATE.format(year=year), county_zip)
    return gpd.read_file(county_zip)


def prepare_county_gdf(county_gdf):
    county_gdf = county_gdf.to_crs('EPSG:4269')
    county_gdf = county_gdf[['STATEFP', 'COUNTYFP', 'GEOID', 'NAME', 'geometry']].copy()
    county_gdf['county_geoid'] = county_gdf['GEOID'].astype(str).str.zfill(5)
    county_gdf['state_abbr'] = county_gdf['STATEFP'].astype(str).map({v: k for k, v in STATE_ABBR_TO_FIPS.items()})
    return county_gdf


def compute_county_assignments(df, county_gdf):
    place_records = df[df['place_geometry'].notna()].copy()
    if place_records.empty:
        df['climate_risk_county'] = None
        df['climate_risk_county_fips'] = None
        return df
    place_gdf = gpd.GeoDataFrame(place_records, geometry='place_geometry', crs='EPSG:4269')
    place_gdf = place_gdf.to_crs('EPSG:5070')
    county_proj = county_gdf.to_crs('EPSG:5070')
    # spatial join intersects
    joined = gpd.sjoin(place_gdf, county_proj[['county_geoid', 'NAME', 'state_abbr', 'geometry']], how='left', predicate='intersects')
    if joined.empty:
        df['climate_risk_county'] = None
        df['climate_risk_county_fips'] = None
        return df
    # ensure same-state join
    joined = joined[joined['state_abbr_left'] == joined['state_abbr_right']]
    # merge county geometry
    county_geoms = county_proj[['county_geoid', 'geometry']].rename(columns={'geometry': 'county_geometry'})
    joined = joined.merge(county_geoms, on='county_geoid', how='left')
    # compute best county by area share
    results = {}
    for row_id, group in joined.groupby('row_id'):
        place_geom = group.iloc[0]['place_geometry']
        if place_geom is None:
            results[row_id] = (None, None)
            continue
        total_area = place_geom.area
        if total_area <= 0:
            results[row_id] = (None, None)
            continue
        candidates = []
        for _, r in group.iterrows():
            cgeom = r.get('county_geometry')
            if cgeom is None:
                continue
            inter = place_geom.intersection(cgeom)
            if inter.is_empty:
                continue
            share = inter.area / total_area
            candidates.append((r['county_geoid'], r['NAME'], share))
        if not candidates:
            results[row_id] = (None, None)
            continue
        candidates.sort(key=lambda x: x[2], reverse=True)
        best = candidates[0]
        if best[2] < 0.5:
            results[row_id] = (None, None)
        else:
            results[row_id] = (best[1], best[0])
    df['climate_risk_county'] = df['row_id'].map(lambda x: results.get(x, (None, None))[0])
    df['climate_risk_county_fips'] = df['row_id'].map(lambda x: results.get(x, (None, None))[1])
    return df


def fetch_fema_nri():
    records = []
    offset = 0
    page = 1000
    while True:
        params = {
            'where': '1=1',
            'outFields': 'STCOFIPS,COUNTY,STATEABBRV,RISK_SCORE,RISK_RATNG',
            'f': 'json',
            'resultOffset': offset,
            'resultRecordCount': page,
        }
        try:
            r = requests.get(FEMA_NRI_QUERY_URL, params=params, timeout=120)
        except requests.RequestException as e:
            raise RuntimeError(f"Failed to fetch FEMA NRI: {e}")
        if r.status_code != 200:
            raise RuntimeError(f"FEMA query failed HTTP {r.status_code}")
        payload = r.json()
        features = payload.get('features', [])
        if not features:
            break
        for f in features:
            a = f.get('attributes', {})
            st = a.get('STCOFIPS')
            if st is None:
                continue
            records.append({
                'STCOFIPS': str(st).zfill(5),
                'COUNTY': a.get('COUNTY'),
                'STATEABBRV': a.get('STATEABBRV'),
                'RISK_SCORE': a.get('RISK_SCORE'),
                'RISK_RATNG': a.get('RISK_RATNG'),
            })
        offset += page
        if len(features) < page:
            break
    if not records:
        raise RuntimeError('No FEMA NRI records fetched')
    return pd.DataFrame(records).set_index('STCOFIPS')


def assign_fema(df, fema_df):
    df['climate_risk_score'] = None
    df['climate_risk_rating'] = None
    for idx, row in df.iterrows():
        fips = row.get('climate_risk_county_fips')
        if not fips or pd.isna(fips):
            continue
        rec = fema_df.loc[fips] if fips in fema_df.index else None
        if rec is None:
            continue
        df.at[idx, 'climate_risk_county'] = rec.get('COUNTY')
        df.at[idx, 'climate_risk_score'] = rec.get('RISK_SCORE')
        df.at[idx, 'climate_risk_rating'] = rec.get('RISK_RATNG')
    return df


def write_output(original_path, df):
    sheets = pd.read_excel(original_path, sheet_name=None, engine='openpyxl', dtype=str)
    # ensure we keep original columns and only add the four climate columns if absent
    for c in ['climate_risk_county', 'climate_risk_county_fips', 'climate_risk_score', 'climate_risk_rating']:
        if c not in df.columns:
            df[c] = None
    sheets['Clean Cities 100k+'] = df.drop(columns=[c for c in ['__city_original','__state_original','city_normalized','place_geometry','match_key','row_id'] if c in df.columns], errors='ignore')
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', datetime_format='yyyy-mm-dd') as writer:
        for name, s in sheets.items():
            s.to_excel(writer, sheet_name=name, index=False)
    # post-process formatting
    wb = load_workbook(OUTPUT_FILE)
    if 'Clean Cities 100k+' in wb.sheetnames:
        sheet = wb['Clean Cities 100k+']
        sheet.freeze_panes = sheet['A2']
        max_row = sheet.max_row
        max_col = sheet.max_column
        sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        # find climate_risk_score column and format
        headers = [sheet.cell(row=1, column=i).value for i in range(1, max_col+1)]
        for idx, h in enumerate(headers, start=1):
            if h == 'climate_risk_score':
                for cell in sheet.iter_cols(min_col=idx, max_col=idx, min_row=2, max_row=max_row):
                    for c in cell:
                        if c.value is not None:
                            try:
                                c.number_format = '0.00'
                            except Exception:
                                pass
    wb.save(OUTPUT_FILE)


def sanity_checks(df):
    expected = {
        ('los angeles', 'CA'): '06037',
        ('chicago', 'IL'): '17031',
        ('houston', 'TX'): '48201',
        ('phoenix', 'AZ'): '04013',
        ('philadelphia', 'PA'): '42101',
        ('san diego', 'CA'): '06073',
        ('san francisco', 'CA'): '06075',
        ('portland', 'OR'): '41051',
    }
    errors = []
    for (city, st), expected_fips in expected.items():
        match = df[(df['city_normalized'] == city) & (df['state_abbr'] == st)]
        if match.empty:
            errors.append(f"Missing row for {city.title()}, {st}")
            continue
        actual = match.iloc[0].get('climate_risk_county_fips')
        if actual != expected_fips:
            errors.append(f"Expected {city.title()}, {st} -> {expected_fips}, got {actual}")
    if errors:
        print('Sanity checks failed:')
        for e in errors:
            print('-', e)
        return False
    return True


def main():
    ensure_dirs()
    df = load_city_sheet()
    needed_states = sorted(df['state_abbr'].dropna().unique())
    place_gdf = None
    county_gdf = None
    used_year = None
    for year in TIGER_YEARS:
        try:
            place_gdf = load_tiger_place_gdf(year, needed_states)
            county_gdf = load_tiger_county_gdf(year)
            used_year = year
            break
        except Exception as e:
            print(f"TIGER {year} load failed: {e}")
            continue
    if place_gdf is None or county_gdf is None:
        print("Unable to load TIGER place or county shapefiles")
        sys.exit(1)
    place_table = prepare_place_match_table(place_gdf)
    df = match_city_to_place(df, place_table)
    place_matched = df['place_geometry'].notna().sum()
    county_gdf = prepare_county_gdf(county_gdf)
    df = compute_county_assignments(df, county_gdf)
    assigned_counties = df['climate_risk_county_fips'].astype(bool).sum()
    # fetch FEMA and join
    fema = fetch_fema_nri()
    df = assign_fema(df, fema)
    matched_fema = df['climate_risk_score'].notna().sum()
    left_blank = len(df) - matched_fema
    # sanitize final columns: keep original columns and add exactly four climate fields (ensure order preserved)
    # run sanity checks
    if not sanity_checks(df):
        print('Aborting save due to failed sanity checks.')
        print(f"Output file (not written): {OUTPUT_FILE}")
        sys.exit(1)
    # drop helper columns before saving
    outdf = df.copy()
    for c in ['place_geometry', 'match_key', 'city_normalized', '__city_original', '__state_original', 'row_id', 'statefp']:
        if c in outdf.columns:
            outdf.drop(columns=[c], inplace=True)
    # ensure climate fields exist
    for c in ['climate_risk_county', 'climate_risk_county_fips', 'climate_risk_score', 'climate_risk_rating']:
        if c not in outdf.columns:
            outdf[c] = None
    write_output(INPUT_FILE, outdf)

    print(f"Output file: {OUTPUT_FILE}")
    print(f"Number of rows: {len(df)}")
    print(f"Number matched to Census PLACE: {place_matched}")
    print(f"Number assigned to county: {assigned_counties}")
    print(f"Number matched to FEMA risk score: {matched_fema}")
    print(f"Number left blank: {left_blank}")

    ranked = df[df['climate_risk_score'].notna()].sort_values('climate_risk_score', ascending=False)
    if not ranked.empty:
        print('\nTop 10 highest risk cities:')
        print(ranked[['__city_original','state_abbr','climate_risk_county','climate_risk_county_fips','climate_risk_score']].head(10).to_string(index=False))
        low = ranked.sort_values('climate_risk_score', ascending=True)
        print('\nTop 10 lowest risk cities:')
        print(low[['__city_original','state_abbr','climate_risk_county','climate_risk_county_fips','climate_risk_score']].head(10).to_string(index=False))


if __name__ == '__main__':
    main()
