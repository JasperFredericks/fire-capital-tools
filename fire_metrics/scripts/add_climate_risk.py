import os
import re
import sys
from pathlib import Path

try:
    import geopandas as gpd
    import numpy as np
    import pandas as pd
    import requests
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    package = exc.name
    print(f"Missing required package: {package}")
    print("Install dependencies with: python3 -m pip install pandas openpyxl requests geopandas shapely pyogrio")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_LANDLORD_AND_POP_CHANGE.xlsx"
OUTPUT_FILE = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_CLIMATE_RISK_FIXED.xlsx"
RAW_DIR = BASE_DIR / "data" / "raw"
TIGER_DIR = RAW_DIR / "tiger"

CLIMATE_COLUMNS_TO_REMOVE = [
    "climate_risk_county_name",
    "climate_risk_county_fips",
    "county_area_share",
    "city_crosses_multiple_counties",
    "climate_risk_score",
    "climate_risk_rating",
    "climate_risk_review_flag",
    "climate_risk_source",
    "county_match_source",
    "climate_risk_notes",
    "county_assignment_notes",
    "fema_match_notes",
    "place_geometry",
    "place_match_ambiguous",
    "place_match_notes",
    "row_id",
    "city_normalized",
    "state_abbr",
    "__city_original",
    "__state_original",
]

FEMA_NRI_QUERY_URL = (
    "https://services.arcgis.com/XG15cJAlne2vxtgt/ArcGIS/rest/services/"
    "National_Risk_Index_Counties/FeatureServer/0/query"
)

TIGER_YEARS = [2025, 2024]
TIGER_PLACE_URL_TEMPLATE = (
    "https://www2.census.gov/geo/tiger/TIGER{year}/PLACE/tl_{year}_us_place.zip"
)
TIGER_PLACE_STATE_URL_TEMPLATE = (
    "https://www2.census.gov/geo/tiger/TIGER{year}/PLACE/tl_{year}_{state_fips}_place.zip"
)
TIGER_COUNTY_URL_TEMPLATE = (
    "https://www2.census.gov/geo/tiger/TIGER{year}/COUNTY/tl_{year}_us_county.zip"
)

STATE_NAME_TO_ABBR = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "District of Columbia": "DC",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "Puerto Rico": "PR",
}
STATE_ABBR_TO_FIPS = {
    "AL": "01",
    "AK": "02",
    "AZ": "04",
    "AR": "05",
    "CA": "06",
    "CO": "08",
    "CT": "09",
    "DE": "10",
    "DC": "11",
    "FL": "12",
    "GA": "13",
    "HI": "15",
    "ID": "16",
    "IL": "17",
    "IN": "18",
    "IA": "19",
    "KS": "20",
    "KY": "21",
    "LA": "22",
    "ME": "23",
    "MD": "24",
    "MA": "25",
    "MI": "26",
    "MN": "27",
    "MS": "28",
    "MO": "29",
    "MT": "30",
    "NE": "31",
    "NV": "32",
    "NH": "33",
    "NJ": "34",
    "NM": "35",
    "NY": "36",
    "NC": "37",
    "ND": "38",
    "OH": "39",
    "OK": "40",
    "OR": "41",
    "PA": "42",
    "RI": "44",
    "SC": "45",
    "SD": "46",
    "TN": "47",
    "TX": "48",
    "UT": "49",
    "VT": "50",
    "VA": "51",
    "WA": "53",
    "WV": "54",
    "WI": "55",
    "WY": "56",
    "PR": "72",
}
STATE_FIPS_TO_ABBR = {value: key for key, value in STATE_ABBR_TO_FIPS.items()}

CLIMATE_RISK_SOURCE = (
    "FEMA National Risk Index Counties ArcGIS REST service"
)
COUNTY_MATCH_SOURCE = (
    "U.S. Census TIGER/Line PLACE and COUNTY shapefiles"
)

CITY_COL_CANDIDATES = ["city", "City", "city_normalized", "Geographic Area"]
STATE_COL_CANDIDATES = ["state_abbr", "State Abbr", "state", "State"]

README_CLIMATE_NOTE = (
    "Climate/NRI score is assigned at the county level using FEMA National Risk Index Counties. "
    "Cities are matched to counties using U.S. Census TIGER/Line PLACE and COUNTY boundaries. "
    "For cities crossing county lines, this workbook assigns the county containing the largest share of the city/place land area when that share is at least 50%. "
    "FEMA NRI is not purely physical climate exposure; it incorporates expected annual loss, social vulnerability, and community resilience."
)

PLACE_SUFFIX_PATTERN = re.compile(r"\b(city|town|village|cdp|municipality|metropolitan government)\b$", flags=re.IGNORECASE)
PUNCTUATION_PATTERN = re.compile(r"[^\w\s]", flags=re.UNICODE)
PAREN_PATTERN = re.compile(r"\s*\([^)]*\)")
MULTI_SPACE_PATTERN = re.compile(r"\s+")


def ensure_directories():
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    TIGER_DIR.mkdir(parents=True, exist_ok=True)
    (TIGER_DIR / "PLACE").mkdir(parents=True, exist_ok=True)
    (TIGER_DIR / "COUNTY").mkdir(parents=True, exist_ok=True)


def download_file(url: str, dest: Path):
    if dest.exists():
        return
    dest.parent.mkdir(parents=True, exist_ok=True)
    print(f"Downloading {url}")
    try:
        response = requests.get(url, stream=True, timeout=120)
    except requests.RequestException as exc:
        print(f"Failed to download URL: {url}")
        print(str(exc))
        sys.exit(1)
    if response.status_code != 200:
        print(f"Failed to download URL: {url}")
        print(f"HTTP status: {response.status_code}")
        sys.exit(1)
    with open(dest, "wb") as handle:
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                handle.write(chunk)


def normalize_for_matching(value):
    if pd.isna(value):
        return None
    text = str(value).strip().lower()
    text = PAREN_PATTERN.sub("", text)
    text = PUNCTUATION_PATTERN.sub(" ", text)
    text = PLACE_SUFFIX_PATTERN.sub("", text).strip()
    text = MULTI_SPACE_PATTERN.sub(" ", text)
    return text


def normalize_state_input(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    if len(text) == 2:
        return text.upper()
    candidate = STATE_NAME_TO_ABBR.get(text)
    if candidate:
        return candidate
    candidate = STATE_NAME_TO_ABBR.get(text.title())
    if candidate:
        return candidate
    return text.upper()


def choose_city_state_columns(columns):
    city_col = next((col for col in CITY_COL_CANDIDATES if col in columns), None)
    state_col = next((col for col in STATE_COL_CANDIDATES if col in columns), None)
    return city_col, state_col


def load_city_workbook():
    if not INPUT_FILE.exists():
        print(f"Input workbook not found: {INPUT_FILE}")
        sys.exit(1)
    df = pd.read_excel(INPUT_FILE, sheet_name="Clean Cities 100k+", engine="openpyxl", dtype=str)
    df = df.drop(columns=[c for c in CLIMATE_COLUMNS_TO_REMOVE if c in df.columns], errors="ignore")
    city_col, state_col = choose_city_state_columns(df.columns)
    if city_col is None or state_col is None:
        print("Unable to find city or state column in the workbook.")
        print("Expected city columns:", CITY_COL_CANDIDATES)
        print("Expected state columns:", STATE_COL_CANDIDATES)
        sys.exit(1)
    df = df.reset_index(drop=True)
    df["row_id"] = df.index.astype(str).str.zfill(6)
    df["__city_original"] = df[city_col].astype(str).fillna("")
    df["__state_original"] = df[state_col].astype(str).fillna("")
    df["city_normalized"] = df["__city_original"].map(normalize_for_matching)
    df["state_abbr"] = df["__state_original"].map(normalize_state_input)
    df["statefp"] = df["state_abbr"].map(STATE_ABBR_TO_FIPS)
    df["match_key"] = df.apply(
        lambda row: f"{row['statefp']}|{row['city_normalized']}"
        if pd.notna(row["statefp"]) and pd.notna(row["city_normalized"]) and row["city_normalized"] != ""
        else None,
        axis=1,
    )
    return df, city_col, state_col


def load_tiger_place_gdf(year, needed_states):
    place_dir = TIGER_DIR / "PLACE"
    best_path = place_dir / f"tl_{year}_us_place.zip"
    if not best_path.exists():
        try:
            download_file(TIGER_PLACE_URL_TEMPLATE.format(year=year), best_path)
        except SystemExit:
            best_path = None
    if best_path and best_path.exists():
        return gpd.read_file(best_path)

    print(f"Falling back to state-level PLACE zips for TIGER {year}")
    state_paths = []
    for abbr in sorted(set(needed_states)):
        fips = STATE_ABBR_TO_FIPS.get(abbr)
        if not fips:
            continue
        state_path = place_dir / f"tl_{year}_{fips}_place.zip"
        if not state_path.exists():
            download_file(
                TIGER_PLACE_STATE_URL_TEMPLATE.format(year=year, state_fips=fips),
                state_path,
            )
        state_paths.append(state_path)
    if not state_paths:
        raise RuntimeError(f"No state-level place shapefiles could be downloaded for TIGER {year}")
    gdf_list = [gpd.read_file(path) for path in state_paths]
    return pd.concat(gdf_list, ignore_index=True)


def load_tiger_county_gdf(year):
    county_dir = TIGER_DIR / "COUNTY"
    county_path = county_dir / f"tl_{year}_us_county.zip"
    if not county_path.exists():
        download_file(TIGER_COUNTY_URL_TEMPLATE.format(year=year), county_path)
    return gpd.read_file(county_path)


def fetch_fema_nri():
    records = []
    offset = 0
    page_size = 1000
    while True:
        params = {
            "where": "1=1",
            "outFields": "STCOFIPS,STATE,STATEABBRV,COUNTY,RISK_SCORE,RISK_RATNG",
            "f": "json",
            "resultRecordCount": page_size,
            "resultOffset": offset,
        }
        try:
            response = requests.get(FEMA_NRI_QUERY_URL, params=params, timeout=120)
        except requests.RequestException as exc:
            print(f"Failed to download URL: {FEMA_NRI_QUERY_URL}")
            print(str(exc))
            sys.exit(1)
        if response.status_code != 200:
            print(f"Failed to download URL: {FEMA_NRI_QUERY_URL}")
            print(f"HTTP status: {response.status_code}")
            sys.exit(1)
        payload = response.json()
        features = payload.get("features", [])
        if not features:
            break
        for feature in features:
            attrs = feature.get("attributes", {})
            stcofips = attrs.get("STCOFIPS")
            if stcofips is None:
                continue
            stcofips = str(stcofips).zfill(5)
            records.append(
                {
                    "STCOFIPS": stcofips,
                    "STATE": attrs.get("STATE"),
                    "STATEABBRV": attrs.get("STATEABBRV"),
                    "COUNTY": attrs.get("COUNTY"),
                    "RISK_SCORE": attrs.get("RISK_SCORE"),
                    "RISK_RATNG": attrs.get("RISK_RATNG"),
                }
            )
        offset += page_size
        if len(features) < page_size:
            break
    if not records:
        print(f"No FEMA NRI data downloaded from {FEMA_NRI_QUERY_URL}")
        sys.exit(1)
    return pd.DataFrame(records)


def prepare_place_match_table(place_gdf):
    place_gdf = place_gdf.to_crs("EPSG:4269")
    place_gdf = place_gdf[["STATEFP", "GEOID", "NAME", "geometry"]].copy()
    place_gdf["place_name_match"] = place_gdf["NAME"].map(normalize_for_matching)
    place_gdf["match_key"] = place_gdf.apply(
        lambda row: f"{row['STATEFP']}|{row['place_name_match']}" if pd.notna(row['place_name_match']) else None,
        axis=1,
    )
    place_gdf = place_gdf[place_gdf["match_key"].notna()].copy()
    return place_gdf


def match_city_to_place(df, place_gdf):
    place_lookup = place_gdf.set_index('match_key')
    matched = []
    for _, row in df.iterrows():
        key = row.get('match_key')
        if key is None:
            matched.append(None)
            continue
        match = place_lookup.loc[key] if key in place_lookup.index else None
        if isinstance(match, pd.DataFrame):
            matched.append(match.iloc[0].geometry)
        elif isinstance(match, pd.Series):
            matched.append(match.geometry)
        else:
            matched.append(None)
    df['place_geometry'] = matched
    df['place_match_key'] = df['match_key']
    return df


def prepare_county_gdf(county_gdf):
    county_gdf = county_gdf.to_crs("EPSG:4269")
    county_gdf = county_gdf[["STATEFP", "COUNTYFP", "GEOID", "NAME", "geometry"]].copy()
    county_gdf["state_abbr"] = county_gdf["STATEFP"].astype(str).map(STATE_FIPS_TO_ABBR)
    county_gdf["county_geoid"] = county_gdf["GEOID"].astype(str).str.zfill(5)
    return county_gdf


def assign_place_geometries(df, place_gdf):
    place_lookup = {}
    for _, row in place_gdf.iterrows():
        key = (row["state_abbr"], row["place_name_match"])
        place_lookup.setdefault(key, []).append(row)

    matched_geoms = []
    multi_match_flags = []
    match_notes = []

    for _, row in df.iterrows():
        key = (row["state_abbr"], row["city_normalized"])
        candidates = place_lookup.get(key, [])
        if not candidates:
            matched_geoms.append(None)
            multi_match_flags.append(False)
            match_notes.append("Place match failed")
            continue
        if len(candidates) > 1:
            matched_geoms.append(candidates[0].geometry)
            multi_match_flags.append(True)
            match_notes.append("Multiple place matches")
        else:
            matched_geoms.append(candidates[0].geometry)
            multi_match_flags.append(False)
            match_notes.append("")
    df["place_geometry"] = matched_geoms
    df["place_match_ambiguous"] = multi_match_flags
    df["place_match_notes"] = match_notes
    return df


def compute_county_assignments(df, county_gdf):
    place_records = df[df["place_geometry"].notna()].copy()
    if place_records.empty:
        return df

    place_gdf = gpd.GeoDataFrame(
        place_records,
        geometry="place_geometry",
        crs="EPSG:4269",
    )
    place_gdf["state_abbr"] = place_gdf["state_abbr"].astype(str)
    place_gdf = place_gdf.to_crs("EPSG:5070")
    county_proj = county_gdf.to_crs("EPSG:5070")
    county_proj = county_proj.rename(columns={"NAME": "county_name"})

    joined = gpd.sjoin(
        place_gdf,
        county_proj[["county_geoid", "county_name", "state_abbr", "geometry"]],
        how="left",
        predicate="intersects",
    )
    if joined.empty:
        return df

    joined = joined[joined["state_abbr_left"] == joined["state_abbr_right"]].copy()
    if joined.empty:
        return df

    county_geoms = county_proj[["county_geoid", "geometry"]].rename(columns={"geometry": "county_geometry"})
    joined = joined.merge(county_geoms, on="county_geoid", how="left")
    joined = gpd.GeoDataFrame(joined, geometry="place_geometry")

    results = []
    for row_id, group in joined.groupby("row_id"):
        place_geom = group.iloc[0]["place_geometry"]
        if place_geom is None:
            results.append((row_id, None, None, 0.0, False, "No place geometry"))
            continue
        total_area = place_geom.area
        if total_area <= 0:
            results.append((row_id, None, None, 0.0, False, "Zero place area"))
            continue
        candidate_rows = []
        for _, row in group.iterrows():
            county_geom = row["county_geometry"]
            if county_geom is None:
                continue
            intersection = place_geom.intersection(county_geom)
            if intersection.is_empty:
                continue
            area_share = intersection.area / total_area
            candidate_rows.append(
                {
                    "county_geoid": row["county_geoid"],
                    "county_name": row["county_name"],
                    "area_share": area_share,
                }
            )
        if not candidate_rows:
            results.append((row_id, None, None, 0.0, False, "No county intersection"))
            continue
        candidate_rows.sort(key=lambda c: c["area_share"], reverse=True)
        best = candidate_rows[0]
        multiple = len(candidate_rows) > 1
        if best["area_share"] < 0.5:
            results.append((row_id, None, None, best["area_share"], True, "No county >= 50%"))
        else:
            results.append((row_id, best["county_geoid"], best["county_name"], best["area_share"], multiple, ""))

    assignment_index = pd.DataFrame(
        results,
        columns=[
            "row_id",
            "climate_risk_county_fips",
            "climate_risk_county_name",
            "county_area_share",
            "city_crosses_multiple_counties",
            "county_assignment_notes",
        ],
    ).set_index("row_id")

    df = df.set_index("row_id")
    df = df.join(assignment_index, how="left")
    df = df.reset_index()
    return df


def assign_fema_scores(df, fema_df):
    fema_df = fema_df.copy()
    fema_df["STCOFIPS"] = fema_df["STCOFIPS"].astype(str).str.zfill(5)
    fema_df = fema_df.set_index("STCOFIPS")
    score_map = fema_df.to_dict(orient="index")

    scores = []
    unmatched_fips = set()
    for _, row in df.iterrows():
        fips = row.get("climate_risk_county_fips")
        if pd.isna(fips) or not isinstance(fips, str) or not fips.strip():
            scores.append((None, None, "No county selected"))
            continue
        match = score_map.get(fips)
        if not match:
            scores.append((None, None, f"FEMA match failed for FIPS {fips}"))
            unmatched_fips.add(fips)
            continue
        scores.append((match.get("RISK_SCORE"), match.get("RISK_RATNG"), ""))

    score_df = pd.DataFrame(scores, columns=["climate_risk_score", "climate_risk_rating", "fema_match_notes"], index=df.index)
    df = df.join(score_df)
    df["climate_risk_score"] = pd.to_numeric(df["climate_risk_score"], errors="coerce")
    return df, sorted(unmatched_fips)


def build_review_flags(df):
    df["climate_risk_county_name"] = df["climate_risk_county_name"].fillna("")
    df["climate_risk_county_fips"] = df["climate_risk_county_fips"].fillna("")
    df["county_area_share"] = pd.to_numeric(df["county_area_share"], errors="coerce")
    df["city_crosses_multiple_counties"] = df["city_crosses_multiple_counties"].fillna(False).astype(bool)
    df["climate_risk_score"] = pd.to_numeric(df["climate_risk_score"], errors="coerce")

    df["climate_risk_source"] = CLIMATE_RISK_SOURCE
    df["county_match_source"] = COUNTY_MATCH_SOURCE

    def review_flag(row):
        if pd.isna(row["city_normalized"]) or row["__city_original"] == "" or row["state_abbr"] == "":
            return True
        if row.get("place_geometry") is None:
            return True
        if row.get("climate_risk_county_fips") in (None, ""):
            return True
        if pd.isna(row.get("climate_risk_score")):
            return True
        return False

    df["climate_risk_review_flag"] = df.apply(review_flag, axis=1)

    notes = []
    for _, row in df.iterrows():
        reasons = []
        if row.get("place_geometry") is None:
            reasons.append("Place match failed")

        assignment_note = row.get("county_assignment_notes")
        if assignment_note is not None and assignment_note != "":
            reasons.append(str(assignment_note))

        fema_note = row.get("fema_match_notes")
        if fema_note is not None and fema_note != "":
            reasons.append(str(fema_note))

        notes.append("; ".join([str(reason) for reason in reasons if reason is not None]))
    df["climate_risk_notes"] = notes
    return df

EXPECTED_ASSIGNMENTS = [
    ("los angeles", "CA", "06037"),
    ("chicago", "IL", "17031"),
    ("houston", "TX", "48201"),
    ("phoenix", "AZ", "04013"),
    ("philadelphia", "PA", "42101"),
    ("san antonio", "TX", "48029"),
    ("san diego", "CA", "06073"),
    ("dallas", "TX", "48113"),
    ("san francisco", "CA", "06075"),
    ("portland", "OR", "41051"),
]


def validate_assignments(df):
    errors = []
    for city, state, expected_fips in EXPECTED_ASSIGNMENTS:
        match = df[(df["city_normalized"] == city) & (df["state_abbr"] == state)]
        if match.empty:
            errors.append(f"Validation row missing for {city.title()}, {state}")
            continue
        actual = match.iloc[0]["climate_risk_county_fips"]
        if actual != expected_fips:
            errors.append(
                f"Expected {city.title()}, {state} -> {expected_fips}, got {actual}"
            )
    ny = df[(df["city_normalized"] == "new york") & (df["state_abbr"] == "NY")]
    if not ny.empty:
        bad_nj = ny[ny["climate_risk_county_fips"].astype(str).str.startswith("34", na=False)]
        if not bad_nj.empty:
            errors.append("New York, NY mapped to a New Jersey county")
    assigned = df[df["climate_risk_county_fips"].astype(bool)].copy()
    for _, row in assigned.iterrows():
        expected_statefp = STATE_ABBR_TO_FIPS.get(row["state_abbr"])
        actual_fips = str(row["climate_risk_county_fips"])
        if not actual_fips.startswith(expected_statefp):
            errors.append(
                f"County FIPS {actual_fips} state prefix mismatch for {row['__city_original']}, {row['state_abbr']}"
            )
        share = row.get("county_area_share")
        if pd.isna(share) or share <= 0 or share < 0.5:
            errors.append(
                f"Assigned county share invalid ({share}) for {row['__city_original']}, {row['state_abbr']}"
            )
        if pd.isna(row.get("climate_risk_score")):
            errors.append(
                f"Missing climate_risk_score for assigned county {actual_fips} on {row['__city_original']}, {row['state_abbr']}"
            )
    if errors:
        print("Validation failed:")
        for err in errors:
            print("-", err)
        sys.exit(1)


def write_output_workbook(original_path, new_df):
    sheets = pd.read_excel(original_path, sheet_name=None, engine="openpyxl")
    sheets["Clean Cities 100k+"] = new_df
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        for name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=name, index=False)
    wb = load_workbook(OUTPUT_FILE)
    if "README" in wb.sheetnames:
        sheet = wb["README"]
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=README_CLIMATE_NOTE)
    if "Clean Cities 100k+" in wb.sheetnames:
        sheet = wb["Clean Cities 100k+"]
        sheet.freeze_panes = sheet["A2"]
        max_row = sheet.max_row
        max_col = sheet.max_column
        sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        for idx, column_cells in enumerate(sheet.iter_cols(min_row=2, max_row=max_row, min_col=1, max_col=max_col), start=1):
            col_letter = get_column_letter(idx)
            if sheet.cell(row=1, column=idx).value in ["climate_risk_score"]:
                for cell in column_cells:
                    if cell.value is not None:
                        cell.number_format = "0.00"
            if sheet.cell(row=1, column=idx).value in ["county_area_share"]:
                for cell in column_cells:
                    if cell.value is not None:
                        cell.number_format = "0.00%"
        for idx in range(1, max_col + 1):
            letter = get_column_letter(idx)
            values = [sheet.cell(row=r, column=idx).value for r in range(1, max_row + 1)]
            max_length = max(
                [len(str(v)) if v is not None else 0 for v in values] + [len(sheet.cell(row=1, column=idx).value or "")]
            )
            sheet.column_dimensions[letter].width = min(max_length + 2, 50)
    wb.save(OUTPUT_FILE)


def print_diagnostics(df, unmatched_city_rows, unmatched_fema_fips=None, place_match_count=None, matched_places=None, total_workbook_valid=None):
    total_rows = len(df)
    if total_workbook_valid is None:
        total_workbook_valid = df['match_key'].notna().sum() if 'match_key' in df.columns else 0
    if matched_places is None:
        matched_places = df['place_geometry'].notna().sum() if 'place_geometry' in df.columns else 0
    assigned_counties = df['climate_risk_county_fips'].astype(bool).sum() if 'climate_risk_county_fips' in df.columns else 0
    matched_fema = df['climate_risk_score'].notna().sum() if 'climate_risk_score' in df.columns else 0
    flagged = df['climate_risk_review_flag'].sum() if 'climate_risk_review_flag' in df.columns else 0
    multi_county = df['city_crosses_multiple_counties'].sum() if 'city_crosses_multiple_counties' in df.columns else 0
    match_rate = (matched_places / total_workbook_valid * 100) if total_workbook_valid > 0 else 0.0

    print(f"Output file: {OUTPUT_FILE}")
    print(f"Total rows: {total_rows}")
    print(f"Rows matched to Census PLACE: {matched_places}")
    print(f"Rows assigned to county: {assigned_counties}")
    print(f"Rows matched to FEMA NRI: {matched_fema}")
    print(f"Rows flagged for review: {flagged}")
    print(f"Multi-county city count: {multi_county}")
    print(f"Match rate: {match_rate:.2f}%")
    if place_match_count is not None:
        print(f"Total Census PLACE rows: {place_match_count}")

    if matched_fema > 0:
        ranked = df[df['climate_risk_score'].notna()].sort_values('climate_risk_score', ascending=False)
        print("\nTop 10 highest climate_risk_score cities:")
        print(ranked[['__city_original', 'state_abbr', 'climate_risk_county_name', 'climate_risk_county_fips', 'climate_risk_score']].head(10).to_string(index=False))
        low = ranked.sort_values('climate_risk_score', ascending=True)
        print("\nTop 10 lowest climate_risk_score cities:")
        print(low[['__city_original', 'state_abbr', 'climate_risk_county_name', 'climate_risk_county_fips', 'climate_risk_score']].head(10).to_string(index=False))

    if 'place_geometry' in df.columns:
        unmatched = df[df['place_geometry'].isna() & df['match_key'].notna()].copy()
        unmatched = unmatched.head(20)
        if not unmatched.empty:
            print("\nFirst 20 unmatched workbook rows:")
            print(unmatched[['__city_original','city_normalized','__state_original','state_abbr','statefp','match_key']].to_string(index=False))
        else:
            print("\nNo unmatched workbook rows with valid city+state.")
    else:
        print("\nNo unmatched workbook rows with valid city+state (place_geometry removed before final diagnostics).")

    if unmatched_city_rows:
        print("\nRows missing valid city/state:")
        print(pd.DataFrame(unmatched_city_rows).to_string(index=False))
    if unmatched_fema_fips:
        print("\nUnmatched county FIPS rows:")
        print(" ".join(unmatched_fema_fips))


def main():
    ensure_directories()
    df, city_col, state_col = load_city_workbook()
    needed_states = sorted(df["state_abbr"].dropna().unique())
    place_gdf = None
    county_gdf = None
    used_year = None
    for year in TIGER_YEARS:
        try:
            place_gdf = load_tiger_place_gdf(year, needed_states)
            county_gdf = load_tiger_county_gdf(year)
            used_year = year
            break
        except SystemExit:
            sys.exit(1)
        except Exception as exc:
            print(f"TIGER {year} load failed: {exc}")
            continue
    if place_gdf is None or county_gdf is None:
        print("Unable to load TIGER place or county shapefiles for any supported year.")
        sys.exit(1)

    place_gdf = prepare_place_match_table(place_gdf)
    place_match_count = len(place_gdf)
    df = match_city_to_place(df, place_gdf)

    unmatched_city_rows = []
    for _, row in df[df['match_key'].isna()].iterrows():
        unmatched_city_rows.append(
            {"city": row["__city_original"], "state": row["__state_original"], "reason": "Invalid city/state"}
        )

    print("PLACE matching diagnostics:")
    matched_places = df['place_geometry'].notna().sum() if 'place_geometry' in df.columns else 0
    total_workbook_valid = df['match_key'].notna().sum() if 'match_key' in df.columns else 0
    print_diagnostics(df, unmatched_city_rows, place_match_count=place_match_count, matched_places=matched_places, total_workbook_valid=total_workbook_valid)

    county_gdf = prepare_county_gdf(county_gdf)
    df = compute_county_assignments(df, county_gdf)

    fema_df = fetch_fema_nri()
    df, unmatched_fema_fips = assign_fema_scores(df, fema_df)
    df = build_review_flags(df)

    unmatched_city_rows = []
    for _, row in df[df['place_geometry'].isna()].iterrows():
        unmatched_city_rows.append(
            {"city": row["__city_original"], "state": row["__state_original"], "reason": "Place match failed"}
        )

    validate_assignments(df)

    drop_columns = [
        c for c in CLIMATE_COLUMNS_TO_REMOVE if c in df.columns
    ] + ["__city_original", "__state_original"]
    output_columns = [c for c in df.columns if c not in drop_columns]
    df = df[output_columns]

    write_output_workbook(INPUT_FILE, df)
    print_diagnostics(df, unmatched_city_rows, unmatched_fema_fips=unmatched_fema_fips, place_match_count=place_match_count)


if __name__ == "__main__":
    main()
