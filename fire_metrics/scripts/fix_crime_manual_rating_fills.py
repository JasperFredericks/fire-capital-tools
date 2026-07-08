#!/usr/bin/env python3
"""Apply targeted manual crime rating fills for unresolved non-Florida rows.

This script updates only the requested crime fields and README methodology notes,
without repulling data or rebuilding other workbook content.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

BASE_DIR = Path(__file__).resolve().parent
INPUT_WB = BASE_DIR / "output" / "us_cities_100k_population_ranked_CRIME_NON_FL_FIXED.xlsx"
OUTPUT_WB = BASE_DIR / "output" / "us_cities_100k_population_ranked_CRIME_MANUAL_RATING_FILLS.xlsx"

HIGHLIGHT_FILL = PatternFill(fill_type="solid", start_color="FFF8E59A", end_color="FFF8E59A")

MANUAL_FILLS = {
    ("New Orleans", "LA"): {
        "rating": "Very High",
        "notes": "Manual Very High crime rating fill. FBI agency match was not reliable enough for calculated FBI crime score; numeric FBI fields left blank pending manual confirmation.",
    },
    ("Baton Rouge", "LA"): {
        "rating": "Very High",
        "notes": "Manual Very High crime rating fill. FBI agency match was not reliable enough for calculated FBI crime score; numeric FBI fields left blank pending manual confirmation.",
    },
    ("Jackson", "MS"): {
        "rating": "Very High",
        "notes": "Manual Very High crime rating fill. FBI agency match was not reliable enough for calculated FBI crime score; numeric FBI fields left blank pending manual confirmation.",
    },
    ("Atlanta", "GA"): {
        "rating": "High",
        "notes": "Manual High crime rating fill. FBI agency match was not reliable enough for calculated FBI crime score; numeric FBI fields left blank pending manual confirmation.",
    },
    ("Augusta-Richmond", "GA"): {
        "rating": "Elevated",
        "notes": "Manual Elevated crime rating fill. Consolidated city-county naming makes FBI agency matching ambiguous; numeric FBI fields left blank pending manual confirmation.",
    },
    ("Macon-Bibb", "GA"): {
        "rating": "High",
        "notes": "Manual High crime rating fill. Consolidated city-county naming makes FBI agency matching ambiguous; numeric FBI fields left blank pending manual confirmation.",
    },
    ("Downey", "CA"): {
        "rating": "Moderate",
        "notes": "Manual Moderate crime rating fill. FBI agency match was not reliable enough for calculated FBI crime score; numeric FBI fields left blank pending manual confirmation.",
    },
    ("Las Vegas", "NV"): {
        "rating": "Elevated",
        "notes": "Manual Elevated crime rating fill. Las Vegas Metropolitan Police Department appears to cover a broader jurisdiction than the Census city, so numeric FBI rates are not directly comparable without manual review.",
    },
}

PROVISIONAL_COLUMNS = {
    "Provisional Crime Rating",
    "Provisional Crime Bracket",
    "Provisional Crime Rating Source",
    "Provisional Crime Rating Notes",
}

NUMERIC_CRIME_COLUMNS_TO_CLEAR = [
    "FBI Population",
    "Coverage Rate",
    "Violent Crime Rate per 100k",
    "Property Crime Rate per 100k",
    "Crime Index Score",
]


def header_index(ws, header_row=1):
    return {
        str(ws.cell(header_row, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(header_row, c).value not in (None, "")
    }


def delete_columns_by_name(ws, columns):
    removed = 0
    while True:
        idx = header_index(ws, 1)
        found = [name for name in columns if name in idx]
        if not found:
            break
        # Delete right-to-left to keep indices stable.
        for name in sorted(found, key=lambda n: idx[n], reverse=True):
            ws.delete_cols(idx[name], 1)
            removed += 1
    return removed


def ensure_column(ws, name):
    idx = header_index(ws, 1)
    if name in idx:
        return idx[name]
    new_col = ws.max_column + 1
    ws.cell(1, new_col).value = name
    return new_col


def find_row(ws, city_col, state_col, city, state):
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, city_col).value or "").strip() == city and str(ws.cell(r, state_col).value or "").strip() == state:
            return r
    return None


def highlight_row(ws, row_num):
    for c in range(1, ws.max_column + 1):
        ws.cell(row_num, c).fill = HIGHLIGHT_FILL


def is_reliable_calculated_row(ws, row_num, idx):
    score = ws.cell(row_num, idx["Crime Index Score"]).value
    rating = ws.cell(row_num, idx["Crime Rating"]).value
    violent = ws.cell(row_num, idx["Violent Crime Rate per 100k"]).value
    prop = ws.cell(row_num, idx["Property Crime Rate per 100k"]).value
    return (score not in (None, "")) and (rating not in (None, "")) and (violent not in (None, "")) and (prop not in (None, ""))


def update_readme(ws):
    lines = [
        "Crime Index methodology update (manual rating fills)",
        "Crime Index uses FBI/UCR-style agency-level data where a reliable city-to-agency match exists. FBI data is agency-based, not clean Census-place-based, so some cities cannot be matched reliably without manual confirmation.",
        "Coverage Rate = FBI Population / Census City Population.",
        "Violent Crime Rate per 100k = Violent Crime Count / FBI Population × 100,000.",
        "Property Crime Rate per 100k = Property Crime Count / FBI Population × 100,000.",
        "Crime Rating bands: 0-20 Very Low, 21-40 Low, 41-60 Moderate, 61-75 Elevated, 76-90 High, 91-100 Very High.",
        "For rows where FBI matching is unreliable, the workbook may include a manual Crime Rating fill. These rows are highlighted, marked Manual Review = TRUE, and do not receive invented FBI population, crime rates, or Crime Index Score.",
    ]

    existing = [str(ws.cell(r, 1).value).strip() for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value not in (None, "")]
    if lines[0] in existing:
        return True

    start = ws.max_row + 1
    ws.cell(start, 1).value = ""
    for i, line in enumerate(lines, start=start + 1):
        ws.cell(i, 1).value = line
    return True


def main():
    if not INPUT_WB.exists():
        raise FileNotFoundError(f"Input workbook not found: {INPUT_WB}")

    wb = load_workbook(INPUT_WB)

    if "Crime Index" not in wb.sheetnames:
        raise RuntimeError("Crime Index sheet missing")

    crime_ws = wb["Crime Index"]
    removed_cols = delete_columns_by_name(crime_ws, PROVISIONAL_COLUMNS)

    idx = header_index(crime_ws, 1)
    required = [
        "City", "State", "Crime Rating", "Manual Review", "FBI Match Method", "FBI Match Notes",
        "FBI Population", "Coverage Rate", "Violent Crime Rate per 100k", "Property Crime Rate per 100k",
        "Crime Index Score",
    ]
    missing = [c for c in required if c not in idx]
    if missing:
        raise RuntimeError(f"Missing required columns in Crime Index: {missing}")

    crime_rating_source_col = ensure_column(crime_ws, "Crime Rating Source")

    idx = header_index(crime_ws, 1)
    notes_col = idx["FBI Match Notes"]

    updated_cities = []
    highlight_count = 0

    for (city, state), payload in MANUAL_FILLS.items():
        row = find_row(crime_ws, idx["City"], idx["State"], city, state)
        if row is None:
            continue

        if (city, state) == ("Las Vegas", "NV") and is_reliable_calculated_row(crime_ws, row, idx):
            continue

        for col_name in NUMERIC_CRIME_COLUMNS_TO_CLEAR:
            crime_ws.cell(row, idx[col_name]).value = None

        crime_ws.cell(row, idx["Crime Rating"]).value = payload["rating"]
        crime_ws.cell(row, idx["Manual Review"]).value = "TRUE"
        crime_ws.cell(row, idx["FBI Match Method"]).value = "manual rating fill"
        crime_ws.cell(row, idx["FBI Match Notes"]).value = payload["notes"]
        crime_ws.cell(row, crime_rating_source_col).value = "Manual rating fill"

        highlight_row(crime_ws, row)
        highlight_count += 1
        updated_cities.append((city, state))

    # Mark source for calculated rows.
    for r in range(2, crime_ws.max_row + 1):
        if crime_ws.cell(r, crime_rating_source_col).value == "Manual rating fill":
            continue
        if is_reliable_calculated_row(crime_ws, r, idx):
            crime_ws.cell(r, crime_rating_source_col).value = "Calculated FBI agency match"

    # Notes formatting: wrap and widen column.
    for r in range(1, crime_ws.max_row + 1):
        crime_ws.cell(r, notes_col).alignment = Alignment(wrap_text=True, vertical="top")
    letter = crime_ws.cell(1, notes_col).column_letter
    current_width = crime_ws.column_dimensions[letter].width or 8.43
    crime_ws.column_dimensions[letter].width = max(current_width, 95)

    # Highlight corresponding rows on Clean Cities 100k+ only if crime columns exist.
    clean_highlighted = 0
    if "Clean Cities 100k+" in wb.sheetnames:
        clean_ws = wb["Clean Cities 100k+"]
        clean_headers = header_index(clean_ws, 2)
        has_crime_cols = any("crime" in h.lower() for h in clean_headers.keys())
        if has_crime_cols and "city" in clean_headers and "state_abbr" in clean_headers:
            for city, state in updated_cities:
                row = find_row(clean_ws, clean_headers["city"], clean_headers["state_abbr"], city, state)
                if row is not None:
                    highlight_row(clean_ws, row)
                    clean_highlighted += 1
                    highlight_count += 1

    # README update.
    readme_updated = False
    if "README" in wb.sheetnames:
        readme_updated = update_readme(wb["README"])

    OUTPUT_WB.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_WB)

    print(f"Number of manual rating fills added: {len(updated_cities)}")
    print("List of manual-fill cities updated:")
    if updated_cities:
        for city, state in updated_cities:
            print(f"- {city}, {state}")
    else:
        print("- none")
    print(f"Number of rows highlighted: {highlight_count}")
    print(f"README crime methodology updated: {'yes' if readme_updated else 'no'}")
    print(f"Removed provisional columns: {removed_cols}")
    print(f"Output file path: {OUTPUT_WB}")


if __name__ == "__main__":
    main()
