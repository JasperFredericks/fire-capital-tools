#!/usr/bin/env python3
"""Formatting and documentation cleanup for the final pre-crime-fix workbook.

This script performs workbook-only cleanup:
- City display name normalization in visible city columns
- Section readability formatting for Clean Cities 100k+
- README sheet rewrite/expansion

No metrics are recalculated and no external data is fetched.
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
INPUT_PATH = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_JOB_GROWTH_FIXED.xlsx"
OUTPUT_PATH = BASE_DIR / "output" / "us_cities_100k_population_ranked_FORMATTED_FINAL_BEFORE_CRIME_FIX.xlsx"

CITY_HEADERS = {"city", "census city", "city_normalized"}

SECTION_COLORS = {
    "City / Location": "DCE6F1",
    "Population": "E2F0D9",
    "Landlord Friendliness": "FCE4D6",
    "Climate Risk": "FFF2CC",
    "Crime Index": "F8CBAD",
    "Median Household Income": "E4DFEC",
    "Median Owner-Occupied Home Value": "D9EAD3",
    "Resident Employment / Job Growth": "D0E0E3",
    "Other": "F2F2F2",
}

THIN = Side(style="thin", color="BFBFBF")
THICK = Side(style="medium", color="7F7F7F")


def clean_city_display_name(value):
    if value is None:
        return value
    text = str(value).strip()
    if not text:
        return text

    suffixes = [
        " metropolitan government (balance)",
        " metro government (balance)",
        " consolidated government (balance)",
        " unified government (balance)",
        " urban county",
        " municipality",
        " Municipality",
        " county",
        " County",
        " city",
        " town",
        " village",
        " CDP",
        " cdp",
    ]

    changed = True
    while changed:
        changed = False
        for suffix in suffixes:
            if text.endswith(suffix):
                candidate = text[: -len(suffix)].rstrip(" ,-/")
                if candidate:
                    text = candidate
                    changed = True
    return text


def classify_section(header):
    h = (header or "").strip().lower()

    if h in {"rank", "rank_2025", "city", "state", "state_abbr", "place_type", "census_place_type"}:
        return "City / Location"

    if (
        "population" in h
        or "absolute_change" in h
        or "percent_change" in h
        or h.startswith("source_")
        or h in {"source_name", "source_year", "source_file", "source_note"}
    ):
        return "Population"

    if "landlord" in h:
        return "Landlord Friendliness"

    if "climate_risk" in h:
        return "Climate Risk"

    crime_terms = [
        "fbi population",
        "coverage rate",
        "violent crime",
        "property crime",
        "crime index",
        "crime rating",
        "manual review",
        "fbi agency",
        "fbi ori",
        "fbi match",
    ]
    if any(t in h for t in crime_terms):
        return "Crime Index"

    if "median household income" in h:
        return "Median Household Income"

    if "median home/condo value" in h or "median owner-occupied home value" in h:
        return "Median Owner-Occupied Home Value"

    if "resident employment" in h or "bls laus" in h or h.startswith("employment growth"):
        return "Resident Employment / Job Growth"

    return "Other"


def set_number_formats(ws, header_row, start_data_row):
    max_col = ws.max_column
    max_row = ws.max_row
    for c in range(1, max_col + 1):
        h = ws.cell(header_row, c).value
        hs = (str(h).strip().lower() if h is not None else "")

        number_fmt = None

        if "fips" in hs or "series id" in hs or hs.endswith(" ori") or " geoid" in hs:
            number_fmt = "@"
        elif "median household income in" in hs or "median home/condo value in" in hs or "median owner-occupied home value in" in hs:
            number_fmt = "$#,##0"
        elif "growth" in hs or "percent_change" in hs or "(%)" in hs:
            number_fmt = "0.00%"
        elif (
            "population" in hs
            or hs.startswith("absolute_change")
            or hs.startswith("resident employment in")
            or hs in {"fbi population"}
            or "rate per 100k" in hs
            or hs.endswith("score")
        ):
            number_fmt = "#,##0"

        if number_fmt is None:
            continue

        for r in range(start_data_row, max_row + 1):
            ws.cell(r, c).number_format = number_fmt


def autofit_columns(ws, max_width=64):
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        best = 10
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            ln = len(str(v))
            if ln > best:
                best = ln
        ws.column_dimensions[letter].width = min(max(best + 2, 10), max_width)


def apply_section_formatting_clean_cities(ws):
    # Insert section row above existing headers.
    ws.insert_rows(1)
    header_row = 2
    section_row = 1
    data_start_row = 3

    max_col = ws.max_column

    sections = []
    for c in range(1, max_col + 1):
        h = ws.cell(header_row, c).value
        sections.append(classify_section(h))

    # Merge contiguous section ranges.
    start = 1
    while start <= max_col:
        sec = sections[start - 1]
        end = start
        while end < max_col and sections[end] == sec:
            end += 1

        ws.cell(section_row, start, sec)
        if end > start:
            ws.merge_cells(start_row=section_row, start_column=start, end_row=section_row, end_column=end)

        fill = PatternFill(fill_type="solid", fgColor=SECTION_COLORS.get(sec, SECTION_COLORS["Other"]))
        for c in range(start, end + 1):
            ws.cell(section_row, c).fill = fill
            ws.cell(header_row, c).fill = fill

            left = THICK if c == start else THIN
            ws.cell(section_row, c).border = Border(left=left, right=THIN, top=THIN, bottom=THIN)
            ws.cell(header_row, c).border = Border(left=left, right=THIN, top=THIN, bottom=THIN)

        ws.cell(section_row, start).font = Font(bold=True, color="1F1F1F")
        ws.cell(section_row, start).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        start = end + 1

    # Header style.
    for c in range(1, max_col + 1):
        cell = ws.cell(header_row, c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[section_row].height = 22
    ws.row_dimensions[header_row].height = 36

    # Keep filter on actual data header row.
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{ws.max_row}"

    # Freeze panes so section+headers stay visible.
    ws.freeze_panes = "A3"

    set_number_formats(ws, header_row=header_row, start_data_row=data_start_row)
    autofit_columns(ws)


def rewrite_readme_sheet(ws):
    ws.delete_rows(1, ws.max_row)

    lines = [
        "Workbook Purpose",
        "This workbook is a city-level real estate market screening tool for U.S. cities/places with population of 100,000+.",
        "",
        "Geography Note",
        "This workbook is city/place level. Some metrics are city/place level, while some are assigned from county-level or agency-level data when no clean city-level equivalent exists.",
        "",
        "Population",
        "Source: U.S. Census Bureau City and Town Population Totals, Vintage 2025.",
        "File: Annual Estimates of the Resident Population for Incorporated Places of 20,000 or More, Ranked by July 1, 2025 Population.",
        "Methodology: Population is Census Population Estimates Program city/place population. The workbook filters to places with 2025 population of 100,000+. Population growth is calculated as latest population / base population - 1. Includes 2020-2025 and 2023-2025 growth where available.",
        "",
        "Landlord Friendliness",
        "Source: Michelle updated landlord-friendliness list / internal scoring framework.",
        "Methodology: State-level qualitative score where +1 is more landlord-friendly, 0 is neutral/moderate, and -1 is more tenant-friendly.",
        "Note: Ohio is county-dependent and may be temporarily set to 0 for state-level screening. This is a subjective/manual framework, not an official government dataset.",
        "",
        "Climate Risk",
        "Source: FEMA National Risk Index Counties.",
        "Methodology: Cities are assigned a county-level FEMA risk score based on matched county. FEMA NRI is not purely physical climate/weather exposure; it also incorporates expected annual loss, social vulnerability, and community resilience.",
        "Columns: climate_risk_county, climate_risk_county_fips, climate_risk_score, climate_risk_rating.",
        "Interpretation: Higher FEMA risk score means higher natural hazard risk.",
        "",
        "Crime Index",
        "Source: FBI crime/UCR-style agency data, with Census city population used for coverage comparison.",
        "Methodology: FBI crime data is agency-based, not perfectly city/place-based. Workflow matches each city to a relevant law-enforcement agency and calculates violent/property crime rates per 100k where reliable.",
        "Coverage Rate: FBI Population / Census City Population.",
        "Note: Missing/uncertain matches or poor coverage should be flagged for manual review. Florida and consolidated-government cities may need additional manual agency matching.",
        "",
        "Median Household Income",
        "Source: U.S. Census ACS 1-year API.",
        "Variable: B19013_001E (Median household income).",
        "Geography: Place/city level.",
        "Years: 2021, latest available ACS 1-year, and the year immediately before latest.",
        "Methodology: Income growth 2021-latest = latest income / 2021 income - 1. Income growth prior-latest = latest income / prior-year income - 1.",
        "Note: 2020 ACS 1-year is skipped because standard 2020 ACS 1-year data is not comparable. Values are nominal, not inflation-adjusted.",
        "",
        "Median Owner-Occupied Home Value",
        "Source: U.S. Census ACS 1-year API.",
        "Variable: B25077_001E (Median value in dollars for owner-occupied housing units).",
        "Geography: Place/city level.",
        "Years: 2021, latest available ACS 1-year, and the year immediately before latest.",
        "Methodology: Home value growth 2021-latest = latest value / 2021 value - 1. Home value growth prior-latest = latest value / prior-year value - 1.",
        "Note: This is ACS median owner-occupied housing value, not Zillow market value or sale price.",
        "",
        "Resident Employment / Job Growth",
        "Source: U.S. Bureau of Labor Statistics Local Area Unemployment Statistics (LAUS) bulk files.",
        "Files: la.area, la.series, la.data.65.City.",
        "Metric: Employment (measure code 05).",
        "Methodology: Uses annual average resident employment, not payroll jobs located in the city. Employment growth is latest employment / base employment - 1 using 2021, latest full annual year, and prior year.",
        "Note: Label reflects resident employment growth (BLS LAUS), not job openings.",
        "",
        "General Notes",
        "Percent growth columns are nominal percentage changes.",
        "Some fields use different geographies because consistent city-level data is not available for every metric.",
        "Review flags indicate rows that should be manually checked before final ranking decisions.",
        "Sources should remain consistent metric-by-metric in future updates.",
    ]

    row = 1
    for line in lines:
        ws.cell(row, 1, line)
        if line and line in {
            "Workbook Purpose",
            "Geography Note",
            "Population",
            "Landlord Friendliness",
            "Climate Risk",
            "Crime Index",
            "Median Household Income",
            "Median Owner-Occupied Home Value",
            "Resident Employment / Job Growth",
            "General Notes",
        }:
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 1).fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
        row += 1

    ws.column_dimensions["A"].width = 150
    for r in range(1, ws.max_row + 1):
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def main():
    errors = []
    skipped = []
    updated_sheets = []
    city_names_cleaned = 0
    readme_updated = False
    section_format_applied = False

    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Input workbook not found: {INPUT_PATH}")

    wb = load_workbook(INPUT_PATH)

    # Clean city display names in every sheet with known city display headers.
    for ws in wb.worksheets:
        try:
            max_col = ws.max_column
            if ws.max_row < 2:
                continue
            headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else "" for c in range(1, max_col + 1)]
            target_cols = [i + 1 for i, h in enumerate(headers) if h.lower() in CITY_HEADERS]
            if not target_cols:
                continue

            changed_this_sheet = 0
            for c in target_cols:
                for r in range(2, ws.max_row + 1):
                    old = ws.cell(r, c).value
                    new = clean_city_display_name(old)
                    if new != old:
                        ws.cell(r, c).value = new
                        changed_this_sheet += 1

            if changed_this_sheet > 0:
                city_names_cleaned += changed_this_sheet
                if ws.title not in updated_sheets:
                    updated_sheets.append(ws.title)
        except Exception as exc:
            errors.append(f"{ws.title}: {exc}")

    # Section formatting for Clean Cities sheet.
    if "Clean Cities 100k+" in wb.sheetnames:
        try:
            ws = wb["Clean Cities 100k+"]
            apply_section_formatting_clean_cities(ws)
            section_format_applied = True
            if ws.title not in updated_sheets:
                updated_sheets.append(ws.title)
        except Exception as exc:
            errors.append(f"Clean Cities 100k+ formatting: {exc}")
    else:
        skipped.append("Clean Cities 100k+ (sheet not found)")

    # README rewrite.
    if "README" in wb.sheetnames:
        try:
            rewrite_readme_sheet(wb["README"])
            readme_updated = True
            if "README" not in updated_sheets:
                updated_sheets.append("README")
        except Exception as exc:
            errors.append(f"README rewrite: {exc}")
    else:
        skipped.append("README (sheet not found)")

    wb.save(OUTPUT_PATH)

    print(f"Output file path: {OUTPUT_PATH}")
    print("Sheets updated:")
    if updated_sheets:
        for name in updated_sheets:
            print(f"- {name}")
    else:
        print("- none")
    print(f"Number of city names cleaned: {city_names_cleaned}")
    print(f"README updated: {'yes' if readme_updated else 'no'}")
    print(f"Section formatting applied to Clean Cities 100k+: {'yes' if section_format_applied else 'no'}")
    print("Any errors or skipped sheets:")
    if errors:
        for e in errors:
            print(f"- ERROR: {e}")
    if skipped:
        for s in skipped:
            print(f"- SKIPPED: {s}")
    if not errors and not skipped:
        print("- none")


if __name__ == "__main__":
    main()
