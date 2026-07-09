#!/usr/bin/env python3
"""Add BLS LAUS resident employment growth columns to Clean Cities 100k+.

Live data source: the BLS Time Series API (api.bls.gov), replacing the
former manual BLS LAUS bulk-file dependency (la.area, la.series,
la.data.65.City). City-level LAUS series IDs are constructed directly from
each city's Census state+place FIPS code (via a Census ACS crosswalk, the
same mechanism used by add_income_growth.py/add_home_value_growth.py) rather
than by fuzzy-matching city names against a bulk-file text catalog.

Measure code used: 05 (employment)
Period used: M13 (annual average, via the API's annualaverage=true option)

A small number of consolidated city-county governments (Nashville-Davidson,
Louisville-Jefferson, Athens-Clarke, Augusta-Richmond, Urban Honolulu) do not
have their own city-level ("CT") LAUS series -- BLS only publishes an
encompassing county-level ("CN") series for these. See COUNTY_FALLBACK
below; each entry was verified individually against the live API.
"""

import datetime as dt
import re
import shutil
import sys
import json
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from fire_metrics_updater.config import get_secret


BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_WORKBOOK = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_HOME_VALUE_GROWTH.xlsx"
OUTPUT_WORKBOOK = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_JOB_GROWTH_FIXED.xlsx"
PREVIOUS_WORKBOOK = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_JOB_GROWTH.xlsx"

# Census key is required (used for the city -> state/place FIPS crosswalk).
# BLS key is optional: unauthenticated access works (verified), a
# registered key just raises the daily query / per-request series limits.
CENSUS_API_KEY = get_secret("CENSUS_API_KEY", "data/cache/census_api_key.txt")
BLS_API_KEY = get_secret("BLS_API_KEY", "data/cache/bls_api_key.txt")

BLS_MEASURE_EMPLOYMENT = "05"
BLS_SERIES_ENDPOINT = "https://api.bls.gov/publicAPI/v2/timeseries/data/"

STATE_TO_ABBR = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR", "CALIFORNIA": "CA",
    "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE", "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID", "ILLINOIS": "IL",
    "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA",
    "MAINE": "ME", "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
    "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK", "OREGON": "OR",
    "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA",
    "WASHINGTON": "WA", "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
}

PUNCT = re.compile(r"[^\w\s]", re.U)
MULTI = re.compile(r"\s+")
TRAILING_DIGITS = re.compile(r"\d+$")
REMOVE_WORDS = {
    "city",
    "town",
    "village",
    "borough",
    "municipality",
    "county",
    "cdp",
    "urban",
    "balance",
    "metro",
    "metropolitan",
    "unified",
    "government",
    "consolidated",
    "county/city",
}

# Historical aliases from the old bulk-file text-matching approach. No
# longer read by this script (there is no bulk-file text catalog to match
# them against anymore -- see COUNTY_FALLBACK and fetch_place_fips_crosswalk
# below for how the same consolidated-government cities are now handled,
# via Census FIPS codes rather than BLS area-name text). Left in place,
# unedited, as Beckett's original domain reference.
ALIASES = {
    ("PA", "philadelphia"): [
        "philadelphia city",
        "philadelphia county city",
    ],
    ("CA", "san francisco"): [
        "san francisco city",
        "san francisco county city",
    ],
    ("CO", "denver"): [
        "denver city",
        "denver county city",
    ],
    ("HI", "urban honolulu"): [
        "honolulu cdp",
        "urban honolulu cdp",
        "honolulu city",
        "honolulu county",
    ],
    ("KY", "lexington fayette"): [
        "lexington fayette urban county",
        "lexington fayette",
        "lexington city",
        "fayette county",
    ],
    ("AK", "anchorage"): [
        "anchorage municipality",
        "anchorage",
        "anchorage borough municipality",
    ],
    ("GA", "columbus"): [
        "columbus city",
        "columbus",
        "muscogee county",
    ],
    ("GA", "augusta richmond"): [
        "augusta richmond county",
        "augusta",
        "richmond county",
    ],
    ("GA", "macon bibb"): [
        "macon bibb county",
        "macon",
        "bibb county",
    ],
    ("GA", "athens clarke"): [
        "athens clarke county",
        "athens",
        "clarke county",
    ],
    ("TN", "nashville davidson"): [
        "nashville davidson",
        "nashville",
        "davidson county",
    ],
    ("KY", "louisville jefferson"): [
        "louisville jefferson county",
        "louisville",
        "jefferson county",
    ],
    ("DC", "washington"): [
        "district of columbia",
        "washington city",
    ],
    ("IN", "indianapolis"): [
        "indianapolis consolidated",
        "indianapolis incorporated",
    ],
    ("CA", "san buenaventura ventura"): ["san buenaventura"],
}

# Consolidated city-county governments with no city-level ("CT") LAUS
# series of their own -- BLS only publishes the encompassing county's
# ("CN") series. Each entry verified individually against the live API
# (2026-07): the CT series for these returns no data, the CN series does.
# Of the 15 cities in the old ALIASES table above, only these 5 actually
# need this fallback -- the other 10 (Philadelphia, San Francisco, Denver,
# Lexington-Fayette, Anchorage, Columbus GA, Macon-Bibb, San Buenaventura,
# Washington DC, Indianapolis) have their own valid CT series directly.
COUNTY_FALLBACK = {
    # Key is base_city_key(normalize_city(...)) of the workbook's own city
    # text -- "urban" is a REMOVE_WORDS stopword, so "Urban Honolulu CDP"
    # reduces to "honolulu", not "urban honolulu".
    ("HI", "honolulu"): ("15", "003"),            # Honolulu County, HI
    ("GA", "augusta richmond"): ("13", "245"),    # Richmond County, GA
    ("GA", "athens clarke"): ("13", "059"),       # Clarke County, GA
    ("TN", "nashville davidson"): ("47", "037"),  # Davidson County, TN
    ("KY", "louisville jefferson"): ("21", "111"),  # Jefferson County, KY
}


def normalize_city(value):
    text = "" if value is None else str(value).strip().lower()
    text = PUNCT.sub(" ", text)
    text = TRAILING_DIGITS.sub("", text).strip()
    text = MULTI.sub(" ", text)
    parts = [p for p in text.split(" ") if p]
    while parts and parts[-1] in REMOVE_WORDS:
        parts.pop()
    return " ".join(parts)


def base_city_key(normalized_city):
    parts = [p for p in normalized_city.split(" ") if p and p not in REMOVE_WORDS]
    return " ".join(parts)


def normalize_state_abbr(value):
    raw = "" if value is None else str(value).strip().upper()
    raw = TRAILING_DIGITS.sub("", raw).strip()
    if len(raw) == 2 and raw.isalpha():
        return raw
    return STATE_TO_ABBR.get(raw, "")


def discover_crosswalk_year(api_key):
    current = dt.date.today().year
    for year in range(current - 1, 2020, -1):
        try:
            params = {"get": "NAME", "for": "place:*", "in": "state:01", "key": api_key}
            query = urllib.parse.urlencode(params)
            url = f"https://api.census.gov/data/{year}/acs/acs1?{query}"
            with urllib.request.urlopen(url, timeout=30) as resp:
                body = resp.read().decode("utf-8")
            if body.strip().startswith("[") and len(json.loads(body)) > 1:
                return year
        except Exception:
            continue
    raise RuntimeError("Unable to find a usable ACS 1-year place list for the BLS crosswalk")


def fetch_place_fips_crosswalk(api_key, year):
    """Census ACS place-level (state_abbr, normalized city) -> (state_fips, place_fips).

    Reuses the same ACS 1-year 'place:*' query pattern as
    add_income_growth.py/add_home_value_growth.py -- the population
    variable here is just a vehicle to get the state/place geography
    columns back; its value is never used.
    """
    params = {"get": "NAME,B01001_001E", "for": "place:*", "in": "state:*", "key": api_key}
    query = urllib.parse.urlencode(params)
    url = f"https://api.census.gov/data/{year}/acs/acs1?{query}"
    with urllib.request.urlopen(url, timeout=60) as resp:
        body = resp.read().decode("utf-8")
    if not body.strip().startswith("["):
        raise RuntimeError(f"Census API error for {year}: {body[:200]}")
    rows = json.loads(body)
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}

    crosswalk = {}
    for row in rows[1:]:
        name = row[idx["NAME"]]
        if "," not in name:
            continue
        place_part, state_name = [x.strip() for x in name.rsplit(",", 1)]
        state_abbr = STATE_TO_ABBR.get(state_name.upper(), "")
        if not state_abbr:
            continue
        key = (state_abbr, base_city_key(normalize_city(place_part)))
        crosswalk.setdefault(key, (row[idx["state"]], row[idx["place"]]))
    return crosswalk


def build_series_id(area_type, state_fips, geo_fips, measure=BLS_MEASURE_EMPLOYMENT):
    geo_code = f"{state_fips}{geo_fips}".ljust(13, "0")
    return f"LAU{area_type}{geo_code}{measure}"


def fetch_bls_series_batch(series_ids, start_year, end_year, api_key=None):
    """POST batched BLS Time Series API requests (with annualaverage=true
    for M13 data), chunked to the per-request series limit. Returns
    {series_id: {year: value}} -- only for series that actually have data.
    """
    chunk_size = 50 if api_key else 25
    results = {}
    ids = list(series_ids)
    for i in range(0, len(ids), chunk_size):
        chunk = ids[i:i + chunk_size]
        payload = {
            "seriesid": chunk,
            "startyear": str(start_year),
            "endyear": str(end_year),
            "annualaverage": True,
        }
        if api_key:
            payload["registrationkey"] = api_key
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            BLS_SERIES_ENDPOINT,
            data=data,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = json.loads(resp.read().decode("utf-8"))
        if body.get("status") not in ("REQUEST_SUCCEEDED",):
            raise RuntimeError(f"BLS API request failed: {body.get('message')}")
        for series in body.get("Results", {}).get("series", []):
            sid = series["seriesID"]
            years = {}
            for point in series.get("data", []):
                if point.get("period") != "M13":
                    continue
                try:
                    years[int(point["year"])] = float(point["value"])
                except (TypeError, ValueError):
                    continue
            if years:
                results[sid] = years
    return results


def find_header_index(ws, names):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    norm = [str(h).strip().lower() if h is not None else "" for h in headers]
    for name in names:
        if name in norm:
            return norm.index(name) + 1
    return None


def remove_existing_job_columns(ws):
    prefixes = (
        "resident employment in ",
        "resident employment growth ",
        "bls laus area name",
        "bls laus series id",
    )
    to_delete = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        text = str(h).strip().lower() if h is not None else ""
        if any(text.startswith(p) for p in prefixes):
            to_delete.append(c)
    for c in reversed(to_delete):
        ws.delete_cols(c, 1)


def resolve_city_series_id(city, st, series_data, crosswalk):
    """Return (series_id, area_label) for a workbook row, or (None, None)."""
    city_norm = normalize_city(city)
    base_key = base_city_key(city_norm)

    fips = crosswalk.get((st, base_key))
    if fips:
        state_fips, place_fips = fips
        ct_id = build_series_id("CT", state_fips, place_fips)
        if ct_id in series_data:
            return ct_id, f"{city}, {st} (Census place {state_fips}{place_fips})"

    county = COUNTY_FALLBACK.get((st, base_key))
    if county:
        county_state_fips, county_fips = county
        cn_id = build_series_id("CN", county_state_fips, county_fips)
        if cn_id in series_data:
            return cn_id, f"{city}, {st} (encompassing county {county_state_fips}{county_fips})"

    return None, None


def append_job_columns(ws, prior_year, latest_year, series_data, crosswalk):
    city_col = find_header_index(ws, ["city"])
    state_col = find_header_index(ws, ["state_abbr", "state"])
    state_name_col = find_header_index(ws, ["state"])

    if city_col is None or state_col is None:
        raise RuntimeError("Could not find city/state columns in Clean Cities 100k+ sheet")

    remove_existing_job_columns(ws)

    start = ws.max_column + 1
    headers = [
        "Resident employment in 2021",
        f"Resident employment in {prior_year}",
        f"Resident employment in {latest_year}",
        f"Employment growth 2021-{latest_year} (%)",
        f"Employment growth {prior_year}-{latest_year} (%)",
        "BLS LAUS area name",
        "BLS LAUS series ID",
    ]
    for i, h in enumerate(headers):
        ws.cell(1, start + i, h)

    matched = 0
    unmatched = []
    sanity = []

    for r in range(2, ws.max_row + 1):
        city = ws.cell(r, city_col).value
        raw_state = ws.cell(r, state_col).value
        state_name = ws.cell(r, state_name_col).value if state_name_col else raw_state

        st = normalize_state_abbr(raw_state)
        if not st:
            st = normalize_state_abbr(state_name)

        c2021 = ws.cell(r, start)
        cprior = ws.cell(r, start + 1)
        clatest = ws.cell(r, start + 2)
        g1 = ws.cell(r, start + 3)
        g2 = ws.cell(r, start + 4)
        area_name = ws.cell(r, start + 5)
        series_id_cell = ws.cell(r, start + 6)

        series_id, area_label = resolve_city_series_id(city, st, series_data, crosswalk)
        if series_id is None:
            unmatched.append((city, st))
            continue

        years = series_data[series_id]
        v2021 = years.get(2021)
        vprior = years.get(prior_year)
        vlast = years.get(latest_year)

        c2021.value = v2021
        cprior.value = vprior
        clatest.value = vlast
        area_name.value = area_label
        series_id_cell.value = series_id

        if v2021 and vlast and v2021 != 0:
            g1.value = (vlast - v2021) / v2021
        else:
            g1.value = None

        if vprior and vlast and vprior != 0:
            g2.value = (vlast - vprior) / vprior
        else:
            g2.value = None

        c2021.number_format = "#,##0"
        cprior.number_format = "#,##0"
        clatest.number_format = "#,##0"
        g1.number_format = "0.00%"
        g2.number_format = "0.00%"

        matched += 1
        if len(sanity) < 5:
            sanity.append((city, st, v2021, vprior, vlast, g1.value, g2.value, area_label))

    return matched, unmatched, sanity


def get_matched_set(workbook_path):
    if not workbook_path.exists():
        return set()
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Clean Cities 100k+" not in wb.sheetnames:
        return set()
    ws = wb["Clean Cities 100k+"]

    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        city_i = headers.index("city")
    except ValueError:
        return set()

    state_i = headers.index("state_abbr") if "state_abbr" in headers else (headers.index("state") if "state" in headers else None)
    series_i = headers.index("bls laus series id") if "bls laus series id" in headers else None
    if state_i is None or series_i is None:
        return set()

    matched = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        city = row[city_i]
        state = row[state_i]
        sid = row[series_i]
        if city is None or state is None or sid in (None, ""):
            continue
        matched.add((str(city).strip(), normalize_state_abbr(state)))
    return matched


def append_readme_note(wb, prior_year, latest_year):
    if "README" not in wb.sheetnames:
        return

    ws = wb["README"]
    note = (
        "Resident employment uses the BLS Time Series API (api.bls.gov) for LAUS "
        "(Local Area Unemployment Statistics), measure code 05 (employment), annual "
        f"average period M13, for years 2021, {prior_year}, and {latest_year}. City-level "
        "series are resolved via each city's Census state/place FIPS code; a small number of "
        "consolidated city-county governments use the encompassing county's LAUS series "
        "instead, since BLS does not publish a separate city-level series for them."
    )

    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and note in str(v):
            return

    ws.append([""])
    ws.append([note])


def add_job_growth(input_path=None, output_path=None, previous_path=None):
    """Add BLS LAUS resident employment + growth columns to a copy of the workbook.

    Returns a summary dict identical in shape to what the CLI used to print.
    """
    input_workbook = Path(input_path) if input_path is not None else INPUT_WORKBOOK
    output_workbook = Path(output_path) if output_path is not None else OUTPUT_WORKBOOK
    previous_workbook = Path(previous_path) if previous_path is not None else PREVIOUS_WORKBOOK

    if not input_workbook.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_workbook}")
    if not CENSUS_API_KEY:
        raise RuntimeError(
            "Census API key not found (needed for the city -> FIPS crosswalk). "
            "Set CENSUS_API_KEY or create data/cache/census_api_key.txt"
        )

    previous_matched = get_matched_set(previous_workbook)

    crosswalk_year = discover_crosswalk_year(CENSUS_API_KEY)
    crosswalk = fetch_place_fips_crosswalk(CENSUS_API_KEY, crosswalk_year)

    wb_probe = load_workbook(input_workbook, read_only=True, data_only=True)
    if "Clean Cities 100k+" not in wb_probe.sheetnames:
        raise RuntimeError("Sheet 'Clean Cities 100k+' not found")
    ws_probe = wb_probe["Clean Cities 100k+"]
    city_col = find_header_index(ws_probe, ["city"])
    state_col = find_header_index(ws_probe, ["state_abbr", "state"])
    state_name_col = find_header_index(ws_probe, ["state"])
    if city_col is None or state_col is None:
        raise RuntimeError("Could not find city/state columns in Clean Cities 100k+ sheet")

    candidate_ids = set()
    for row in ws_probe.iter_rows(min_row=2, values_only=True):
        city = row[city_col - 1]
        raw_state = row[state_col - 1]
        state_name = row[state_name_col - 1] if state_name_col else raw_state
        st = normalize_state_abbr(raw_state) or normalize_state_abbr(state_name)
        base_key = base_city_key(normalize_city(city))

        fips = crosswalk.get((st, base_key))
        if fips:
            candidate_ids.add(build_series_id("CT", fips[0], fips[1]))
        county = COUNTY_FALLBACK.get((st, base_key))
        if county:
            candidate_ids.add(build_series_id("CN", county[0], county[1]))
    wb_probe.close()

    current_year = dt.date.today().year
    series_data = fetch_bls_series_batch(candidate_ids, 2021, current_year, api_key=BLS_API_KEY)

    years_seen = sorted({y for years in series_data.values() for y in years})
    if not years_seen:
        raise RuntimeError("BLS API returned no annual (M13) employment data for any requested series")
    latest_year = years_seen[-1]
    prior_year = latest_year - 1

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = input_workbook.with_name(f"{input_workbook.stem}_backup_{stamp}{input_workbook.suffix}")
    shutil.copy2(input_workbook, backup_path)

    shutil.copy2(input_workbook, output_workbook)
    wb = load_workbook(output_workbook)
    if "Clean Cities 100k+" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Clean Cities 100k+' not found")

    ws = wb["Clean Cities 100k+"]
    matched, unmatched, sanity = append_job_columns(ws, prior_year, latest_year, series_data, crosswalk)
    append_readme_note(wb, prior_year, latest_year)
    wb.save(output_workbook)

    current_matched = get_matched_set(output_workbook)
    newly_matched = sorted(current_matched - previous_matched)

    total_rows = ws.max_row - 1

    return {
        "output_path": str(output_workbook),
        "backup_path": str(backup_path),
        "prior_year": prior_year,
        "latest_year": latest_year,
        "total_rows": total_rows,
        "matched": matched,
        "unmatched": unmatched,
        "newly_matched": newly_matched,
        "sanity": sanity,
    }


def main():
    result = add_job_growth()

    prior_year = result["prior_year"]
    latest_year = result["latest_year"]
    total_rows = result["total_rows"]
    matched = result["matched"]
    unmatched = result["unmatched"]
    newly_matched = result["newly_matched"]
    failed = len(unmatched)
    match_rate = (matched / total_rows * 100.0) if total_rows else 0.0

    print(f"Output file path: {result['output_path']}")
    print(f"BLS years used: 2021, {prior_year}, {latest_year}")
    print(f"Total city rows: {total_rows}")
    print(f"Matched count: {matched}")
    print(f"Failed count: {failed}")
    print(f"Match rate: {match_rate:.2f}%")
    print(f"Newly matched cities compared with previous run: {len(newly_matched)}")
    for city, st in newly_matched:
        print(f"- {city}, {st}")
    print("Remaining unmatched cities:")
    for city, st in unmatched:
        print(f"- {city}, {st}")
    print("5 sanity-check rows:")
    for city, st, v1, vp, vl, g1, g2, area in result["sanity"]:
        print(
            f"- {city}, {st} | 2021: {v1} | {prior_year}: {vp} | {latest_year}: {vl} "
            f"| growth 2021-{latest_year}: {g1} | growth {prior_year}-{latest_year}: {g2} | area: {area}"
        )


if __name__ == "__main__":
    main()
