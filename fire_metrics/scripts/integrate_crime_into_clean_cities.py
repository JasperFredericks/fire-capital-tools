#!/usr/bin/env python3

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_FILE = BASE_DIR / "output" / "us_cities_100k_population_ranked_CRIME_MANUAL_RATING_FILLS.xlsx"
OUTPUT_FILE = BASE_DIR / "output" / "us_cities_100k_population_ranked_ALL_METRICS_CLEAN.xlsx"

CRIME_COLUMNS = [
    "Crime Index Score",
    "Crime Rating",
    "Density-Adjusted Crime Score",
    "Density-Adjusted Crime Rating",
    "Manual Review",
    "Crime Rating Source",
]

CRIME_SECTION_FILL = PatternFill(fill_type="solid", start_color="F8CBAD", end_color="F8CBAD")
MANUAL_CELL_FILL = PatternFill(fill_type="solid", start_color="FFF8E59A", end_color="FFF8E59A")

STATE_TO_ABBR = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR", "CALIFORNIA": "CA",
    "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE", "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID", "ILLINOIS": "IL",
    "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA",
    "MAINE": "ME", "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
    "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK", "OREGON": "OR",
    "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA",
    "WASHINGTON": "WA", "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
}

MULTISPACE = re.compile(r"\s+")
PUNCT = re.compile(r"[^\w\s]")
TRAILING_SUFFIX = re.compile(
    r"\b(city|town|village|cdp|municipality|balance|metropolitan government|metro government|consolidated government|unified government|urban county)\b$",
    flags=re.I,
)


def normalize_state(value):
    raw = str(value or "").strip().upper()
    if len(raw) == 2 and raw.isalpha():
        return raw
    return STATE_TO_ABBR.get(raw, raw)


def normalize_city(value):
    text = str(value or "").strip().lower()
    text = PUNCT.sub(" ", text)
    text = MULTISPACE.sub(" ", text).strip()
    # Repeatedly peel trailing suffixes.
    while True:
        new_text = TRAILING_SUFFIX.sub("", text).strip()
        new_text = MULTISPACE.sub(" ", new_text).strip()
        if new_text == text:
            break
        text = new_text
    return text


def header_map(ws, row):
    return {
        str(ws.cell(row=row, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=row, column=c).value not in (None, "")
    }


def ensure_clean_crime_columns(clean_ws):
    row2_headers = header_map(clean_ws, 2)
    start_col = clean_ws.max_column + 1
    for name in CRIME_COLUMNS:
        if name not in row2_headers:
            col = clean_ws.max_column + 1
            clean_ws.cell(row=2, column=col, value=name)
            row2_headers[name] = col
    # Reset start based on first crime column now present.
    crime_cols = [row2_headers[name] for name in CRIME_COLUMNS]
    start_col = min(crime_cols)
    end_col = max(crime_cols)

    # Section row label and header styling.
    for c in range(start_col, end_col + 1):
        clean_ws.cell(row=1, column=c, value="Crime Index")
        clean_ws.cell(row=1, column=c).fill = CRIME_SECTION_FILL
        clean_ws.cell(row=1, column=c).font = Font(bold=True)

        clean_ws.cell(row=2, column=c).fill = CRIME_SECTION_FILL
        clean_ws.cell(row=2, column=c).font = Font(bold=True)
        clean_ws.cell(row=2, column=c).alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    return row2_headers


def update_readme_crime_section(readme_ws):
    lines = [
        "Crime Index",
        "Crime Index Source:",
        "The Crime Index uses FBI/UCR-style law-enforcement agency crime data where a reliable match exists between a Census city/place and the relevant FBI agency. Census city population is used as the reference population for coverage checks.",
        "Geography / Matching Note:",
        "FBI crime data is agency-based, not clean Census-place-based. The workbook attempts to match each city to the relevant police department, sheriff's office, metro police department, consolidated-government agency, or equivalent FBI reporting agency. Because agency jurisdictions do not always line up perfectly with Census city boundaries, some rows require manual review.",
        "Coverage Rate:",
        "Coverage Rate = FBI Population / Census City Population",
        "Coverage Interpretation:",
        "A coverage rate near 1.00 suggests the FBI agency population roughly matches the Census city population.",
        "A coverage rate far below 1.00 may indicate the agency covers only part of the city or that the wrong agency was matched.",
        "A coverage rate far above 1.00 may indicate the agency covers a broader jurisdiction than the Census city.",
        "Rows with poor or uncertain coverage should be marked Manual Review = TRUE.",
        "Crime Rate Formulas:",
        "Violent Crime Rate per 100k = Violent Crime Count / FBI Population x 100,000",
        "Property Crime Rate per 100k = Property Crime Count / FBI Population x 100,000",
        "Base Crime Index Score:",
        "Crime Index Score is calculated as: 0.75 * violent-crime-rate percentile + 0.25 * property-crime-rate percentile. Scores are standardized to 0-100. Higher scores indicate worse crime conditions.",
        "Density-Adjusted Crime Score:",
        "Density-Adjusted Crime Score is a separate adjusted score that modifies the base Crime Index Score based on population density. The purpose is to avoid over-penalizing dense urban markets solely because they naturally experience more reported incidents per square mile or per resident interaction. Density adjustment does not replace the base Crime Index Score. Both should be kept. Higher Density-Adjusted Crime Score still means worse crime conditions after density adjustment.",
        "Density Adjustment Methodology:",
        "Population Density = Census City Population / Land Area Sq Mi.",
        "Density Percentile is percentile rank of Population Density across valid rows.",
        "Density Adjustment = clamp(-5 * ((Density Percentile - 50) / 50), -5, +5).",
        "Density-Adjusted Crime Score = clamp(Crime Index Score + Density Adjustment, 0, 100).",
        "Crime Rating Bands:",
        "0-20 = Very Low",
        "21-40 = Low",
        "41-60 = Moderate",
        "61-75 = Elevated",
        "76-90 = High",
        "91-100 = Very High",
        "Manual Rating Fill:",
        "Some cities could not be reliably auto-matched to FBI agency data. For those rows, the workbook may include a manual Crime Rating fill.",
        "Manual-fill rows are marked Manual Review = TRUE, highlighted in the workbook, and do not receive invented FBI Population, Coverage Rate, crime rates, Crime Index Score, or Density-Adjusted Crime Score.",
        "Manual-fill rows should not be treated as final calculated FBI scores.",
        "Florida Note:",
        "Florida cities are handled cautiously because agency reporting and jurisdiction matching can be inconsistent. Florida rows that cannot be reliably matched remain blank/manual review rather than being force-filled.",
        "Clean Cities 100k+ Note:",
        "The Clean Cities 100k+ sheet includes only the decision-useful Crime Index fields: Crime Index Score, Crime Rating, Density-Adjusted Crime Score, Density-Adjusted Crime Rating, Manual Review, and Crime Rating Source.",
        "The full detailed FBI agency/audit fields remain in the Crime Index sheet.",
    ]

    existing = [readme_ws.cell(row=r, column=1).value for r in range(1, readme_ws.max_row + 1)]

    start = None
    end = None
    for i, val in enumerate(existing, start=1):
        if str(val).strip() == "Crime Index":
            start = i
            break

    if start is None:
        readme_ws.cell(row=readme_ws.max_row + 2, column=1, value="Crime Index")
        start = readme_ws.max_row

    for i in range(start + 1, len(existing) + 1):
        val = str(existing[i - 1]).strip() if existing[i - 1] is not None else ""
        if val == "Median Household Income":
            end = i - 1
            break
    if end is None:
        end = readme_ws.max_row

    block_len = end - start + 1
    desired_len = len(lines)

    if desired_len > block_len:
        readme_ws.insert_rows(end + 1, amount=desired_len - block_len)
    elif desired_len < block_len:
        readme_ws.delete_rows(start + desired_len, amount=block_len - desired_len)

    for offset, line in enumerate(lines):
        readme_ws.cell(row=start + offset, column=1, value=line)

    return True


def integrate_crime_into_clean_cities(input_path=None, output_path=None):
    """Copy decision-useful Crime Index columns onto the Clean Cities 100k+ sheet.

    Returns a summary dict identical in shape to what the CLI used to print.
    """
    input_file = Path(input_path) if input_path is not None else INPUT_FILE
    output_file = Path(output_path) if output_path is not None else OUTPUT_FILE

    if not input_file.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_file}")

    wb = load_workbook(input_file)
    clean_ws = wb["Clean Cities 100k+"]
    crime_ws = wb["Crime Index"]
    readme_ws = wb["README"]

    clean_headers = ensure_clean_crime_columns(clean_ws)
    crime_headers = header_map(crime_ws, 1)

    # Build lookup from Crime Index.
    crime_lookup = {}
    for r in range(2, crime_ws.max_row + 1):
        city = crime_ws.cell(row=r, column=crime_headers["City"]).value
        state = crime_ws.cell(row=r, column=crime_headers["State"]).value
        key = (normalize_city(city), normalize_state(state))
        if key == ("", ""):
            continue
        crime_lookup[key] = r

    total_rows = clean_ws.max_row - 2
    matched = 0
    unmatched = 0
    unmatched_rows = []
    copied_score = False
    copied_rating = False
    copied_density_score = False
    copied_density_rating = False
    manual_fill_copied = 0

    manual_col_start = clean_headers[CRIME_COLUMNS[0]]
    manual_col_end = clean_headers[CRIME_COLUMNS[-1]]

    for r in range(3, clean_ws.max_row + 1):
        city = clean_ws.cell(row=r, column=clean_headers["city"]).value
        state = clean_ws.cell(row=r, column=clean_headers["state_abbr"]).value
        st_abbr = normalize_state(state)
        key = (normalize_city(city), st_abbr)

        src_row = crime_lookup.get(key)
        if src_row is None:
            unmatched += 1
            if len(unmatched_rows) < 10:
                unmatched_rows.append((str(city or "").strip(), st_abbr))
            continue

        matched += 1

        # Do not modify Florida rows.
        if st_abbr == "FL":
            continue

        src_method = str(crime_ws.cell(row=src_row, column=crime_headers.get("FBI Match Method", 0)).value or "").strip().lower()
        src_source = str(crime_ws.cell(row=src_row, column=crime_headers.get("Crime Rating Source", 0)).value or "").strip().lower()
        is_manual_fill = (src_method == "manual rating fill") or (src_source == "manual rating fill")

        for col_name in CRIME_COLUMNS:
            dst_col = clean_headers[col_name]
            src_col = crime_headers[col_name]
            value = crime_ws.cell(row=src_row, column=src_col).value

            if col_name == "Manual Review":
                if is_manual_fill:
                    value = "TRUE"
                else:
                    value = "TRUE" if str(value or "").strip().upper() in {"TRUE", "YES", "Y"} else "FALSE"

            if col_name == "Crime Rating Source" and is_manual_fill:
                value = "Manual rating fill"

            if is_manual_fill and col_name in ["Crime Index Score", "Density-Adjusted Crime Score", "Density-Adjusted Crime Rating"]:
                # Keep blank if blank in source for manual rows.
                if value in (None, ""):
                    value = None

            clean_ws.cell(row=r, column=dst_col, value=value)

            if col_name == "Crime Index Score" and value not in (None, ""):
                copied_score = True
            if col_name == "Crime Rating" and value not in (None, ""):
                copied_rating = True
            if col_name == "Density-Adjusted Crime Score" and value not in (None, ""):
                copied_density_score = True
            if col_name == "Density-Adjusted Crime Rating" and value not in (None, ""):
                copied_density_rating = True

        if is_manual_fill:
            manual_fill_copied += 1
            for c in range(manual_col_start, manual_col_end + 1):
                clean_ws.cell(row=r, column=c).fill = MANUAL_CELL_FILL

    # Score formatting.
    for col_name in ["Crime Index Score", "Density-Adjusted Crime Score"]:
        c = clean_headers[col_name]
        for r in range(3, clean_ws.max_row + 1):
            clean_ws.cell(row=r, column=c).number_format = "0.00"

    # Keep filters and extend to include new columns.
    last_col_letter = get_column_letter(clean_ws.max_column)
    clean_ws.auto_filter.ref = f"A2:{last_col_letter}{clean_ws.max_row}"

    # Freeze panes: preserve existing if present.
    if clean_ws.freeze_panes is None:
        clean_ws.freeze_panes = "A3"

    # Widen new columns for readability.
    preferred_widths = {
        "Crime Index Score": 16,
        "Crime Rating": 14,
        "Density-Adjusted Crime Score": 24,
        "Density-Adjusted Crime Rating": 25,
        "Manual Review": 14,
        "Crime Rating Source": 22,
    }
    for name, width in preferred_widths.items():
        col = clean_headers[name]
        letter = get_column_letter(col)
        current = clean_ws.column_dimensions[letter].width or 8.43
        clean_ws.column_dimensions[letter].width = max(current, width)

    readme_updated = update_readme_crime_section(readme_ws)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    print(f"Number of Clean Cities 100k+ rows: {total_rows}")
    print(f"Number matched to Crime Index: {matched}")
    print(f"Number not matched: {unmatched}")
    print("First 10 unmatched rows, if any:")
    if unmatched_rows:
        for city, st in unmatched_rows:
            print(f"- {city}, {st}")
    else:
        print("- none")
    print(f"Whether Crime Index Score was copied: {'yes' if copied_score else 'no'}")
    print(f"Whether Crime Rating was copied: {'yes' if copied_rating else 'no'}")
    print(f"Whether Density-Adjusted Crime Score was copied: {'yes' if copied_density_score else 'no'}")
    print(f"Whether Density-Adjusted Crime Rating was copied: {'yes' if copied_density_rating else 'no'}")
    print(f"Number of manual rating fill rows copied: {manual_fill_copied}")
    print(f"README Crime Index section updated: {'yes' if readme_updated else 'no'}")
    print(f"Output file path: {output_file}")

    return {
        "output_path": str(output_file),
        "total_rows": total_rows,
        "matched": matched,
        "unmatched": unmatched,
        "unmatched_rows": unmatched_rows,
        "copied_score": copied_score,
        "copied_rating": copied_rating,
        "copied_density_score": copied_density_score,
        "copied_density_rating": copied_density_rating,
        "manual_fill_copied": manual_fill_copied,
        "readme_updated": readme_updated,
    }


def main():
    integrate_crime_into_clean_cities()


if __name__ == "__main__":
    main()
