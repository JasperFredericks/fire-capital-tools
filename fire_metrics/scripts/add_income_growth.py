#!/usr/bin/env python3
"""Add ACS 1-year median household income and growth columns to workbook city sheets.

Variable: B19013_001E (Median household income)
Years: 2021, latest ACS 1-year available, and year immediately before latest.
"""

import argparse
import datetime as dt
import json
import re
import shutil
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from fire_metrics_updater.config import get_secret


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "output" / "us_cities_100k_population_ranked_WITH_CRIME_INDEX.xlsx"
CACHE_DIR = BASE_DIR / "data" / "cache"
CENSUS_API_KEY = get_secret("CENSUS_API_KEY", "data/cache/census_api_key.txt")
ACS_DATASET = "acs/acs1"
ACS_VAR = "B19013_001E"

STATE_TO_ABBR = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR", "CALIFORNIA": "CA",
    "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE", "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID", "ILLINOIS": "IL",
    "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA",
    "MAINE": "ME", "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
    "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK", "OREGON": "OR",
    "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA",
    "WASHINGTON": "WA", "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
}

ABBR_TO_STATE = {v: k for k, v in STATE_TO_ABBR.items()}

PLACE_SUFFIX = re.compile(r"\b(city|town|village|cdp|municipality|metro government|balance)\b$", re.I)
TRAILING_DIGITS = re.compile(r"\d+$")
PUNCT = re.compile(r"[^\w\s]", re.U)
MULTI = re.compile(r"\s+")


def normalize_city(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = PUNCT.sub(" ", text)
    text = TRAILING_DIGITS.sub("", text).strip()
    text = PLACE_SUFFIX.sub("", text).strip()
    text = MULTI.sub(" ", text)
    return text


def normalize_state_to_abbr(value):
    if value is None:
        return ""
    raw = str(value).strip().upper()
    raw = TRAILING_DIGITS.sub("", raw).strip()
    if len(raw) == 2 and raw.isalpha():
        return raw
    return STATE_TO_ABBR.get(raw, raw)


def load_census_api_key():
    if CENSUS_API_KEY:
        return CENSUS_API_KEY
    raise RuntimeError(
        "Census API key not found. Set CENSUS_API_KEY or create data/cache/census_api_key.txt"
    )


def api_get_json(year, params, api_key):
    params = dict(params)
    params["key"] = api_key
    query = urllib.parse.urlencode(params)
    url = f"https://api.census.gov/data/{year}/{ACS_DATASET}?{query}"
    with urllib.request.urlopen(url, timeout=60) as resp:
        body = resp.read().decode("utf-8")
    # Census can return HTML for invalid key; ensure JSON.
    if not body.strip().startswith("["):
        raise RuntimeError(f"Census API error for {year}: {body[:200]}")
    return json.loads(body)


def discover_latest_year(api_key):
    current = dt.date.today().year
    for year in range(current - 1, 2020, -1):
        try:
            data = api_get_json(
                year,
                {"get": f"NAME,{ACS_VAR}", "for": "place:*", "in": "state:01"},
                api_key,
            )
            if len(data) > 1:
                return year
        except Exception:
            continue
    raise RuntimeError("Unable to find a usable ACS 1-year for B19013_001E")


def fetch_income_by_place(year, api_key):
    rows = api_get_json(year, {"get": f"NAME,{ACS_VAR}", "for": "place:*", "in": "state:*"}, api_key)
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}

    out = {}
    for row in rows[1:]:
        name = row[idx["NAME"]]
        val = row[idx[ACS_VAR]]
        state_fips = row[idx["state"]]
        place = row[idx["place"]]
        if val in (None, "", "-666666666", "-222222222", "-333333333", "null"):
            continue
        try:
            income = int(float(val))
        except Exception:
            continue

        # NAME is usually "Place name, State Name"
        if "," in name:
            city_part, state_name = [x.strip() for x in name.rsplit(",", 1)]
            state_abbr = STATE_TO_ABBR.get(state_name.upper(), "")
        else:
            city_part = name.strip()
            state_abbr = ""

        if not state_abbr:
            continue

        key = (state_abbr, normalize_city(city_part))
        if key not in out:
            out[key] = income
        # keep first seen income for deterministic behavior
    return out


def backup_workbook(path):
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_name(f"{path.stem}_backup_{stamp}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def find_city_state_columns(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    lower = [str(h).strip().lower() if h is not None else "" for h in headers]

    city_col = None
    state_col = None
    for i, h in enumerate(lower, start=1):
        if city_col is None and h == "city":
            city_col = i
        if state_col is None and h in ("state", "state_abbr"):
            state_col = i
    return city_col, state_col


def remove_old_income_columns(ws):
    patterns = (
        "median household income in ",
        "median household income growth ",
    )
    to_delete = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        text = str(v).strip().lower() if v is not None else ""
        if any(text.startswith(p) for p in patterns):
            to_delete.append(c)
    for c in reversed(to_delete):
        ws.delete_cols(c, 1)


def add_income_columns_to_sheet(ws, years, incomes_by_year):
    city_col, state_col = find_city_state_columns(ws)
    if not city_col or not state_col:
        return None

    y2021, yprior, ylatest = years
    headers = [
        f"Median household income in {y2021}",
        f"Median household income in {yprior}",
        f"Median household income in {ylatest}",
        f"Median household income growth {y2021}\u2013{ylatest} (%)",
        f"Median household income growth {yprior}\u2013{ylatest} (%)",
    ]

    remove_old_income_columns(ws)

    start_col = ws.max_column + 1
    for i, h in enumerate(headers):
        ws.cell(1, start_col + i, h)

    matched = 0
    unmatched = []
    sanity = []

    for r in range(2, ws.max_row + 1):
        city = ws.cell(r, city_col).value
        state = ws.cell(r, state_col).value
        if city is None or state is None:
            continue

        st = normalize_state_to_abbr(state)
        ct = normalize_city(city)
        key = (st, ct)

        inc_2021 = incomes_by_year[y2021].get(key)
        inc_prior = incomes_by_year[yprior].get(key)
        inc_latest = incomes_by_year[ylatest].get(key)

        c2021 = ws.cell(r, start_col)
        cprior = ws.cell(r, start_col + 1)
        clatest = ws.cell(r, start_col + 2)
        cg1 = ws.cell(r, start_col + 3)
        cg2 = ws.cell(r, start_col + 4)

        c2021.value = inc_2021
        cprior.value = inc_prior
        clatest.value = inc_latest

        if inc_2021 and inc_latest and inc_2021 != 0:
            cg1.value = (inc_latest - inc_2021) / inc_2021
        else:
            cg1.value = None

        if inc_prior and inc_latest and inc_prior != 0:
            cg2.value = (inc_latest - inc_prior) / inc_prior
        else:
            cg2.value = None

        for c in (c2021, cprior, clatest):
            c.number_format = "$#,##0"
        for c in (cg1, cg2):
            c.number_format = "0.00%"

        if inc_latest is not None:
            matched += 1
            if len(sanity) < 5:
                sanity.append((city, st, inc_2021, inc_latest, inc_prior, cg1.value, cg2.value))
        else:
            unmatched.append((city, st))

    return {
        "sheet": ws.title,
        "matched": matched,
        "unmatched": len(unmatched),
        "unmatched_list": unmatched[:10],
        "sanity": sanity,
    }


def update_readme_sheet(wb, years):
    if "README" not in wb.sheetnames:
        return
    y2021, yprior, ylatest = years
    ws = wb["README"]
    marker = "Income growth methodology update"
    existing = "\n".join(str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1))
    if marker in existing:
        return
    lines = [
        "",
        marker,
        f"ACS 1-year median household income variable: {ACS_VAR}.",
        f"Years used: {y2021}, {yprior}, {ylatest}. 2020 excluded due to comparability guidance.",
        f"Growth {y2021}\u2013{ylatest} = (Income {ylatest} - Income {y2021}) / Income {y2021}.",
        f"Growth {yprior}\u2013{ylatest} = (Income {ylatest} - Income {yprior}) / Income {yprior}.",
    ]
    for line in lines:
        ws.append([line])


def choose_target_sheets(wb):
    preferred = ["Clean Cities 100k+", "Crime Index"]
    targets = [s for s in preferred if s in wb.sheetnames]
    if targets:
        return targets
    # Fallback: any sheet with City+State headers.
    out = []
    for s in wb.sheetnames:
        ws = wb[s]
        city_col, state_col = find_city_state_columns(ws)
        if city_col and state_col and ws.max_row > 2:
            out.append(s)
    return out


def main():
    parser = argparse.ArgumentParser(description="Add ACS income growth columns to workbook")
    parser.add_argument("--workbook", default=str(WORKBOOK_PATH), help="Workbook path")
    args = parser.parse_args()

    workbook = Path(args.workbook)
    if not workbook.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook}")

    api_key = load_census_api_key()
    latest = discover_latest_year(api_key)
    prior = latest - 1
    years = (2021, prior, latest)

    incomes_by_year = {
        2021: fetch_income_by_place(2021, api_key),
        prior: fetch_income_by_place(prior, api_key),
        latest: fetch_income_by_place(latest, api_key),
    }

    backup = backup_workbook(workbook)
    wb = load_workbook(workbook)

    targets = choose_target_sheets(wb)
    if not targets:
        raise RuntimeError("No target city/state sheets found")

    summaries = []
    for sheet_name in targets:
        ws = wb[sheet_name]
        summary = add_income_columns_to_sheet(ws, years, incomes_by_year)
        if summary:
            summaries.append(summary)

    update_readme_sheet(wb, years)
    wb.save(workbook)

    # report summary based on primary sheet
    primary = next((s for s in summaries if s["sheet"] == "Clean Cities 100k+"), summaries[0])
    print("ACS years used:", years)
    print("Backup created:", backup)
    print("Updated sheets:", ", ".join(s["sheet"] for s in summaries))
    print("Cities matched:", primary["matched"])
    print("Cities failed to match:", primary["unmatched"])
    print("First 10 unmatched cities:")
    for city, st in primary["unmatched_list"]:
        print(f"- {city}, {st}")
    print("5 sanity-check rows:")
    y2021, yprior, ylatest = years
    for city, st, inc21, incl, incp, g1, g2 in primary["sanity"]:
        print(
            f"- {city}, {st} | {y2021}: {inc21} | {ylatest}: {incl} | {yprior}: {incp} "
            f"| growth {y2021}\u2013{ylatest}: {g1} | growth {yprior}\u2013{ylatest}: {g2}"
        )


if __name__ == "__main__":
    main()
