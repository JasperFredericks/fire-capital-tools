"""SQLite ingestion for the FIRE Metrics search dashboard.

Rewrite of the old JSON-index builder. That version read one merged
workbook and fuzzy-guessed which column meant what (FIELD_PATTERNS token
matching), because it had no other way to know the shape of whatever
workbook it was handed. That doesn't apply anymore: each pipeline script in
fire_metrics/scripts/ now produces its own output workbook with known,
exact column names (see each script's own append_*_columns function), so
this module reads each one directly by name and upserts into SQLite --
no guessing needed. See db.py for the schema this writes into.
"""

from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
import urllib.request
import zipfile

import openpyxl

from . import db as db_module
from .city_search import build_city_aliases, make_search_key, normalize_city_tokens, normalize_query

POPULATION_THRESHOLD = 100_000
GAZETTEER_DIR = Path(__file__).resolve().parent.parent / "data" / "cache" / "census_gazetteer"
GAZETTEER_URLS = [
    "https://www2.census.gov/geo/docs/maps-data/data/gazetteer/Gaz_places_national.zip",
    "https://www2.census.gov/geo/docs/maps-data/data/gazetteer/2024_Gaz_place_national.zip",
    "https://www2.census.gov/geo/docs/maps-data/data/gazetteer/2023_Gaz_place_national.zip",
]
_COORD_CITY_ALIASES = {
    "nyc": "new york",
    "new york city": "new york",
    "la": "los angeles",
    "l a": "los angeles",
    "los angeles city": "los angeles",
    "st louis": "st louis",
    "saint louis": "st louis",
    "st. louis": "st louis",
    "st. louis city": "st louis",
    "phoenix city": "phoenix",
    "las vegas city": "las vegas",
    "atlanta city": "atlanta",
}
_CITY_SUFFIX_EXCEPTIONS = {"oklahoma city", "kansas city", "salt lake city", "carson city"}

# Some pipeline outputs (e.g. add_climate_risk.py, which correctly drops its
# own internally-recomputed state_abbr column as internal working state --
# see the Priority 2 output-column bug fix) only carry the full state name
# forward, not the 2-letter abbreviation the DB is keyed by. Resolve it here
# rather than mis-treating the full name as if it were already an abbreviation.
STATE_TO_ABBR = {
    "ALABAMA": "AL", "ALASKA": "AK", "ARIZONA": "AZ", "ARKANSAS": "AR", "CALIFORNIA": "CA",
    "COLORADO": "CO", "CONNECTICUT": "CT", "DELAWARE": "DE", "DISTRICT OF COLUMBIA": "DC",
    "FLORIDA": "FL", "GEORGIA": "GA", "HAWAII": "HI", "IDAHO": "ID", "ILLINOIS": "IL",
    "INDIANA": "IN", "IOWA": "IA", "KANSAS": "KS", "KENTUCKY": "KY", "LOUISIANA": "LA",
    "MAINE": "ME", "MARYLAND": "MD", "MASSACHUSETTS": "MA", "MICHIGAN": "MI", "MINNESOTA": "MN",
    "MISSISSIPPI": "MS", "MISSOURI": "MO", "MONTANA": "MT", "NEBRASKA": "NE", "NEVADA": "NV",
    "NEW HAMPSHIRE": "NH", "NEW JERSEY": "NJ", "NEW MEXICO": "NM", "NEW YORK": "NY",
    "NORTH CAROLINA": "NC", "NORTH DAKOTA": "ND", "OHIO": "OH", "OKLAHOMA": "OK", "OREGON": "OR",
    "PENNSYLVANIA": "PA", "RHODE ISLAND": "RI", "SOUTH CAROLINA": "SC", "SOUTH DAKOTA": "SD",
    "TENNESSEE": "TN", "TEXAS": "TX", "UTAH": "UT", "VERMONT": "VT", "VIRGINIA": "VA",
    "WASHINGTON": "WA", "WEST VIRGINIA": "WV", "WISCONSIN": "WI", "WYOMING": "WY",
    "PUERTO RICO": "PR",
}


def _resolve_state_abbr(value: Any) -> str:
    raw = str(value or "").strip()
    if len(raw) == 2 and raw.isalpha():
        return raw.upper()
    return STATE_TO_ABBR.get(raw.upper(), raw.upper())

LANDLORD_LABELS = {
    1: "Landlord-friendly",
    1.0: "Landlord-friendly",
    0: "Neutral",
    0.0: "Neutral",
    -1: "Tenant-friendly",
    -1.0: "Tenant-friendly",
}


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sheet_headers(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value not in (None, "")
    }


def _find_col(headers: dict[str, int], *candidates: str) -> int | None:
    for name in candidates:
        if name in headers:
            return headers[name]
    return None


def _find_col_startswith(headers: dict[str, int], prefix: str) -> int | None:
    prefix_lower = prefix.lower()
    for name, idx in headers.items():
        if name.lower().startswith(prefix_lower):
            return idx
    return None


def _find_growth_col(headers: dict[str, int], label_prefix: str, span: str) -> int | None:
    """Find a "<label> growth <span> (%)"-style column without needing the
    exact years (which change every refresh) -- span is e.g. "2021" to
    match "...growth 2021-2025 (%)" regardless of the second year.
    """
    prefix_lower = label_prefix.lower()
    for name, idx in headers.items():
        lname = name.lower()
        if lname.startswith(prefix_lower) and span in lname and "%" in lname:
            return idx
    return None


def _landlord_label(score: Any) -> str | None:
    if score is None:
        return None
    try:
        return LANDLORD_LABELS.get(float(score))
    except (TypeError, ValueError):
        return None


_DISPLAY_SUFFIXES = [
    " metropolitan government (balance)",
    " metro government (balance)",
    " consolidated government (balance)",
    " unified government (balance)",
    " urban county",
    " municipality",
    " county",
    " city",
    " town",
    " village",
    " CDP",
    " cdp",
]


def _clean_display_city(city: str) -> str:
    """Same suffix-stripping intent as format_workbook_final.py's
    clean_city_display_name() -- duplicated in miniature here rather than
    cross-importing fire_metrics/scripts/format_workbook_final.py, since
    fire_metrics/scripts/ isn't set up as an importable package and this is
    a handful of lines.
    """
    text = str(city).strip()
    changed = True
    while changed:
        changed = False
        for suffix in _DISPLAY_SUFFIXES:
            if text.endswith(suffix):
                candidate = text[: -len(suffix)].rstrip(" ,-/")
                if candidate:
                    text = candidate
                    changed = True
    return text


def _identity_row(city: str, state_abbr: str) -> dict[str, Any]:
    clean_city = _clean_display_city(city)
    display_name = f"{clean_city}, {state_abbr}"

    # Build aliases from BOTH the raw Census text ("Los Angeles city") and
    # the cleaned display form ("Los Angeles") -- most users search the
    # cleaned form (that's what's actually shown to them), so it needs to
    # be an exact alias match too, not just fuzzy-close.
    search_keys = set(build_city_aliases(city, state_abbr))
    search_keys |= build_city_aliases(clean_city, state_abbr)

    return {
        "city": city,
        "state": state_abbr,
        "display_name": display_name,
        "normalized_city": normalize_city_tokens(clean_city),
        "normalized_display_name": normalize_query(display_name),
        "search_key": make_search_key(clean_city, state_abbr),
        "search_keys": sorted(search_keys),
        "latitude": None,
        "longitude": None,
    }


def _coord_city_token(value: str) -> str:
    normalized = normalize_city_tokens(value)
    if normalized.startswith("city of "):
        normalized = normalized[len("city of "):].strip()
    if normalized.endswith(" city") and normalized not in _CITY_SUFFIX_EXCEPTIONS:
        normalized = normalized[: -len(" city")].strip()
    return _COORD_CITY_ALIASES.get(normalized, normalized)


def _ensure_gazetteer_places_file() -> Path:
    GAZETTEER_DIR.mkdir(parents=True, exist_ok=True)

    existing = sorted(GAZETTEER_DIR.glob("*Gaz_place*_national.txt"))
    if not existing:
        existing = sorted(GAZETTEER_DIR.glob("Gaz_places_national.txt"))
    if existing:
        return existing[0]

    for url in GAZETTEER_URLS:
        zip_name = url.split("/")[-1]
        zip_path = GAZETTEER_DIR / zip_name
        try:
            urllib.request.urlretrieve(url, zip_path)
            with zipfile.ZipFile(zip_path, "r") as archive:
                archive.extractall(GAZETTEER_DIR)
            extracted = sorted(GAZETTEER_DIR.glob("*Gaz_place*_national.txt"))
            if not extracted:
                extracted = sorted(GAZETTEER_DIR.glob("Gaz_places_national.txt"))
            if extracted:
                return extracted[0]
        except Exception:
            continue

    raise RuntimeError(
        f"Could not download Census Gazetteer place file into {GAZETTEER_DIR}."
    )


def _load_gazetteer_coordinate_lookup() -> dict[tuple[str, str], tuple[float, float]]:
    path = _ensure_gazetteer_places_file()
    lookup: dict[tuple[str, str], tuple[float, float]] = {}

    with path.open("r", encoding="latin-1", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for raw_row in reader:
            row = {str(k).strip().upper(): v for k, v in raw_row.items() if k is not None}
            state = _resolve_state_abbr(row.get("USPS"))
            name = str(row.get("NAME") or "").strip()
            if not state or not name:
                continue

            lat_raw = row.get("INTPTLAT")
            lon_raw = row.get("INTPTLONG")
            if lat_raw in (None, "") or lon_raw in (None, ""):
                continue
            try:
                latitude = float(lat_raw)
                longitude = float(lon_raw)
            except (TypeError, ValueError):
                continue

            raw_token = _coord_city_token(name)
            clean_token = _coord_city_token(_clean_display_city(name))
            lookup[(state, raw_token)] = (latitude, longitude)
            lookup[(state, clean_token)] = (latitude, longitude)

    return lookup


def _coordinates_for_city(city: str, state_abbr: str, lookup: dict[tuple[str, str], tuple[float, float]]) -> tuple[float | None, float | None]:
    state = _resolve_state_abbr(state_abbr)
    if not state:
        return (None, None)

    candidates = [
        _coord_city_token(city),
        _coord_city_token(_clean_display_city(city)),
        _coord_city_token(normalize_city_tokens(city)),
    ]
    for token in candidates:
        match = lookup.get((state, token))
        if match:
            return match
    return (None, None)


def backfill_city_coordinates(conn) -> dict[str, Any]:
    rows = conn.execute(
        """
        SELECT city, state
        FROM cities
        WHERE include_flag = 1
          AND (latitude IS NULL OR longitude IS NULL)
        """
    ).fetchall()
    if not rows:
        return {"rows_checked": 0, "rows_updated": 0}

    lookup = _load_gazetteer_coordinate_lookup()
    updates = []
    for row in rows:
        latitude, longitude = _coordinates_for_city(row["city"], row["state"], lookup)
        if latitude is None or longitude is None:
            continue
        updates.append((latitude, longitude, row["city"], row["state"]))

    if updates:
        conn.executemany(
            """
            UPDATE cities
            SET latitude = ?, longitude = ?
            WHERE city = ? AND state = ?
            """,
            updates,
        )
        conn.commit()

    return {"rows_checked": len(rows), "rows_updated": len(updates)}


def ingest_population_and_landlord(workbook_path: Path, conn) -> dict[str, Any]:
    """Read update_fire_metrics.py's own output: Clean Cities 100k+ (population,
    population change) + Landlord Friendliness (per-state score) sheets.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["Clean Cities 100k+"]
    headers = _sheet_headers(ws)

    city_col = _find_col(headers, "city")
    state_abbr_col = _find_col(headers, "state_abbr")
    state_name_col = _find_col(headers, "state")
    rank_col = _find_col(headers, "rank_2025")
    pop_col = _find_col(headers, "population_2025")
    growth_2020_col = _find_col(headers, "percent_change_2020_to_2025")
    growth_recent_col = _find_col(headers, "percent_change_2024_to_2025")

    if not city_col or not (state_abbr_col or state_name_col):
        raise RuntimeError("Could not find city/state columns in Clean Cities 100k+ sheet")
    state_col = state_abbr_col or state_name_col

    # Landlord Friendliness is keyed by (city, full state NAME) -- e.g.
    # "New York city", "New York" -- not the 2-letter abbreviation used
    # elsewhere, so it needs its own lookup key.
    landlord_scores: dict[tuple[str, str], Any] = {}
    if "Landlord Friendliness" in wb.sheetnames:
        lws = wb["Landlord Friendliness"]
        lheaders = _sheet_headers(lws)
        lcity_col = _find_col(lheaders, "City")
        lstate_col = _find_col(lheaders, "State")
        lscore_col = _find_col(lheaders, "Landlord Score")
        if lcity_col and lstate_col and lscore_col:
            for r in range(2, lws.max_row + 1):
                city = lws.cell(r, lcity_col).value
                state = lws.cell(r, lstate_col).value
                score = lws.cell(r, lscore_col).value
                if city and state:
                    landlord_scores[(str(city).strip(), str(state).strip())] = score

    coord_lookup = _load_gazetteer_coordinate_lookup()
    identity_rows = []
    metric_rows = []
    excluded_rows = []

    for r in range(2, ws.max_row + 1):
        city = ws.cell(r, city_col).value
        state_abbr = ws.cell(r, state_col).value
        state_name = ws.cell(r, state_name_col).value if state_name_col else None
        if not city or not state_abbr:
            continue
        city = str(city).strip()
        state_abbr = str(state_abbr).strip().upper()

        identity = _identity_row(city, state_abbr)
        latitude, longitude = _coordinates_for_city(city, state_abbr, coord_lookup)
        identity["latitude"] = latitude
        identity["longitude"] = longitude
        identity_rows.append(identity)

        pop_current = ws.cell(r, pop_col).value if pop_col else None
        landlord_score = landlord_scores.get((city, str(state_name).strip() if state_name else ""))

        include_flag = 1 if (pop_current is not None and pop_current >= POPULATION_THRESHOLD) else 0
        metric_rows.append({
            "city": city, "state": state_abbr,
            "population_rank": rank_col and ws.cell(r, rank_col).value,
            "population_current": pop_current,
            "population_growth_2020_2025": growth_2020_col and ws.cell(r, growth_2020_col).value,
            "population_growth_recent": growth_recent_col and ws.cell(r, growth_recent_col).value,
            "landlord_friendliness_score": landlord_score,
            "landlord_friendliness_label": _landlord_label(landlord_score),
            "include_flag": include_flag,
            "threshold_reason": None if include_flag else "Below 100,000 population threshold.",
        })

        if not include_flag:
            excluded_rows.append({
                "city": city, "state": state_abbr,
                "normalized_city": normalize_city_tokens(city),
                "normalized_key": make_search_key(city, state_abbr),
                "latest_population": pop_current,
                "threshold_reason": "Below 100,000 population threshold.",
            })

    updated_at = _utc_now()
    db_module.upsert_city_identity(conn, identity_rows)
    count = db_module.upsert_metric_family(conn, "population", metric_rows, updated_at)
    db_module.replace_excluded_cities(conn, excluded_rows)

    return {"metric_family": "population", "rows_updated": count, "updated_at": updated_at, "total_rows": len(identity_rows)}


def _ingest_generic(workbook_path: Path, conn, family: str, sheet_name: str, column_builder) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
    headers = _sheet_headers(ws)

    city_col = _find_col(headers, "city")
    state_col = _find_col(headers, "state_abbr", "state", "State")
    if not city_col or not state_col:
        raise RuntimeError(f"Could not find city/state columns in {sheet_name} sheet")

    metric_rows = []
    for r in range(2, ws.max_row + 1):
        city = ws.cell(r, city_col).value
        state_raw = ws.cell(r, state_col).value
        if not city or not state_raw:
            continue
        city = str(city).strip()
        state_abbr = _resolve_state_abbr(state_raw)
        row = column_builder(ws, r, headers)
        if row is None:
            continue
        row["city"] = city
        row["state"] = state_abbr
        metric_rows.append(row)

    updated_at = _utc_now()
    count = db_module.upsert_metric_family(conn, family, metric_rows, updated_at)
    return {"metric_family": family, "rows_updated": count, "updated_at": updated_at, "total_rows": len(metric_rows)}


def ingest_income(workbook_path: Path, conn) -> dict[str, Any]:
    """Read add_income_growth.py's own output columns."""
    def build(ws, r, headers):
        current_col = _find_col_startswith(headers, "Median household income in ")
        # There are 3 "in <year>" columns (2021, prior, latest); the latest
        # one is whichever has the highest year in its header text.
        year_cols = [(name, idx) for name, idx in headers.items() if name.lower().startswith("median household income in ")]
        if not year_cols:
            return None
        latest_col = max(year_cols, key=lambda item: item[0])[1]
        growth_2021_col = _find_growth_col(headers, "median household income growth 2021", "2021")
        growth_recent_col = None
        for name, idx in headers.items():
            if name.lower().startswith("median household income growth") and not name.startswith("Median household income growth 2021"):
                growth_recent_col = idx
        return {
            "median_income_current": ws.cell(r, latest_col).value,
            "median_income_growth_2021_2024": ws.cell(r, growth_2021_col).value if growth_2021_col else None,
            "median_income_growth_recent": ws.cell(r, growth_recent_col).value if growth_recent_col else None,
        }

    return _ingest_generic(workbook_path, conn, "income", "Clean Cities 100k+", build)


def ingest_home_value(workbook_path: Path, conn) -> dict[str, Any]:
    """Read add_home_value_growth.py's own output columns."""
    def build(ws, r, headers):
        year_cols = [(name, idx) for name, idx in headers.items() if name.lower().startswith("median home/condo value in ")]
        if not year_cols:
            return None
        latest_col = max(year_cols, key=lambda item: item[0])[1]
        growth_2021_col = _find_growth_col(headers, "median home/condo value growth 2021", "2021")
        growth_recent_col = None
        for name, idx in headers.items():
            if name.lower().startswith("median home/condo value growth") and not name.startswith("Median home/condo value growth 2021"):
                growth_recent_col = idx
        return {
            "median_home_value_current": ws.cell(r, latest_col).value,
            "median_home_value_growth_2021_2024": ws.cell(r, growth_2021_col).value if growth_2021_col else None,
            "median_home_value_growth_recent": ws.cell(r, growth_recent_col).value if growth_recent_col else None,
        }

    return _ingest_generic(workbook_path, conn, "home_value", "Clean Cities 100k+", build)


def ingest_employment(workbook_path: Path, conn) -> dict[str, Any]:
    """Read add_job_growth.py's own output columns."""
    def build(ws, r, headers):
        year_cols = [(name, idx) for name, idx in headers.items() if name.lower().startswith("resident employment in ")]
        if not year_cols:
            return None
        latest_col = max(year_cols, key=lambda item: item[0])[1]
        growth_2021_col = None
        growth_recent_col = None
        for name, idx in headers.items():
            lname = name.lower()
            if lname.startswith("employment growth 2021-"):
                growth_2021_col = idx
            elif lname.startswith("employment growth") and "2021" not in lname:
                growth_recent_col = idx
        return {
            "employment_current": ws.cell(r, latest_col).value,
            "employment_growth_2021_2025": ws.cell(r, growth_2021_col).value if growth_2021_col else None,
            "employment_growth_recent": ws.cell(r, growth_recent_col).value if growth_recent_col else None,
        }

    return _ingest_generic(workbook_path, conn, "employment", "Clean Cities 100k+", build)


def ingest_climate_risk(workbook_path: Path, conn) -> dict[str, Any]:
    """Read add_climate_risk.py's own output columns (snake_case)."""
    def build(ws, r, headers):
        score_col = _find_col(headers, "climate_risk_score")
        rating_col = _find_col(headers, "climate_risk_rating")
        if score_col is None and rating_col is None:
            return None
        return {
            "climate_risk_score": ws.cell(r, score_col).value if score_col else None,
            "climate_risk_rating": ws.cell(r, rating_col).value if rating_col else None,
        }

    return _ingest_generic(workbook_path, conn, "climate", "Clean Cities 100k+", build)


def ingest_crime(workbook_path: Path, conn) -> dict[str, Any]:
    """Read crime_pipeline.py's final output (Crime Index columns merged
    onto Clean Cities 100k+ by integrate_crime_into_clean_cities.py).
    """
    def build(ws, r, headers):
        score_col = _find_col(headers, "Crime Index Score")
        rating_col = _find_col(headers, "Crime Rating")
        density_score_col = _find_col(headers, "Density-Adjusted Crime Score")
        density_rating_col = _find_col(headers, "Density-Adjusted Crime Rating")
        review_col = _find_col(headers, "Manual Review")
        if score_col is None and rating_col is None:
            return None
        return {
            "crime_index_score": ws.cell(r, score_col).value if score_col else None,
            "crime_rating": ws.cell(r, rating_col).value if rating_col else None,
            "density_adjusted_crime_score": ws.cell(r, density_score_col).value if density_score_col else None,
            "density_adjusted_crime_rating": ws.cell(r, density_rating_col).value if density_rating_col else None,
            "crime_manual_review": ws.cell(r, review_col).value if review_col else None,
        }

    return _ingest_generic(workbook_path, conn, "crime", "Clean Cities 100k+", build)
