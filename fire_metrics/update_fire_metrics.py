import csv
import json
import re
import urllib.request
from pathlib import Path

import geopandas as gpd
import pandas as pd
import requests
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "data" / "raw"
PROCESSED_DIR = BASE_DIR / "data" / "processed"
OUTPUT_DIR = BASE_DIR / "output"
RAW_FILE = RAW_DIR / "SUB-IP-EST2025-ANNRNK.xlsx"
OUTPUT_FILE = OUTPUT_DIR / "us_cities_100k_population_ranked_WITH_CLIMATE_RISK.xlsx"
CENSUS_URL = "https://www2.census.gov/programs-surveys/popest/tables/2020-2025/cities/totals/SUB-IP-EST2025-ANNRNK.xlsx"
FEMA_NRI_QUERY_URL = "https://services.arcgis.com/XG15cJAlne2vxtgt/ArcGIS/rest/services/National_Risk_Index_Counties/FeatureServer/0/query"

TIGER_YEAR = 2025
TIGER_PLACE_URL_TEMPLATE = "https://www2.census.gov/geo/tiger/TIGER{year}/PLACE/tl_{year}_{state_fips}_place.zip"
TIGER_COUNTY_URL_TEMPLATE = "https://www2.census.gov/geo/tiger/TIGER{year}/COUNTY/tl_{year}_us_county.zip"
TIGER_PLACE_DIR = RAW_DIR / "tiger" / "PLACE"
TIGER_COUNTY_DIR = RAW_DIR / "tiger" / "COUNTY"

STATE_ABBR_TO_FIPS = {
    "AL": "01",
    "AK": "02",
    "AZ": "04",
    "AR": "05",
    "CA": "06",
    "CO": "08",
    "CT": "09",
    "DE": "10",
    "DC": "11",
    "FL": "12",
    "GA": "13",
    "HI": "15",
    "ID": "16",
    "IL": "17",
    "IN": "18",
    "IA": "19",
    "KS": "20",
    "KY": "21",
    "LA": "22",
    "ME": "23",
    "MD": "24",
    "MA": "25",
    "MI": "26",
    "MN": "27",
    "MS": "28",
    "MO": "29",
    "MT": "30",
    "NE": "31",
    "NV": "32",
    "NH": "33",
    "NJ": "34",
    "NM": "35",
    "NY": "36",
    "NC": "37",
    "ND": "38",
    "OH": "39",
    "OK": "40",
    "OR": "41",
    "PA": "42",
    "RI": "44",
    "SC": "45",
    "SD": "46",
    "TN": "47",
    "TX": "48",
    "UT": "49",
    "VT": "50",
    "VA": "51",
    "WA": "53",
    "WV": "54",
    "WI": "55",
    "WY": "56",
    "PR": "72",
}
STATE_FIPS_TO_ABBR = {value: key for key, value in STATE_ABBR_TO_FIPS.items()}

SOURCE_NAME = (
    "U.S. Census Bureau City and Town Population Totals: 2020-2025"
)
SOURCE_YEAR = 2025
SOURCE_FILE_NAME = RAW_FILE.name
SOURCE_NOTE = (
    "Annual Estimates of the Resident Population for Incorporated Places of 20,000 or More, "
    "Ranked by July 1, 2025 Population: April 1, 2020 to July 1, 2025"
)

STATE_ABBREVIATIONS = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "District of Columbia": "DC",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
}

PLACE_TYPE_KEYWORDS = [
    "city",
    "town",
    "village",
    "borough",
    "municipality",
    "census designated place",
    "cdp",
    "township",
]

LANDLORD_SCORES = {
    "Alabama": 1,
    "Alaska": 1,
    "Arizona": 1,
    "Arkansas": 1,
    "California": -1,
    "Colorado": 1,
    "Connecticut": -1,
    "Delaware": None,
    "District of Columbia": None,
    "Florida": 1,
    "Georgia": 1,
    "Hawaii": -1,
    "Idaho": 0,
    "Illinois": 0,
    "Indiana": 1,
    "Iowa": 1,
    "Kansas": 0,
    "Kentucky": 1,
    "Louisiana": 1,
    "Maine": None,
    "Maryland": 1,
    "Massachusetts": -1,
    "Michigan": 1,
    "Minnesota": 0,
    "Mississippi": 1,
    "Missouri": 0,
    "Montana": 0,
    "Nebraska": 0,
    "Nevada": 0,
    "New Hampshire": -1,
    "New Jersey": -1,
    "New Mexico": 0,
    "New York": -1,
    "North Carolina": 1,
    "North Dakota": 0,
    "Ohio": 1,
    "Oklahoma": 0,
    "Oregon": -1,
    "Pennsylvania": 0,
    "Rhode Island": -1,
    "South Carolina": 0,
    "South Dakota": 0,
    "Tennessee": 1,
    "Texas": 1,
    "Utah": 0,
    "Vermont": None,
    "Virginia": 0,
    "Washington": -1,
    "West Virginia": None,
    "Wisconsin": 0,
    "Wyoming": None,
}


def ensure_directories():
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def download_source_file():
    if RAW_FILE.exists():
        print(f"Source file already exists: {RAW_FILE}")
        return

    print(f"Downloading Census source file from {CENSUS_URL}")
    urllib.request.urlretrieve(CENSUS_URL, RAW_FILE)
    print(f"Downloaded raw source to {RAW_FILE}")


def normalize_column_name(column):
    if isinstance(column, str):
        column = column.strip()
        if "April 1, 2020" in column and "Estimates Base" in column:
            return "population_2020_base"
        if column.startswith("Population Estimate (as of July 1)_"):
            year_part = column.split("_", 1)[1]
            if year_part.isdigit():
                return f"population_{year_part}"
        return column
    if isinstance(column, float) and column.is_integer():
        return f"population_{int(column)}"
    if isinstance(column, int):
        return f"population_{column}"
    return column


def parse_place_field(text):
    if not isinstance(text, str):
        return None, None

    parts = [part.strip() for part in text.split(",") if part.strip()]
    if len(parts) >= 2:
        state_name = parts[-1]
        city_name = ", ".join(parts[:-1])
        return city_name, state_name

    return text.strip(), None


def build_state_abbreviation(state_name):
    if not isinstance(state_name, str):
        return None
    return STATE_ABBREVIATIONS.get(state_name)


def extract_census_place_type(city_name):
    if not isinstance(city_name, str):
        return None

    normalized = city_name.lower().strip()
    for keyword in PLACE_TYPE_KEYWORDS:
        if normalized.endswith(keyword):
            if keyword == "cdp":
                return "CDP"
            if keyword == "census designated place":
                return "Census Designated Place"
            return keyword.title()
    return None


def flatten_multiindex_columns(columns):
    flat_columns = []
    for col in columns:
        if isinstance(col, tuple):
            primary = str(col[0]).strip()
            secondary = str(col[1]).strip()
            if primary and secondary and "Unnamed" not in secondary:
                flat_columns.append(f"{primary}_{secondary}")
            else:
                flat_columns.append(primary)
        else:
            flat_columns.append(str(col).strip())
    return flat_columns


def load_and_clean_data():
    raw_df = pd.read_excel(RAW_FILE, sheet_name=0, header=[2, 3], engine="openpyxl")
    raw_df.columns = [normalize_column_name(col) for col in flatten_multiindex_columns(raw_df.columns)]

    expected_columns = [
        "Rank",
        "Geographic Area",
        "population_2020_base",
        "population_2020",
        "population_2021",
        "population_2022",
        "population_2023",
        "population_2024",
        "population_2025",
    ]

    raw_df = raw_df.loc[
        raw_df["Rank"].notna() & raw_df["Geographic Area"].notna(), expected_columns
    ]
    raw_df = raw_df.copy()

    numeric_columns = [
        "Rank",
        "population_2020_base",
        "population_2020",
        "population_2021",
        "population_2022",
        "population_2023",
        "population_2024",
        "population_2025",
    ]
    for col in numeric_columns:
        raw_df[col] = pd.to_numeric(raw_df[col], errors="coerce").astype("Int64")

    raw_df["city"], raw_df["state"] = zip(*raw_df["Geographic Area"].map(parse_place_field))
    raw_df["state_abbr"] = raw_df["state"].map(build_state_abbreviation)
    raw_df["census_place_type"] = raw_df["city"].map(extract_census_place_type)

    clean_df = raw_df.loc[raw_df["population_2025"] >= 100_000].copy()
    clean_df = clean_df.sort_values("population_2025", ascending=False).reset_index(drop=True)
    clean_df["rank_2025"] = clean_df.index + 1

    clean_df["absolute_change_2024_to_2025"] = (
        clean_df["population_2025"] - clean_df["population_2024"]
    )
    clean_df["percent_change_2024_to_2025"] = (
        clean_df["absolute_change_2024_to_2025"] / clean_df["population_2024"]
    )

    clean_df["absolute_change_2020_to_2025"] = (
        clean_df["population_2025"] - clean_df["population_2020_base"]
    )
    clean_df["percent_change_2020_to_2025"] = (
        clean_df["absolute_change_2020_to_2025"] / clean_df["population_2020_base"]
    )

    output_columns = [
        "rank_2025",
        "city",
        "state",
        "state_abbr",
        "population_2025",
        "population_2024",
        "population_2020_base",
        "absolute_change_2024_to_2025",
        "percent_change_2024_to_2025",
        "absolute_change_2020_to_2025",
        "percent_change_2020_to_2025",
        "census_place_type",
        "source_name",
        "source_year",
        "source_file",
        "source_note",
    ]

    clean_df = clean_df.assign(
        source_name=SOURCE_NAME,
        source_year=SOURCE_YEAR,
        source_file=SOURCE_FILE_NAME,
        source_note=SOURCE_NOTE,
    )
    clean_df = clean_df[output_columns]

    return raw_df, clean_df


def format_excel_sheet(writer, clean_df):
    workbook = writer.book
    worksheet = writer.sheets["Clean Cities 100k+"]
    worksheet.freeze_panes = worksheet["A2"]

    max_col = worksheet.max_column
    max_row = worksheet.max_row
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    population_cols = [
        "population_2025",
        "population_2024",
        "population_2020_base",
        "absolute_change_2024_to_2025",
        "absolute_change_2020_to_2025",
    ]
    percent_cols = [
        "percent_change_2024_to_2025",
        "percent_change_2020_to_2025",
    ]

    header_names = list(clean_df.columns)
    for col_name in population_cols + percent_cols:
        if col_name not in header_names:
            continue
        column_index = header_names.index(col_name) + 1
        column_letter = get_column_letter(column_index)
        for row in range(2, max_row + 1):
            cell = worksheet[f"{column_letter}{row}"]
            if cell.value is None:
                continue
            if col_name in percent_cols:
                cell.number_format = "0.00%"
            else:
                cell.number_format = "#,##0"

    for idx, column_title in enumerate(header_names, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = max(
            14,
            min(40, len(column_title) + 2),
        )


def add_readme_sheet(writer):
    workbook = writer.book
    if "README" in workbook.sheetnames:
        readme_sheet = workbook["README"]
    else:
        readme_sheet = workbook.create_sheet("README")

    readme_sheet["A1"] = "U.S. Census Bureau City/Place Population Estimates"
    readme_sheet["A2"] = (
        "Source: Annual Estimates of the Resident Population for Incorporated Places "
        "of 20,000 or More, Ranked by July 1, 2025 Population: April 1, 2020 to July 1, 2025."
    )
    readme_sheet["A3"] = f"Census file: {SOURCE_FILE_NAME}"
    readme_sheet["A4"] = "Population estimates are for July 1, 2025."
    readme_sheet["A5"] = "The Clean Cities 100k+ sheet includes U.S. incorporated places with population 100,000 or more."
    readme_sheet["A6"] = "Raw Data contains the imported Census table used to build the cleaned output."
    readme_sheet.column_dimensions["A"].width = 120


def add_landlord_friendliness_sheet(writer, clean_df):
    """Create a Landlord Friendliness sheet with cities that have landlord scores."""
    # Add landlord score to clean_df if not already present
    if "landlord_score" not in clean_df.columns:
        clean_df = clean_df.copy()
        clean_df["landlord_score"] = clean_df["state"].map(
            lambda s: LANDLORD_SCORES.get(s) if isinstance(s, str) else None
        )
    
    # Filter to only rows with a landlord score (not None)
    landlord_df = clean_df[clean_df["landlord_score"].notna()].copy()
    
    # Create output dataframe with normalized names and proper score format
    landlord_output = pd.DataFrame({
        "City": landlord_df["city"].str.strip(),
        "State": landlord_df["state"].str.strip(),
        "Landlord Score": landlord_df["landlord_score"].astype(int).astype(str)
    })
    
    # Convert +1, 0, -1 to proper format
    landlord_output["Landlord Score"] = landlord_output["Landlord Score"].apply(
        lambda x: f"+{x}" if x == "1" else x
    )
    
    landlord_output.to_excel(writer, sheet_name="Landlord Friendliness", index=False)
    
    # Format the sheet
    workbook = writer.book
    worksheet = writer.sheets["Landlord Friendliness"]
    worksheet.freeze_panes = worksheet["A2"]
    
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # Set column widths
    for idx, col_name in enumerate(["City", "State", "Landlord Score"], start=1):
        col_letter = get_column_letter(idx)
        if col_name == "City":
            worksheet.column_dimensions[col_letter].width = 30
        elif col_name == "State":
            worksheet.column_dimensions[col_letter].width = 20
        else:
            worksheet.column_dimensions[col_letter].width = 15


def export_to_excel(raw_df, clean_df):
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        raw_df.to_excel(writer, sheet_name="Raw Data", index=False)
        clean_df.to_excel(writer, sheet_name="Clean Cities 100k+", index=False)
        add_readme_sheet(writer)
        add_landlord_friendliness_sheet(writer, clean_df)
        format_excel_sheet(writer, clean_df)

    print(f"Saved output workbook to {OUTPUT_FILE}")


def main():
    ensure_directories()
    download_source_file()
    raw_df, clean_df = load_and_clean_data()
    export_to_excel(raw_df, clean_df)

    print(f"Total U.S. cities/places with population >= 100,000: {len(clean_df)}")
    print("Top 10 cities by 2025 population:")
    print(clean_df[["rank_2025", "city", "state", "population_2025"]].head(10).to_string(index=False))
    print("\nBottom 10 cities included:")
    print(clean_df[["rank_2025", "city", "state", "population_2025"]].tail(10).to_string(index=False))


if __name__ == "__main__":
    main()
