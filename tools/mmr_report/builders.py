from openpyxl.styles import Font
from datetime import datetime
import re

from tools.mmr_report.charts import add_expiring_leases_chart, add_projected_occupancy_chart
from tools.mmr_report.detection import detect_source_system, source_system_display
from tools.mmr_report.helpers import fmt_date, fmt_month, fmt_pct
from tools.mmr_report.sheets import default_box_score
from tools.mmr_report.styles import C, DARK_BLUE, L, LIGHT_BLUE, PALE_BLUE, R, TL, _FONTS, _box, _fill, col_hdr, data_row, merge_band, merge_wc, section_hdr, wc




# ══════════════════════════════════════════════════════════════════════════
#  SUMMARY BUILDER
# ══════════════════════════════════════════════════════════════════════════

# Column layout (1-indexed):
#  A=1  spacer
#  B=2  labels / table col 1
#  C=3  values / table col 2
#  D=4  table col 3
#  E=5  table col 4
#  F=6  WO description / table col 5
#  G=7  Projected Occ – Week
#  H=8  Projected Occ – Units
#  I=9  Projected Occ – %

COL_WIDTHS = {
    "A": 2,
    "B": 18,
    "C": 13,
    "D": 16,
    "E": 13,
    "F": 18,
    "G": 13,
    "H": 2,
    "I": 13,
    "J": 13,
    "K": 13,
    "L": 13,
    "M": 13,
    "N": 13,
    "O": 13,
    "P": 13,
}




def build_summary_legacy(wb, data):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    bs = data["box_score"]
    dl = data["delinquency"]
    rr = data["rent_roll"]
    au = data["available_units"]
    el = data["expiring_leases"]
    ps = data["prospect_sources"]
    wo = data["work_orders"]

    r = 1  # running row cursor

    # ── TITLE BLOCK ───────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 24
    merge_wc(ws, r, 2, 9, bs["property_name"], font=_FONTS["title"], align=C)
    r += 1

    merge_wc(ws, r, 2, 9, "Weekly Property Summary",
             font=_FONTS["sub"], fill=None, align=C)
    r += 1
    merge_wc(ws, r, 2, 9, bs["date_range"],
             font=_FONTS["data"], fill=None, align=C)
    r += 1
    merge_wc(ws, r, 2, 9, f"Date Printed: {bs['printed']}",
             font=_FONTS["meta"], fill=None, align=C)
    r += 1
    ws.row_dimensions[r].height = 6
    r += 1

    top_row = r  # start of two-column section

    # ── LEFT: OCCUPANCY ───────────────────────────────────────────────────
    section_hdr(ws, r, "OCCUPANCY", c1=2, c2=5)
    r += 1

    occ_rows = [
        ("% Occupancy",    fmt_pct(bs["pct_occ"])),
        ("Occupied (Occ)", bs["occupied"]),
        ("Vacant",         bs["vacant"]),
        ("Preleases (Vacant + On-Notice)", au["prelease_count"]),
        ("On-Notice",      bs["on_notice"]),
    ]
    for label, val in occ_rows:
        wc(ws, r, 2, label, font="label", align=L)
        wc(ws, r, 3, val,   font="data",  align=R)
        r += 1

    ws.row_dimensions[r].height = 6
    r += 1

    # ── LEFT: LEASING ACTIVITY ────────────────────────────────────────────
    section_hdr(ws, r, "LEASING ACTIVITY", c1=2, c2=5)
    r += 1

    for label, val in [("Applied", bs["applied"]), ("Approved", bs["approved"]), ("Signed", bs["signed"])]:
        wc(ws, r, 2, label, font="label", align=L)
        wc(ws, r, 3, val,   font="data",  align=R)
        r += 1

    ws.row_dimensions[r].height = 6
    r += 1

    # ── LEFT: DELINQUENCY ─────────────────────────────────────────────────
    section_hdr(ws, r, "DELINQUENCY", c1=2, c2=5)
    r += 1
    wc(ws, r, 2, "Total Delinquency", font="label", align=L)
    wc(ws, r, 3, dl["total"],          font="data",  align=R, num_fmt='"$"#,##0.00')
    r += 1

    ws.row_dimensions[r].height = 6
    r += 1

    # ── LEFT: RENTAL INCOME ───────────────────────────────────────────────
    section_hdr(ws, r, "RENTAL INCOME TO DATE", c1=2, c2=5)
    r += 1
    wc(ws, r, 2, "Total Rental Revenue",     font="label", align=L)
    wc(ws, r, 3, rr["total_rental"],          font="data",  align=R, num_fmt='"$"#,##0.00')
    r += 1
    wc(ws, r, 2, "Average Rent / Unit / Mo", font="label", align=L)
    wc(ws, r, 3, rr["avg_rent"],              font="data",  align=R, num_fmt='"$"#,##0.00')
    r += 1

    bottom_left = r  # track where the left column ends

    # ── RIGHT: PROJECTED OCCUPANCY ────────────────────────────────────────
    pr = top_row
    section_hdr(ws, pr, "PROJECTED OCCUPANCY", c1=7, c2=9)
    pr += 1

    col_hdr(ws, pr, 7, "Week")
    col_hdr(ws, pr, 8, "Occ Units")
    col_hdr(ws, pr, 9, "% Occupied")
    pr += 1

    for i, entry in enumerate(bs["proj_occ"]):
        z = i % 2 == 0
        data_row(ws, pr, 7, fmt_date(entry["date"]), zebra=z, align=C)
        data_row(ws, pr, 8, entry["occ"],             zebra=z, align=C)
        data_row(ws, pr, 9, entry["pct"],             zebra=z, align=C, num_fmt="0.0%")
        pr += 1

    r = max(r, pr)
    ws.row_dimensions[r].height = 8
    r += 1

    # ══════════════════════════════════════════════════════════════════════
    #  FULL-WIDTH TABLES
    # ══════════════════════════════════════════════════════════════════════

    # ── READY UNITS ───────────────────────────────────────────────────────
    section_hdr(ws, r, "READY UNITS — VACANT & VACANT PRE-LEASED")
    r += 1

    for col, label in zip([2, 3, 4, 5], ["Unit", "Unit Type", "Section / Status", "Unit Status"]):
        col_hdr(ws, r, col, label)
    r += 1

    if au["ready_units"]:
        for i, unit in enumerate(au["ready_units"]):
            z = i % 2 == 0
            data_row(ws, r, 2, unit["unit"],    zebra=z, align=C)
            data_row(ws, r, 3, unit["type"],    zebra=z, align=L)
            data_row(ws, r, 4, unit["section"], zebra=z, align=L)
            data_row(ws, r, 5, unit["status"],  zebra=z, align=C)
            r += 1
    else:
        merge_wc(ws, r, 2, 5, "No ready units found", font="data", fill=None, align=C)
        r += 1

    ws.row_dimensions[r].height = 8
    r += 1

    # ── EXPIRING LEASES BY MONTH ──────────────────────────────────────────
    section_hdr(ws, r, "EXPIRING LEASES BY MONTH (NEXT 10 MONTHS)")
    r += 1

    for col, label in zip([2, 3, 4], ["Month", "Lease Expirations", "Renewal Starts"]):
        col_hdr(ws, r, col, label)
    r += 1

    if el:
        for i, m in enumerate(el):
            z = i % 2 == 0
            data_row(ws, r, 2, fmt_month(m["dt"]),  zebra=z, align=L)
            data_row(ws, r, 3, m["expirations"],    zebra=z, align=C)
            data_row(ws, r, 4, m["renewals"],       zebra=z, align=C)
            r += 1
    else:
        merge_wc(ws, r, 2, 4, "No expiring lease data", font="data", fill=None, align=C)
        r += 1

    ws.row_dimensions[r].height = 8
    r += 1

    # ── TOP 2 PROSPECT SOURCES ────────────────────────────────────────────
    section_hdr(ws, r, "TOP 2 PROSPECT SOURCES")
    r += 1

    for col, label in zip([2, 3, 4, 5, 6],
                          ["Category", "#1 Source", "#1 Count", "#2 Source", "#2 Count"]):
        col_hdr(ws, r, col, label)
    r += 1

    METRIC_LABELS = {
        "New Prospects":    "New Prospects",
        "Return Prospects": "Return Prospects",
        "New Apps":         "New Applications",
        "Net Leases":       "Net Leases",
    }
    for i, (key, label) in enumerate(METRIC_LABELS.items()):
        z = i % 2 == 0
        ranked = ps.get(key, [])
        s1, c1_v = ranked[0] if len(ranked) > 0 else ("—", 0)
        s2, c2_v = ranked[1] if len(ranked) > 1 else ("—", 0)
        data_row(ws, r, 2, label, zebra=z, align=L)
        data_row(ws, r, 3, s1,    zebra=z, align=L)
        data_row(ws, r, 4, c1_v,  zebra=z, align=C)
        data_row(ws, r, 5, s2,    zebra=z, align=L)
        data_row(ws, r, 6, c2_v,  zebra=z, align=C)
        r += 1

    ws.row_dimensions[r].height = 8
    r += 1

    # ── OPEN WORK ORDERS ──────────────────────────────────────────────────
    section_hdr(ws, r, f"OPEN WORK ORDERS  ({len(wo['work_orders'])} total)")
    r += 1

    # Issue type summary mini-table
    ic = wo["issue_counts"]
    if any(ic.values()):
        wc(ws, r, 2, "Issue Type", font="col_hdr", fill=_fill(LIGHT_BLUE), align=C, border=_box())
        wc(ws, r, 3, "Count",      font="col_hdr", fill=_fill(LIGHT_BLUE), align=C, border=_box())
        r += 1
        for j, (issue, cnt) in enumerate(ic.items()):
            z = j % 2 == 0
            data_row(ws, r, 2, issue, zebra=z, align=L)
            data_row(ws, r, 3, cnt,   zebra=z, align=C)
            r += 1
        ws.row_dimensions[r].height = 6
        r += 1

    # WO detail table
    for col, label in zip([2, 3, 4, 5, 6],
                          ["WO #", "Unit / Location", "Date Reported", "Category", "Description"]):
        col_hdr(ws, r, col, label)
    r += 1

    if wo["work_orders"]:
        for i, order in enumerate(wo["work_orders"]):
            z = i % 2 == 0
            desc = order["description"]
            if len(desc) > 300:
                desc = desc[:297] + "..."

            data_row(ws, r, 2, order["number"],   zebra=z, align=C)
            data_row(ws, r, 3, order["location"], zebra=z, align=C)
            data_row(ws, r, 4, fmt_date(order["reported"]), zebra=z, align=C)
            data_row(ws, r, 5, order["category"], zebra=z, align=L)

            # Description cell — allow text wrap and auto-height
            fill = _fill(PALE_BLUE) if z else None
            cell = ws.cell(row=r, column=6, value=desc)
            cell.font      = _FONTS["data"]
            cell.alignment = TL
            cell.border    = _box()
            if fill:
                cell.fill = fill

            ws.row_dimensions[r].height = 45
            r += 1
    else:
        merge_wc(ws, r, 2, 6, "No open work orders", font="data", fill=None, align=C)
        r += 1

    # Footer
    ws.row_dimensions[r].height = 6
    r += 1
    merge_wc(ws, r, 2, 9,
             f"Generated by FIRE Capital MMR Summary Tool  |  {datetime.now().strftime('%m/%d/%Y %I:%M %p')}",
             font=_FONTS["meta"], fill=None, align=C)

    # Freeze panes below title block
    ws.freeze_panes = ws.cell(row=6, column=1)

    return ws




def section_band(ws, row, title, c1=2, c2=7):
    return merge_band(ws, row, c1, c2, title, font="hdr", fill_color=DARK_BLUE, align=C)




def write_kv(ws, row, label_col, value_col, label, value, num_fmt=None):
    wc(ws, row, label_col, label, font="label", align=L)
    wc(ws, row, value_col, value, font="data", align=R, num_fmt=num_fmt)




def na_if_none(value):
    return "N/A" if value is None else value




def write_pair_row(ws, row, left_label, left_value, right_label=None, right_value=None, left_fmt=None, right_fmt=None):
    write_kv(ws, row, 2, 3, left_label, left_value, left_fmt)
    if right_label is not None:
        write_kv(ws, row, 5, 7, right_label, right_value, right_fmt)




def setup_summary_print(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes     = None   # BUG2: no frozen rows
    ws.print_title_rows = None   # BUG2: no repeated header on print
    ws.print_area = "A1:P54"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False




_PROPERTY_ABBREVS = {
    "oxford pointe":  "OXPT",
    "eagle rock":     "ERA",
    "the canyon":     "Canyon",
    "canyon":         "Canyon",
    "maple valley":   "Maple Valley",
}




def make_download_filename(property_name: str, date_range: str, printed: str = "") -> str:
    """Return e.g. 'OXPT Summary 06.22.26.xlsx' from property name + date range."""
    pn = (property_name or "").strip()
    pn_lower = pn.lower()
    range_text = str(date_range or "")
    range_lower = range_text.lower()
    is_appfolio_range = (
        "maple valley" in pn_lower
        or (" to " in range_lower and "trailing" in range_lower)
        or range_lower.startswith("period range:")
    )

    abbrev = None
    for key, val in _PROPERTY_ABBREVS.items():
        if key in pn_lower:
            abbrev = val
            break
    if abbrev is None:
        # First word, skipping "The" prefix
        words = pn.split()
        abbrev = words[1] if words and words[0].lower() == "the" and len(words) > 1 else (words[0] if words else "Property")

    def parse_printed_date(value):
        clean_printed = str(value or "").replace("Printed", "").replace("Exported On:", "").strip()
        date_match = re.search(r"\d{1,2}/\d{1,2}/\d{4}", clean_printed)
        if date_match:
            return datetime.strptime(date_match.group(0), "%m/%d/%Y")
        return None

    def parse_resman_range_end(value):
        end_str = str(value or "").split(" - ")[-1].strip()
        return datetime.strptime(end_str, "%m/%d/%Y")

    def parse_appfolio_range_end(value):
        import calendar as _cal
        match = re.search(r"to\s+([A-Za-z]+\s+\d{4})", str(value or ""))
        if not match:
            return None
        month_label = match.group(1).strip()
        for fmt in ("%b %Y", "%B %Y"):
            try:
                dt = datetime.strptime(month_label, fmt)
                last_day = _cal.monthrange(dt.year, dt.month)[1]
                return dt.replace(day=last_day)
            except ValueError:
                pass
        return None

    dt = None
    for parser, value in (
        (parse_printed_date, printed),
        (parse_resman_range_end, None if is_appfolio_range else range_text),
        (parse_appfolio_range_end, None if not is_appfolio_range else range_text),
    ):
        if dt is not None or not value:
            continue
        try:
            dt = parser(value)
        except Exception:
            dt = None

    if dt is None:
        dt = datetime.now()

    return f"{abbrev} Summary {dt.strftime('%m.%d.%y')}.xlsx"




def build_summary(wb, data):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    # Columns R–U no longer hold visible data; chart data lives in chart_data sheet

    bs = data.get("box_score", default_box_score())
    dl = data.get("delinquency") or {"total": None}
    rr = data.get("rent_roll") or {"total_rental": None, "avg_rent": None}
    au = data.get("available_units") or {"ready_units": None, "prelease_count": None, "holding_count": None, "eviction_count": None}
    el = data.get("expiring_leases", [])
    ps = data.get("prospect_sources")
    wo = data.get("work_orders") or {"work_orders": None, "issue_counts": {}}
    source_system = data.get("source_system") or detect_source_system(wb)
    source_text, source_color, source_fill = source_system_display(source_system)

    setup_summary_print(ws)
    for row in range(1, 55):
        ws.row_dimensions[row].height = 15
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 19
    ws.row_dimensions[5].height = 4
    ws.row_dimensions[30].height = 4
    ws.row_dimensions[53].height = 4

    title = bs.get("property_name") or "MMR Summary"
    period = bs.get("date_range") or ""
    printed = bs.get("printed") or ""

    merge_band(ws, 1, 2, 16, title, font=_FONTS["title"], fill_color=None, align=C)
    merge_band(ws, 2, 2, 16, f"Weekly Property Summary  |  {period}  |  Date Printed: {printed}".strip(" |"),
               font=_FONTS["sub"], fill_color=None, align=C)
    merge_band(
        ws,
        3,
        2,
        16,
        f"Source System: {source_text}",
        font=Font(name="Calibri", bold=True, color=source_color, size=11),
        fill_color=source_fill,
        align=C,
        border=_box(),
    )

    # Left side: key stats and compact detail tables.
    section_band(ws, 6, "OCCUPANCY", 2, 7)
    occupancy_value = fmt_pct(bs.get("pct_occ")) if bs.get("pct_occ") is not None else "N/A"
    write_pair_row(ws, 7, "% Occupancy", occupancy_value, "Occupied", na_if_none(bs.get("occupied")))
    # Vacant is Box Score's own total_units - occupied -- units on notice
    # (resident gave notice, but hasn't moved out yet) are still physically
    # occupied and already counted in "Occupied", so they must not also be
    # added here. Doing so previously double-counted them, inflating
    # Vacant by exactly the on-notice count (confirmed against Eagle
    # Rock's 07.20.26 file: Total 92 - Occupied 86 = 6 real vacant units,
    # cross-checked against the raw Available Units report's Vacant (5)
    # + Vacant Preleased (1) sections; the prior code showed 15, exactly
    # 6 + the file's 9 on-notice units).
    vacant_total = bs.get("vacant")
    write_pair_row(ws, 8, "Vacant", na_if_none(vacant_total), "Total Units", na_if_none(bs.get("total_units")))
    # Prefer Box Score "Vacant Pre-Leased" column value; fall back to Available Units count.
    # This already includes Notice-to-Vacate-Preleased units (see _PRELEASE_SECTIONS).
    prelease_val = bs.get("prelease_count") if bs.get("prelease_count") is not None else au.get("prelease_count", 0)
    write_pair_row(ws, 9, "Vacant Preleased", na_if_none(prelease_val))
    write_pair_row(ws, 10, "Holding", na_if_none(au.get("holding_count")))
    write_pair_row(ws, 11, "Eviction", na_if_none(au.get("eviction_count")))

    section_band(ws, 14, "LEASING / FINANCIAL", 2, 7)
    write_pair_row(ws, 15, "Applied", na_if_none(bs.get("applied")), "Approved", na_if_none(bs.get("approved")))
    write_pair_row(ws, 16, "Signed", na_if_none(bs.get("signed")), "Total Delinquency", na_if_none(dl.get("total")), right_fmt='"$"#,##0.00')
    write_pair_row(ws, 17, "Total Rental Revenue", na_if_none(rr.get("total_rental")), "Average Rent / Unit", na_if_none(rr.get("avg_rent")),
                   left_fmt='"$"#,##0.00', right_fmt='"$"#,##0.00')

    ready_units = au.get("ready_units")
    ready_count = "N/A" if ready_units is None else len(ready_units)
    section_band(ws, 20, f"READY UNITS ({ready_count} total)", 2, 7)
    for col, label in zip([2, 3, 4, 5, 6, 7], ["Unit", "Section", "Status", "Unit", "Section", "Status"]):
        col_hdr(ws, 21, col, label)
    if ready_units is None:
        merge_band(ws, 22, 2, 7, "N/A", font="data", fill_color=None, align=C, border=_box())
    elif ready_units:
        for idx, unit in enumerate(ready_units[:18]):
            row = 22 + (idx % 9)
            base_col = 2 if idx < 9 else 5
            z = (row - 22) % 2 == 0
            data_row(ws, row, base_col, unit.get("unit", ""), zebra=z, align=C)
            data_row(ws, row, base_col + 1, unit.get("section", ""), zebra=z, align=L)
            data_row(ws, row, base_col + 2, unit.get("status", ""), zebra=z, align=C)
    else:
        merge_band(ws, 22, 2, 7, "No ready units found", font="data", fill_color=None, align=C, border=_box())

    section_band(ws, 32, "TOP 2 PROSPECT SOURCES", 2, 7)
    for col, label in zip([2, 3, 4, 5, 7], ["Category", "#1 Source", "#1 Count", "#2 Source", "#2 Count"]):
        col_hdr(ws, 33, col, label)
    metric_labels = {
        "New Prospects": "New Prospects",
        "Return Prospects": "Return Prospects",
        "New Apps": "New Applications",
        "Net Leases": "Net Leases",
    }
    if ps is None:
        merge_band(ws, 34, 2, 7, "N/A", font="data", fill_color=None, align=C, border=_box())
        metric_labels = {}
    for i, (key, label) in enumerate(metric_labels.items(), 34):
        ranked = ps.get(key, [])
        s1, c1_v = ranked[0] if len(ranked) > 0 else ("—", 0)
        s2, c2_v = ranked[1] if len(ranked) > 1 else ("—", 0)
        z = (i - 34) % 2 == 0
        data_row(ws, i, 2, label, zebra=z, align=L)
        data_row(ws, i, 3, s1, zebra=z, align=L)
        data_row(ws, i, 4, c1_v, zebra=z, align=C)
        data_row(ws, i, 5, s2, zebra=z, align=L)
        data_row(ws, i, 7, c2_v, zebra=z, align=C)

    work_orders = wo.get("work_orders")
    work_order_count = "N/A" if work_orders is None else len(work_orders)
    section_band(ws, 41, f"EMERGENCY WORK ORDERS ({work_order_count} total)", 2, 7)
    issue_counts = wo.get("issue_counts", {})
    summary_text = "N/A" if work_orders is None else (" | ".join(f"{issue}: {count}" for issue, count in issue_counts.items()) or "No emergency work orders")
    wc(ws, 42, 2, "Issue Types", font="col_hdr", fill=_fill(LIGHT_BLUE), align=C, border=_box())
    merge_band(ws, 42, 3, 7, summary_text, font="data", fill_color=LIGHT_BLUE, align=L, border=_box())

    for col, label in zip([2, 3, 4, 5], ["WO #", "Unit/Location", "Date Reported", "Category"]):
        col_hdr(ws, 43, col, label)
    merge_band(ws, 43, 6, 7, "Description", font="col_hdr", fill_color=LIGHT_BLUE, align=C, border=_box())
    if work_orders is None:
        merge_band(ws, 44, 2, 7, "N/A", font="data", fill_color=None, align=C, border=_box())
    elif work_orders:
        for i, order in enumerate(work_orders[:8], 44):
            z = (i - 44) % 2 == 0
            data_row(ws, i, 2, order.get("number", ""), zebra=z, align=C)
            data_row(ws, i, 3, order.get("location", ""), zebra=z, align=C)
            data_row(ws, i, 4, order.get("date_reported") or fmt_date(order.get("reported")), zebra=z, align=C)
            data_row(ws, i, 5, order.get("category", ""), zebra=z, align=L)
            merge_band(ws, i, 6, 7, order.get("description", ""), font="data",
                       fill_color=PALE_BLUE if z else None, align=L, border=_box())
    else:
        merge_band(ws, 44, 2, 7, "No emergency work orders", font="data", fill_color=None, align=C, border=_box())

    merge_band(ws, 52, 2, 7, f"Generated by FIRE Capital MMR Summary Tool | {datetime.now().strftime('%m/%d/%Y %I:%M %p')}",
               font=_FONTS["meta"], fill_color=None, align=C)

    # Clean up any leftover chart_data sheet from previous runs
    if "chart_data" in wb.sheetnames:
        del wb["chart_data"]

    # Right side: charts (matplotlib PNGs embedded as images).
    merge_band(ws, 6, 9, 16, "PROJECTED OCCUPANCY", font="hdr", fill_color=DARK_BLUE, align=C)
    add_projected_occupancy_chart(ws, bs.get("proj_occ", []))
    merge_band(ws, 25, 9, 16, "EXPIRING LEASES BY MONTH", font="hdr", fill_color=DARK_BLUE, align=C)
    add_expiring_leases_chart(ws, el)

    # Open directly to this sheet instead of whatever sheet was active before.
    wb.active = wb.sheetnames.index("Summary")
    ws.sheet_view.tabSelected = True

    return ws
