from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side



# ── Colours ────────────────────────────────────────────────────────────────

DARK_BLUE  = "1F4E79"


LIGHT_BLUE = "D6E4F0"


PALE_BLUE  = "EBF5FB"


MID_GRAY   = "595959"


LT_GRAY    = "808080"


GREEN      = "548235"


PALE_GREEN = "E2F0D9"


AMBER      = "C65911"


PALE_AMBER = "FCE4D6"


RED        = "C00000"


PALE_RED   = "F4CCCC"



# ── Style helpers ──────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")



def _side(style="thin", color="BFBFBF"):
    return Side(style=style, color=color)



def _box():
    s = _side()
    return Border(top=s, left=s, right=s, bottom=s)



_FONTS = {
    "title":   Font(name="Calibri", bold=True,  color=DARK_BLUE, size=16),
    "sub":     Font(name="Calibri", bold=True,  color=MID_GRAY,  size=12),
    "meta":    Font(name="Calibri", italic=True, color=LT_GRAY,  size=9),
    "hdr":     Font(name="Calibri", bold=True,  color="FFFFFF",  size=10),
    "col_hdr": Font(name="Calibri", bold=True,  color=DARK_BLUE, size=10),
    "label":   Font(name="Calibri", bold=True,                   size=10),
    "data":    Font(name="Calibri",                               size=10),
}



C = Alignment(horizontal="center", vertical="center", wrap_text=True)


L = Alignment(horizontal="left",   vertical="center", wrap_text=False)


R = Alignment(horizontal="right",  vertical="center")


TL = Alignment(horizontal="left",  vertical="top",    wrap_text=True)




def wc(ws, row, col, value, font="data", fill=None, align=L, border=None, num_fmt=None):
    """Write a single cell with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _FONTS.get(font, _FONTS["data"]) if isinstance(font, str) else font
    cell.alignment = align
    if fill:    cell.fill   = fill
    if border:  cell.border = border
    if num_fmt: cell.number_format = num_fmt
    return cell




def merge_wc(ws, row, c1, c2, value, font="hdr", fill=None, align=C, border=None):
    """Write a merged cell."""
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    return wc(ws, row, c1, value, font=font, fill=fill, align=align, border=border)




def section_hdr(ws, row, title, c1=2, c2=9):
    """Full-width dark-blue section banner."""
    merge_wc(ws, row, c1, c2, title, font="hdr", fill=_fill(DARK_BLUE), align=C)




def col_hdr(ws, row, col, label):
    """Light-blue column header cell."""
    wc(ws, row, col, label, font="col_hdr", fill=_fill(LIGHT_BLUE), align=C, border=_box())




def data_row(ws, row, col, value, zebra=False, align=L, num_fmt=None):
    """Standard data cell with optional zebra shading."""
    fill = _fill(PALE_BLUE) if zebra else None
    wc(ws, row, col, value, font="data", fill=fill, align=align, border=_box(), num_fmt=num_fmt)




def merge_band(ws, row, c1, c2, value, font="hdr", fill_color=DARK_BLUE, align=C, border=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    fill = _fill(fill_color) if fill_color else None
    font_obj = _FONTS.get(font, font) if isinstance(font, str) else font
    for col in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_obj
        cell.alignment = align
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
    ws.cell(row=row, column=c1, value=value)
    return ws.cell(row=row, column=c1)
