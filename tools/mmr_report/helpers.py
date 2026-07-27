from datetime import datetime
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter
import re
from datetime import timedelta



# ── Format helpers ─────────────────────────────────────────────────────────

def fmt_pct(v):
    if isinstance(v, (int, float)):
        return f"{v:.1%}"
    return str(v or "")




def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime("%m/%d/%Y")
    if isinstance(v, (int, float)) and v > 0:
        try:
            return from_excel(v).strftime("%m/%d/%Y")
        except Exception:
            pass
    if isinstance(v, str) and re.match(r"\d{1,2}/\d{1,2}/\d{4}", v):
        return v
    return str(v or "")




def fmt_month(v):
    if isinstance(v, datetime):
        return v.strftime("%B %Y")
    return str(v or "")



# ── Sheet reader ───────────────────────────────────────────────────────────

def rows_of(ws):
    """Return all cell values as a flat list-of-lists (0-indexed)."""
    return [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]



# ── Parse helpers ──────────────────────────────────────────────────────────

def norm(s):
    """Normalize for comparison: collapse whitespace, strip, and lowercase."""
    return re.sub(r"\s+", " ", str(s or "").replace("\xa0", " ")).strip().lower()



def safe_get(row, i, default=None):
    """Access row[i] without raising IndexError."""
    try:
        return row[i]
    except (IndexError, TypeError):
        return default




def safe_row(rows, i):
    """Access rows[i] without raising IndexError."""
    try:
        return rows[i]
    except (IndexError, TypeError):
        return []



def coerce_pct(v):
    """
    Convert various percent representations to a 0–1 float.
      0.947   → 0.947
      "94.7%" → 0.947
      "0.947" → 0.947
      94.7    → 0.947  (value > 1 treated as already-in-percent form)
    Returns None if conversion fails.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f <= 1.0 else f / 100.0
    s = str(v).strip().rstrip("%")
    try:
        f = float(s)
        return f if f <= 1.0 else f / 100.0
    except ValueError:
        return None



def coerce_num(v, default=0.0):
    """Convert a value to float, returning default on failure."""
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v or "").strip()
    if not s:
        return default
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        val = float(s)
        return -val if neg else val
    except (ValueError, TypeError):
        return default



def find_col(header_row, *names):
    """First column index (0-based) whose normalized header exactly matches any of names."""
    name_set = {n.lower() for n in names}
    for c, h in enumerate(header_row):
        if norm(h) in name_set:
            return c
    return None



def find_col_contains(header_row, *substrings):
    """First column index whose normalized header contains any of the given substrings."""
    subs = [s.lower() for s in substrings]
    for c, h in enumerate(header_row):
        hn = norm(h)
        if any(sub in hn for sub in subs):
            return c
    return None



def debug_box_score_preleases(header_row, total_row, vacant_prelease_col, notice_prelease_col):
    """Print Box Score occupancy headers and the chosen Preleases sources."""
    print("  Box Score occupancy headers:")
    for c, header in enumerate(header_row, 1):
        if header is not None and str(header).strip():
            print(f"    {get_column_letter(c)}: {str(header).replace(chr(10), ' ')}")

    def describe(col):
        if col is None:
            return "missing", 0
        header = str(safe_get(header_row, col) or "").replace(chr(10), " ")
        value = coerce_num(safe_get(total_row, col), default=0)
        return f"{int(value)} from {get_column_letter(col + 1)} ({header})", int(value)

    vacant_desc, vacant_value = describe(vacant_prelease_col)
    notice_desc, notice_value = describe(notice_prelease_col)
    print(f"  Preleases picked: Vacant Pre-Leased = {vacant_desc}")
    print(f"  Preleases picked: On-Notice Pre-Leased = {notice_desc}")
    print(f"  Preleases total: {vacant_value} + {notice_value} = {vacant_value + notice_value}")




def coerce_excel_date(value):
    """Convert Excel date values, date strings, or serials to datetime."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and value > 0:
        try:
            return from_excel(value)
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                pass
    return None




def describe_raw(value):
    return f"{value!r}<{type(value).__name__}>"




def find_projected_occupancy_header(rows, start_idx):
    for i in range(start_idx, min(len(rows), start_idx + 8)):
        row = rows[i]
        if find_col(row, "date") is None:
            continue
        if find_col(row, "occupied units") is None and find_col(row, "occupancy", "% occupancy", "% occupied") is None:
            continue
        return i, row
    return None, None




def parse_projected_occupancy(rows, start_idx, total_units):
    header_idx, header = find_projected_occupancy_header(rows, start_idx)
    if header_idx is None:
        print("WARNING: Projected Occupancy header row not found.")
        return []

    date_col = find_col(header, "date")
    occ_u_col = find_col(header, "occupied units")
    pct_col = find_col(header, "occupancy", "% occupancy", "% occupied")

    print("  Projected Occupancy headers:")
    for c, value in enumerate(header, 1):
        if value is not None and str(value).strip():
            print(f"    {get_column_letter(c)}: {str(value).replace(chr(10), ' ')}")

    proj_occ = []
    for row_num, drow in enumerate(rows[header_idx + 1:], header_idx + 2):
        raw_date = safe_get(drow, date_col)
        raw_occ_units = safe_get(drow, occ_u_col) if occ_u_col is not None else None
        raw_pct = safe_get(drow, pct_col) if pct_col is not None else None

        dt = coerce_excel_date(raw_date)
        if dt is None:
            if raw_date and not is_junk_row(drow):
                print(f"    stop row {row_num}: date={describe_raw(raw_date)}")
            break

        occ_u = coerce_num(raw_occ_units, default=None) if occ_u_col is not None else None
        pct = coerce_pct(raw_pct) if pct_col is not None else None
        if pct is None and total_units and occ_u is not None:
            pct = occ_u / total_units
        if occ_u is None and total_units and pct is not None:
            occ_u = pct * total_units

        print(
            f"    row {row_num}: raw_date={describe_raw(raw_date)}, "
            f"parsed_date={fmt_date(dt)}, raw_occupied_units={describe_raw(raw_occ_units)}, "
            f"raw_occupancy={describe_raw(raw_pct)}, parsed_pct={pct}"
        )

        if pct is None:
            continue
        occ_display = int(occ_u) if isinstance(occ_u, float) and occ_u.is_integer() else occ_u
        proj_occ.append({"date": dt, "occ": occ_display, "pct": pct})

    print(f"  Projected Occupancy rows extracted: {len(proj_occ)}")
    if 0 < len(proj_occ) < 20:
        last = proj_occ[-1]
        print(
            "  Projected Occupancy extension: carrying forward "
            f"{last['occ']} occupied units / {last['pct']:.4%} through 20 weekly points"
        )
        while len(proj_occ) < 20:
            next_date = proj_occ[-1]["date"] + timedelta(days=7)
            proj_occ.append({
                "date": next_date,
                "occ": last["occ"],
                "pct": last["pct"],
                "carried_forward": True,
            })
            print(
                f"    extended row {len(proj_occ)}: parsed_date={fmt_date(next_date)}, "
                f"occupied_units={last['occ']}, parsed_pct={last['pct']}"
            )

    return proj_occ




def is_junk_row(row):
    """True for blank rows, copyright lines, or ResMan footer rows."""
    first = safe_get(row, 0)
    if first is None:
        return True
    s = str(first).strip()
    return not s or s.startswith("©") or s.startswith("*") or "ResMan" in s or s.startswith("Printed")




def nonempty_values(row):
    """Return non-empty values from a row."""
    return [v for v in row if v is not None and str(v).strip()]




def looks_like_group_header(row):
    """A ResMan section/status header usually has text only in the first cell."""
    first = safe_get(row, 0)
    if not isinstance(first, str) or not first.strip():
        return False
    return len(nonempty_values(row[1:])) == 0




def looks_like_unit_value(v):
    """Accept string or numeric unit IDs while rejecting obvious labels/footers."""
    if v is None:
        return False
    s = str(v).strip()
    if not s or s.startswith("*") or s.startswith("©"):
        return False
    return norm(s) not in {"unit", "total", "totals", "grand total"}




def worksheet_contains(ws, *phrases, max_rows=30, max_cols=30):
    """True if any of the phrases appears in the top-left area of a worksheet."""
    wanted = [norm(p) for p in phrases]
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, max_rows),
        min_col=1,
        max_col=min(ws.max_column, max_cols),
        values_only=True,
    ):
        for value in row:
            text = norm(value)
            if text and any(p in text for p in wanted):
                return True
    return False




def find_first_text(wb, predicate, max_rows=30, max_cols=12):
    """Find the first top-left workbook cell whose text satisfies predicate."""
    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, max_rows),
            min_col=1,
            max_col=min(ws.max_column, max_cols),
            values_only=True,
        ):
            for value in row:
                text = str(value or "").strip()
                if text and predicate(text):
                    return text
    return ""
