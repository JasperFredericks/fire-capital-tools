"""
FIRE Capital Tools - Scorecard Pro.

Flask Blueprint plus the core Scorecard Pro parsing, KPI, reporting, and
export logic ported from Michelle Jeong's standalone Streamlit prototype.
"""

from __future__ import annotations

import base64
import csv
import datetime
import io
import json
import math
import os
import re
import secrets
import time
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.image as mpimg
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
from flask import (
    Blueprint,
    abort,
    current_app,
    jsonify,
    render_template,
    request,
    send_file,
    session,
)
from flask_login import login_required
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.patches import Polygon
from openpyxl.styles import Font, PatternFill
from werkzeug.utils import secure_filename

from tools import scorecard_history


MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_INDEX = {month: idx + 1 for idx, month in enumerate(MONTHS)}
ALLOWED_PNL_EXT = {".csv", ".xlsx", ".xlsm"}
ALLOWED_SCORECARD_EXT = {".xlsx", ".xlsm"}
MAX_PENDING = 8


class PnLParser:
    def __init__(self, filepath):
        self.property_name = "Unknown Property"
        self.period = "Unknown Period"
        self.accounts = {}
        self.detected_format = "Unknown"
        self.warnings = []

        if Path(filepath).suffix.lower() in (".xlsx", ".xlsm"):
            self.filepath = self._convert_workbook_to_csv(filepath)
        else:
            self.filepath = filepath

        # Standard Mapping for "The View" (Name -> Code)
        self.name_map = {
            "Gross potential rent": "4110",
            "loss to vacancy & other": "4220",
            "Net rental income": "4000",
            "Other income": "4300",
            "Personnel": "6400",
            "Maintenance & repairs": "6530",
            "Turn expenses": "6500",
            "Marketing & resident retention": "6300",
            "Insurance": "6700",
            "Property taxes": "6800",
            "Utilities subtotal": "6600",
            "Trash": "6500",
            "Revenues": "Total Income",
        }

        # Standard Mapping for "Paresh" (Paresh Code -> ERA Code)
        self.code_map = {
            "40210": "4110",
            "40310": "4220",
            "40200": "4000",
        }

        # Standard Mapping for "OXPT" (Name -> Code)
        self.oxpt_map = {
            "Gross Potential Rent": "4110",
            "Vacancy": "4220",
            "Total RENTS": "4000",
            "Cost Recovery Fee": "4300",
            "Total FEES": "4300",
            "Laundry Income": "4300",
            "Parking Income": "4300",
            "Total MANAGEMENT FEES": "6113",
            "Total INSURANCE": "6700",
            "Total GROUNDS & LAWN MAINTENANCE": "6515",
            "Total OFFICE EXPENSE": "6100",
            "Total PAYROLL EXPENSE": "6400",
            "Total CLEANING & TRASH REMOVAL": "6520",
            "Total REPAIRS": "6530",
            "Total TAXES": "6800",
            "Total UTILITIES": "6600",
            "Total OUTSIDE CONTRACTORS": "6500",
            "Total OTHER EXPENSES": "6100",
            "Total Operating Income": "9998",
            "Total Operating Expense": "9999",
        }

        # Standard Mapping for "Canyon"
        self.canyon_map = {
            "Gross Potential Rent (Scheduled)": "4110",
            "Vacancy Loss": "4220",
            "NET RENTAL REVENUE": "4000",
            "TOTAL OTHER INCOME": "4300",
            "TOTAL PERSONNEL EXPENSES": "6400",
            "TOTAL MANAGEMENT FEES": "6113",
            "TOTAL ADMINISTRATIVE EXPENSES": "6100",
            "TOTAL LEGAL & PROFESSIONAL": "6200",
            "TOTAL MARKETING & LEASING": "6300",
            "TOTAL UTILITIES": "6600",
            "TOTAL CONTRACT SERVICES": "6500",
            "TOTAL TURNOVER / CLEANING": "6520",
            "TOTAL REPAIRS & MAINTENANCE": "6530",
            "TOTAL TAXES & INSURANCE": "6800",
            "TOTAL INCOME": "9998",
            "TOTAL OPERATING EXPENSES": "9999",
        }

    def _convert_workbook_to_csv(self, filepath):
        """
        Convert an uploaded .xlsx/.xlsm P&L export to an equivalent CSV file
        on disk so the rest of PnLParser (format detection, all parse_*
        methods) can keep operating on self.filepath unchanged.

        Prefers a sheet named "Accounting Tree Report" (ResMan's T12 P&L
        export), falling back to the workbook's first sheet for any other
        xlsx P&L layout we haven't seen yet.
        """
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
        sheet_name = next(
            (name for name in wb.sheetnames if name.strip().lower() == "accounting tree report"),
            wb.sheetnames[0],
        )
        ws = wb[sheet_name]

        out_path = Path(filepath).with_suffix(".converted.csv")
        with open(out_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
        return out_path

    def parse(self):
        try:
            header_lines = self._read_head_lines(25)
            content = "".join(header_lines).lower()

            if "cash flow" in content and "account name" in content:
                self.detected_format = "Cash Flow (Generic)"
                self.parse_cash_flow()
            elif "exported on" in content and "account name" in content:
                self.detected_format = "Cash Flow (Generic)"
                self.parse_cash_flow()
            elif header_lines and "category" in header_lines[0].lower() and "canyon apartments" in content:
                self.detected_format = "Canyon"
                self.parse_canyon()
            elif header_lines and "category" in header_lines[0].lower():
                self.detected_format = "Paresh"
                self.parse_paresh()
            elif "account name" in content:
                self.detected_format = "OXPT"
                self.parse_oxpt()
            elif len(header_lines) > 1 and "ltm" in header_lines[1].lower():
                self.detected_format = "The View"
                self.parse_the_view()
            else:
                self.detected_format = "ResMan (Standard)"
                self.parse_resman()
        except Exception as exc:
            self.warnings.append(f"Error parsing CSV: {exc}")

    def _read_head_lines(self, max_lines=20):
        encodings = ["utf-8-sig", "utf-16", "latin-1"]
        for enc in encodings:
            try:
                with open(self.filepath, "r", encoding=enc, errors="ignore") as handle:
                    return [handle.readline() for _ in range(max_lines)]
            except Exception:
                continue
        return []

    def _read_csv_robust(self, header="infer", skiprows=0, sep=None):
        encodings = ["utf-8-sig", "utf-16", "latin-1"]
        last_err = None
        for enc in encodings:
            try:
                return pd.read_csv(
                    self.filepath,
                    header=header,
                    skiprows=skiprows,
                    sep=sep,
                    engine="python",
                    on_bad_lines="skip",
                    encoding=enc,
                )
            except Exception as exc:
                last_err = exc
                continue
        try:
            return pd.read_csv(self.filepath, header=header, skiprows=skiprows, encoding="utf-8-sig")
        except Exception as exc:
            self.warnings.append(f"CSV read failed: {exc}")
            raise exc if last_err is None else last_err

    def _find_line_index(self, needle):
        needle_lower = needle.lower()
        encodings = ["utf-8-sig", "utf-16", "latin-1"]
        for enc in encodings:
            try:
                with open(self.filepath, "r", encoding=enc, errors="ignore") as handle:
                    for idx, line in enumerate(handle):
                        if needle_lower in line.lower():
                            return idx
            except Exception:
                continue
        return 0

    def _clean_columns(self, df):
        new_cols = []
        seen = {}
        for col in df.columns:
            raw = str(col).replace("\ufeff", "").replace("\n", " ").replace("\r", " ")
            clean = re.sub(r"\s+", " ", raw).strip()
            if clean in seen:
                seen[clean] += 1
                clean = f"{clean}_{seen[clean]}"
            else:
                seen[clean] = 0
            new_cols.append(clean)
        df.columns = new_cols
        return df

    def _infer_default_year(self, labels):
        for label in labels:
            match = re.search(r"(20\d{2})", str(label))
            if match:
                return int(match.group(1))
        return datetime.date.today().year

    def normalize_month(self, raw_month, default_year=None):
        raw = str(raw_month).strip()
        if not raw:
            return None

        raw = raw.replace("\n", " ").replace("\r", " ")
        raw = raw.replace("-", " ").replace("_", " ").replace("/", " ").replace("\\", " ")
        raw = re.sub(r"\s+", " ", raw)

        month_map = {
            "jan": ("Jan", 1),
            "january": ("Jan", 1),
            "feb": ("Feb", 2),
            "february": ("Feb", 2),
            "mar": ("Mar", 3),
            "march": ("Mar", 3),
            "apr": ("Apr", 4),
            "april": ("Apr", 4),
            "may": ("May", 5),
            "jun": ("Jun", 6),
            "june": ("Jun", 6),
            "jul": ("Jul", 7),
            "july": ("Jul", 7),
            "aug": ("Aug", 8),
            "august": ("Aug", 8),
            "sep": ("Sep", 9),
            "sept": ("Sep", 9),
            "september": ("Sep", 9),
            "oct": ("Oct", 10),
            "october": ("Oct", 10),
            "nov": ("Nov", 11),
            "november": ("Nov", 11),
            "dec": ("Dec", 12),
            "december": ("Dec", 12),
        }

        raw_lower = raw.lower()
        month_abbr = None

        for key, (abbr, _) in month_map.items():
            if re.search(rf"\b{key}\b", raw_lower):
                month_abbr = abbr
                break

        if not month_abbr:
            num_match = re.search(r"\b([01]?\d)\b", raw_lower)
            if num_match:
                num = int(num_match.group(1))
                if 1 <= num <= 12:
                    month_abbr = MONTHS[num - 1]

        yr_match = re.search(r"(20\d{2}|\d{2})", raw_lower)
        year = None
        if yr_match:
            yr = yr_match.group(1)
            year = int("20" + yr) if len(yr) == 2 else int(yr)
        elif default_year:
            year = int(default_year)

        if month_abbr and year:
            return f"{month_abbr} {year}"
        if month_abbr and not year:
            return month_abbr
        return None

    def _parse_amount(self, raw_val):
        if raw_val is None:
            return 0.0
        value = str(raw_val).strip()
        if not value or value.lower() in ("nan", "none", "null"):
            return 0.0
        if re.fullmatch(r"[-\u2013\u2014]+", value):
            return 0.0
        value = value.replace("$", "").replace(",", "").replace('"', "").strip()
        negative = False
        if value.startswith("(") and value.endswith(")"):
            negative = True
            value = value[1:-1].strip()
        if value.endswith("-"):
            negative = True
            value = value[:-1].strip()
        try:
            parsed = float(value)
        except Exception:
            return 0.0
        return -parsed if negative else parsed

    def _merge_account(self, code, name, monthly_data, depth=None):
        if code in self.accounts:
            existing = self.accounts[code]
            existing_data = existing["data"]
            for key, value in monthly_data.items():
                existing_data[key] = existing_data.get(key, 0.0) + (value if value is not None else 0.0)
            if depth is not None:
                existing["depth"] = depth if existing.get("depth") is None else min(existing["depth"], depth)
        else:
            self.accounts[code] = {"name": name, "data": monthly_data, "depth": depth}

    def parse_canyon(self):
        df = self._clean_columns(self._read_csv_robust(header=0))
        self.property_name = "Canyon Apartments"
        self.period = "T12"

        default_year = self._infer_default_year(df.columns)
        month_cols = {
            col: month
            for col in df.columns
            if (month := self.normalize_month(col, default_year=default_year))
            and "20" in month
            and any(mon in month for mon in MONTHS)
        }

        for _, row in df.iterrows():
            cat = str(row["Category"]).strip()
            code = self.canyon_map.get(cat)
            if not code:
                continue
            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, cat, monthly_data)

    def parse_oxpt(self):
        header_row_idx = self._find_line_index("Account Name")
        df = self._clean_columns(self._read_csv_robust(header=0, skiprows=header_row_idx))
        self.property_name = "Oxford Pointe"
        self.period = "2025"

        default_year = self._infer_default_year(df.columns)
        month_cols = {
            col: month
            for col in df.columns
            if (month := self.normalize_month(col, default_year=default_year))
            and "20" in month
            and any(mon in month for mon in MONTHS)
        }

        for _, row in df.iterrows():
            name_raw = str(row["Account Name"]).strip()
            if not name_raw or name_raw == "nan":
                continue

            code = self.oxpt_map.get(name_raw)
            if not code and "Total" in name_raw:
                if "Payroll" in name_raw:
                    code = "6400"
                elif "Utilities" in name_raw:
                    code = "6600"
                elif "Repairs" in name_raw:
                    code = "6500"
                elif "Marketing" in name_raw:
                    code = "6300"

            if not code:
                continue

            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, name_raw, monthly_data)

    def parse_paresh(self):
        df = self._clean_columns(self._read_csv_robust(header=0))
        self.property_name = "Paresh Property"
        self.period = "T12"

        default_year = self._infer_default_year(df.columns)
        month_cols = {
            col: month
            for col in df.columns
            if (month := self.normalize_month(col, default_year=default_year))
            and "20" in month
            and any(mon in month for mon in MONTHS)
        }

        for _, row in df.iterrows():
            cat = str(row["Category"]).strip()
            match = re.match(r"^(\d{5})\.\d{4}-(.+)$", cat)
            if not match:
                continue
            p_code = match.group(1)
            name = match.group(2)
            code = self.code_map.get(p_code, p_code[:4])
            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, name, monthly_data)

    def parse_the_view(self):
        df = self._clean_columns(self._read_csv_robust(header=0, skiprows=1))
        self.property_name = "The View"
        self.period = "LTM 2025"

        default_year = self._infer_default_year(df.columns)
        month_cols = {
            col: month
            for col in df.columns
            if (month := self.normalize_month(col, default_year=default_year))
            and "20" in month
            and any(mon in month for mon in MONTHS)
        }

        for _, row in df.iterrows():
            name = str(row.iloc[0]).strip()
            if not name or name == "nan":
                continue

            code = self.name_map.get(name)
            if not code:
                continue

            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, name, monthly_data)

    def parse_resman(self):
        df = self._read_csv_robust(header=None, sep=",")
        self.property_name = df.iloc[0, 0]
        self.period = df.iloc[3, 0]

        header_row = df.iloc[5]
        month_col_indices = {}
        for idx in range(8, 20):
            if idx < len(header_row):
                month = self.normalize_month(str(header_row[idx]), default_year=self._infer_default_year(header_row))
                if month and "20" in month:
                    month_col_indices[idx] = month

        for idx in range(6, len(df)):
            row = df.iloc[idx]
            account_str = None
            account_depth = None
            for col_idx in range(8):
                val = row[col_idx]
                if pd.notna(val) and str(val).strip() != "":
                    account_str = str(val).strip()
                    account_depth = col_idx
                    break

            if not account_str:
                continue

            code_match = re.match(r"^(\d{4})\s+(.+)$", account_str)
            if not code_match:
                continue

            code = code_match.group(1)
            name = code_match.group(2)
            monthly_values = {
                month_name: self._parse_amount(row[col_idx])
                for col_idx, month_name in month_col_indices.items()
            }
            self._merge_account(code, name, monthly_values, depth=account_depth)

    def parse_cash_flow(self):
        header_row_idx = 0
        header_lines = self._read_head_lines(15)
        for line in header_lines:
            line_clean = line.strip().strip(",")
            if not line_clean:
                continue
            if line_clean.lower().startswith("exported on"):
                continue
            if line_clean.lower().startswith("period range:"):
                self.period = line_clean.split(":", 1)[-1].strip()
                continue
            if self.property_name == "Unknown Property":
                self.property_name = line_clean

        header_row_idx = self._find_line_index("Account Name")
        df = self._clean_columns(self._read_csv_robust(header=0, skiprows=header_row_idx))

        if self.property_name == "Unknown Property":
            self.property_name = "Property"
        if self.period == "Unknown Period":
            self.period = "Cash Flow"

        if "Account Name" not in df.columns:
            self.warnings.append("Cash Flow parser: 'Account Name' column not found.")
            return

        default_year = self._infer_default_year(df.columns)
        month_cols = {}
        for col in df.columns:
            if col == "Account Name":
                continue
            month = self.normalize_month(col, default_year=default_year)
            if month and "20" in month and any(mon in month for mon in MONTHS):
                month_cols[col] = month

        def map_keyword_to_code(label):
            lowered = label.lower()
            if any(key in lowered for key in ["vacancy", "concession", "loss to vacancy"]):
                return "4220"
            if any(key in lowered for key in ["rent", "rents"]) and "vacancy" not in lowered:
                return "4000"
            if any(
                key in lowered
                for key in [
                    "other income",
                    "fee",
                    "fees",
                    "misc",
                    "forfeit",
                    "application",
                    "utility reimbursement",
                    "pet rent",
                    "late",
                ]
            ):
                return "4300"
            if any(key in lowered for key in ["payroll", "wages", "salary", "salaries", "benefits"]):
                return "6400"
            if any(key in lowered for key in ["utilities", "electric", "water", "gas", "sewer"]):
                return "6600"
            if any(key in lowered for key in ["repair", "repairs", "maintenance", "supplies", "cleaning"]):
                return "6530"
            if any(key in lowered for key in ["contract", "landscaping", "trash", "pest", "grounds"]):
                return "6500"
            if any(key in lowered for key in ["marketing", "leasing", "advertising"]):
                return "6300"
            if "insurance" in lowered:
                return "6700"
            if any(key in lowered for key in ["tax", "taxes"]):
                return "6800"
            if any(key in lowered for key in ["legal", "professional", "accounting"]):
                return "6200"
            if any(key in lowered for key in ["office", "administrative", "admin"]):
                return "6100"
            return None

        for _, row in df.iterrows():
            name_raw = str(row["Account Name"]).strip()
            if not name_raw or name_raw.lower() == "nan":
                continue
            name_clean = re.sub(r"^\s+", "", name_raw)
            lowered = name_clean.lower()

            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in month_cols.keys()):
                continue

            code = None
            if "total operating income" in lowered:
                code = "9998"
            elif "total operating expense" in lowered:
                code = "9999"
            elif lowered.startswith("total "):
                continue
            else:
                code = map_keyword_to_code(lowered)

            if not code:
                continue

            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, name_clean, monthly_data)

    def get_data(self):
        return {
            "property": str(self.property_name or "Property"),
            "period": str(self.period or "Period"),
            "accounts": self.accounts,
            "meta": {"format": self.detected_format, "warnings": self.warnings},
        }


class ScorecardTargetParser:
    def __init__(self, filepath):
        self.filepath = filepath
        self.targets = {"UW": {}, "PM": {}}
        self.diagnostics = {
            "scorecard_sheet_found": False,
            "found_columns": {},
            "missing_columns": [],
            "found_rows": [],
            "missing_rows": [],
            "warnings": [],
        }

    def parse(self):
        try:
            wb = openpyxl.load_workbook(self.filepath, data_only=True)
            if "Scorecard" not in wb.sheetnames:
                self.diagnostics["warnings"].append("Target parser: 'Scorecard' sheet not found.")
                return

            self.diagnostics["scorecard_sheet_found"] = True
            sheet = wb["Scorecard"]
            uw_col = None
            pm_col = None

            for row_idx in range(1, 30):
                row_vals = []
                for col_idx in range(1, 20):
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    row_vals.append((col_idx, str(val).strip().lower()) if val else (col_idx, ""))

                row_pm_col = next(
                    (col_idx for col_idx, val in row_vals if "pm budget" in val or "manager budget" in val),
                    None,
                )
                if row_pm_col is None:
                    continue

                # Found the real data-table header row. Scope the UW column
                # search to this same row only — an unrelated "UW" mention
                # elsewhere on the sheet (e.g. a property-info label like
                # "OPERATING ASSUMPTIONS - UW YR1") must never be mistaken
                # for the actual UW/underwriting column of this table.
                pm_col = row_pm_col
                uw_col = next(
                    (col_idx for col_idx, val in row_vals
                     if "uw" in val and "per unit" not in val and "variance" not in val),
                    None,
                )
                if uw_col is None:
                    # Some scorecards label their UW column "Year N" / "Yr N"
                    # (the underwritten target for that scorecard year)
                    # rather than spelling out "UW".
                    uw_col = next(
                        (col_idx for col_idx, val in row_vals
                         if re.search(r"\b(?:year|yr)\.?\s*\d+\b", val)
                         and "variance" not in val and "per unit" not in val),
                        None,
                    )
                break

            if uw_col:
                self.diagnostics["found_columns"]["UW"] = uw_col
            else:
                self.diagnostics["missing_columns"].append("UW")
                self.diagnostics["warnings"].append(
                    "Target parser: could not find a UW/underwriting column in the Scorecard "
                    "header row — UW targets will show as $0 and should not be trusted."
                )
            if pm_col:
                self.diagnostics["found_columns"]["PM"] = pm_col
            else:
                self.diagnostics["missing_columns"].append("PM Budget")

            if not uw_col and not pm_col:
                self.diagnostics["warnings"].append("Target parser: could not find UW or PM Budget columns.")
                return

            row_map = {}
            for row_idx in range(1, sheet.max_row + 1):
                val_a = sheet.cell(row=row_idx, column=1).value
                val_b = sheet.cell(row=row_idx, column=2).value
                label = (str(val_a) + " " + str(val_b)).lower()

                if "variance" in label:
                    continue
                if ("total operating income" in label or "total income" in label) and "Income" not in row_map:
                    row_map["Income"] = row_idx
                elif ("total operating expenses" in label or "total expenses" in label) and "Expenses" not in row_map:
                    row_map["Expenses"] = row_idx
                elif ("net operating income" in label or "noi" in label) and "NOI" not in row_map:
                    row_map["NOI"] = row_idx

            for metric in ("Income", "Expenses", "NOI"):
                if metric in row_map:
                    self.diagnostics["found_rows"].append({"metric": metric, "row": row_map[metric]})
                else:
                    self.diagnostics["missing_rows"].append(metric)

            def get_monthly(row_idx, col_idx):
                if row_idx and col_idx:
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        return float(val) / 12.0
                return 0.0

            if uw_col:
                self.targets["UW"]["Income"] = get_monthly(row_map.get("Income"), uw_col)
                self.targets["UW"]["Expenses"] = get_monthly(row_map.get("Expenses"), uw_col)
                self.targets["UW"]["NOI"] = get_monthly(row_map.get("NOI"), uw_col)

            if pm_col:
                self.targets["PM"]["Income"] = get_monthly(row_map.get("Income"), pm_col)
                self.targets["PM"]["Expenses"] = get_monthly(row_map.get("Expenses"), pm_col)
                self.targets["PM"]["NOI"] = get_monthly(row_map.get("NOI"), pm_col)
        except Exception as exc:
            self.diagnostics["warnings"].append(f"Error parsing Scorecard targets: {exc}")

    def get_data(self):
        return self.targets

    def get_diagnostics(self):
        return self.diagnostics


# OXPT-specific write-back category mapping (Michelle's explicit decisions).
# Maps existing Scorecard T12 row labels — already present in the sheet,
# never newly created or renamed — to the P&L account code(s) whose values
# should be summed into that row. Scoped to OXPT only (see the property-name
# check in ScorecardUpdater.update()); must not affect Eagle Rock or Canyon.
_OXPT_ROW_GROUPS = {
    "total office expense": ["6100"],                # Administration Costs ("G&A")
    "total legal & professional fees": ["6200", "4418"],  # Legal & Professional + Attorney/Court Fees (4418 sits
                                                       # under Other Income in the P&L, but Michelle wants it
                                                       # tracked against Legal regardless of that placement)
    "total advertising": ["6300"],                    # Marketing & Leasing ("Marketing")
    "total payroll expense": ["6400"],                 # Salaries & Payroll
    "total cleaning & trash removal": ["6520"],        # Contract Services: housekeeping-specific
    "total outside contractors": ["6540", "6545", "6555", "6565"],  # Contract Services: trades
    "total other expenses": ["6560"],                  # Contract Services: pest control
    "total repairs": ["6600", "6700"],                  # Maintenance Related + Turnover Costs (6700 is a
                                                        # Repairs & Maintenance subcategory, not its own category)
    "total grounds & lawn maintenance": ["6800"],       # Grounds
    "total utilities": ["6900"],                        # Utilities
    "total insurance": ["7120"],                        # Insurance
    "total taxes": ["7130"],                            # Property Taxes
    "asset mgmt fee": ["7210"],                         # Asset Management Fee — kept in its existing row/location
    "total management fees": ["7220", "7250"],          # Management Company Charges minus Asset Mgmt Fee
    "total debt service": ["7300"],                     # Debt Service
    "total capital expenses": ["7500"],                 # Big-ticket capital repairs
    # Other Income (4300 family) splits across several existing rows rather
    # than one bucket — fee-type income to the relevant Fees row, utility
    # reimbursement to Utility Recovery, parking-related to Parking Income.
    "electric": ["4320"],
    "late fees": ["4400"],
    "insurance services": ["4402"],
    "admin fee": ["4405"],
    "application fee income": ["4415"],
    "cleaning fee": ["4420"],
    "early termination fee": ["4448"],
    "month-to-month fee": ["4450"],
    "damages": ["4452"],
    "pet fee-non refundable": ["4455"],
    "nsf fees collected": ["4460"],
    "parking income": ["4508"],
    # Oddball Other Income lines with no dedicated row of their own — verified
    # against real OXPT data as revenue (all three post positive values nested
    # under 4300 Other Income), so they land in the existing catch-all fee
    # row alongside the ten name-matched lines below, rather than getting a
    # new row each. 4295 (Miscellaneous Credit) and 4580 (High Risk Fee, a
    # sibling income line outside the 4000/4300 tree) join the same bucket.
    "miscellaneous fees": ["4341", "4470", "4500", "4295", "4580"],
    # Rents family (4000/4100/4200 children) — the rollups themselves
    # (4000, 4100, 4200) are excluded below since these children fully
    # capture their values.
    "loss/gain to market": ["4120"],           # "(Loss) / Gain to Old Lease" — best-fit match despite wording
    "delinquency": ["4210"],                   # Bad Debt / Write-Off Uncollectable Rent
    "vacancy": ["4220"],                        # Vacancy Loss
    "concessions": ["4250", "4258", "4260"],    # Rent Concessions + Utility Credit + Rent Discount
}

# Other Income detail lines with no assigned P&L code known ahead of time —
# matched by account name (case-insensitive) rather than code, and routed to
# the same "Miscellaneous Fees" row as the codes above.
_OXPT_MISC_FEE_NAMES = {
    "auto charge",
    "auto payment fee",
    "charge off recovery",
    "corporate housing",
    "furniture rental income",
    "housing assistance payment",
    "insurance - waived",
    "insurance – waived",
    "move in charge",
    "renters insurance income",
    "storage income",
}

# Codes whose value is already fully captured by a group above (parent
# rollups distributed via children, or children merged into a parent
# rollup via a group) — excluded from digit/name-based matching so they
# can't also write into the same or a different cell a second time.
_OXPT_EXCLUDED_CODES = {
    "6000", "7000",  # grand rollups, fully distributed via sub-categories
    "6112", "6113", "6115", "6118", "6138", "6139", "6140", "6157", "6164", "6178", "6187",  # 6100 children
    "6205", "6210",  # 6200 children
    "6305", "6315", "6350", "6355", "6360",  # 6300 children
    "6405", "6415", "6430", "6450", "6465",  # 6400 children
    "6500",  # Contract Services rollup, distributed via children individually
    "6606", "6627", "6636", "6639", "6651", "6654", "6660", "6663", "6669", "6675",  # 6600 children
    "6810",  # 6800 child
    "6910", "6911", "6915", "6930", "6940", "6955", "6960",  # 6900 children
    "7200",  # Management Company Charges rollup, distributed via 7210 + [7220, 7250]
    "7330", "7350",  # 7300 children
    "7502", "7505", "7511", "7516", "7518", "7520", "7536", "7543", "7544", "7545",
    "7547", "7549", "7550", "7556", "7560", "7564", "7568", "7570", "7573", "7578", "7595",  # 7500 children
    "4300",  # Other Income rollup, distributed via children individually
    "6750",  # 6700 (Turnover Costs) child, folded into "total repairs" via the 6700 code above
    "7400",  # Other Expenses rollup — its only known child (7210, Asset Mgmt Fee) is mapped separately
             # above and kept below the NOI line; no other 7400 child has appeared in any real OXPT
             # export seen so far. If one does, it will surface here as unmatched rather than being
             # silently miscounted — a genuine "Misc Expense" row should be added against real data
             # at that point, not built speculatively now.
    "4000", "4100", "4200",  # Rents rollups (Net Rental Income / Gross Possible Rent / Deductions),
                             # fully distributed via their 4110/4120/4210/4220/4250/4258/4260 children.
    # 7110 Replacement Reserve Escrow is deliberately NOT excluded or mapped here — no existing T12
    # row fits it (it's a peer to Insurance/Taxes/Management Fees, not nested under any of them), and
    # per the precedent already set for Eagle Rock/Canyon, genuinely homeless items are left flagged
    # as unmatched rather than getting a new row invented for them.
}

# Eagle Rock's Scorecard T12 sheet mirrors the P&L's own tree structure (every
# "Total X" rollup gets its own code-prefixed row, e.g. "6100 Total
# Administration Costs"), which the existing code-prefix scan already matches
# directly — no explicit row-group mapping is needed the way OXPT required.
# The only gap is leaf-level detail codes whose parent rollup IS matched but
# who don't get their own separate line; excluding them here (so they don't
# show as unmatched) is safe because their dollar value is already fully
# captured in that matched parent total. Verified against the real Eagle
# Rock T12 (annual figures): each child's parent total equals the sum of its
# section's children exactly, including 7330 Mortgage Payment, which is
# fully captured by "7300 Total Debt Service" ($18,847.04/month = $6,460.10
# Mortgage Payment + $12,386.94 Mortgage Interest for May 2026, matching
# exactly) — unlike OXPT's 7110, this is not a homeless financing item, so
# it does not need a below-NOI or new-row decision.
_EAGLE_ROCK_EXCLUDED_CODES = {
    "4490",  # 4300 (Other Income) child
    "6125", "6128", "6169",  # 6100 (Administration Costs) children
    "6639",  # 6600 (Maintenance Related) child
    "6780",  # 6700 (Turnover Costs) child
    "7551",  # 7500 (Repairs) child
    "7330",  # 7300 (Debt Service) child — captured by "7300 Total Debt Service", already matched
}

# Canyon's Scorecard T12 sheet uses the same mirrored-tree structure as Eagle
# Rock (confirmed against the real file), in an unrelated older "12 Month
# Rolling" report's columns (1-14) sitting alongside the real account-code
# region (column 16+) that the parser actually scans — the two must not be
# confused. Same logic as Eagle Rock: leaf children of an already-matched
# parent rollup are excluded here since their value is already captured.
# 6800 (Grounds) and its only child 6810 (Landscape - Maintenance) are
# deliberately NOT excluded — verified no "6800 Total Grounds" row (or any
# other row referencing 6800/6810/Grounds/Landscape) exists anywhere in this
# Scorecard, so unlike the rest of this list, that one dollar amount
# ($17,561.50/yr) isn't captured anywhere and is left flagged as unmatched.
# 4341/6320/6370/6605/6785/7544/7595 were added later -- newer, more
# granular leaf accounts a subsequent T12 P&L export broke out under the
# same already-matched parent rollups (4300/6300/6600/6700/7500) as
# everything else in this set; not a new gap.
_CANYON_EXCLUDED_CODES = {
    "4305", "4428", "4470", "4341",  # 4300 (Other Income) children
    "6106", "6178", "6184", "6187",  # 6100 (Administration Costs) children
    "6330", "6335", "6343", "6320", "6370",  # 6300 (Marketing & Leasing) children
    "6415",  # 6400 (Salaries & Payroll Related) child
    "6530", "6545",  # 6500 (Contract Services) children
    "6606", "6612", "6636", "6642", "6654", "6605",  # 6600 (Maintenance Related) children
    "6740", "6780", "6785",  # 6700 (Turnover Costs) children
    "6905",  # 6900 (Utilities) child
    "7240",  # 7200 (Management Company Charges) child
    "7502", "7518", "7534", "7536", "7537", "7541", "7564", "7570", "7573", "7544", "7595",  # 7500 (Repairs) children
}


class ScorecardUpdater:
    def __init__(self, scorecard_path, data):
        self.scorecard_path = scorecard_path
        self.data = data
        self.wb = None
        self.sheet = None
        self.diagnostics = {"updated_cells": 0, "warnings": []}

    def update(self, output_path):
        try:
            suffix = str(self.scorecard_path).lower()
            self.wb = openpyxl.load_workbook(self.scorecard_path, keep_vba=suffix.endswith(".xlsm"))
        except Exception as exc:
            self.diagnostics["warnings"].append(f"Error loading scorecard: {exc}")
            return None

        if "T12" not in self.wb.sheetnames:
            self.diagnostics["warnings"].append("Scorecard updater: 'T12' sheet not found.")
            return None

        self.sheet = self.wb["T12"]
        excel_month_map = {}
        header_row = 6
        first_month_col = None
        last_month_col = None
        account_header_col = None
        for col_idx in range(1, 50):
            cell_val = self.sheet.cell(row=header_row, column=col_idx).value
            if not isinstance(cell_val, str):
                continue
            stripped = cell_val.strip()
            if account_header_col is None and stripped.lower() in ("account", "account name"):
                account_header_col = col_idx
                continue
            # Month headers vary by export: "Jan 2025 Actual" (4-digit year)
            # or "Jan-25" (2-digit year). Normalize both to the same
            # "Mon YYYY" key the parsed P&L data already uses.
            match = re.search(r"([A-Za-z]{3}).*?(\d{4}|\d{2})\b", stripped)
            if match:
                month_abbr = match.group(1)[:3].title()
                year_str = match.group(2)
                year = int(year_str) if len(year_str) == 4 else int("20" + year_str)
                if 2000 <= year <= 2099 and month_abbr in MONTHS:
                    excel_month_map[f"{month_abbr} {year}"] = col_idx
                    if first_month_col is None:
                        first_month_col = col_idx
                    last_month_col = col_idx

        # Extend the sheet with new month columns if the uploaded P&L covers
        # months the workbook's own T12 tab doesn't have yet (e.g. a newer
        # T12 refreshing a Scorecard whose T12 tab stops months earlier) —
        # rather than silently writing only the overlapping months.
        source_months = set()
        for acc_data in self.data["accounts"].values():
            source_months.update(acc_data["data"].keys())
        missing_months = sorted(
            (m for m in source_months if m not in excel_month_map),
            key=month_sort_key,
        )

        skipped_months = []
        skip_reason = None
        if missing_months and last_month_col is not None:
            insertion_col = last_month_col + 1
            # openpyxl shifts cell values on insert_cols() but does not
            # rewrite formula text, so inserting into a column range that
            # any formula elsewhere on the sheet references (e.g. an
            # "Adjusted Total" that sums across the month columns) would
            # silently produce a wrong total rather than an obviously
            # incomplete one. Only extend the sheet if no formulas exist
            # anywhere from the insertion point onward.
            has_formula_in_insert_region = any(
                isinstance(self.sheet.cell(r, c).value, str) and self.sheet.cell(r, c).value.startswith("=")
                for r in range(1, self.sheet.max_row + 1)
                for c in range(insertion_col, self.sheet.max_column + 1)
            )
            if has_formula_in_insert_region:
                skipped_months = missing_months
                skip_reason = (
                    "the T12 sheet has formulas in or after the column where new months "
                    "would be inserted, and inserting there could silently break those totals"
                )
            else:
                self.sheet.insert_cols(insertion_col, amount=len(missing_months))
                for offset, month_key in enumerate(missing_months):
                    col = insertion_col + offset
                    self.sheet.cell(row=header_row, column=col).value = f"{month_key} Actual"
                    excel_month_map[month_key] = col
        elif missing_months:
            skipped_months = missing_months
            skip_reason = "no existing month columns could be identified in the T12 sheet at all"

        # Different Scorecard T12 layouts put account-code labels in
        # different columns (Eagle Rock: column 1; Canyon: column 16+, after
        # an unrelated older "12 Month Rolling" report occupies columns
        # 1-14). Locate the label region dynamically — from the "Account"
        # header column through the column just before the first month
        # column — instead of assuming a fixed range.
        label_scan_start = account_header_col or 1
        label_scan_end = (first_month_col - 1) if first_month_col else (label_scan_start + 4)
        label_scan_end = max(label_scan_end, label_scan_start)

        # Label-column index (label -> row indices that actually carry data,
        # i.e. not a blank section-header row) — reused by both the OXPT
        # explicit category mapping and the generic name-based fallback.
        label_rows: dict = {}
        for row_idx in range(7, self.sheet.max_row + 1):
            label_val = self.sheet.cell(row_idx, label_scan_start).value
            if not isinstance(label_val, str) or not label_val.strip():
                continue
            has_data = any(
                self.sheet.cell(row_idx, c).value is not None
                and str(self.sheet.cell(row_idx, c).value).strip() != ""
                for c in range(label_scan_start + 1, self.sheet.max_column + 1)
            )
            if not has_data:
                continue
            label_rows.setdefault(label_val.strip().lower(), []).append(row_idx)

        account_row_map = {}
        for row_idx in range(7, self.sheet.max_row + 1):
            for col_idx in range(label_scan_start, label_scan_end + 1):
                val = self.sheet.cell(row=row_idx, column=col_idx).value
                if not val:
                    continue
                match = re.match(r"^(\d{4})\b", str(val).strip())
                if match:
                    account_row_map[match.group(1)] = row_idx
                    break

        # OXPT-specific explicit category mapping (Michelle's decisions —
        # see _OXPT_ROW_GROUPS above). Scoped strictly to OXPT by property
        # name so Eagle Rock/Canyon's matching is never affected.
        excluded_codes = set()
        oxpt_ambiguous_groups = []
        is_oxpt = "oxford pointe" in str(self.data.get("property") or "").strip().lower()
        if is_oxpt:
            excluded_codes |= _OXPT_EXCLUDED_CODES
            for label, codes in _OXPT_ROW_GROUPS.items():
                rows = label_rows.get(label)
                if not rows:
                    continue
                if len(rows) > 1:
                    oxpt_ambiguous_groups.append((label, codes, rows))
                    continue
                row_idx = rows[0]
                for code in codes:
                    if code in self.data["accounts"]:
                        account_row_map[code] = row_idx

            # Other Income detail lines with no fixed P&L code (see
            # _OXPT_MISC_FEE_NAMES) — matched by account name instead,
            # into the same "Miscellaneous Fees" row as the coded group above.
            misc_fee_rows = label_rows.get("miscellaneous fees")
            if misc_fee_rows:
                if len(misc_fee_rows) > 1:
                    oxpt_ambiguous_groups.append(("miscellaneous fees (name-matched)", [], misc_fee_rows))
                else:
                    for code, acc_data in self.data["accounts"].items():
                        name = str(acc_data.get("name") or "").strip().lower()
                        if name in _OXPT_MISC_FEE_NAMES:
                            account_row_map[code] = misc_fee_rows[0]

        # Eagle Rock/Canyon: no explicit row-group mapping needed (see
        # _EAGLE_ROCK_EXCLUDED_CODES / _CANYON_EXCLUDED_CODES above) — their
        # Scorecard T12 sheets already code-prefix-match every rollup
        # directly, so this only excludes already-captured leaf children.
        is_eagle_rock = "eagle rock" in str(self.data.get("property") or "").strip().lower()
        is_canyon = "canyon" in str(self.data.get("property") or "").strip().lower()
        if is_eagle_rock:
            excluded_codes |= _EAGLE_ROCK_EXCLUDED_CODES
        if is_canyon:
            excluded_codes |= _CANYON_EXCLUDED_CODES

        # Fallback for accounts not resolved above and not part of an
        # OXPT-excluded group. Match by the parsed P&L account's own name
        # against label-column rows that carry data, only when unambiguous.
        ambiguous_names = list(oxpt_ambiguous_groups)
        unmatched_codes = [
            code for code in self.data["accounts"]
            if code not in account_row_map and code not in excluded_codes
        ]
        for code in unmatched_codes:
            name = str(self.data["accounts"][code].get("name") or "").strip().lower()
            if not name:
                continue
            rows = label_rows.get(name)
            if not rows:
                continue
            if len(rows) == 1:
                account_row_map[code] = rows[0]
            else:
                ambiguous_names.append((code, self.data["accounts"][code].get("name"), rows))

        # Aggregate per (row, month-column) before writing — some rows
        # receive the summed value of multiple P&L codes (e.g. OXPT's
        # Contract Services split). For every other layout this is a
        # no-op: each code maps to a distinct row, so the "sum" is just
        # that one value, identical to a direct overwrite.
        cell_totals: dict = {}
        for code, acc_data in self.data["accounts"].items():
            if code not in account_row_map:
                continue
            row_idx = account_row_map[code]
            for month_key, value in acc_data["data"].items():
                col_idx = excel_month_map.get(month_key)
                if col_idx is None:
                    continue
                key = (row_idx, col_idx)
                cell_totals[key] = cell_totals.get(key, 0.0) + (value or 0.0)

        updates_count = 0
        for (row_idx, col_idx), total in cell_totals.items():
            self.sheet.cell(row=row_idx, column=col_idx).value = total
            updates_count += 1

        if skipped_months:
            self.diagnostics["warnings"].append(
                "Scorecard updater: could not add columns for "
                + ", ".join(skipped_months)
                + f" — {skip_reason}. These months were not written to the updated "
                  "scorecard; extend the T12 sheet's month columns manually and re-run."
            )

        # Codes in excluded_codes are intentionally not matched individually
        # because their value is already fully captured via a mapped parent
        # rollup/group — they're not failures, so they're left out of both
        # the denominator and the "not written" list below.
        reportable_codes = [code for code in self.data.get("accounts", {}) if code not in excluded_codes]
        total_accounts = len(reportable_codes)
        matched_accounts = sum(1 for code in reportable_codes if code in account_row_map)
        if total_accounts and matched_accounts < total_accounts:
            unresolved = sorted(code for code in reportable_codes if code not in account_row_map)
            preview = ", ".join(unresolved[:15])
            more = f" (+{len(unresolved) - 15} more)" if len(unresolved) > 15 else ""
            self.diagnostics["warnings"].append(
                f"Scorecard updater: matched {matched_accounts} of {total_accounts} parsed "
                f"P&L accounts to rows in the T12 sheet; {len(unresolved)} could not be "
                "confidently matched — likely different account naming/grouping between "
                f"the P&L export and this Scorecard's own T12 tab — and were not written: "
                f"{preview}{more}."
            )
        if ambiguous_names:
            desc = "; ".join(f"{code} ({name!r} matched rows {rows})" for code, name, rows in ambiguous_names[:10])
            self.diagnostics["warnings"].append(
                "Scorecard updater: could not confidently place these accounts because "
                f"their name matched more than one row in the T12 sheet: {desc}."
            )

        self.diagnostics["updated_cells"] = updates_count
        self.wb.save(output_path)
        return output_path

    def get_diagnostics(self):
        return self.diagnostics


class KPICalculator:
    def __init__(self, pnl_data):
        self.accounts = pnl_data["accounts"]
        available_months = set()
        for acc in self.accounts.values():
            available_months.update(acc["data"].keys())
        self.available_months = sorted(list(available_months), key=month_sort_key)
        self.expense_fallback_codes = sorted(
            code for code in self.accounts if re.fullmatch(r"6\d{3}", str(code))
        )

        # Additional top-level income categories beyond GPR/NRI (4000) and
        # Other Income (4300) — e.g. a tree-report P&L with a sibling income
        # line like "4580 High Risk Fee" that isn't nested under either.
        # Scoped by tree depth (the column an account code was found in, set
        # by parse_resman()) rather than by code range, since a leaf code
        # can numerically fall outside 4300's range while still being a
        # nested sub-line already counted in 4300's own total (e.g. "4500
        # Credit Builder" nested one level deeper than the 4580 sibling).
        # Formats without depth info (flat CSVs) leave this empty, which
        # preserves the exact previous nri + other_income behavior for them.
        income_depths = [
            acc.get("depth")
            for code, acc in self.accounts.items()
            if re.fullmatch(r"4\d{3}", str(code)) and acc.get("depth") is not None
        ]
        if income_depths:
            shallowest_income_depth = min(income_depths)
            self.income_fallback_codes = sorted(
                code
                for code, acc in self.accounts.items()
                if re.fullmatch(r"4\d{3}", str(code))
                and acc.get("depth") == shallowest_income_depth
                and code not in ("4000", "4300")
            )
        else:
            self.income_fallback_codes = []

        # OXPT-specific: Asset Management Fees (code 7210) is treated as a
        # below-NOI item in the app's own NOI math, per Michelle's explicit
        # decision — the exported Scorecard spreadsheet is unaffected (7210
        # still gets written to its existing row by ScorecardUpdater).
        # Scoped to OXPT by property name, not by bare code number: Canyon's
        # chart of accounts also happens to use code 7210 for the same
        # concept, but that's a separate decision Michelle hasn't made yet,
        # and Eagle Rock uses a different code (7270) entirely.
        self.below_noi_codes = set()
        property_name = str(pnl_data.get("property") or "").strip().lower()
        if "oxford pointe" in property_name:
            self.below_noi_codes.add("7210")

    def get_val(self, code, month):
        if code in self.accounts:
            return float(self.accounts[code]["data"].get(month, 0.0) or 0.0)
        return 0.0

    def calculate(self):
        kpis = {
            "income": {},
            "expenses": {},
            "noi": {},
            "physical_occupancy": {},
            "economic_occupancy": {},
            "expense_ratio": {},
            "noi_margin": {},
            "occupancy_status": {},
            "expense_fallback_codes": self.expense_fallback_codes,
            "income_fallback_codes": self.income_fallback_codes,
        }

        for month in self.available_months:
            gpr = self.get_val("4110", month)
            vacancy_loss = self.get_val("4220", month)
            nri = self.get_val("4000", month)
            other_income = self.get_val("4300", month)

            # Only reconstruct NRI from GPR + Vacancy Loss when code 4000
            # was never captured in this file at all — a genuinely-parsed
            # 4000 value of exactly 0 (a real accounting outcome some
            # months) must be trusted, not silently overridden.
            if "4000" not in self.accounts and gpr != 0:
                nri = gpr + vacancy_loss

            override_income = self.get_val("9998", month)
            if override_income != 0:
                total_income = override_income
            else:
                additional_income = sum(self.get_val(code, month) for code in self.income_fallback_codes)
                total_income = nri + other_income + additional_income

            controllable = self.get_val("6000", month)
            non_controllable = self.get_val("7000", month)
            for code in self.below_noi_codes:
                non_controllable -= self.get_val(code, month)
            override_expenses = self.get_val("9999", month)
            if override_expenses != 0:
                total_expenses = override_expenses
            else:
                if controllable == 0 and non_controllable == 0:
                    for code in self.expense_fallback_codes:
                        controllable += self.get_val(code, month)
                total_expenses = controllable + non_controllable

            noi = total_income - total_expenses

            if gpr == 0:
                phys_occ = None
                econ_occ = None
                occ_status = "missing_gpr"
            else:
                phys_occ = 1 - (abs(vacancy_loss) / gpr)
                econ_occ = nri / gpr
                occ_status = "zero" if phys_occ == 0 else "ok"

            exp_ratio = total_expenses / total_income if total_income != 0 else None
            noi_margin = noi / total_income if total_income != 0 else None

            kpis["income"][month] = total_income
            kpis["expenses"][month] = total_expenses
            kpis["noi"][month] = noi
            kpis["physical_occupancy"][month] = phys_occ
            kpis["economic_occupancy"][month] = econ_occ
            kpis["expense_ratio"][month] = exp_ratio
            kpis["noi_margin"][month] = noi_margin
            kpis["occupancy_status"][month] = occ_status

        return kpis


class ReportGenerator:
    def __init__(self, kpis):
        self.kpis = kpis
        self.months = list(kpis["income"].keys())

    def generate(self):
        total_income = sum(float(v or 0.0) for v in self.kpis["income"].values())
        total_noi = sum(float(v or 0.0) for v in self.kpis["noi"].values())
        valid_occupancies = [
            value
            for month, value in self.kpis["physical_occupancy"].items()
            if isinstance(value, (int, float)) and self.kpis["occupancy_status"].get(month) != "missing_gpr"
        ]
        avg_occ = sum(valid_occupancies) / len(valid_occupancies) if valid_occupancies else None

        report = []
        report.append("=== PROPERTY FINANCIAL SCORECARD REPORT ===")
        report.append(f"Period Analysis: {len(self.months)} Months")
        report.append("\n-- KEY METRICS --")
        report.append(f"Total Income: ${total_income:,.2f}")
        report.append(f"Total NOI:    ${total_noi:,.2f}")
        report.append(f"Avg Physical Occupancy: {format_percent(avg_occ)}")

        report.append("\n-- MONTHLY TRENDS --")
        header = f"{'Month':<10} {'Income':<15} {'NOI':<15} {'Occ%':<10}"
        report.append(header)
        report.append("-" * len(header))
        for month in self.months:
            inc = self.kpis["income"][month]
            noi = self.kpis["noi"][month]
            occ = self.kpis["physical_occupancy"][month]
            occ_text = "No GPR" if self.kpis["occupancy_status"].get(month) == "missing_gpr" else format_percent(occ)
            report.append(f"{month:<10} ${inc:,.0f}       ${noi:,.0f}       {occ_text:<10}")

        q1_months = [m for m in self.months if m.split()[0] in ["Jan", "Feb", "Mar"]]
        q4_months = [m for m in self.months if m.split()[0] in ["Oct", "Nov", "Dec"]]

        if q1_months and q4_months:
            q1_noi = sum(self.kpis["noi"][month] for month in q1_months)
            q4_noi = sum(self.kpis["noi"][month] for month in q4_months)

            report.append("\n-- TREND ANALYSIS --")
            report.append(f"Q1 Total NOI: ${q1_noi:,.0f}")
            report.append(f"Q4 Total NOI: ${q4_noi:,.0f}")
            delta = q4_noi - q1_noi
            report.append(f"Change: {'+' if delta >= 0 else ''}${delta:,.0f}")

        report.append("\n-- RECOMMENDATIONS --")
        if avg_occ is not None and avg_occ < 0.90:
            report.append("1. Focus on leasing strategies to boost occupancy above 90%.")
        if total_noi < 0:
            report.append("2. CRITICAL: Review expenses immediately, NOI is negative.")
        if report[-1] == "\n-- RECOMMENDATIONS --":
            report.append("1. Continue monitoring monthly performance against budget.")

        return "\n".join(report)


def generate_advanced_insights(df_filtered, accounts, targets=None):
    def get_category_metrics(code_prefixes, name):
        relevant_codes = [code for code in accounts.keys() if any(str(code).startswith(prefix) for prefix in code_prefixes)]
        if not relevant_codes or df_filtered.empty:
            return None

        series = []
        for month in df_filtered["Month"]:
            value = sum(accounts[code]["data"].get(month, 0) for code in relevant_codes)
            series.append(value)
        series_pd = pd.Series(series, dtype="float64")
        total_val = float(series_pd.sum())

        if len(series_pd) >= 2:
            mid_point = len(series_pd) // 2
            first_half_avg = series_pd.iloc[:mid_point].mean()
            last_half_avg = series_pd.iloc[mid_point:].mean()
            pct_change = (last_half_avg - first_half_avg) / abs(first_half_avg) if first_half_avg != 0 else 0.0
        else:
            pct_change = 0.0

        return {"name": name, "total": total_val, "pct_change": float(pct_change)}

    categories = [
        (["4000", "4100", "4110"], "Rental Income"),
        (["4300"], "Other Income"),
        (["6600", "66"], "Utilities"),
        (["6500", "65"], "Contract Services & R&M"),
        (["6400", "64"], "Payroll"),
        (["6300"], "Marketing"),
        (["6100", "6200"], "Admin & Professional"),
    ]

    analyzed_cats = [get_category_metrics(prefixes, name) for prefixes, name in categories]
    analyzed_cats = [cat for cat in analyzed_cats if cat]

    key_trends = []
    for cat in analyzed_cats:
        change = cat["pct_change"]
        if abs(change) >= 0.03:
            direction = "increased" if change > 0 else "decreased"
            is_income = "Income" in cat["name"]
            is_good = (change > 0) if is_income else (change < 0)
            key_trends.append((f"{cat['name']} {direction} by {abs(change):.1%}.", is_good))

    green_flags = []
    red_flags = []
    occ_values = df_filtered["Occupancy"].dropna() if "Occupancy" in df_filtered else pd.Series(dtype="float64")
    occ_avg = float(occ_values.mean()) if not occ_values.empty else None
    if occ_avg is not None and occ_avg >= 0.93:
        green_flags.append(f"Excellent Occupancy: {occ_avg:.1%}")
    elif occ_avg is not None and occ_avg < 0.90:
        red_flags.append(f"Low Occupancy: {occ_avg:.1%}")

    income_sum = df_filtered["Income"].sum() if "Income" in df_filtered else 0
    noi_margin = df_filtered["NOI"].sum() / income_sum if income_sum else 0
    if noi_margin > 0.55:
        green_flags.append(f"Strong NOI Margin: {noi_margin:.1%}")
    elif noi_margin < 0.40:
        red_flags.append(f"Low NOI Margin: {noi_margin:.1%}")

    # Aggregate (sum expenses / sum income) rather than averaging the monthly
    # ExpenseRatio column directly — matches the NOI Margin calc above, and
    # avoids a single near-zero-income lease-up month from dominating the
    # average the way a mean-of-ratios would (confirmed against real OXPT
    # data: a mean-of-ratios gave 348% off one such month vs. a real 65%).
    expenses_sum = df_filtered["Expenses"].sum() if "Expenses" in df_filtered else 0
    expense_ratio_avg = (expenses_sum / income_sum) if income_sum else None
    if expense_ratio_avg is not None and expense_ratio_avg > 0.65:
        red_flags.append(f"High Expense Ratio: {expense_ratio_avg:.1%}")
    elif expense_ratio_avg is not None and expense_ratio_avg < 0.50:
        green_flags.append(f"Low Expense Ratio: {expense_ratio_avg:.1%}")

    # NOI vs UW/PM Budget, rolled up over the selected months (same +/-10%
    # red / +/-3% green thresholds used for the per-month Comparison table
    # flags — see noi_variance_flag() — applied here to the period total).
    months_count = len(df_filtered) if not df_filtered.empty else 0
    actual_noi_total = float(df_filtered["NOI"].sum()) if "NOI" in df_filtered and months_count else 0.0
    for label, target_dict in (("UW Budget", (targets or {}).get("UW") or {}), ("PM Budget", (targets or {}).get("PM") or {})):
        noi_target_monthly = float(target_dict.get("NOI") or 0.0)
        if not noi_target_monthly or not months_count:
            continue
        noi_target_total = noi_target_monthly * months_count
        flag = noi_variance_flag(actual_noi_total - noi_target_total, noi_target_total)
        variance_pct = (actual_noi_total - noi_target_total) / abs(noi_target_total)
        if flag == "red":
            red_flags.append(f"NOI vs {label} off by {variance_pct:+.1%}")
        elif flag == "green":
            green_flags.append(f"NOI on track vs {label} ({variance_pct:+.1%})")

    for cat in analyzed_cats:
        if "Utilities" in cat["name"] and cat["pct_change"] > 0.10:
            red_flags.append(f"Utilities spiked {cat['pct_change']:.1%}")
        if "Payroll" in cat["name"] and cat["pct_change"] > 0.10:
            red_flags.append(f"Payroll up {cat['pct_change']:.1%}")
        if "Rental Income" in cat["name"] and cat["pct_change"] > 0.05:
            green_flags.append(f"Rental Income up {cat['pct_change']:.1%}")

    recommendations = []
    if occ_avg is not None and occ_avg < 0.90:
        recommendations.append(f"Leasing: Increase marketing outreach and referral incentives (avg occupancy {occ_avg:.1%}).")
    elif occ_avg is not None and occ_avg > 0.95:
        recommendations.append(f"Revenue: Test modest rent increases or premium add-ons (avg occupancy {occ_avg:.1%}).")

    for cat in analyzed_cats:
        change = cat["pct_change"]
        if "Utilities" in cat["name"] and change > 0.05:
            recommendations.append(f"Utilities: Audit water/HVAC usage and validate vendor billing (trend {change:+.1%}).")
        if "Contract" in cat["name"] and change > 0.10:
            recommendations.append(f"Maintenance: Validate CapEx vs OpEx coding and review vendor scope (trend {change:+.1%}).")
        if "Other Income" in cat["name"] and change < -0.05:
            recommendations.append(f"Ancillary: Audit fee collections and enforce add-on compliance (trend {change:+.1%}).")

    if not recommendations:
        recommendations.append("General: Monitor weekly leasing traffic.")

    return {
        "trends": key_trends,
        "green_flags": green_flags,
        "red_flags": red_flags,
        "recommendations": recommendations,
        "analyzed_cats": analyzed_cats,
    }


scorecard_bp = Blueprint("scorecard", __name__)


@scorecard_bp.route("/")
@login_required
def index():
    return render_template("tools/scorecard_pro.html")


@scorecard_bp.route("/upload", methods=["POST"])
@login_required
def upload():
    _cleanup_old_uploads()

    if "pnl_file" not in request.files:
        return jsonify({"error": "No P&L file included in the request."}), 400

    pnl_file = request.files["pnl_file"]
    if not pnl_file or not pnl_file.filename:
        return jsonify({"error": "No P&L file selected."}), 400

    pnl_name = secure_filename(pnl_file.filename)
    pnl_ext = Path(pnl_name).suffix.lower()
    if pnl_ext not in ALLOWED_PNL_EXT:
        return jsonify({"error": "P&L upload must be a .csv, .xlsx, or .xlsm file."}), 400

    scorecard_file = request.files.get("scorecard_file")
    scorecard_name = ""
    if scorecard_file and scorecard_file.filename:
        scorecard_name = secure_filename(scorecard_file.filename)
        if Path(scorecard_name).suffix.lower() not in ALLOWED_SCORECARD_EXT:
            return jsonify({"error": "Scorecard upload must be an .xlsx or .xlsm file."}), 400

    token = secrets.token_urlsafe(16)
    upload_dir = _upload_dir()
    pnl_path = upload_dir / f"{token}_pnl{pnl_ext}"
    scorecard_path = None

    try:
        pnl_file.save(str(pnl_path))
        if scorecard_file and scorecard_name:
            scorecard_path = upload_dir / f"{token}_scorecard{Path(scorecard_name).suffix.lower()}"
            scorecard_file.save(str(scorecard_path))
    except Exception as exc:
        _delete_token_files(token)
        return jsonify({"error": f"Could not save upload: {exc}"}), 500

    try:
        record = process_scorecard(token, pnl_path, pnl_name, scorecard_path, scorecard_name)
    except ValueError as exc:
        _delete_token_files(token)
        return jsonify({"error": str(exc)}), 422
    except Exception as exc:
        _delete_token_files(token)
        return jsonify({"error": f"Processing failed: {exc}"}), 500

    pending = session.get("pending_scorecard_downloads", {})
    if len(pending) >= MAX_PENDING:
        oldest = next(iter(pending))
        _delete_token_files(oldest)
        del pending[oldest]
    pending[token] = record["download_names"]
    session["pending_scorecard_downloads"] = pending
    session.modified = True

    return jsonify({"token": token, "original_name": pnl_name, "analysis": build_payload(record)})


@scorecard_bp.route("/analysis/<token>", methods=["POST"])
@login_required
def analysis(token):
    _assert_pending_token(token)
    record = _load_record(token)
    payload = request.get_json(silent=True) or {}
    selected_months = payload.get("months")
    return jsonify({"analysis": build_payload(record, selected_months)})


@scorecard_bp.route("/download/<token>/<kind>")
@login_required
def download(token, kind):
    _assert_pending_token(token)
    record = _load_record(token)
    files = record.get("files", {})
    download_names = record.get("download_names", {})
    if kind not in files or kind not in download_names:
        abort(404)

    file_path = _upload_dir() / files[kind]
    if not file_path.exists():
        abort(404)

    return send_file(
        io.BytesIO(file_path.read_bytes()),
        as_attachment=True,
        download_name=download_names[kind],
        mimetype=_mimetype_for_kind(kind, file_path),
    )


def _history_months_from_kpis(kpis):
    """Build the list of per-month dicts scorecard_history.upsert_months()
    expects, using this module's own month-label parsing (month_sort_key)
    so the standalone history module stays generic and doesn't need to
    know anything about "Mon YYYY"-style labels."""
    months = []
    for month in kpis["income"]:
        months.append(
            {
                "month": month,
                "month_start": month_sort_key(month).isoformat(),
                "income": kpis["income"].get(month),
                "expenses": kpis["expenses"].get(month),
                "noi": kpis["noi"].get(month),
                "occupancy": kpis["physical_occupancy"].get(month),
                "expense_ratio": kpis["expense_ratio"].get(month),
            }
        )
    return months


def _pct_change(current, previous):
    if current is None or previous is None or previous == 0:
        return None
    return (current - previous) / abs(previous)


def save_history_and_build_comparison(pnl_data, kpis):
    """Save this upload's monthly KPIs to the property's history, and
    compare its most recent month against whatever was on file for this
    property before this save (across all prior uploads, not just the
    prior upload's own trailing window). Never raises -- a history/DB
    hiccup should not break the upload itself."""
    property_name = pnl_data.get("property") or "Unknown Property"
    months = _history_months_from_kpis(kpis)
    if not months:
        return [], None, None

    current_month = max(months, key=lambda m: m["month_start"])
    uploaded_at = datetime.datetime.utcnow().isoformat()

    try:
        with scorecard_history.get_connection() as conn:
            property_key = scorecard_history.normalize_property_key(property_name)
            previous_latest = scorecard_history.get_latest(conn, property_key)
            scorecard_history.upsert_months(conn, property_name, months, uploaded_at)
            full_history = scorecard_history.get_history(conn, property_key)
    except Exception as exc:
        return [], None, f"Scorecard history: could not save/compare this upload ({exc})."

    comparison = {"available": False}
    if previous_latest:
        comparison = {
            "available": True,
            "previous_month": previous_latest["month"],
            "previous_uploaded_at": previous_latest["uploaded_at"],
            "current_month": current_month["month"],
            "metrics": {
                "noi": {
                    "previous": previous_latest["noi"],
                    "current": current_month["noi"],
                    "pct_change": _pct_change(current_month["noi"], previous_latest["noi"]),
                },
                "occupancy": {
                    "previous": previous_latest["occupancy"],
                    "current": current_month["occupancy"],
                    "point_change": (
                        None
                        if current_month["occupancy"] is None or previous_latest["occupancy"] is None
                        else current_month["occupancy"] - previous_latest["occupancy"]
                    ),
                },
                "expense_ratio": {
                    "previous": previous_latest["expense_ratio"],
                    "current": current_month["expense_ratio"],
                    "point_change": (
                        None
                        if current_month["expense_ratio"] is None or previous_latest["expense_ratio"] is None
                        else current_month["expense_ratio"] - previous_latest["expense_ratio"]
                    ),
                },
            },
        }
        noi_pct = comparison["metrics"]["noi"]["pct_change"]
        if noi_pct is not None:
            direction = "up" if noi_pct >= 0 else "down"
            comparison["summary_text"] = (
                f"NOI {direction} {abs(noi_pct):.1%} since your last upload "
                f"({previous_latest['month']} -> {current_month['month']})."
            )
        else:
            comparison["summary_text"] = (
                f"Compared against your last upload ({previous_latest['month']}), "
                f"but NOI wasn't available for one of the two months."
            )
    return full_history, comparison, None


def process_scorecard(token, pnl_path, pnl_name, scorecard_path=None, scorecard_name=""):
    parser = PnLParser(pnl_path)
    parser.parse()
    pnl_data = parser.get_data()
    if not pnl_data["accounts"]:
        raise ValueError("No recognizable accounts were parsed from the P&L CSV.")

    calc = KPICalculator(pnl_data)
    kpis = calc.calculate()
    report_text = ReportGenerator(kpis).generate()

    targets = {"UW": {}, "PM": {}}
    target_diagnostics = None
    update_diagnostics = None
    files = {}

    if scorecard_path:
        target_parser = ScorecardTargetParser(scorecard_path)
        target_parser.parse()
        targets = target_parser.get_data()
        target_diagnostics = target_parser.get_diagnostics()

        updated_ext = Path(scorecard_path).suffix.lower()
        updated_path = _upload_dir() / f"{token}_updated_scorecard{updated_ext}"
        updater = ScorecardUpdater(scorecard_path, pnl_data)
        updated = updater.update(updated_path)
        update_diagnostics = updater.get_diagnostics()
        if updated:
            files["scorecard"] = updated_path.name

    full_df = build_kpi_dataframe(kpis)
    property_slug = slugify(pnl_data.get("property") or Path(pnl_name).stem or "property")
    base_name = f"{property_slug}_scorecard"

    report_path = _upload_dir() / f"{token}_report.txt"
    report_path.write_text(report_text, encoding="utf-8")
    files["report"] = report_path.name

    csv_path = _upload_dir() / f"{token}_kpi_data.csv"
    write_kpi_csv(csv_path, full_df)
    files["csv"] = csv_path.name

    xlsx_path = _upload_dir() / f"{token}_scorecard_data.xlsx"
    write_export_xlsx(xlsx_path, full_df, pnl_data["accounts"], targets)
    files["xlsx"] = xlsx_path.name

    pdf_path = _upload_dir() / f"{token}_scorecard_report.pdf"
    create_pdf_report(pdf_path, pnl_data, kpis, targets, generate_advanced_insights(full_df, pnl_data["accounts"], targets), full_df)
    files["pdf"] = pdf_path.name

    download_names = {
        "report": f"{base_name}_report.txt",
        "csv": f"{base_name}_kpi_data.csv",
        "xlsx": f"{base_name}_data.xlsx",
        "pdf": f"{base_name}_report.pdf",
    }
    if "scorecard" in files:
        ext = Path(scorecard_name or files["scorecard"]).suffix.lower() or ".xlsx"
        download_names["scorecard"] = f"{base_name}_updated{ext}"

    history_trend, history_comparison, history_error = save_history_and_build_comparison(pnl_data, kpis)

    warnings = list(pnl_data.get("meta", {}).get("warnings", []))
    if target_diagnostics:
        warnings.extend(target_diagnostics.get("warnings", []))
    if update_diagnostics:
        warnings.extend(update_diagnostics.get("warnings", []))
    if history_error:
        warnings.append(history_error)

    record = {
        "token": token,
        "pnl_name": pnl_name,
        "scorecard_name": scorecard_name,
        "pnl_data": pnl_data,
        "kpis": kpis,
        "targets": targets,
        "target_diagnostics": target_diagnostics,
        "update_diagnostics": update_diagnostics,
        "report_text": report_text,
        "files": files,
        "download_names": download_names,
        "warnings": warnings,
        "history_trend": history_trend,
        "history_comparison": history_comparison,
        "created_at": datetime.datetime.utcnow().isoformat(),
    }
    _save_record(token, record)
    return record


def build_payload(record, selected_months=None):
    pnl_data = record["pnl_data"]
    kpis = record["kpis"]
    months = list(kpis["income"].keys())
    selected = [month for month in (selected_months or months) if month in months]
    if not selected:
        selected = months

    df_full = build_kpi_dataframe(kpis)
    df_filtered = df_full[df_full["Month"].isin(selected)]
    targets = record.get("targets") or {"UW": {}, "PM": {}}
    insights = generate_advanced_insights(df_filtered, pnl_data["accounts"], targets)
    comparison_rows = build_comparison_rows(kpis, targets, selected)

    payload = {
        "property": pnl_data.get("property", "Property"),
        "period": pnl_data.get("period", "Period"),
        "format": pnl_data.get("meta", {}).get("format", "Unknown"),
        "warnings": record.get("warnings", []),
        "months": months,
        "selected_months": selected,
        "summary": summarize_dataframe(df_filtered, kpis),
        "latest_quarter": latest_full_quarter(df_full),
        "kpi_rows": _records_for_json(df_filtered),
        "accounts": build_account_payload(pnl_data["accounts"], selected),
        "comparison": comparison_rows,
        "targets": targets,
        "target_diagnostics": record.get("target_diagnostics"),
        "update_diagnostics": record.get("update_diagnostics"),
        "insights": insights,
        "report_text": record.get("report_text", ""),
        "downloads": sorted(record.get("download_names", {}).keys()),
        "charts": build_charts(df_filtered),
        "history": {
            "trend": record.get("history_trend") or [],
            "comparison": record.get("history_comparison"),
        },
    }
    return clean_for_json(payload)


def build_kpi_dataframe(kpis):
    months = list(kpis["income"].keys())
    return pd.DataFrame(
        {
            "Month": months,
            "Income": [kpis["income"].get(month, 0.0) for month in months],
            "Expenses": [kpis["expenses"].get(month, 0.0) for month in months],
            "NOI": [kpis["noi"].get(month, 0.0) for month in months],
            "Occupancy": [kpis["physical_occupancy"].get(month) for month in months],
            "EconomicOccupancy": [kpis["economic_occupancy"].get(month) for month in months],
            "ExpenseRatio": [kpis["expense_ratio"].get(month) for month in months],
            "NOIMargin": [kpis["noi_margin"].get(month) for month in months],
            "OccupancyStatus": [kpis["occupancy_status"].get(month, "ok") for month in months],
        }
    )


def summarize_dataframe(df, kpis):
    if df.empty:
        return {
            "total_income": 0.0,
            "total_expenses": 0.0,
            "total_noi": 0.0,
            "avg_occupancy": None,
            "avg_expense_ratio": None,
            "missing_gpr_months": [],
            "zero_occupancy_months": [],
            "expense_fallback_codes": kpis.get("expense_fallback_codes", []),
        }
    occ_values = df["Occupancy"].dropna()
    ratio_values = df["ExpenseRatio"].dropna()
    return {
        "total_income": float(df["Income"].sum()),
        "total_expenses": float(df["Expenses"].sum()),
        "total_noi": float(df["NOI"].sum()),
        "avg_occupancy": float(occ_values.mean()) if not occ_values.empty else None,
        "avg_expense_ratio": float(ratio_values.mean()) if not ratio_values.empty else None,
        "missing_gpr_months": df.loc[df["OccupancyStatus"] == "missing_gpr", "Month"].tolist(),
        "zero_occupancy_months": df.loc[df["OccupancyStatus"] == "zero", "Month"].tolist(),
        "expense_fallback_codes": kpis.get("expense_fallback_codes", []),
    }


def latest_full_quarter(df):
    if df.empty:
        return None

    working = df.copy()
    working["QuarterKey"] = working["Month"].apply(quarter_key)
    working = working[working["QuarterKey"].notna()]
    if working.empty:
        return None

    counts = working.groupby("QuarterKey")["Month"].count()
    complete = [key for key, count in counts.items() if count == 3]
    if not complete:
        return None

    latest_key = sorted(complete)[-1]
    q_data = working[working["QuarterKey"] == latest_key]
    occ_values = q_data["Occupancy"].dropna()
    return {
        "label": f"Q{latest_key[1]} {latest_key[0]}",
        "income": float(q_data["Income"].sum()),
        "expenses": float(q_data["Expenses"].sum()),
        "noi": float(q_data["NOI"].sum()),
        "occupancy": float(occ_values.mean()) if not occ_values.empty else None,
    }


def build_account_payload(accounts, selected_months):
    rows = []
    for code in sorted(accounts):
        data = accounts[code]
        monthly = [{"month": month, "amount": float(data["data"].get(month, 0.0) or 0.0)} for month in selected_months]
        amounts = [item["amount"] for item in monthly]
        rows.append(
            {
                "code": code,
                "name": data.get("name", code),
                "total": float(sum(amounts)),
                "average": float(sum(amounts) / len(amounts)) if amounts else 0.0,
                "monthly": monthly,
            }
        )
    return rows


def noi_variance_flag(variance, target):
    """Red beyond +/-10% of budget, green within +/-3%, otherwise unflagged."""
    if not target:
        return None
    pct = variance / abs(target)
    if abs(pct) > 0.10:
        return "red"
    if abs(pct) <= 0.03:
        return "green"
    return None


def build_comparison_rows(kpis, targets, selected_months):
    rows = []
    for month in selected_months:
        for metric, key in (("Revenue", "income"), ("Expenses", "expenses"), ("NOI", "noi")):
            target_key = "Income" if metric == "Revenue" else metric
            actual = float(kpis[key].get(month, 0.0) or 0.0)
            uw = float((targets.get("UW") or {}).get(target_key, 0.0) or 0.0)
            pm = float((targets.get("PM") or {}).get(target_key, 0.0) or 0.0)
            variance_uw = actual - uw
            variance_pm = actual - pm
            row = {
                "month": month,
                "metric": metric,
                "actual": actual,
                "uw": uw,
                "pm": pm,
                "variance_uw": variance_uw,
                "variance_pm": variance_pm,
            }
            # Flags are scoped to NOI (the metric Michelle asked to have
            # flagged against budget) — Revenue/Expenses variance is shown
            # but intentionally left unflagged here.
            if metric == "NOI":
                row["flag_uw"] = noi_variance_flag(variance_uw, uw)
                row["flag_pm"] = noi_variance_flag(variance_pm, pm)
            rows.append(row)
    return rows


def build_charts(df):
    if df.empty:
        return {}
    return {
        "trend": chart_trend(df),
        "waterfall": chart_waterfall(df),
        "occupancy": chart_occupancy(df),
        "expense_ratio": chart_expense_ratio(df),
    }


def chart_trend(df):
    fig, ax = plt.subplots(figsize=(8.5, 3.8))
    x = range(len(df))
    income_bars = ax.bar(x, df["Income"], label="Income", color="#1e40af", alpha=0.70)
    expense_bars = ax.bar(x, df["Expenses"], label="Expenses", color="#f97316", alpha=0.62)
    noi_line = ax.plot(x, df["NOI"], label="NOI", color="#4cbb17", linewidth=2.6, marker="o")[0]
    ax.set_xticks(list(x), df["Month"], rotation=35, ha="right")
    ax.yaxis.set_major_formatter(lambda val, _: money_axis(val))
    ax.grid(axis="y", alpha=0.18)

    # Data labels: dollar figure on each bar and each NOI point. Font kept
    # small (7pt) since a full 12-month history means three labels per
    # month competing for the same width; the NOI label gets a translucent
    # white backing since the line crosses in front of both bar series.
    # Months with no GPR data at all (Income/Expenses/NOI all genuinely
    # $0, not just small) are skipped entirely -- three stacked "$0"
    # labels add no information and are the one case dense enough to
    # actually overlap illegibly.
    has_data = df["OccupancyStatus"] != "missing_gpr"
    ax.bar_label(income_bars, labels=[money_label(v) if keep else "" for keep, v in zip(has_data, df["Income"])], padding=2, fontsize=7)
    ax.bar_label(expense_bars, labels=[money_label(v) if keep else "" for keep, v in zip(has_data, df["Expenses"])], padding=2, fontsize=7)
    for xi, yi, keep in zip(x, df["NOI"], has_data):
        if not keep:
            continue
        ax.annotate(
            money_label(yi), (xi, yi), textcoords="offset points", xytext=(0, 8),
            ha="center", fontsize=7, fontweight="bold", color="#2f6b0e",
            bbox=dict(facecolor="white", edgecolor="none", alpha=0.65, pad=1),
        )

    # Below the chart rather than overlapping the plotted bars/line -- a
    # tall Income month previously sat right under the upper-left legend.
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.32), ncols=3, frameon=False)
    ax.set_title("Financial Performance Trend", loc="left", fontweight="bold")
    fig.tight_layout()
    return fig_to_data_uri(fig)


def chart_waterfall(df):
    total_inc = float(df["Income"].sum())
    total_exp = float(df["Expenses"].sum())
    total_noi = float(df["NOI"].sum())
    fig, ax = plt.subplots(figsize=(5.4, 3.8))
    labels = ["Income", "Expenses", "NOI"]
    # Floating bars: Income stands at full height, Expenses is drawn as a
    # positive-magnitude segment bridging down from Income to NOI (the
    # portion of Income it consumes), and NOI stands at its own full
    # height as the remaining value -- not a negative expense bar summed
    # against some other total.
    bottoms = [0, total_noi, 0]
    heights = [total_inc, total_exp, total_noi]
    colors = ["#1e40af", "#f97316", "#4cbb17" if total_noi >= 0 else "#dc2626"]
    bars = ax.bar(labels, heights, bottom=bottoms, color=colors, alpha=0.85)
    ax.bar_label(
        bars, labels=[money_label(v) for v in [total_inc, total_exp, total_noi]],
        label_type="center", fontsize=10, fontweight="bold", color="white",
    )
    ax.axhline(0, color="#1f2937", linewidth=0.8)
    ax.yaxis.set_major_formatter(lambda val, _: money_axis(val))
    ax.grid(axis="y", alpha=0.18)
    ax.set_title("NOI Breakdown", loc="left", fontweight="bold")
    fig.tight_layout()
    return fig_to_data_uri(fig)


def chart_occupancy(df):
    fig, ax = plt.subplots(figsize=(6.4, 3.8))
    values = [None if pd.isna(value) else value * 100 for value in df["Occupancy"]]
    colors = ["#9ca3af" if value is None else ("#dc2626" if value < 80 else "#f59e0b" if value < 90 else "#059669") for value in values]
    display_values = [0 if value is None else value for value in values]
    bars = ax.bar(df["Month"], display_values, color=colors, alpha=0.78)
    labels = ["N/A" if value is None else f"{value:.1f}%" for value in values]
    ax.bar_label(bars, labels=labels, padding=3, fontsize=7)
    ax.axhline(90, color="#1f2937", linestyle="--", linewidth=1)
    ax.set_ylim(0, 112)
    ax.set_ylabel("Occupancy %")
    ax.tick_params(axis="x", rotation=35)
    ax.grid(axis="y", alpha=0.18)
    ax.set_title("Occupancy Health", loc="left", fontweight="bold")
    fig.tight_layout()
    return fig_to_data_uri(fig)


def chart_expense_ratio(df):
    fig, ax = plt.subplots(figsize=(5.4, 3.8))
    ratios = [None if pd.isna(value) else value * 100 for value in df["ExpenseRatio"]]
    ax.plot(df["Month"], ratios, color="#7c3aed", linewidth=2.6, marker="o")
    for xi, yi in zip(df["Month"], ratios):
        if yi is None:
            continue
        ax.annotate(
            f"{yi:.1f}%", (xi, yi), textcoords="offset points", xytext=(0, 8),
            ha="center", fontsize=7, fontweight="bold", color="#5b21b6",
            bbox=dict(facecolor="white", edgecolor="none", alpha=0.65, pad=1),
        )
    ax.axhline(55, color="#f59e0b", linestyle="--", linewidth=1)
    ax.margins(y=0.15)
    ax.set_ylabel("Expense Ratio %")
    ax.tick_params(axis="x", rotation=35)
    ax.grid(axis="y", alpha=0.18)
    ax.set_title("Expense Ratio Analysis", loc="left", fontweight="bold")
    fig.tight_layout()
    return fig_to_data_uri(fig)


def fig_to_data_uri(fig):
    output = io.BytesIO()
    fig.savefig(output, format="png", dpi=140, bbox_inches="tight")
    plt.close(fig)
    return "data:image/png;base64," + base64.b64encode(output.getvalue()).decode("ascii")


def write_kpi_csv(path, df):
    export_df = df.copy()
    export_df.to_csv(path, index=False)


def write_export_xlsx(path, df, accounts, targets):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Data"
    write_sheet_rows(ws, [df.columns.tolist()] + df.fillna("").values.tolist())

    ws_accounts = wb.create_sheet("Accounts")
    account_rows = [["Code", "Name", "Month", "Amount"]]
    for code in sorted(accounts):
        acc = accounts[code]
        for month, amount in acc["data"].items():
            account_rows.append([code, acc.get("name", code), month, amount])
    write_sheet_rows(ws_accounts, account_rows)

    if targets and ((targets.get("UW") or {}) or (targets.get("PM") or {})):
        ws_targets = wb.create_sheet("Targets")
        rows = [["Metric", "UW Monthly", "PM Budget Monthly"]]
        for metric in ("Income", "Expenses", "NOI"):
            rows.append([metric, (targets.get("UW") or {}).get(metric, 0), (targets.get("PM") or {}).get(metric, 0)])
        write_sheet_rows(ws_targets, rows)

    wb.save(path)


def write_sheet_rows(ws, rows):
    header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for row_idx, row in enumerate(rows, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18


def create_pdf_report(path, pnl_data, kpis, targets, insights, df):
    with PdfPages(path) as pdf:
        fig = plt.figure(figsize=(11, 8.5))
        add_pdf_header(fig, pnl_data)
        fig.text(0.06, 0.76, "Dashboard Summary", fontsize=15, fontweight="bold", color="#1a2744")
        summary = summarize_dataframe(df, kpis)
        metric_lines = [
            ("Total Income", format_currency(summary["total_income"])),
            ("Total Expenses", format_currency(summary["total_expenses"])),
            ("Total NOI", format_currency(summary["total_noi"])),
            ("Average Occupancy", format_percent(summary["avg_occupancy"])),
        ]
        for idx, (label, value) in enumerate(metric_lines):
            x = 0.06 + idx * 0.225
            fig.text(x, 0.69, label, fontsize=9, color="#6b7280", fontweight="bold")
            fig.text(x, 0.65, value, fontsize=16, color="#111827", fontweight="bold")

        ax = fig.add_axes([0.08, 0.15, 0.84, 0.38])
        x = range(len(df))
        ax.bar(x, df["Income"], label="Income", color="#5b8def", alpha=0.70)
        ax.bar(x, df["Expenses"], label="Expenses", color="#f59e0b", alpha=0.62)
        ax.plot(x, df["NOI"], label="NOI", color="#059669", linewidth=2.6, marker="o")
        ax.set_xticks(list(x), df["Month"], rotation=35, ha="right")
        ax.yaxis.set_major_formatter(lambda val, _: money_axis(val))
        ax.grid(axis="y", alpha=0.18)
        ax.legend(loc="upper left", ncols=3, frameon=False)
        ax.set_title("Financial Performance Trend", loc="left", fontweight="bold")
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        fig = plt.figure(figsize=(11, 8.5))
        add_pdf_header(fig, pnl_data)
        fig.text(0.06, 0.76, "Insights & Recommendations", fontsize=15, fontweight="bold", color="#1a2744")
        y = 0.69
        for title, items in (
            ("Key Trends", [item[0] for item in insights.get("trends", [])] or ["Metrics are relatively stable."]),
            ("Green Flags", insights.get("green_flags", []) or ["None identified."]),
            ("Red Flags", insights.get("red_flags", []) or ["None identified."]),
            ("Recommendations", insights.get("recommendations", [])),
        ):
            fig.text(0.06, y, title, fontsize=11, fontweight="bold", color="#1a2744")
            y -= 0.035
            for item in items[:5]:
                fig.text(0.08, y, f"- {item}", fontsize=9.5, color="#111827")
                y -= 0.032
            y -= 0.025
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        if targets and ((targets.get("UW") or {}) or (targets.get("PM") or {})):
            fig = plt.figure(figsize=(11, 8.5))
            add_pdf_header(fig, pnl_data)
            fig.text(0.06, 0.76, "NOI Target Comparison", fontsize=15, fontweight="bold", color="#1a2744")
            ax = fig.add_axes([0.09, 0.18, 0.84, 0.48])
            months = df["Month"].tolist()
            actual = [kpis["noi"].get(month, 0) for month in months]
            uw = [(targets.get("UW") or {}).get("NOI", 0) for _ in months]
            pm = [(targets.get("PM") or {}).get("NOI", 0) for _ in months]
            x = range(len(months))
            ax.bar(x, actual, color="#5b8def", label="Actual", alpha=0.80)
            ax.plot(x, uw, color="#6b7280", label="UW", linestyle="--", linewidth=2.0)
            ax.plot(x, pm, color="#f59e0b", label="PM Budget", linestyle=":", linewidth=2.4)
            ax.set_xticks(list(x), months, rotation=35, ha="right")
            ax.yaxis.set_major_formatter(lambda val, _: money_axis(val))
            ax.grid(axis="y", alpha=0.18)
            ax.legend(frameon=False)
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)


def add_pdf_header(fig, pnl_data):
    logo_path = _logo_png_path()
    if logo_path and logo_path.exists():
        try:
            logo_ax = fig.add_axes([0.06, 0.86, 0.24, 0.08])
            logo_ax.imshow(mpimg.imread(str(logo_path)))
            logo_ax.axis("off")
        except Exception:
            pass
    fig.text(0.94, 0.91, "Property Scorecard Report", ha="right", fontsize=14, fontweight="bold", color="#1a2744")
    fig.text(0.94, 0.875, str(pnl_data.get("property", "Property")), ha="right", fontsize=10, color="#4b5563")
    fig.text(0.94, 0.845, str(pnl_data.get("period", "Period")), ha="right", fontsize=9, color="#6b7280")


def _logo_png_path():
    path = Path(current_app.root_path) / "static" / "fire_logo.png"
    return path if path.exists() else None


def _records_for_json(df):
    return df.where(pd.notna(df), None).to_dict(orient="records")


def month_sort_key(month_str):
    try:
        parts = str(month_str).split()
        if len(parts) == 2:
            mon = parts[0]
            year = int(parts[1])
            return datetime.date(year, MONTH_INDEX[mon], 1)
    except Exception:
        pass
    return datetime.date(1900, 1, 1)


def quarter_key(month_str):
    try:
        parts = str(month_str).split()
        if len(parts) != 2:
            return None
        mon = parts[0]
        year = int(parts[1])
        quarter = ((MONTH_INDEX[mon] - 1) // 3) + 1
        return (year, quarter)
    except Exception:
        return None


def format_currency(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "-"
    return f"${float(value):,.0f}"


def format_percent(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "N/A"
    return f"{float(value) * 100:.1f}%"


def money_axis(value):
    value = float(value)
    sign = "-" if value < 0 else ""
    value = abs(value)
    if value >= 1_000_000:
        return f"{sign}${value / 1_000_000:.1f}M"
    if value >= 1_000:
        return f"{sign}${value / 1_000:.0f}K"
    return f"{sign}${value:.0f}"


def money_label(value):
    """Full-precision, comma-formatted dollar figure for on-chart data
    labels -- unlike money_axis's K/M abbreviation (meant for axis ticks),
    this matches the exact formatting already used for dollar figures
    elsewhere in the tool (report text, dashboard summary cards)."""
    value = float(value)
    sign = "-" if value < 0 else ""
    return f"{sign}${abs(value):,.0f}"


def slugify(value):
    slug = re.sub(r"[^A-Za-z0-9]+", "_", str(value or "").strip()).strip("_").lower()
    return slug or "property"


def clean_for_json(value):
    if isinstance(value, dict):
        return {key: clean_for_json(val) for key, val in value.items()}
    if isinstance(value, list):
        return [clean_for_json(item) for item in value]
    if isinstance(value, tuple):
        return [clean_for_json(item) for item in value]
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def _upload_dir():
    path = Path(current_app.config["UPLOAD_FOLDER"]) / "scorecard-pro"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _record_path(token):
    return _upload_dir() / f"{token}_analysis.json"


def _save_record(token, record):
    _record_path(token).write_text(json.dumps(clean_for_json(record), ensure_ascii=False), encoding="utf-8")


def _load_record(token):
    path = _record_path(token)
    if not path.exists():
        abort(404)
    return json.loads(path.read_text(encoding="utf-8"))


def _assert_pending_token(token):
    pending = session.get("pending_scorecard_downloads", {})
    if token not in pending:
        abort(403)


def _cleanup_old_uploads(max_age=None):
    if max_age is None:
        max_age = int(current_app.permanent_session_lifetime.total_seconds())
    cutoff = time.time() - max_age
    for file_path in _upload_dir().glob("*"):
        if not file_path.is_file():
            continue
        try:
            if file_path.stat().st_mtime < cutoff:
                file_path.unlink(missing_ok=True)
        except OSError:
            pass


def _delete_token_files(token):
    for file_path in _upload_dir().glob(f"{token}_*"):
        if not file_path.is_file():
            continue
        try:
            file_path.unlink(missing_ok=True)
        except OSError:
            pass


def _mimetype_for_kind(kind, file_path):
    if kind == "pdf":
        return "application/pdf"
    if kind == "csv":
        return "text/csv"
    if kind == "report":
        return "text/plain"
    if file_path.suffix.lower() == ".xlsm":
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
