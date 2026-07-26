from __future__ import annotations

import re

import openpyxl
from openpyxl.cell.cell import MergedCell

from tools.scorecard_pro.constants import (
    MONTHS,
    _CANYON_EXCLUDED_CODES,
    _EAGLE_ROCK_EXCLUDED_CODES,
    _OXPT_EXCLUDED_CODES,
    _OXPT_MISC_FEE_NAMES,
    _OXPT_ROW_GROUPS,
)
from tools.scorecard_pro.utils import month_sort_key


def _safe_set_cell_value(sheet, row, col, value):
    """Write a value at (row, col), unmerging first if that coordinate
    currently falls on a non-anchor cell of a merged range.

    openpyxl's Worksheet.insert_cols()/insert_rows() shift cell *values*
    but do not shift merged_cells.ranges to match (confirmed against a
    real Canyon Scorecard: after inserting 8 missing month columns, the
    sheet's pre-existing decorative merges like "V50:AB50" still claimed
    columns 22-28 even though columns 26-28 now hold real newly-inserted
    month data) -- so a merge that used to sit harmlessly off to the side
    of the real data columns can end up silently overlapping them after
    insertion. Writing into the covered-but-not-anchor part of that stale
    merge raises "'MergedCell' object attribute 'value' is read-only".
    Once unmerged here, the range is not re-merged -- the merge's
    coordinates are already stale/meaningless post-insert, so re-applying
    it would just perpetuate the misalignment.
    """
    cell = sheet.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                sheet.unmerge_cells(str(merged_range))
                break
        cell = sheet.cell(row=row, column=col)
    cell.value = value


class ScorecardUpdater:
    def __init__(self, scorecard_path, data):
        self.scorecard_path = scorecard_path
        self.data = data
        self.wb = None
        self.sheet = None
        self.diagnostics = {"updated_cells": 0, "warnings": []}

    def update(self, output_path):
        try:
            suffix = str(self.scorecard_path).lower()
            self.wb = openpyxl.load_workbook(self.scorecard_path, keep_vba=suffix.endswith(".xlsm"))
        except Exception as exc:
            self.diagnostics["warnings"].append(f"Error loading scorecard: {exc}")
            return None

        if "T12" not in self.wb.sheetnames:
            self.diagnostics["warnings"].append("Scorecard updater: 'T12' sheet not found.")
            return None

        self.sheet = self.wb["T12"]
        excel_month_map = {}
        header_row = 6
        first_month_col = None
        last_month_col = None
        account_header_col = None
        for col_idx in range(1, 50):
            cell_val = self.sheet.cell(row=header_row, column=col_idx).value
            if not isinstance(cell_val, str):
                continue
            stripped = cell_val.strip()
            if account_header_col is None and stripped.lower() in ("account", "account name"):
                account_header_col = col_idx
                continue
            # Month headers vary by export: "Jan 2025 Actual" (4-digit year)
            # or "Jan-25" (2-digit year). Normalize both to the same
            # "Mon YYYY" key the parsed P&L data already uses.
            match = re.search(r"([A-Za-z]{3}).*?(\d{4}|\d{2})\b", stripped)
            if match:
                month_abbr = match.group(1)[:3].title()
                year_str = match.group(2)
                year = int(year_str) if len(year_str) == 4 else int("20" + year_str)
                if 2000 <= year <= 2099 and month_abbr in MONTHS:
                    excel_month_map[f"{month_abbr} {year}"] = col_idx
                    if first_month_col is None:
                        first_month_col = col_idx
                    last_month_col = col_idx

        # Extend the sheet with new month columns if the uploaded P&L covers
        # months the workbook's own T12 tab doesn't have yet (e.g. a newer
        # T12 refreshing a Scorecard whose T12 tab stops months earlier) —
        # rather than silently writing only the overlapping months.
        source_months = set()
        for acc_data in self.data["accounts"].values():
            source_months.update(acc_data["data"].keys())
        missing_months = sorted(
            (m for m in source_months if m not in excel_month_map),
            key=month_sort_key,
        )

        skipped_months = []
        skip_reason = None
        if missing_months and last_month_col is not None:
            insertion_col = last_month_col + 1
            # openpyxl shifts cell values on insert_cols() but does not
            # rewrite formula text, so inserting into a column range that
            # any formula elsewhere on the sheet references (e.g. an
            # "Adjusted Total" that sums across the month columns) would
            # silently produce a wrong total rather than an obviously
            # incomplete one. Only extend the sheet if no formulas exist
            # anywhere from the insertion point onward.
            has_formula_in_insert_region = any(
                isinstance(self.sheet.cell(r, c).value, str) and self.sheet.cell(r, c).value.startswith("=")
                for r in range(1, self.sheet.max_row + 1)
                for c in range(insertion_col, self.sheet.max_column + 1)
            )
            if has_formula_in_insert_region:
                skipped_months = missing_months
                skip_reason = (
                    "the T12 sheet has formulas in or after the column where new months "
                    "would be inserted, and inserting there could silently break those totals"
                )
            else:
                self.sheet.insert_cols(insertion_col, amount=len(missing_months))
                for offset, month_key in enumerate(missing_months):
                    col = insertion_col + offset
                    _safe_set_cell_value(self.sheet, header_row, col, f"{month_key} Actual")
                    excel_month_map[month_key] = col
        elif missing_months:
            skipped_months = missing_months
            skip_reason = "no existing month columns could be identified in the T12 sheet at all"

        # Different Scorecard T12 layouts put account-code labels in
        # different columns (Eagle Rock: column 1; Canyon: column 16+, after
        # an unrelated older "12 Month Rolling" report occupies columns
        # 1-14). Locate the label region dynamically — from the "Account"
        # header column through the column just before the first month
        # column — instead of assuming a fixed range.
        label_scan_start = account_header_col or 1
        label_scan_end = (first_month_col - 1) if first_month_col else (label_scan_start + 4)
        label_scan_end = max(label_scan_end, label_scan_start)

        # Label-column index (label -> row indices that actually carry data,
        # i.e. not a blank section-header row) — reused by both the OXPT
        # explicit category mapping and the generic name-based fallback.
        label_rows: dict = {}
        for row_idx in range(7, self.sheet.max_row + 1):
            label_val = self.sheet.cell(row_idx, label_scan_start).value
            if not isinstance(label_val, str) or not label_val.strip():
                continue
            has_data = any(
                self.sheet.cell(row_idx, c).value is not None
                and str(self.sheet.cell(row_idx, c).value).strip() != ""
                for c in range(label_scan_start + 1, self.sheet.max_column + 1)
            )
            if not has_data:
                continue
            label_rows.setdefault(label_val.strip().lower(), []).append(row_idx)

        account_row_map = {}
        for row_idx in range(7, self.sheet.max_row + 1):
            for col_idx in range(label_scan_start, label_scan_end + 1):
                val = self.sheet.cell(row=row_idx, column=col_idx).value
                if not val:
                    continue
                match = re.match(r"^(\d{4})\b", str(val).strip())
                if match:
                    account_row_map[match.group(1)] = row_idx
                    break

        # OXPT-specific explicit category mapping (Michelle's decisions —
        # see _OXPT_ROW_GROUPS above). Scoped strictly to OXPT by property
        # name so Eagle Rock/Canyon's matching is never affected.
        excluded_codes = set()
        oxpt_ambiguous_groups = []
        is_oxpt = "oxford pointe" in str(self.data.get("property") or "").strip().lower()
        if is_oxpt:
            excluded_codes |= _OXPT_EXCLUDED_CODES
            for label, codes in _OXPT_ROW_GROUPS.items():
                rows = label_rows.get(label)
                if not rows:
                    continue
                if len(rows) > 1:
                    oxpt_ambiguous_groups.append((label, codes, rows))
                    continue
                row_idx = rows[0]
                for code in codes:
                    if code in self.data["accounts"]:
                        account_row_map[code] = row_idx

            # Other Income detail lines with no fixed P&L code (see
            # _OXPT_MISC_FEE_NAMES) — matched by account name instead,
            # into the same "Miscellaneous Fees" row as the coded group above.
            misc_fee_rows = label_rows.get("miscellaneous fees")
            if misc_fee_rows:
                if len(misc_fee_rows) > 1:
                    oxpt_ambiguous_groups.append(("miscellaneous fees (name-matched)", [], misc_fee_rows))
                else:
                    for code, acc_data in self.data["accounts"].items():
                        name = str(acc_data.get("name") or "").strip().lower()
                        if name in _OXPT_MISC_FEE_NAMES:
                            account_row_map[code] = misc_fee_rows[0]

        # Eagle Rock/Canyon: no explicit row-group mapping needed (see
        # _EAGLE_ROCK_EXCLUDED_CODES / _CANYON_EXCLUDED_CODES above) — their
        # Scorecard T12 sheets already code-prefix-match every rollup
        # directly, so this only excludes already-captured leaf children.
        is_eagle_rock = "eagle rock" in str(self.data.get("property") or "").strip().lower()
        is_canyon = "canyon" in str(self.data.get("property") or "").strip().lower()
        if is_eagle_rock:
            excluded_codes |= _EAGLE_ROCK_EXCLUDED_CODES
        if is_canyon:
            excluded_codes |= _CANYON_EXCLUDED_CODES

        # Fallback for accounts not resolved above and not part of an
        # OXPT-excluded group. Match by the parsed P&L account's own name
        # against label-column rows that carry data, only when unambiguous.
        ambiguous_names = list(oxpt_ambiguous_groups)
        unmatched_codes = [
            code for code in self.data["accounts"]
            if code not in account_row_map and code not in excluded_codes
        ]
        for code in unmatched_codes:
            name = str(self.data["accounts"][code].get("name") or "").strip().lower()
            if not name:
                continue
            rows = label_rows.get(name)
            if not rows:
                continue
            if len(rows) == 1:
                account_row_map[code] = rows[0]
            else:
                ambiguous_names.append((code, self.data["accounts"][code].get("name"), rows))

        # Aggregate per (row, month-column) before writing — some rows
        # receive the summed value of multiple P&L codes (e.g. OXPT's
        # Contract Services split). For every other layout this is a
        # no-op: each code maps to a distinct row, so the "sum" is just
        # that one value, identical to a direct overwrite.
        cell_totals: dict = {}
        for code, acc_data in self.data["accounts"].items():
            if code not in account_row_map:
                continue
            row_idx = account_row_map[code]
            for month_key, value in acc_data["data"].items():
                col_idx = excel_month_map.get(month_key)
                if col_idx is None:
                    continue
                key = (row_idx, col_idx)
                cell_totals[key] = cell_totals.get(key, 0.0) + (value or 0.0)

        updates_count = 0
        for (row_idx, col_idx), total in cell_totals.items():
            _safe_set_cell_value(self.sheet, row_idx, col_idx, total)
            updates_count += 1

        if skipped_months:
            self.diagnostics["warnings"].append(
                "Scorecard updater: could not add columns for "
                + ", ".join(skipped_months)
                + f" — {skip_reason}. These months were not written to the updated "
                  "scorecard; extend the T12 sheet's month columns manually and re-run."
            )

        # Codes in excluded_codes are intentionally not matched individually
        # because their value is already fully captured via a mapped parent
        # rollup/group — they're not failures, so they're left out of both
        # the denominator and the "not written" list below.
        reportable_codes = [code for code in self.data.get("accounts", {}) if code not in excluded_codes]
        total_accounts = len(reportable_codes)
        matched_accounts = sum(1 for code in reportable_codes if code in account_row_map)
        if total_accounts and matched_accounts < total_accounts:
            unresolved = sorted(code for code in reportable_codes if code not in account_row_map)
            # Plain-language summary for the Parsing Notes UI -- the raw
            # code list isn't dropped, just moved out of the sentence
            # itself; it's still available in full via the Accounts tab
            # and the exported KPI CSV/XLSX (both list every account code).
            self.diagnostics["warnings"].append(
                f"Scorecard update: matched {matched_accounts} of {total_accounts} P&L accounts "
                f"to rows in the scorecard; {len(unresolved)} line item(s) could not be confidently "
                "matched (likely different account naming between the P&L export and this "
                "scorecard's own layout) and were not updated. See the Accounts tab or exported "
                "KPI data for the full account list."
            )
        if ambiguous_names:
            self.diagnostics["warnings"].append(
                f"Scorecard update: {len(ambiguous_names)} account(s) could not be confidently "
                "placed because their name matched more than one row in the scorecard. See the "
                "Accounts tab or exported KPI data for details."
            )

        self.diagnostics["updated_cells"] = updates_count
        self.wb.save(output_path)
        return output_path

    def get_diagnostics(self):
        return self.diagnostics
