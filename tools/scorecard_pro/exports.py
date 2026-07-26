from __future__ import annotations

from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.image as mpimg
import matplotlib.pyplot as plt
import openpyxl
from flask import current_app
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.styles import Font, PatternFill

from tools.scorecard_pro.utils import (
    format_currency,
    format_percent,
    money_axis,
    summarize_dataframe,
)


def write_kpi_csv(path, df):
    export_df = df.copy()
    export_df.to_csv(path, index=False)


def write_export_xlsx(path, df, accounts, targets):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Data"
    write_sheet_rows(ws, [df.columns.tolist()] + df.fillna("").values.tolist())

    ws_accounts = wb.create_sheet("Accounts")
    account_rows = [["Code", "Name", "Month", "Amount"]]
    for code in sorted(accounts):
        acc = accounts[code]
        for month, amount in acc["data"].items():
            account_rows.append([code, acc.get("name", code), month, amount])
    write_sheet_rows(ws_accounts, account_rows)

    if targets and ((targets.get("UW") or {}) or (targets.get("PM") or {})):
        ws_targets = wb.create_sheet("Targets")
        rows = [["Metric", "UW Monthly", "PM Budget Monthly"]]
        for metric in ("Income", "Expenses", "NOI"):
            rows.append([metric, (targets.get("UW") or {}).get(metric, 0), (targets.get("PM") or {}).get(metric, 0)])
        write_sheet_rows(ws_targets, rows)

    wb.save(path)


def write_sheet_rows(ws, rows):
    header_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for row_idx, row in enumerate(rows, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18


def create_pdf_report(path, pnl_data, kpis, targets, insights, df):
    with PdfPages(path) as pdf:
        fig = plt.figure(figsize=(11, 8.5))
        add_pdf_header(fig, pnl_data)
        fig.text(0.06, 0.76, "Dashboard Summary", fontsize=15, fontweight="bold", color="#1a2744")
        summary = summarize_dataframe(df, kpis)
        metric_lines = [
            ("Total Income", format_currency(summary["total_income"])),
            ("Total Expenses", format_currency(summary["total_expenses"])),
            ("Total NOI", format_currency(summary["total_noi"])),
            ("Average Occupancy", format_percent(summary["avg_occupancy"])),
        ]
        for idx, (label, value) in enumerate(metric_lines):
            x = 0.06 + idx * 0.225
            fig.text(x, 0.69, label, fontsize=9, color="#6b7280", fontweight="bold")
            fig.text(x, 0.65, value, fontsize=16, color="#111827", fontweight="bold")

        ax = fig.add_axes([0.08, 0.15, 0.84, 0.38])
        x = range(len(df))
        ax.bar(x, df["Income"], label="Income", color="#5b8def", alpha=0.70)
        ax.bar(x, df["Expenses"], label="Expenses", color="#f59e0b", alpha=0.62)
        ax.plot(x, df["NOI"], label="NOI", color="#059669", linewidth=2.6, marker="o")
        ax.set_xticks(list(x), df["Month"], rotation=35, ha="right")
        ax.yaxis.set_major_formatter(lambda val, _: money_axis(val))
        ax.grid(axis="y", alpha=0.18)
        ax.legend(loc="upper left", ncols=3, frameon=False)
        ax.set_title("Financial Performance Trend", loc="left", fontweight="bold")
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        fig = plt.figure(figsize=(11, 8.5))
        add_pdf_header(fig, pnl_data)
        fig.text(0.06, 0.76, "Insights & Recommendations", fontsize=15, fontweight="bold", color="#1a2744")
        y = 0.69
        for title, items in (
            ("Key Trends", [item[0] for item in insights.get("trends", [])] or ["Metrics are relatively stable."]),
            ("Green Flags", insights.get("green_flags", []) or ["None identified."]),
            ("Red Flags", insights.get("red_flags", []) or ["None identified."]),
            ("Recommendations", insights.get("recommendations", [])),
        ):
            fig.text(0.06, y, title, fontsize=11, fontweight="bold", color="#1a2744")
            y -= 0.035
            for item in items[:5]:
                fig.text(0.08, y, f"- {item}", fontsize=9.5, color="#111827")
                y -= 0.032
            y -= 0.025
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        if targets and ((targets.get("UW") or {}) or (targets.get("PM") or {})):
            fig = plt.figure(figsize=(11, 8.5))
            add_pdf_header(fig, pnl_data)
            fig.text(0.06, 0.76, "NOI Target Comparison", fontsize=15, fontweight="bold", color="#1a2744")
            ax = fig.add_axes([0.09, 0.18, 0.84, 0.48])
            months = df["Month"].tolist()
            actual = [kpis["noi"].get(month, 0) for month in months]
            uw = [(targets.get("UW") or {}).get("NOI", 0) for _ in months]
            pm = [(targets.get("PM") or {}).get("NOI", 0) for _ in months]
            x = range(len(months))
            ax.bar(x, actual, color="#5b8def", label="Actual", alpha=0.80)
            ax.plot(x, uw, color="#6b7280", label="UW", linestyle="--", linewidth=2.0)
            ax.plot(x, pm, color="#f59e0b", label="PM Budget", linestyle=":", linewidth=2.4)
            ax.set_xticks(list(x), months, rotation=35, ha="right")
            ax.yaxis.set_major_formatter(lambda val, _: money_axis(val))
            ax.grid(axis="y", alpha=0.18)
            ax.legend(frameon=False)
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)


def add_pdf_header(fig, pnl_data):
    logo_path = _logo_png_path()
    if logo_path and logo_path.exists():
        try:
            logo_ax = fig.add_axes([0.06, 0.86, 0.24, 0.08])
            logo_ax.imshow(mpimg.imread(str(logo_path)))
            logo_ax.axis("off")
        except Exception:
            pass
    fig.text(0.94, 0.91, "Property Scorecard Report", ha="right", fontsize=14, fontweight="bold", color="#1a2744")
    fig.text(0.94, 0.875, str(pnl_data.get("property", "Property")), ha="right", fontsize=10, color="#4b5563")
    fig.text(0.94, 0.845, str(pnl_data.get("period", "Period")), ha="right", fontsize=9, color="#6b7280")


def _logo_png_path():
    path = Path(current_app.root_path) / "static" / "fire_logo.png"
    return path if path.exists() else None
