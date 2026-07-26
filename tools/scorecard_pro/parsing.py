from __future__ import annotations

import csv
import datetime
import re
from pathlib import Path

import openpyxl
import pandas as pd

from tools.scorecard_pro.constants import MONTHS


class PnLParser:
    def __init__(self, filepath):
        self.property_name = "Unknown Property"
        self.period = "Unknown Period"
        self.accounts = {}
        self.detected_format = "Unknown"
        self.warnings = []

        if Path(filepath).suffix.lower() in (".xlsx", ".xlsm"):
            self.filepath = self._convert_workbook_to_csv(filepath)
        else:
            self.filepath = filepath

        # Standard Mapping for "The View" (Name -> Code)
        self.name_map = {
            "Gross potential rent": "4110",
            "loss to vacancy & other": "4220",
            "Net rental income": "4000",
            "Other income": "4300",
            "Personnel": "6400",
            "Maintenance & repairs": "6530",
            "Turn expenses": "6500",
            "Marketing & resident retention": "6300",
            "Insurance": "6700",
            "Property taxes": "6800",
            "Utilities subtotal": "6600",
            "Trash": "6500",
            "Revenues": "Total Income",
        }

        # Standard Mapping for "Paresh" (Paresh Code -> ERA Code)
        self.code_map = {
            "40210": "4110",
            "40310": "4220",
            "40200": "4000",
        }

        # Standard Mapping for "OXPT" (Name -> Code)
        self.oxpt_map = {
            "Gross Potential Rent": "4110",
            "Vacancy": "4220",
            "Total RENTS": "4000",
            "Cost Recovery Fee": "4300",
            "Total FEES": "4300",
            "Laundry Income": "4300",
            "Parking Income": "4300",
            "Total MANAGEMENT FEES": "6113",
            "Total INSURANCE": "6700",
            "Total GROUNDS & LAWN MAINTENANCE": "6515",
            "Total OFFICE EXPENSE": "6100",
            "Total PAYROLL EXPENSE": "6400",
            "Total CLEANING & TRASH REMOVAL": "6520",
            "Total REPAIRS": "6530",
            "Total TAXES": "6800",
            "Total UTILITIES": "6600",
            "Total OUTSIDE CONTRACTORS": "6500",
            "Total OTHER EXPENSES": "6100",
            "Total Operating Income": "9998",
            "Total Operating Expense": "9999",
        }

        # Standard Mapping for "Canyon"
        self.canyon_map = {
            "Gross Potential Rent (Scheduled)": "4110",
            "Vacancy Loss": "4220",
            "NET RENTAL REVENUE": "4000",
            "TOTAL OTHER INCOME": "4300",
            "TOTAL PERSONNEL EXPENSES": "6400",
            "TOTAL MANAGEMENT FEES": "6113",
            "TOTAL ADMINISTRATIVE EXPENSES": "6100",
            "TOTAL LEGAL & PROFESSIONAL": "6200",
            "TOTAL MARKETING & LEASING": "6300",
            "TOTAL UTILITIES": "6600",
            "TOTAL CONTRACT SERVICES": "6500",
            "TOTAL TURNOVER / CLEANING": "6520",
            "TOTAL REPAIRS & MAINTENANCE": "6530",
            "TOTAL TAXES & INSURANCE": "6800",
            "TOTAL INCOME": "9998",
            "TOTAL OPERATING EXPENSES": "9999",
        }

    def _convert_workbook_to_csv(self, filepath):
        """
        Convert an uploaded .xlsx/.xlsm P&L export to an equivalent CSV file
        on disk so the rest of PnLParser (format detection, all parse_*
        methods) can keep operating on self.filepath unchanged.

        Prefers a sheet named "Accounting Tree Report" (ResMan's T12 P&L
        export), falling back to the workbook's first sheet for any other
        xlsx P&L layout we haven't seen yet.
        """
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
        sheet_name = next(
            (name for name in wb.sheetnames if name.strip().lower() == "accounting tree report"),
            wb.sheetnames[0],
        )
        ws = wb[sheet_name]

        out_path = Path(filepath).with_suffix(".converted.csv")
        with open(out_path, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
        return out_path

    def parse(self):
        try:
            header_lines = self._read_head_lines(25)
            content = "".join(header_lines).lower()

            if "cash flow" in content and "account name" in content:
                self.detected_format = "Cash Flow (Generic)"
                self.parse_cash_flow()
            elif "exported on" in content and "account name" in content:
                self.detected_format = "Cash Flow (Generic)"
                self.parse_cash_flow()
            elif header_lines and "category" in header_lines[0].lower() and "canyon apartments" in content:
                self.detected_format = "Canyon"
                self.parse_canyon()
            elif header_lines and "category" in header_lines[0].lower():
                self.detected_format = "Paresh"
                self.parse_paresh()
            elif "account name" in content:
                self.detected_format = "OXPT"
                self.parse_oxpt()
            elif len(header_lines) > 1 and "ltm" in header_lines[1].lower():
                self.detected_format = "The View"
                self.parse_the_view()
            else:
                self.detected_format = "ResMan (Standard)"
                self.parse_resman()
        except Exception as exc:
            self.warnings.append(f"Error parsing CSV: {exc}")

    def _read_head_lines(self, max_lines=20):
        encodings = ["utf-8-sig", "utf-16", "latin-1"]
        for enc in encodings:
            try:
                with open(self.filepath, "r", encoding=enc, errors="ignore") as handle:
                    return [handle.readline() for _ in range(max_lines)]
            except Exception:
                continue
        return []

    def _read_csv_robust(self, header="infer", skiprows=0, sep=None):
        encodings = ["utf-8-sig", "utf-16", "latin-1"]
        last_err = None
        for enc in encodings:
            try:
                return pd.read_csv(
                    self.filepath,
                    header=header,
                    skiprows=skiprows,
                    sep=sep,
                    engine="python",
                    on_bad_lines="skip",
                    encoding=enc,
                )
            except Exception as exc:
                last_err = exc
                continue
        try:
            return pd.read_csv(self.filepath, header=header, skiprows=skiprows, encoding="utf-8-sig")
        except Exception as exc:
            self.warnings.append(f"CSV read failed: {exc}")
            raise exc if last_err is None else last_err

    def _find_line_index(self, needle):
        needle_lower = needle.lower()
        encodings = ["utf-8-sig", "utf-16", "latin-1"]
        for enc in encodings:
            try:
                with open(self.filepath, "r", encoding=enc, errors="ignore") as handle:
                    for idx, line in enumerate(handle):
                        if needle_lower in line.lower():
                            return idx
            except Exception:
                continue
        return 0

    def _clean_columns(self, df):
        new_cols = []
        seen = {}
        for col in df.columns:
            raw = str(col).replace("\ufeff", "").replace("\n", " ").replace("\r", " ")
            clean = re.sub(r"\s+", " ", raw).strip()
            if clean in seen:
                seen[clean] += 1
                clean = f"{clean}_{seen[clean]}"
            else:
                seen[clean] = 0
            new_cols.append(clean)
        df.columns = new_cols
        return df

    def _infer_default_year(self, labels):
        for label in labels:
            match = re.search(r"(20\d{2})", str(label))
            if match:
                return int(match.group(1))
        return datetime.date.today().year

    def normalize_month(self, raw_month, default_year=None):
        raw = str(raw_month).strip()
        if not raw:
            return None

        raw = raw.replace("\n", " ").replace("\r", " ")
        raw = raw.replace("-", " ").replace("_", " ").replace("/", " ").replace("\\", " ")
        raw = re.sub(r"\s+", " ", raw)

        month_map = {
            "jan": ("Jan", 1),
            "january": ("Jan", 1),
            "feb": ("Feb", 2),
            "february": ("Feb", 2),
            "mar": ("Mar", 3),
            "march": ("Mar", 3),
            "apr": ("Apr", 4),
            "april": ("Apr", 4),
            "may": ("May", 5),
            "jun": ("Jun", 6),
            "june": ("Jun", 6),
            "jul": ("Jul", 7),
            "july": ("Jul", 7),
            "aug": ("Aug", 8),
            "august": ("Aug", 8),
            "sep": ("Sep", 9),
            "sept": ("Sep", 9),
            "september": ("Sep", 9),
            "oct": ("Oct", 10),
            "october": ("Oct", 10),
            "nov": ("Nov", 11),
            "november": ("Nov", 11),
            "dec": ("Dec", 12),
            "december": ("Dec", 12),
        }

        raw_lower = raw.lower()
        month_abbr = None

        for key, (abbr, _) in month_map.items():
            if re.search(rf"\b{key}\b", raw_lower):
                month_abbr = abbr
                break

        if not month_abbr:
            num_match = re.search(r"\b([01]?\d)\b", raw_lower)
            if num_match:
                num = int(num_match.group(1))
                if 1 <= num <= 12:
                    month_abbr = MONTHS[num - 1]

        yr_match = re.search(r"(20\d{2}|\d{2})", raw_lower)
        year = None
        if yr_match:
            yr = yr_match.group(1)
            year = int("20" + yr) if len(yr) == 2 else int(yr)
        elif default_year:
            year = int(default_year)

        if month_abbr and year:
            return f"{month_abbr} {year}"
        if month_abbr and not year:
            return month_abbr
        return None

    def _parse_amount(self, raw_val):
        if raw_val is None:
            return 0.0
        value = str(raw_val).strip()
        if not value or value.lower() in ("nan", "none", "null"):
            return 0.0
        if re.fullmatch(r"[-\u2013\u2014]+", value):
            return 0.0
        value = value.replace("$", "").replace(",", "").replace('"', "").strip()
        negative = False
        if value.startswith("(") and value.endswith(")"):
            negative = True
            value = value[1:-1].strip()
        if value.endswith("-"):
            negative = True
            value = value[:-1].strip()
        try:
            parsed = float(value)
        except Exception:
            return 0.0
        return -parsed if negative else parsed

    def _merge_account(self, code, name, monthly_data, depth=None):
        if code in self.accounts:
            existing = self.accounts[code]
            existing_data = existing["data"]
            for key, value in monthly_data.items():
                existing_data[key] = existing_data.get(key, 0.0) + (value if value is not None else 0.0)
            if depth is not None:
                existing["depth"] = depth if existing.get("depth") is None else min(existing["depth"], depth)
        else:
            self.accounts[code] = {"name": name, "data": monthly_data, "depth": depth}

    def parse_canyon(self):
        df = self._clean_columns(self._read_csv_robust(header=0))
        self.property_name = "Canyon Apartments"
        self.period = "T12"

        default_year = self._infer_default_year(df.columns)
        month_cols = {
            col: month
            for col in df.columns
            if (month := self.normalize_month(col, default_year=default_year))
            and "20" in month
            and any(mon in month for mon in MONTHS)
        }

        for _, row in df.iterrows():
            cat = str(row["Category"]).strip()
            code = self.canyon_map.get(cat)
            if not code:
                continue
            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, cat, monthly_data)

    def parse_oxpt(self):
        header_row_idx = self._find_line_index("Account Name")
        df = self._clean_columns(self._read_csv_robust(header=0, skiprows=header_row_idx))
        self.property_name = "Oxford Pointe"
        self.period = "2025"

        default_year = self._infer_default_year(df.columns)
        month_cols = {
            col: month
            for col in df.columns
            if (month := self.normalize_month(col, default_year=default_year))
            and "20" in month
            and any(mon in month for mon in MONTHS)
        }

        for _, row in df.iterrows():
            name_raw = str(row["Account Name"]).strip()
            if not name_raw or name_raw == "nan":
                continue

            code = self.oxpt_map.get(name_raw)
            if not code and "Total" in name_raw:
                if "Payroll" in name_raw:
                    code = "6400"
                elif "Utilities" in name_raw:
                    code = "6600"
                elif "Repairs" in name_raw:
                    code = "6500"
                elif "Marketing" in name_raw:
                    code = "6300"

            if not code:
                continue

            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, name_raw, monthly_data)

    def parse_paresh(self):
        df = self._clean_columns(self._read_csv_robust(header=0))
        self.property_name = "Paresh Property"
        self.period = "T12"

        default_year = self._infer_default_year(df.columns)
        month_cols = {
            col: month
            for col in df.columns
            if (month := self.normalize_month(col, default_year=default_year))
            and "20" in month
            and any(mon in month for mon in MONTHS)
        }

        for _, row in df.iterrows():
            cat = str(row["Category"]).strip()
            match = re.match(r"^(\d{5})\.\d{4}-(.+)$", cat)
            if not match:
                continue
            p_code = match.group(1)
            name = match.group(2)
            code = self.code_map.get(p_code, p_code[:4])
            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, name, monthly_data)

    def parse_the_view(self):
        df = self._clean_columns(self._read_csv_robust(header=0, skiprows=1))
        self.property_name = "The View"
        self.period = "LTM 2025"

        default_year = self._infer_default_year(df.columns)
        month_cols = {
            col: month
            for col in df.columns
            if (month := self.normalize_month(col, default_year=default_year))
            and "20" in month
            and any(mon in month for mon in MONTHS)
        }

        for _, row in df.iterrows():
            name = str(row.iloc[0]).strip()
            if not name or name == "nan":
                continue

            code = self.name_map.get(name)
            if not code:
                continue

            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, name, monthly_data)

    def parse_resman(self):
        df = self._read_csv_robust(header=None, sep=",")
        self.property_name = df.iloc[0, 0]
        self.period = df.iloc[3, 0]

        header_row = df.iloc[5]
        month_col_indices = {}
        for idx in range(8, 20):
            if idx < len(header_row):
                month = self.normalize_month(str(header_row[idx]), default_year=self._infer_default_year(header_row))
                if month and "20" in month:
                    month_col_indices[idx] = month

        for idx in range(6, len(df)):
            row = df.iloc[idx]
            account_str = None
            account_depth = None
            for col_idx in range(8):
                val = row[col_idx]
                if pd.notna(val) and str(val).strip() != "":
                    account_str = str(val).strip()
                    account_depth = col_idx
                    break

            if not account_str:
                continue

            code_match = re.match(r"^(\d{4})\s+(.+)$", account_str)
            if not code_match:
                continue

            code = code_match.group(1)
            name = code_match.group(2)
            monthly_values = {
                month_name: self._parse_amount(row[col_idx])
                for col_idx, month_name in month_col_indices.items()
            }
            self._merge_account(code, name, monthly_values, depth=account_depth)

    def parse_cash_flow(self):
        header_row_idx = 0
        header_lines = self._read_head_lines(15)
        for line in header_lines:
            line_clean = line.strip().strip(",")
            if not line_clean:
                continue
            if line_clean.lower().startswith("exported on"):
                continue
            if line_clean.lower().startswith("period range:"):
                self.period = line_clean.split(":", 1)[-1].strip()
                continue
            if self.property_name == "Unknown Property":
                self.property_name = line_clean

        header_row_idx = self._find_line_index("Account Name")
        df = self._clean_columns(self._read_csv_robust(header=0, skiprows=header_row_idx))

        if self.property_name == "Unknown Property":
            self.property_name = "Property"
        if self.period == "Unknown Period":
            self.period = "Cash Flow"

        if "Account Name" not in df.columns:
            self.warnings.append("Cash Flow parser: 'Account Name' column not found.")
            return

        default_year = self._infer_default_year(df.columns)
        month_cols = {}
        for col in df.columns:
            if col == "Account Name":
                continue
            month = self.normalize_month(col, default_year=default_year)
            if month and "20" in month and any(mon in month for mon in MONTHS):
                month_cols[col] = month

        def map_keyword_to_code(label):
            lowered = label.lower()
            if any(key in lowered for key in ["vacancy", "concession", "loss to vacancy"]):
                return "4220"
            if any(key in lowered for key in ["rent", "rents"]) and "vacancy" not in lowered:
                return "4000"
            if any(
                key in lowered
                for key in [
                    "other income",
                    "fee",
                    "fees",
                    "misc",
                    "forfeit",
                    "application",
                    "utility reimbursement",
                    "pet rent",
                    "late",
                ]
            ):
                return "4300"
            if any(key in lowered for key in ["payroll", "wages", "salary", "salaries", "benefits"]):
                return "6400"
            if any(key in lowered for key in ["utilities", "electric", "water", "gas", "sewer"]):
                return "6600"
            if any(key in lowered for key in ["repair", "repairs", "maintenance", "supplies", "cleaning"]):
                return "6530"
            if any(key in lowered for key in ["contract", "landscaping", "trash", "pest", "grounds"]):
                return "6500"
            if any(key in lowered for key in ["marketing", "leasing", "advertising"]):
                return "6300"
            if "insurance" in lowered:
                return "6700"
            if any(key in lowered for key in ["tax", "taxes"]):
                return "6800"
            if any(key in lowered for key in ["legal", "professional", "accounting"]):
                return "6200"
            if any(key in lowered for key in ["office", "administrative", "admin"]):
                return "6100"
            return None

        for _, row in df.iterrows():
            name_raw = str(row["Account Name"]).strip()
            if not name_raw or name_raw.lower() == "nan":
                continue
            name_clean = re.sub(r"^\s+", "", name_raw)
            lowered = name_clean.lower()

            if all(pd.isna(row[col]) or str(row[col]).strip() == "" for col in month_cols.keys()):
                continue

            code = None
            if "total operating income" in lowered:
                code = "9998"
            elif "total operating expense" in lowered:
                code = "9999"
            elif lowered.startswith("total "):
                continue
            else:
                code = map_keyword_to_code(lowered)

            if not code:
                continue

            monthly_data = {month_std: self._parse_amount(row[col]) for col, month_std in month_cols.items()}
            self._merge_account(code, name_clean, monthly_data)

    def get_data(self):
        return {
            "property": str(self.property_name or "Property"),
            "period": str(self.period or "Period"),
            "accounts": self.accounts,
            "meta": {"format": self.detected_format, "warnings": self.warnings},
        }


class ScorecardTargetParser:
    def __init__(self, filepath):
        self.filepath = filepath
        self.targets = {"UW": {}, "PM": {}}
        self.diagnostics = {
            "scorecard_sheet_found": False,
            "found_columns": {},
            "missing_columns": [],
            "found_rows": [],
            "missing_rows": [],
            "warnings": [],
        }

    def parse(self):
        try:
            wb = openpyxl.load_workbook(self.filepath, data_only=True)
            if "Scorecard" not in wb.sheetnames:
                self.diagnostics["warnings"].append("Target parser: 'Scorecard' sheet not found.")
                return

            self.diagnostics["scorecard_sheet_found"] = True
            sheet = wb["Scorecard"]
            uw_col = None
            pm_col = None

            for row_idx in range(1, 30):
                row_vals = []
                for col_idx in range(1, 20):
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    row_vals.append((col_idx, str(val).strip().lower()) if val else (col_idx, ""))

                row_pm_col = next(
                    (col_idx for col_idx, val in row_vals if "pm budget" in val or "manager budget" in val),
                    None,
                )
                if row_pm_col is None:
                    continue

                # Found the real data-table header row. Scope the UW column
                # search to this same row only — an unrelated "UW" mention
                # elsewhere on the sheet (e.g. a property-info label like
                # "OPERATING ASSUMPTIONS - UW YR1") must never be mistaken
                # for the actual UW/underwriting column of this table.
                pm_col = row_pm_col
                uw_col = next(
                    (col_idx for col_idx, val in row_vals
                     if "uw" in val and "per unit" not in val and "variance" not in val),
                    None,
                )
                if uw_col is None:
                    # Some scorecards label their UW column "Year N" / "Yr N"
                    # (the underwritten target for that scorecard year)
                    # rather than spelling out "UW".
                    uw_col = next(
                        (col_idx for col_idx, val in row_vals
                         if re.search(r"\b(?:year|yr)\.?\s*\d+\b", val)
                         and "variance" not in val and "per unit" not in val),
                        None,
                    )
                break

            if uw_col:
                self.diagnostics["found_columns"]["UW"] = uw_col
            else:
                self.diagnostics["missing_columns"].append("UW")
                self.diagnostics["warnings"].append(
                    "Target parser: could not find a UW/underwriting column in the Scorecard "
                    "header row — UW targets will show as $0 and should not be trusted."
                )
            if pm_col:
                self.diagnostics["found_columns"]["PM"] = pm_col
            else:
                self.diagnostics["missing_columns"].append("PM Budget")

            if not uw_col and not pm_col:
                self.diagnostics["warnings"].append("Target parser: could not find UW or PM Budget columns.")
                return

            row_map = {}
            for row_idx in range(1, sheet.max_row + 1):
                val_a = sheet.cell(row=row_idx, column=1).value
                val_b = sheet.cell(row=row_idx, column=2).value
                label = (str(val_a) + " " + str(val_b)).lower()

                if "variance" in label:
                    continue
                if ("total operating income" in label or "total income" in label) and "Income" not in row_map:
                    row_map["Income"] = row_idx
                elif ("total operating expenses" in label or "total expenses" in label) and "Expenses" not in row_map:
                    row_map["Expenses"] = row_idx
                elif ("net operating income" in label or "noi" in label) and "NOI" not in row_map:
                    row_map["NOI"] = row_idx

            for metric in ("Income", "Expenses", "NOI"):
                if metric in row_map:
                    self.diagnostics["found_rows"].append({"metric": metric, "row": row_map[metric]})
                else:
                    self.diagnostics["missing_rows"].append(metric)

            def get_monthly(row_idx, col_idx):
                if row_idx and col_idx:
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        return float(val) / 12.0
                return 0.0

            if uw_col:
                self.targets["UW"]["Income"] = get_monthly(row_map.get("Income"), uw_col)
                self.targets["UW"]["Expenses"] = get_monthly(row_map.get("Expenses"), uw_col)
                self.targets["UW"]["NOI"] = get_monthly(row_map.get("NOI"), uw_col)

            if pm_col:
                self.targets["PM"]["Income"] = get_monthly(row_map.get("Income"), pm_col)
                self.targets["PM"]["Expenses"] = get_monthly(row_map.get("Expenses"), pm_col)
                self.targets["PM"]["NOI"] = get_monthly(row_map.get("NOI"), pm_col)
        except Exception as exc:
            self.diagnostics["warnings"].append(f"Error parsing Scorecard targets: {exc}")

    def get_data(self):
        return self.targets

    def get_diagnostics(self):
        return self.diagnostics
