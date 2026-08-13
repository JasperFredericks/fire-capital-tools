"""
FIRE Capital Tools - Underwriting scenario export.

A one-document summary of a scenario: property info, market context,
loans and debt service, the capex budget, the rent-roll/T12 cross-check,
and the headline returns.

NOT A NEW EXPORT MECHANISM

It reuses underwriting_pnl_export's page geometry, header, colours and
money formatter directly -- imported, not copied -- so this document and
the P&L look like the same document family, and a change to the header
moves both. Scorecard Pro and Site DD already share that geometry, which
makes this the fourth user of one pattern rather than a fourth pattern.

Nothing here computes a figure. Every number written comes from the
analyze_scenario() result and the panels the detail page already
rendered, so the export cannot disagree with the screen. Same discipline
as underwriting_pnl_export, stated there and kept here.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from tools.underwriting_pnl_export import (
    BODY, INK, MUTED, NEGATIVE, PAGE_SIZE, RULE, _header, _money,
)

# Vertical rhythm for a text page. Matches the P&L's row pitch so the two
# documents scan the same way.
TOP_Y = 0.78
LINE_H = 0.030
SECTION_GAP = 0.022


def _pct(v: Any, dp: int = 1) -> str:
    if v is None:
        return "—"
    return f"{v:,.{dp}f}%"


def _ratio(v: Any) -> str:
    return "—" if v is None else f"{v:,.2f}x"


def _fig(logo_path, subtitle, meta, title="Underwriting Summary"):
    fig = plt.figure(figsize=PAGE_SIZE)
    _header(fig, title, subtitle, meta, logo_path)
    return fig


def _section(fig, y: float, heading: str) -> float:
    fig.text(0.06, y, heading.upper(), fontsize=9.5, fontweight="bold", color=INK)
    fig.add_artist(plt.Line2D([0.06, 0.94], [y - 0.008, y - 0.008],
                              color=RULE, linewidth=0.6))
    return y - LINE_H


def _row(fig, y: float, label: str, value: str, *,
         indent: float = 0.0, bold: bool = False, negative: bool = False,
         note: str | None = None) -> float:
    fig.text(0.06 + indent, y, label, fontsize=8.5,
             fontweight="bold" if bold else "normal", color=BODY)
    fig.text(0.62, y, value, fontsize=8.5, ha="right",
             fontweight="bold" if bold else "normal",
             color=NEGATIVE if negative else BODY)
    if note:
        fig.text(0.64, y, note, fontsize=7.5, color=MUTED)
    return y - LINE_H


def _wrapped(fig, y: float, text: str, width: int = 118) -> float:
    """Naive wrap. matplotlib has no text flow, so long warning sentences
    are broken by character count rather than measured width -- adequate
    for a fixed font at a fixed size, and a wrong break is cosmetic."""
    import textwrap
    for line in textwrap.wrap(text, width):
        fig.text(0.06, y, line, fontsize=8, color=BODY)
        y -= LINE_H * 0.82
    return y


def build_pdf(path, data: dict[str, Any], logo_path: Path | None = None) -> Path:
    """Write the scenario summary PDF to `path` and return it."""
    path = Path(path)
    scenario = data["scenario"]
    result = data.get("result") or {}
    returns = result.get("returns") or {}
    prop = data.get("property_info") or {}
    market = data.get("market") or {}
    capex = result.get("capex") or {}
    crosscheck = data.get("crosscheck") or {}
    loans = data.get("loans") or []
    stack = result.get("debt_stack")

    subtitle = f"{scenario.get('property_label')} · {scenario.get('name')}"
    units = (prop.get("unit_count") or {}).get("value")
    meta = (f"{units:,.0f} units · " if units else "") + \
           f"{scenario.get('hold_years') or 0}-year hold"

    with PdfPages(str(path)) as pdf:
        # ── Page 1: property, market, returns ────────────────────────────
        fig = _fig(logo_path, subtitle, meta)
        y = TOP_Y

        y = _section(fig, y, "Property")
        uc = prop.get("unit_count") or {}
        oc = prop.get("occupancy") or {}
        y = _row(fig, y, "Units", f"{uc.get('value'):,.0f}" if uc.get("value") is not None else "—",
                 note=_source_note(uc))
        y = _row(fig, y, "Occupancy", _pct(oc.get("value")), note=_source_note(oc))
        y = _row(fig, y, "Parking spaces",
                 f"{prop.get('parking_spaces'):,.0f}" if prop.get("parking_spaces") is not None else "—",
                 note=prop.get("parking_notes") or None)
        y = _row(fig, y, "Market",
                 f"{prop.get('city')}, {prop.get('state')}" if prop.get("city") else "—")
        y -= SECTION_GAP

        y = _section(fig, y, "Market context")
        if market.get("available"):
            for m in market.get("metrics", []):
                y = _row(fig, y, m["label"], _market_value(m),
                         note=m.get("rating") or None)
        else:
            y = _wrapped(fig, y, market.get("reason") or
                         "No market context available for this scenario.")
        y -= SECTION_GAP

        y = _section(fig, y, "Returns")
        y = _row(fig, y, "Purchase price", _money(scenario.get("purchase_price")))
        y = _row(fig, y, "Cash to close (equity invested)",
                 _money(returns.get("equity_invested")), bold=True)
        y = _row(fig, y, "Going-in cap rate", _pct(_x100(returns.get("going_in_cap_rate")), 2))
        y = _row(fig, y, "Cash-on-cash (Yr 1)", _pct(_x100(returns.get("cash_on_cash")), 2))
        y = _row(fig, y, "DSCR (Yr 1)", _ratio(returns.get("dscr")),
                 note=returns.get("dscr_reason") or None)
        y = _row(fig, y, "Levered IRR", _pct(_x100(returns.get("levered_irr")), 2),
                 note=returns.get("levered_irr_reason") or None)
        y = _row(fig, y, "Equity multiple", _ratio(returns.get("equity_multiple")))
        _footer(fig, 1)
        pdf.savefig(fig)
        plt.close(fig)

        # ── Page 2: debt, capex, cross-check ─────────────────────────────
        fig = _fig(logo_path, subtitle, meta)
        y = TOP_Y

        y = _section(fig, y, "Debt")
        y = _row(fig, y, "Loan amount", _money(returns.get("loan_amount")), bold=True)
        y = _row(fig, y, "Annual debt service", _money(returns.get("annual_debt_service")), bold=True)
        y = _row(fig, y, "Monthly debt service", _money(returns.get("monthly_debt_service")))
        if stack:
            y = _row(fig, y, "Implied LTV", _pct(stack.get("implied_ltv_pct"), 2))
            for ln in stack.get("loans", []):
                y = _row(fig, y,
                         f"{ln.get('name')} — {_pct(ln.get('rate_pct'), 2)}, "
                         f"{ln.get('amort_years')}yr",
                         _money(ln.get("amount")), indent=0.02,
                         note=f"DSCR {_ratio(ln.get('dscr'))}")
        elif loans:
            y = _row(fig, y, f"{len(loans)} loans on file", "")
        else:
            y = _row(fig, y, "Single loan sized from LTV", _pct(scenario.get("ltv_pct"), 2))
        y -= SECTION_GAP

        y = _section(fig, y, "Capex budget")
        if capex.get("has_lines"):
            for line in capex.get("lines", []):
                if line.get("is_contingency"):
                    continue
                scope = (line.get("scope") or "").title()
                y = _row(fig, y, f"{scope} — {line.get('label')}",
                         _money(_line_total(line)), indent=0.02,
                         note=(line.get("category") or None))
            y = _row(fig, y, "Itemized subtotal", _money(capex.get("itemized_total")), bold=True)
            y = _row(fig, y, f"Contingency ({_pct(capex.get('contingency_pct'))})",
                     _money(capex.get("contingency_total")))
            y = _row(fig, y, "Total capex", _money(capex.get("total")), bold=True)
            y = _row(fig, y, "Per unit",
                     _money(capex.get("per_unit")) if capex.get("per_unit") is not None else "—")
        else:
            y = _wrapped(fig, y, "No capex budget entered for this scenario.")
        y -= SECTION_GAP

        y = _section(fig, y, "Rent roll vs T12")
        if not crosscheck.get("available"):
            y = _wrapped(fig, y, crosscheck.get("reason") or "Not available.")
        else:
            for c in crosscheck.get("checks", []):
                y = _row(fig, y, c["label"], c["summary"],
                         negative=bool(c.get("fires")),
                         note=(f"{c['gap_pct']:+,.1f}%" if c.get("gap_pct") is not None else None))
            for c in crosscheck.get("firing", []):
                y -= LINE_H * 0.4
                y = _wrapped(fig, y, c["message"])
        _footer(fig, 2)
        pdf.savefig(fig)
        plt.close(fig)

    return path


def _line_total(line):
    total = line.get("total_cost")
    if total not in (None, ""):
        return total
    q, u = line.get("quantity"), line.get("unit_cost")
    if q is not None and u is not None:
        return q * u
    return 0.0


def _x100(v):
    return None if v is None else v * 100.0


def _market_value(m):
    v = m.get("value")
    if v is None:
        return "—"
    kind = m.get("kind")
    if kind == "money":
        return _money(v)
    if kind == "count":
        return f"{v:,.0f}"
    if kind == "pct":
        return f"{v * 100:,.2f}%"
    return f"{v:,.1f}"


def _source_note(field):
    src = (field or {}).get("source")
    return {
        "derived": "from the rent roll",
        "override_agrees": "entered, matches the rent roll",
        "override_disagrees": "entered — differs from the rent roll",
    }.get(src)


def _footer(fig, page_no: int) -> None:
    fig.text(0.06, 0.05,
             "Figures come from the same calculation the Underwriting page shows.",
             fontsize=7.5, color=MUTED)
    fig.text(0.94, 0.05, f"Page {page_no}", ha="right", fontsize=7.5, color=MUTED)


# ── Excel ────────────────────────────────────────────────────────────────

def build_xlsx(path, data: dict[str, Any]) -> Path:
    """Write the scenario summary workbook to `path` and return it.

    Numbers are written as numbers, not preformatted strings, so the sheet
    stays usable as a model rather than a picture of one -- the same rule
    underwriting_pnl_export states for the P&L workbook.
    """
    path = Path(path)
    scenario = data["scenario"]
    result = data.get("result") or {}
    returns = result.get("returns") or {}
    prop = data.get("property_info") or {}
    market = data.get("market") or {}
    capex = result.get("capex") or {}
    crosscheck = data.get("crosscheck") or {}

    wb = openpyxl.Workbook()
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1A2744")
    bold = Font(bold=True)

    def sheet(title, headers):
        ws = wb.create_sheet(title)
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i, value=h)
            c.font, c.fill = head, fill
            c.alignment = Alignment(horizontal="left")
            ws.column_dimensions[get_column_letter(i)].width = 34 if i == 1 else 20
        return ws

    wb.remove(wb.active)

    ws = sheet("Summary", ["Item", "Value", "Note"])
    rows = [
        ("Property", scenario.get("property_label"), scenario.get("name")),
        ("City", prop.get("city"), prop.get("state")),
        ("Units", (prop.get("unit_count") or {}).get("value"), _source_note(prop.get("unit_count"))),
        ("Occupancy %", (prop.get("occupancy") or {}).get("value"), _source_note(prop.get("occupancy"))),
        ("Parking spaces", prop.get("parking_spaces"), prop.get("parking_notes")),
        ("Purchase price", scenario.get("purchase_price"), None),
        ("Cash to close", returns.get("equity_invested"), "price - loan + costs + capex"),
        ("Loan amount", returns.get("loan_amount"), None),
        ("Annual debt service", returns.get("annual_debt_service"), None),
        ("Going-in cap rate", returns.get("going_in_cap_rate"), None),
        ("Cash-on-cash (Yr 1)", returns.get("cash_on_cash"), None),
        ("DSCR (Yr 1)", returns.get("dscr"), returns.get("dscr_reason")),
        ("Levered IRR", returns.get("levered_irr"), returns.get("levered_irr_reason")),
        ("Equity multiple", returns.get("equity_multiple"), None),
    ]
    for r, (a, b, c) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=a).font = bold
        ws.cell(row=r, column=2, value=b)
        ws.cell(row=r, column=3, value=c)

    ws = sheet("Market", ["Metric", "Value", "Rating"])
    if market.get("available"):
        for r, m in enumerate(market.get("metrics", []), start=2):
            ws.cell(row=r, column=1, value=m["label"])
            ws.cell(row=r, column=2, value=m.get("value"))
            ws.cell(row=r, column=3, value=m.get("rating"))
    else:
        ws.cell(row=2, column=1, value=market.get("reason"))

    ws = sheet("Capex", ["Item", "Scope", "Category", "Qty", "Unit cost", "Total", "Source"])
    r = 2
    for line in capex.get("lines", []):
        ws.cell(row=r, column=1, value=line.get("label"))
        ws.cell(row=r, column=2, value=line.get("scope"))
        ws.cell(row=r, column=3, value=line.get("category"))
        ws.cell(row=r, column=4, value=line.get("quantity"))
        ws.cell(row=r, column=5, value=line.get("unit_cost"))
        ws.cell(row=r, column=6, value=_line_total(line))
        ws.cell(row=r, column=7, value=line.get("source"))
        r += 1
    for label, value in (("Itemized subtotal", capex.get("itemized_total")),
                         (f"Contingency ({capex.get('contingency_pct')}%)",
                          capex.get("contingency_total")),
                         ("Total capex", capex.get("total")),
                         ("Per unit", capex.get("per_unit"))):
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=6, value=value).font = bold
        r += 1

    ws = sheet("Cross-check", ["Comparison", "Model", "T12", "Difference %", "Status", "Message"])
    if crosscheck.get("available"):
        for r, c in enumerate(crosscheck.get("checks", []), start=2):
            ws.cell(row=r, column=1, value=c["label"])
            ws.cell(row=r, column=2, value=c["model"])
            ws.cell(row=r, column=3, value=c["actual"])
            ws.cell(row=r, column=4, value=c.get("gap_pct"))
            ws.cell(row=r, column=5, value="worth a look" if c.get("fires") else "within tolerance")
            ws.cell(row=r, column=6, value=c.get("message") if c.get("fires") else None)
    else:
        ws.cell(row=2, column=1, value=crosscheck.get("reason"))

    wb.save(str(path))
    return path


def export_filename(scenario: dict[str, Any], ext: str) -> str:
    """Filesystem-safe download name. Same shape as the P&L export's."""
    label = "".join(ch if (ch.isalnum() or ch in " -_") else "_"
                    for ch in (scenario.get("property_label") or "scenario")).strip()
    label = "_".join(label.split()) or "scenario"
    return f"{label}_Underwriting_Summary.{ext}"
