"""
PDF and Excel export of the pro-forma P&L.

Same mechanism as Scorecard Pro and Site DD -- matplotlib PdfPages, one
figure per page, table text drawn directly onto the figure -- so the three
exports stay visually and structurally consistent.

Following site_dd_report rather than scorecard_pro.exports for the logo:
Site DD takes `logo_path` as an argument, keeping this module free of
Flask imports (scorecard_pro.exports reaches for current_app). The route
resolves the path via tools.branding and passes it down, which is the
split used everywhere else -- pure module, Flask-aware caller.

Nothing here computes a figure. Every number written comes from the
already-reconciled dict that underwriting_pnl.build_pnl() returned.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")
import matplotlib.image as mpimg
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PAGE_SIZE = (11, 8.5)          # landscape, same as Site DD
INK = "#1a2744"
MUTED = "#6b7280"
BODY = "#111827"
RULE = "#d1d5db"
NEGATIVE = "#b91c1c"

# Rows of table body that fit on one page under the header block.
ROWS_PER_PAGE = 28
LABEL_TRUNCATE_AT = 38


def _money(v: Any) -> str:
    """Accounting-style: thousands separated, negatives parenthesised."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return "—"
    if abs(f) < 0.005:
        f = 0.0
    return f"({abs(f):,.0f})" if f < 0 else f"{f:,.0f}"


def _truncate(text: str, limit: int = LABEL_TRUNCATE_AT) -> str:
    text = " ".join(str(text or "").split())
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "…"


def _header(fig, title: str, subtitle: str, meta: str, logo_path: Path | None) -> None:
    """Same geometry as site_dd_report._header / scorecard_pro's
    add_pdf_header, reimplemented for the same reason Site DD did: the
    Scorecard version hard-codes its own title string and needs Flask."""
    if logo_path and Path(logo_path).exists():
        try:
            ax = fig.add_axes([0.06, 0.86, 0.24, 0.08])
            ax.imshow(mpimg.imread(str(logo_path)))
            ax.axis("off")
        except Exception:
            pass  # a missing/unreadable logo must never fail the export
    fig.text(0.94, 0.91, title, ha="right", fontsize=14, fontweight="bold", color=INK)
    fig.text(0.94, 0.875, subtitle, ha="right", fontsize=10, color="#4b5563")
    fig.text(0.94, 0.845, meta, ha="right", fontsize=9, color=MUTED)


# ── Row model shared by both exports ─────────────────────────────────────

def flatten_rows(pnl: dict[str, Any]) -> list[dict[str, Any]]:
    """The P&L as ordered display rows.

    Built once and used by both the PDF and the workbook so the two
    exports can never show a different statement.

    kind: section | line | subtotal | total | spacer
    """
    rows: list[dict[str, Any]] = []
    n = len(pnl["years"])

    rows.append({"kind": "section", "label": "REVENUE", "amounts": None})
    for r in pnl["revenue"]:
        rows.append({"kind": "line", "label": r["label"], "amounts": r["amounts"]})
    rows.append({"kind": "total", "label": "Effective Gross Income",
                 "amounts": pnl["revenue_totals"]})
    rows.append({"kind": "spacer", "label": "", "amounts": None})

    rows.append({"kind": "section", "label": "OPERATING EXPENSES", "amounts": None})
    for g in pnl["expenses"]:
        rows.append({"kind": "subheading", "label": g["category"], "amounts": None})
        for l in g["lines"]:
            label = f"{l['gl_code']}  {l['label']}" if l["gl_code"] else l["label"]
            rows.append({"kind": "line", "label": label, "amounts": l["amounts"]})
        rows.append({"kind": "subtotal", "label": f"Total {g['category']}",
                     "amounts": g["subtotals"]})
    rows.append({"kind": "total", "label": "Total Operating Expenses",
                 "amounts": pnl["expense_totals"]})
    rows.append({"kind": "spacer", "label": "", "amounts": None})

    rows.append({"kind": "total", "label": "NET OPERATING INCOME",
                 "amounts": pnl["noi"]})
    rows.append({"kind": "line", "label": "Operating Margin",
                 "amounts": [None if m is None else m for m in pnl["margin"]],
                 "is_pct": True})

    if pnl["excluded"]:
        rows.append({"kind": "spacer", "label": "", "amounts": None})
        rows.append({"kind": "section",
                     "label": "EXCLUDED FROM NOI (shown, not counted)",
                     "amounts": None})
        for x in pnl["excluded"]:
            label = f"{x['gl_code']}  {x['label']}" if x["gl_code"] else x["label"]
            rows.append({"kind": "line", "label": label,
                         "amounts": [x["annual_amount"]] + [None] * (n - 1),
                         "muted": True})
    return rows


# ── PDF ──────────────────────────────────────────────────────────────────

def _paginate(rows: list[dict[str, Any]],
              rows_per_page: int = ROWS_PER_PAGE) -> list[list[dict[str, Any]]]:
    """Split rows into pages of roughly equal length.

    Naive chunking leaves whatever is left over on the final page, which on
    a 109-line expense set produced a page carrying one orphaned row under
    a full header. Choosing the page count first and then dividing evenly
    keeps the last page substantial, at no cost to the others.

    A leading spacer is dropped from each page after the first: a blank row
    under a page header reads as a rendering fault rather than as spacing.
    """
    if not rows:
        return [[]]
    n_pages = max(1, -(-len(rows) // rows_per_page))     # ceil
    per_page = -(-len(rows) // n_pages)                  # ceil, evenly spread
    pages = [rows[i:i + per_page] for i in range(0, len(rows), per_page)]
    return [p[1:] if (idx and p and p[0]["kind"] == "spacer") else p
            for idx, p in enumerate(pages)]


def build_pdf(path, pnl: dict[str, Any], logo_path: Path | None = None) -> Path:
    """Write the P&L PDF to `path` and return it."""
    path = Path(path)
    years = pnl["years"]
    n = len(years)
    rows = flatten_rows(pnl)

    subtitle = f"{pnl['property_label']} · {pnl['scenario_name']}"
    meta = (f"{pnl['unit_count']} units · {pnl['hold_years']}-year pro forma · "
            f"rent growth {pnl['rent_growth_pct'] or 0:g}% · "
            f"expense growth {pnl['expense_growth_pct'] or 0:g}%")

    # Column geometry: label column then one column per year.
    x_label = 0.06
    x_first = 0.34
    col_w = (0.94 - x_first) / max(1, n)

    pages = _paginate(rows)

    with PdfPages(str(path)) as pdf:
        for page_idx, page_rows in enumerate(pages):
            fig = plt.figure(figsize=PAGE_SIZE)
            _header(fig, "Pro-Forma Profit & Loss", subtitle, meta, logo_path)

            y = 0.78
            fig.text(x_label, y, "", fontsize=9)
            for i, yr in enumerate(years):
                fig.text(x_first + i * col_w + col_w - 0.01, y, f"Year {yr}",
                         ha="right", fontsize=9, fontweight="bold", color=INK)
            y -= 0.018
            fig.add_artist(plt.Line2D([x_label, 0.94], [y, y], color=RULE, lw=0.8))
            y -= 0.022

            for row in page_rows:
                kind = row["kind"]
                if kind == "spacer":
                    y -= 0.014
                    continue

                if kind == "section":
                    fig.text(x_label, y, row["label"], fontsize=9.5,
                             fontweight="bold", color=INK)
                elif kind == "subheading":
                    fig.text(x_label + 0.008, y, _truncate(row["label"]), fontsize=8.5,
                             fontweight="bold", color="#374151")
                elif kind in ("subtotal", "total"):
                    fig.text(x_label + (0.008 if kind == "subtotal" else 0), y,
                             _truncate(row["label"]), fontsize=8.5,
                             fontweight="bold", color=INK)
                else:
                    fig.text(x_label + 0.02, y, _truncate(row["label"]), fontsize=8,
                             color=MUTED if row.get("muted") else BODY)

                if row["amounts"]:
                    bold = kind in ("subtotal", "total")
                    for i in range(n):
                        val = row["amounts"][i] if i < len(row["amounts"]) else None
                        if val is None:
                            text = ""
                        elif row.get("is_pct"):
                            text = f"{val * 100:.1f}%"
                        else:
                            text = _money(val)
                        colour = BODY
                        if not row.get("is_pct") and isinstance(val, (int, float)) and val < 0:
                            colour = NEGATIVE
                        if kind in ("subtotal", "total"):
                            colour = INK
                        if row.get("muted"):
                            colour = MUTED
                        fig.text(x_first + i * col_w + col_w - 0.01, y, text,
                                 ha="right", fontsize=8,
                                 fontweight="bold" if bold else "normal", color=colour)

                if kind == "total":
                    fig.add_artist(plt.Line2D([x_label, 0.94], [y - 0.008, y - 0.008],
                                              color=RULE, lw=0.8))
                y -= 0.0235

            fig.text(x_label, 0.04,
                     "Pro-forma projection — not trailing actuals. "
                     "Figures are this scenario's underwriting model.",
                     fontsize=7.5, color=MUTED)
            fig.text(0.94, 0.04, f"Page {page_idx + 1} of {len(pages)}",
                     ha="right", fontsize=7.5, color=MUTED)
            pdf.savefig(fig)
            plt.close(fig)

    return path


# ── Excel ────────────────────────────────────────────────────────────────

def build_xlsx(path, pnl: dict[str, Any]) -> Path:
    """Write the P&L workbook to `path` and return it.

    Numbers are written as numbers, not preformatted strings, so the sheet
    stays usable as a model rather than a picture of one.
    """
    path = Path(path)
    years = pnl["years"]
    rows = flatten_rows(pnl)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pro-Forma P&L"

    head_fill = PatternFill(start_color="1A2744", end_color="1A2744", fill_type="solid")
    head_font = Font(color="FFFFFF", bold=True)
    money_fmt = '#,##0;(#,##0)'

    ws.cell(row=1, column=1, value="Pro-Forma Profit & Loss").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{pnl['property_label']} · {pnl['scenario_name']}")
    ws.cell(row=3, column=1,
            value=(f"{pnl['unit_count']} units · {pnl['hold_years']}-year pro forma · "
                   f"rent growth {pnl['rent_growth_pct'] or 0:g}% · "
                   f"expense growth {pnl['expense_growth_pct'] or 0:g}%"))
    ws.cell(row=4, column=1,
            value="Pro-forma projection — not trailing actuals.").font = Font(italic=True)

    header_row = 6
    c = ws.cell(row=header_row, column=1, value="")
    c.fill, c.font = head_fill, head_font
    for i, yr in enumerate(years):
        c = ws.cell(row=header_row, column=2 + i, value=f"Year {yr}")
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(horizontal="right")

    r_idx = header_row + 1
    for row in rows:
        if row["kind"] == "spacer":
            r_idx += 1
            continue
        bold = row["kind"] in ("section", "subheading", "subtotal", "total")
        indent = {"section": 0, "subheading": 1, "subtotal": 1, "total": 0}.get(row["kind"], 2)
        c = ws.cell(row=r_idx, column=1, value=row["label"])
        c.font = Font(bold=bold)
        c.alignment = Alignment(indent=indent)

        if row["amounts"]:
            for i in range(len(years)):
                val = row["amounts"][i] if i < len(row["amounts"]) else None
                if val is None:
                    r_idx_cell = ws.cell(row=r_idx, column=2 + i, value=None)
                else:
                    r_idx_cell = ws.cell(row=r_idx, column=2 + i, value=float(val))
                    r_idx_cell.number_format = '0.0%' if row.get("is_pct") else money_fmt
                r_idx_cell.font = Font(bold=bold)
        r_idx += 1

    ws.column_dimensions["A"].width = 46
    for i in range(len(years)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 16
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)

    # A second sheet carrying the reconciliation, so the workbook itself
    # evidences that the statement ties to the scenario it came from.
    ws2 = wb.create_sheet("Reconciliation")
    hdr = ["Check", "Year", "P&L", "Scenario projection", "Difference", "Passed"]
    for i, h in enumerate(hdr, 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.fill, c.font = head_fill, head_font
    for j, chk in enumerate(pnl.get("reconciliation") or [], start=2):
        ws2.cell(row=j, column=1, value=chk["name"])
        ws2.cell(row=j, column=2, value=chk["year"])
        ws2.cell(row=j, column=3, value=float(chk["got"])).number_format = money_fmt
        ws2.cell(row=j, column=4, value=float(chk["want"])).number_format = money_fmt
        ws2.cell(row=j, column=5, value=float(chk["diff"]))
        ws2.cell(row=j, column=6, value="PASS" if chk["passed"] else "FAIL")
    for col, w in zip("ABCDEF", (30, 8, 18, 22, 14, 10)):
        ws2.column_dimensions[col].width = w

    wb.save(str(path))
    return path


def export_filename(pnl: dict[str, Any], ext: str) -> str:
    """Filesystem-safe download name."""
    label = "".join(ch if (ch.isalnum() or ch in " -_") else "_"
                    for ch in (pnl.get("property_label") or "scenario")).strip()
    label = "_".join(label.split()) or "scenario"
    return f"{label}_ProForma_PL.{ext}"
