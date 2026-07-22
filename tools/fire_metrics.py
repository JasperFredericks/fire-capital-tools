"""
FIRE Capital Tools — FIRE Metric search dashboard.

Search UI/frontend originally built by Beckett (BeckettTest branch, commit
04d3b47) on top of a JSON-file index and a validate-and-copy scaffold.
This version keeps his templates/city_search.py matching logic/dashboard
card layout, but the backend underneath is now SQLite (fire_metrics_updater
/db.py) fed by the real, working pipeline scripts (fire_metrics/scripts/,
fixed in JasperTest Priorities 1-4) via a real orchestration function
(fire_metrics_updater/orchestrator.py) instead of a scaffold that only
copied the uploaded file.

Changes from Beckett's version, and why (see final report for full detail):
- Upload-a-workbook-and-process-it is removed. The live pipeline no longer
  needs a user-uploaded workbook to gather fresh data (population/income/
  home-value/employment/climate all come from live APIs); there's no
  coherent thing left for an arbitrary uploaded workbook to do.
- Format Only / Dry Run options are removed along with it -- both were
  specific to the old single-workbook run_update() scaffold and don't map
  to the new per-metric-family orchestration.
- Refresh All Data now runs the real orchestrator in a separate OS process
  (fire_metrics_updater/refresh_worker.py, launched via subprocess.Popen),
  with status tracked in the refresh_metadata table rather than an
  in-process thread/lock -- a prior in-process-thread version of this
  caused 502s in production, since the CPU-heavy climate-risk step shares
  this process's GIL with (and starves) the request that triggered it.
  (Beckett's version called it synchronously, guarded only by a
  re-entrancy lock -- not actually a background job.)
- Rebuild Search Index is repurposed: it now re-ingests from whatever
  pipeline output files already exist on disk, without calling any live
  API -- useful after running a script by hand from the CLI.
- Download Latest Workbook now exports the current database state to a
  fresh .xlsx on demand, since there's no longer one static "latest
  workbook" file -- SQLite is the source of truth.
"""

from __future__ import annotations

import io
import json
import os
import re
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from flask import Blueprint, current_app, jsonify, render_template, request
from flask_login import login_required

from fire_metrics.fire_metrics_updater import db as db_module
from fire_metrics.fire_metrics_updater.city_search import find_city_match
from tools import fire_metrics_ai_summary as ai_summary

fire_metrics_bp = Blueprint("fire_metrics", __name__)

REPO_ROOT = Path(__file__).resolve().parent.parent

# A refresh in the "running" state for longer than this is treated as
# crashed/stuck rather than genuinely in progress, so one dead subprocess
# (e.g. killed by an OOM or a Railway restart) can't permanently block
# every future refresh. Generous relative to the real chain (climate risk
# alone is documented elsewhere in this file as "several minutes" on a
# cold cache) but still bounded.
REFRESH_STALE_AFTER_SECONDS = 60 * 60


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _is_refresh_running(metadata: dict) -> bool:
    """Derive "is a refresh actually running" from persisted state alone --
    no in-process thread/lock object, since the real work now happens in a
    separate OS process (possibly started by a different web request than
    the one asking) and status must be readable regardless of which
    process/request is checking.
    """
    if metadata.get("refresh_running") != "1":
        return False
    started_at = metadata.get("refresh_started_at")
    if not started_at:
        return False
    try:
        started = datetime.fromisoformat(started_at)
    except ValueError:
        return False
    age_seconds = (datetime.now(timezone.utc) - started).total_seconds()
    return age_seconds < REFRESH_STALE_AFTER_SECONDS


def _refresh_status() -> dict:
    with db_module.get_connection() as conn:
        metadata = db_module.get_metadata(conn)
        total_cities = conn.execute("SELECT COUNT(*) FROM cities").fetchone()[0]

    running = _is_refresh_running(metadata)
    status = "running" if running else metadata.get("last_refresh_status", "missing" if total_cities == 0 else "current")

    return {
        "status": status,
        "running": running,
        "last_refresh_at": metadata.get("last_refresh_at"),
        "last_refresh_error": metadata.get("last_refresh_error"),
        "city_count": total_cities,
    }


def _summary_enabled() -> bool:
    return bool(current_app.config.get("FIRE_METRICS_AI_SUMMARIES_ENABLED", False))


def _summary_model_name() -> str:
    return str(current_app.config.get("FIRE_METRICS_SUMMARY_MODEL") or "").strip()


def _summary_api_key() -> str:
    return str(current_app.config.get("OPENAI_API_KEY") or "").strip()


def _summary_unavailable_response(
    *,
    selected_city: dict[str, Any] | None,
    benchmark_data: dict[str, Any] | None,
    reason: str,
    data_refreshed_at: str | None = None,
):
    if selected_city and benchmark_data:
        structured = ai_summary.fallback_summary(selected_city, benchmark_data)
        combined = ai_summary.combined_summary(structured)
        return {
            "status": "ready",
            "summary": combined,
            "summary_structured": structured,
            "generated_at": ai_summary.utc_now_iso(),
            "data_refreshed_at": data_refreshed_at,
            "cached": False,
            "city_key": ai_summary.city_key(selected_city),
            "relative_market_profile_score": benchmark_data.get("relative_market_profile_score"),
            "relative_market_profile_percentile": benchmark_data.get("relative_market_profile_percentile"),
            "tracked_city_relative_market_profile_average": benchmark_data.get("tracked_city_relative_market_profile_average"),
            "recommendation_category": benchmark_data.get("recommendation_category"),
            "score": benchmark_data.get("selected_overall_score"),
            "computed_composite_score": benchmark_data.get("selected_overall_score"),
            "tracked_city_average": benchmark_data.get("tracked_city_average"),
            "tracked_city_composite_average": benchmark_data.get("tracked_city_average"),
            "tracked_city_count": benchmark_data.get("tracked_city_count"),
            "percentile": benchmark_data.get("selected_percentile"),
            "source": "fallback",
            "note": reason,
        }
    return {
        "status": "unavailable",
        "summary": "AI market overview is unavailable for this city.",
        "summary_structured": {
            "strength_sentence": "The strongest currently available signals are limited by missing values.",
            "weakness_sentence": "The largest currently available risks are limited by missing values.",
            "comparison_sentence": "The computed FIRE Metrics composite score assessment is limited because too many component values are missing.",
        },
        "generated_at": ai_summary.utc_now_iso(),
        "data_refreshed_at": data_refreshed_at,
        "cached": False,
        "city_key": ai_summary.city_key(selected_city) if selected_city else None,
        "relative_market_profile_score": None,
        "relative_market_profile_percentile": None,
        "tracked_city_relative_market_profile_average": None,
        "recommendation_category": None,
        "score": None,
        "computed_composite_score": None,
        "tracked_city_average": None,
        "tracked_city_composite_average": None,
        "tracked_city_count": 0,
        "percentile": None,
        "source": "fallback",
        "note": reason,
    }


_SCRIPTS_DIR = REPO_ROOT / "fire_metrics" / "scripts"

# The real FBI Table 8 workbook is a few MB; this is generous headroom
# while still catching an obviously-wrong file quickly with a clear
# message. Flask's own global MAX_CONTENT_LENGTH (20 MB, see config.py) is
# a hard backstop above this for the whole app, independent of this check.
MAX_CRIME_WORKBOOK_BYTES = 10 * 1024 * 1024

# Matches add_crime_index.load_fbi_table_8's header=3 (0-indexed) -- the
# workbook has a few title rows before the real header.
CRIME_WORKBOOK_HEADER_ROW = 4
CRIME_WORKBOOK_REQUIRED_COLUMNS = {"state", "city", "population", "violent crime", "property crime"}


def _normalize_crime_workbook_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _get_crime_workbook_path() -> Path:
    """fire_metrics/scripts isn't a real package -- it's a directory of
    standalone scripts imported via sys.path insertion, the same way
    orchestrator.py does it. Mirrored here (rather than importing
    orchestrator itself) since this web process deliberately doesn't
    import the pipeline/orchestration modules otherwise -- only the
    refresh subprocess does; see _start_refresh()'s docstring for why.
    """
    if str(_SCRIPTS_DIR) not in sys.path:
        sys.path.insert(0, str(_SCRIPTS_DIR))
    from add_crime_index import get_fbi_crime_workbook_path
    return get_fbi_crime_workbook_path()


def _crime_workbook_status() -> dict:
    path = _get_crime_workbook_path()
    if not path.exists():
        return {"exists": False, "uploaded_at": None, "path": str(path)}
    uploaded_at = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()
    return {"exists": True, "uploaded_at": uploaded_at, "path": str(path)}


def _validate_crime_workbook_bytes(data: bytes) -> str | None:
    """Return an error message if `data` doesn't look like a real FBI
    Table 8 workbook, or None if it looks valid enough to save.

    This is a structural sanity check, not a re-implementation of the
    crime pipeline's own matching/scoring logic -- just enough to reject
    an obviously-wrong file (wrong format, wrong sheet layout, wrong
    columns) with a clear reason up front, instead of silently accepting
    it and having the actual pipeline run fail confusingly later, or
    worse, "succeed" on garbage data.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        return f"Could not read this file as an Excel workbook: {exc}"

    try:
        ws = wb.worksheets[0]
        header_cells = next(
            ws.iter_rows(min_row=CRIME_WORKBOOK_HEADER_ROW, max_row=CRIME_WORKBOOK_HEADER_ROW, values_only=True),
            None,
        )
        if header_cells is None:
            return (
                f"This workbook doesn't have a row {CRIME_WORKBOOK_HEADER_ROW} -- the FBI "
                f"Table 8 workbook has a few title rows before its real header row, which "
                f"is expected there."
            )

        normalized = {_normalize_crime_workbook_header(cell) for cell in header_cells if cell is not None}
        missing = CRIME_WORKBOOK_REQUIRED_COLUMNS - normalized
        if missing:
            found_preview = ", ".join(sorted(normalized)[:15]) or "(no column headers found)"
            return (
                f"This doesn't look like an FBI Table 8 workbook. Expected a header row at "
                f"row {CRIME_WORKBOOK_HEADER_ROW} including columns "
                f"{sorted(CRIME_WORKBOOK_REQUIRED_COLUMNS)}, but couldn't find: {sorted(missing)}. "
                f"Found instead: {found_preview}"
            )
    finally:
        wb.close()

    return None


def _start_refresh(skip_climate: bool = False, skip_crime: bool = True) -> bool:
    """Start the refresh as a real, separate OS process (fire_metrics/
    fire_metrics_updater/refresh_worker.py) -- not a thread. Threads share
    this process's GIL, so the CPU-heavy climate-risk step (geopandas/GDAL
    processing) was starving this same process's ability to answer the
    request that triggered it, until Railway's proxy gave up and returned
    a 502 -- confirmed empirically, and confirmed NOT fixed by
    threaded=True (that only helps connection-accept concurrency, not
    GIL/CPU contention). A real subprocess has its own GIL.

    Returns False if a refresh is already running and not stale (the
    caller should show that as a message, not start a second overlapping
    one). There's a small theoretical check-then-write race if two
    requests hit this within microseconds of each other -- acceptable for
    an admin button a human clicks, not worth the extra complexity of a
    manual SQLite write-lock for.
    """
    with db_module.get_connection() as conn:
        metadata = db_module.get_metadata(conn)
        if _is_refresh_running(metadata):
            return False

        args = [sys.executable, "-m", "fire_metrics.fire_metrics_updater.refresh_worker"]
        if skip_climate:
            args.append("--skip-climate")
        if skip_crime:
            args.append("--skip-crime")

        proc = subprocess.Popen(
            args,
            cwd=str(REPO_ROOT),
            env=os.environ.copy(),
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )

        # Written here (the parent), not by the worker at its own startup:
        # this must be visible to a second request arriving microseconds
        # later, before the subprocess has even finished interpreter
        # startup and its own imports -- Popen() already returns proc.pid
        # synchronously, so there's no need to wait on the child to report
        # it back.
        db_module.set_metadata(
            conn,
            refresh_running="1",
            refresh_started_at=_utc_now(),
            refresh_pid=str(proc.pid),
        )
    return True


def _reingest_from_disk() -> dict:
    """Re-ingest whatever pipeline output files are already on disk, with
    no live API calls -- for picking up a script that was run by hand.
    """
    from fire_metrics.fire_metrics_updater import index_builder, orchestrator as orch

    results = {}
    with db_module.get_connection() as conn:
        if orch.POP_LANDLORD_FILE.exists():
            results["population"] = index_builder.ingest_population_and_landlord(orch.POP_LANDLORD_FILE, conn)
            results["income"] = index_builder.ingest_income(orch.POP_LANDLORD_FILE, conn)
        if orch.HOME_VALUE_FILE.exists():
            results["home_value"] = index_builder.ingest_home_value(orch.HOME_VALUE_FILE, conn)
        if orch.JOB_GROWTH_FILE.exists():
            results["employment"] = index_builder.ingest_employment(orch.JOB_GROWTH_FILE, conn)
        if orch.CLIMATE_RISK_FILE.exists():
            results["climate"] = index_builder.ingest_climate_risk(orch.CLIMATE_RISK_FILE, conn)
        if orch.CRIME_FINAL_FILE.exists():
            results["crime"] = index_builder.ingest_crime(orch.CRIME_FINAL_FILE, conn)
        results["coordinates"] = index_builder.backfill_city_coordinates(conn)
    return results


def _export_workbook() -> bytes:
    import openpyxl

    with db_module.get_connection() as conn:
        cities = db_module.fetch_all_cities(conn)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "City Metrics"
    if cities:
        headers = [k for k in cities[0].keys() if k != "search_keys"]
        ws.append(headers)
        for city in cities:
            ws.append([city.get(h) for h in headers])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@fire_metrics_bp.route("/", methods=["GET", "POST"])
@login_required
def index():
    # Computed first (depends only on request headers, can't itself raise)
    # so the outermost except below always knows whether this caller's own
    # JS is going to do res.json() unconditionally on whatever comes back.
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def safe_refresh_status() -> dict:
        # _refresh_status() opens its own SQLite connection -- and, right
        # after _start_refresh() below, so does the refresh subprocess it
        # just spawned (schema-init runs on every get_connection() call).
        # A transient "database is locked" race between those two is
        # possible. If it happens, fall back to a status payload built
        # from what we already know rather than letting the exception
        # propagate out of this view. "running" defaults to False here --
        # if we can't even read the DB, we genuinely don't know, and
        # assuming "not running" lets the user retry rather than looking
        # permanently stuck.
        try:
            return _refresh_status()
        except Exception as exc:
            return {
                "status": "error",
                "running": False,
                "last_refresh_at": None,
                "last_refresh_error": f"Could not read refresh status: {exc}",
                "city_count": 0,
            }

    try:
        status = safe_refresh_status()
        context = {
            "status": status,
            "crime_workbook": _crime_workbook_status(),
            "success_message": None,
            "error_message": None,
            "search_query": "",
            "search_payload": None,
            "city_preview": [],
            "google_maps_api_key": current_app.config.get("GOOGLE_MAPS_API_KEY") or "",
            "google_maps_map_id": current_app.config.get("GOOGLE_MAPS_MAP_ID") or "",
        }

        with db_module.get_connection() as conn:
            context["city_preview"] = db_module.fetch_all_cities(conn)[:5]

        if request.method == "GET":
            query = request.args.get("q", "").strip()
            if query:
                context["search_query"] = query
                with db_module.get_connection() as conn:
                    city_index = db_module.build_city_index_payload(conn)
                    excluded_index = db_module.build_excluded_index_payload(conn)
                context["search_payload"] = find_city_match(query, city_index, excluded_index)
            return render_template("tools/fire_metrics.html", **context)

        action = request.form.get("action", "")

        def respond(status_code: int = 200):
            if is_ajax:
                # Admin actions are triggered via fetch() from the page's
                # own JS specifically so they don't navigate away -- a
                # full-page render/redirect here would reload the page and
                # discard whatever search result the user currently has on
                # screen (client-side only state, never persisted
                # server-side).
                payload = {
                    "success_message": context["success_message"],
                    "error_message": context["error_message"],
                    "crime_workbook": _crime_workbook_status(),
                }
                payload.update(safe_refresh_status())
                return jsonify(payload), status_code
            return render_template("tools/fire_metrics.html", **context), status_code

        if action == "refresh_all":
            started = _start_refresh(skip_climate=False, skip_crime=False)
            if started:
                context["success_message"] = "Refresh started in the background. This can take several minutes (climate risk especially, on a cold cache)."
            else:
                context["error_message"] = "A refresh is already running. Check back shortly."
            context["status"] = safe_refresh_status()
            return respond()

        if action == "refresh_live_only":
            # Population/income/home-value/employment only -- skips the slow
            # climate step and the manual/periodic crime step.
            started = _start_refresh(skip_climate=True, skip_crime=True)
            if started:
                context["success_message"] = "Refreshing live metrics (population, income, home value, employment) in the background."
            else:
                context["error_message"] = "A refresh is already running. Check back shortly."
            context["status"] = safe_refresh_status()
            return respond()

        if action == "rebuild_index":
            try:
                results = _reingest_from_disk()
                if not results:
                    context["error_message"] = "No pipeline output files found on disk yet. Run a refresh first."
                else:
                    context["success_message"] = f"Re-ingested from disk: {', '.join(results.keys())}."
            except Exception as exc:
                context["error_message"] = f"Could not re-ingest from disk: {exc}"
            context["status"] = safe_refresh_status()
            return respond()

        context["error_message"] = "Unknown action."
        return respond(status_code=400)
    except Exception as exc:
        # Last-resort guard covering the whole view, including the
        # earliest calls above (status/context/city_preview) that run
        # before respond() even exists yet: if anything in this view
        # raises unexpectedly and the caller is this page's own AJAX JS,
        # it must still get valid JSON back -- otherwise Flask's default
        # HTML error page reaches the browser and res.json() throws a
        # confusing "Unexpected token '<'" instead of showing the real
        # problem. Non-AJAX (plain GET/POST) callers keep the normal Flask
        # error behavior.
        if is_ajax:
            return jsonify({
                "success_message": None,
                "error_message": f"Unexpected error: {exc}",
                "status": "error",
                "running": False,
                "last_refresh_at": None,
                "last_refresh_error": str(exc),
                "city_count": 0,
                "crime_workbook": {"exists": False, "uploaded_at": None},
            }), 500
        raise


@fire_metrics_bp.route("/search")
@login_required
def search():
    query = request.args.get("q", "").strip()
    try:
        with db_module.get_connection() as conn:
            city_index = db_module.build_city_index_payload(conn)
            excluded_index = db_module.build_excluded_index_payload(conn)
        payload = find_city_match(query, city_index, excluded_index)
        payload["query"] = query
        payload["status_meta"] = _refresh_status()
        return jsonify(payload)
    except Exception as exc:
        return jsonify({"status": "error", "query": query, "user_message": f"Search failed: {exc}"}), 500


@fire_metrics_bp.route("/api/city-summary", methods=["POST"])
@login_required
def city_summary():
    payload = request.get_json(silent=True) or {}
    city_key = str(payload.get("city_key") or "").strip()
    city = str(payload.get("city") or "").strip()
    state = str(payload.get("state") or "").strip().upper()

    if not city_key and (not city or not state):
        return jsonify({
            "status": "error",
            "error_code": "invalid_city_identifier",
            "user_message": "City identifier is required.",
        }), 400

    try:
        with db_module.get_connection() as conn:
            selected_city = db_module.fetch_city_by_summary_identity(
                conn,
                city_key=city_key or None,
                city=city or None,
                state=state or None,
            )
            if not selected_city:
                return jsonify({
                    "status": "error",
                    "error_code": "city_not_found",
                    "user_message": "City not found in tracked FIRE Metrics data.",
                }), 404

            all_cities = db_module.fetch_all_included_cities(conn)
            metadata = db_module.get_metadata(conn)

            benchmarks = ai_summary.compute_benchmarks(selected_city, all_cities)

            if not _summary_enabled():
                return jsonify(_summary_unavailable_response(
                    selected_city=selected_city,
                    benchmark_data=benchmarks,
                    reason="AI summaries are disabled.",
                    data_refreshed_at=metadata.get("last_refresh_at"),
                ))

            model_name = _summary_model_name()
            fingerprint_input = ai_summary.fingerprint_payload(
                selected_city=selected_city,
                benchmarks=benchmarks,
                model_name=model_name,
                refresh_last_at=metadata.get("last_refresh_at"),
            )
            data_fingerprint = ai_summary.build_fingerprint(fingerprint_input)

            try:
                cache_row = db_module.fetch_cached_city_summary(
                    conn,
                    city=selected_city["city"],
                    state=selected_city["state"],
                    data_fingerprint=data_fingerprint,
                    model_name=model_name,
                    prompt_version=ai_summary.PROMPT_VERSION,
                )
            except Exception as exc:
                current_app.logger.warning(
                    "FIRE Metrics city-summary cache read failed: %s",
                    exc.__class__.__name__,
                )
                cache_row = None
            if cache_row:
                return jsonify({
                    "status": "ready",
                    "summary": cache_row["summary_text"],
                    "summary_structured": {
                        "strength_sentence": cache_row["strength_sentence"],
                        "weakness_sentence": cache_row["weakness_sentence"],
                        "comparison_sentence": cache_row["comparison_sentence"],
                    },
                    "generated_at": cache_row["generated_at"],
                    "data_refreshed_at": metadata.get("last_refresh_at"),
                    "cached": True,
                    "city_key": cache_row["city_key"],
                    "relative_market_profile_score": benchmarks.get("relative_market_profile_score"),
                    "relative_market_profile_percentile": benchmarks.get("relative_market_profile_percentile"),
                    "tracked_city_relative_market_profile_average": benchmarks.get("tracked_city_relative_market_profile_average"),
                    "recommendation_category": benchmarks.get("recommendation_category"),
                    "score": benchmarks.get("selected_overall_score"),
                    "computed_composite_score": benchmarks.get("selected_overall_score"),
                    "tracked_city_average": benchmarks.get("tracked_city_average"),
                    "tracked_city_composite_average": benchmarks.get("tracked_city_average"),
                    "tracked_city_count": benchmarks.get("tracked_city_count"),
                    "percentile": benchmarks.get("selected_percentile"),
                    "source": "cache",
                })

            generated_at = ai_summary.utc_now_iso()
            api_key = _summary_api_key()
            if not api_key:
                return jsonify(_summary_unavailable_response(
                    selected_city=selected_city,
                    benchmark_data=benchmarks,
                    reason="OPENAI_API_KEY is not configured.",
                    data_refreshed_at=metadata.get("last_refresh_at"),
                ))

            if not model_name:
                return jsonify(_summary_unavailable_response(
                    selected_city=selected_city,
                    benchmark_data=benchmarks,
                    reason="FIRE_METRICS_SUMMARY_MODEL is not configured.",
                    data_refreshed_at=metadata.get("last_refresh_at"),
                ))

            try:
                structured = ai_summary.openai_summary(
                    api_key=api_key,
                    model_name=model_name,
                    selected_city=selected_city,
                    benchmarks=benchmarks,
                )
                structured = ai_summary.normalize_summary(structured, selected_city, benchmarks)
            except Exception:
                structured = ai_summary.fallback_summary(selected_city, benchmarks)

            summary_text = ai_summary.combined_summary(structured)
            cache_payload = {
                "city": selected_city["city"],
                "state": selected_city["state"],
                "city_key": ai_summary.city_key(selected_city),
                "data_fingerprint": data_fingerprint,
                "model_name": model_name,
                "prompt_version": ai_summary.PROMPT_VERSION,
                "summary_text": summary_text,
                "strength_sentence": structured["strength_sentence"],
                "weakness_sentence": structured["weakness_sentence"],
                "comparison_sentence": structured["comparison_sentence"],
                "generated_at": generated_at,
            }

            try:
                db_module.upsert_city_summary_cache(conn, cache_payload)
            except Exception as exc:
                current_app.logger.warning(
                    "FIRE Metrics city-summary cache write failed: %s",
                    exc.__class__.__name__,
                )

            return jsonify({
                "status": "ready",
                "summary": summary_text,
                "summary_structured": structured,
                "generated_at": generated_at,
                "data_refreshed_at": metadata.get("last_refresh_at"),
                "cached": False,
                "city_key": cache_payload["city_key"],
                "relative_market_profile_score": benchmarks.get("relative_market_profile_score"),
                "relative_market_profile_percentile": benchmarks.get("relative_market_profile_percentile"),
                "tracked_city_relative_market_profile_average": benchmarks.get("tracked_city_relative_market_profile_average"),
                "recommendation_category": benchmarks.get("recommendation_category"),
                "score": benchmarks.get("selected_overall_score"),
                "computed_composite_score": benchmarks.get("selected_overall_score"),
                "tracked_city_average": benchmarks.get("tracked_city_average"),
                "tracked_city_composite_average": benchmarks.get("tracked_city_average"),
                "tracked_city_count": benchmarks.get("tracked_city_count"),
                "percentile": benchmarks.get("selected_percentile"),
                "source": "generated",
            })
    except Exception as exc:
        current_app.logger.exception("FIRE Metrics city-summary endpoint failed: %s", exc.__class__.__name__)
        response = _summary_unavailable_response(
            selected_city=None,
            benchmark_data=None,
            reason="Summary generation is currently unavailable.",
            data_refreshed_at=None,
        )
        response["error_code"] = "summary_endpoint_failed"
        return jsonify(response), 500


@fire_metrics_bp.route("/refresh-status")
@login_required
def refresh_status():
    return jsonify(_refresh_status())


@fire_metrics_bp.route("/upload-crime-workbook", methods=["POST"])
@login_required
def upload_crime_workbook():
    def safe_crime_workbook_status() -> dict:
        try:
            return _crime_workbook_status()
        except Exception:
            return {"exists": False, "uploaded_at": None}

    def respond(success: bool, message: str, status_code: int = 200):
        return jsonify({
            "success": success,
            "message": message,
            "crime_workbook": safe_crime_workbook_status(),
        }), status_code

    try:
        file = request.files.get("crime_workbook")
        if file is None or not file.filename:
            return respond(False, "No file selected.", 400)

        if not file.filename.lower().endswith(".xlsx"):
            return respond(False, "File must be a .xlsx workbook.", 400)

        data = file.read()
        if not data:
            return respond(False, "File is empty. Upload the .xlsx workbook exactly as downloaded from the FBI.", 400)

        if len(data) > MAX_CRIME_WORKBOOK_BYTES:
            size_mb = len(data) / (1024 * 1024)
            return respond(
                False,
                f"File is too large ({size_mb:.1f} MB) -- the real FBI Table 8 workbook is "
                f"only a few MB. Check you selected the right file.",
                400,
            )

        validation_error = _validate_crime_workbook_bytes(data)
        if validation_error:
            return respond(False, validation_error, 400)

        # Uses the same resolver the crime pipeline uses. In production
        # this should be FBI_CRIME_WORKBOOK_PATH on the persistent /data
        # volume, so the file survives redeploys the same way the SQLite
        # DB now does via FIRE_METRICS_DB_PATH.
        target_path = _get_crime_workbook_path()
        target_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile("wb", dir=target_path.parent, delete=False) as tmp:
                tmp.write(data)
                tmp_path = Path(tmp.name)
            os.replace(tmp_path, target_path)
        finally:
            if tmp_path and tmp_path.exists():
                tmp_path.unlink()

        return respond(
            True,
            "Crime workbook uploaded. It will be picked up the next time you run "
            "\"Refresh All Data\".",
        )
    except Exception as exc:
        return respond(False, f"Unexpected error while uploading: {exc}", 500)


@fire_metrics_bp.route("/debug-refresh")
@login_required
def debug_refresh():
    # TEMPORARY diagnostic route -- added specifically to inspect the raw
    # refresh_metadata table (including per-step results the normal status
    # payload doesn't surface) without needing direct Railway console/DB
    # access. Not linked from any page; remove once the climate-risk
    # never-populates investigation is resolved.
    with db_module.get_connection() as conn:
        metadata = db_module.get_metadata(conn)
    steps_raw = metadata.get("refresh_steps_json")
    try:
        parsed_steps = json.loads(steps_raw) if steps_raw else None
    except json.JSONDecodeError as exc:
        parsed_steps = f"<could not parse refresh_steps_json: {exc}>"
    return jsonify({
        "raw_metadata": metadata,
        "parsed_steps": parsed_steps,
    })


@fire_metrics_bp.route("/download-latest")
@login_required
def download_latest():
    data = _export_workbook()
    from flask import send_file

    return send_file(
        io.BytesIO(data),
        as_attachment=True,
        download_name="fire_metrics_city_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
