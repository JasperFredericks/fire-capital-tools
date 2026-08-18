"""
FIRE Capital Tools - Site DD capex export.

Turns an assessment's findings into a capital budget: PDF via matplotlib
PdfPages and XLSX via openpyxl, the same two mechanisms every other
export here uses.

THE PROVENANCE COLUMN IS THE POINT

Every line says where its number came from:

    Inspector estimate    somebody stood in the room and judged it
    Researched average    a national figure from tools/site_dd_reference_costs
    No estimate           nothing priced it, and the line says so

Those are three different kinds of claim and a budget that rendered them
identically would let a national average acquire the authority of a
site visit. The three are separated in the totals as well as the rows,
because "of this $84,000, $61,000 is national averages nobody has
checked" is the sentence that decides whether the number is usable.

WHAT IS NOT INCLUDED

A finding with no cost contributes nothing to the total, but IS still
listed with "no estimate". Dropping it would make the budget look
complete when it is not, which is the failure mode that matters: an
underlying repair that never got priced is invisible, and the total
reads as the whole job.
"""

from __future__ import annotations

import textwrap
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt                       # noqa: E402
from matplotlib.backends.backend_pdf import PdfPages   # noqa: E402

from tools import site_dd_checklist as cl              # noqa: E402
from tools import site_dd_costs as costs               # noqa: E402
from tools import site_dd_reference_costs as refcosts  # noqa: E402
from tools import underwriting_capex as ucx            # noqa: E402

PAGE_SIZE = (11, 8.5)
INK = "#1a2744"
MUTED = "#6b7280"
BODY = "#111827"
WARN = "#b45309"

ROWS_PER_PAGE = 26

SOURCE_COLUMN = {
    costs.SOURCE_MANUAL: "Inspector estimate",
    costs.SOURCE_REFERENCE: "Researched average",
    costs.SOURCE_NONE: "No estimate",
}


def build_lines(findings: list[dict[str, Any]], labels: dict[str, str] | None = None
                ) -> list[dict[str, Any]]:
    """Budget rows, with quantity as the instance count.

    Forty toilets are one line of quantity 40, not forty lines of one.
    This is the grouping site_dd_costs.to_capex_lines() has always
    implemented for the Underwriting hand-off; the export was written
    separately and hard-coded quantity to 1, so a unit cost entered by
    hand produced a line total of exactly that unit cost however many of
    the thing there were. The two paths now agree.

    WHAT IS IN THE GROUPING KEY, AND WHY IT IS MORE THAN (area, room, item)

    to_capex_lines() groups on (area, room, item) alone and takes the
    first non-null cost it finds. That is safe when every instance is
    priced the same and silently wrong when they are not: two toilets at
    $450 and $600 would become "Toilet x2" at whichever price came first,
    and $300 would leave the budget without a trace.

    So condition, unit cost and provenance join the key. Instances that
    are genuinely the same collapse into one line with a quantity;
    instances that differ in what is wrong with them or what they cost
    stay visible as separate lines. Nothing can be absorbed into a
    quantity unless it is interchangeable with the rows beside it.

    The total is not computed here. It comes from
    underwriting_capex.line_total(), which is the function that already
    owns quantity x unit cost for the whole app -- writing a second
    multiplication here would create two numbers that can disagree.
    """
    labels = labels or {}
    groups: dict[tuple, dict[str, Any]] = {}
    order: list[tuple] = []

    for f in findings or []:
        described = costs.describe(f)
        key = (f.get("area_id"), f.get("room_id"), f.get("item_key"),
               f.get("condition"), described["cost"], described["source"],
               (f.get("instance_label") or "").strip())
        if key not in groups:
            groups[key] = {"rows": [], "first": f, "described": described}
            order.append(key)
        groups[key]["rows"].append(f)

    out = []
    for key in order:
        group = groups[key]
        f, described = group["first"], group["described"]
        cat = costs.capex_category(f)

        # WHICH KIND OF QUANTITY THIS LINE TAKES
        #
        # A per-item figure is multiplied by the instance count, which is
        # what the grouping above produced and is unchanged.
        #
        # A rate -- dollars per square foot, per linear foot -- cannot be.
        # Multiplying $5.75/sqft by "one wall" yields $5.75, which is the
        # rate wearing a total's clothing. Rates are therefore priced only
        # from a measured quantity recorded against the finding, and when
        # there is none the line carries its rate, no total, and a reason
        # naming the measurement it needs.
        # THE UNIT BELONGS TO THE ITEM, NOT TO WHO PRICED IT
        #
        # This used to read the unit off the _reference object, which is
        # attached only when a researched cost was APPLIED. An inspector
        # typing their own figure on walls_ceiling therefore produced no
        # _reference, fell through to "each", and got $5.75 multiplied by
        # an instance count -- the original bug, arriving through the
        # manual door instead of the researched one.
        #
        # walls_ceiling is a per-square-foot item whoever prices it. So
        # the unit is looked up from the item, and a manual figure on a
        # rate item is a rate.
        ref = f.get("_reference")
        unit = getattr(ref, "unit", None)
        if unit is None:
            known = refcosts.for_item(f.get("item_key"),
                                      f.get("detail") if f.get("item_key") == "flooring"
                                      else None)
            unit = getattr(known, "unit", None)
        if unit is None and described["cost"] is not None:
            measure = (f.get("measure") or "").strip().lower()
            if measure in refcosts.UNITS:
                unit = measure
            else:
                # A FREEFORM item with a hand-typed cost and no answer to
                # the per-job / per-sq-ft toggle.
                #
                # Nothing in the table describes the item, so there is no
                # unit to inherit, and the toggle -- which exists to ask
                # exactly this -- was left blank. The honest answer is
                # that we do not know what the number means, so the line
                # is not totalled at all.
                #
                # It used to be classified by magnitude: under $15 it was
                # called a rate, over $15 a job price. That guess is what
                # the toggle replaces. Guessing and then silently
                # totalling is the failure this whole line of work exists
                # to remove -- a $5.75 that means "per sq ft" became a
                # $5.75 repaint budget for a whole kitchen.
                unit = None

        instances = float(len(group["rows"]))
        measured = [costs.clean_cost(r.get("quantity")) for r in group["rows"]]
        measured = [m for m in measured if m is not None]
        if refcosts.is_rate(unit):
            quantity = sum(measured) if measured else None
        elif unit is None and described["cost"] is not None:
            # A cost in unstated units. There is no quantity that means
            # anything here: multiplying by the instance count would
            # assume "per job", which is the assumption the toggle exists
            # to stop being made silently.
            quantity = None
        else:
            quantity = instances

        needs = ""
        if described["cost"] is not None and unit is None:
            # Priced by a person, in units nobody stated. The magnitude
            # hint is a HINT: it tells the inspector which answer is
            # likely without letting that likelihood become a total.
            hint = ("It looks like a rate" if refcosts.looks_like_a_rate(described["cost"])
                    else "It looks like a job price")
            needs = (f"Cost entered, unit not specified. ${described['cost']:,.2f} "
                     f"could be a price for the whole job or a rate per square "
                     f"foot, and those differ by orders of magnitude. {hint}, "
                     f"but say so on the finding rather than leaving it to be "
                     f"guessed. Not included in the total until then.")
        elif refcosts.is_rate(unit) and quantity is None:
            # A whole sentence, because this lands in a "Why no estimate"
            # cell that has to be readable on its own: it must say what
            # the figure is, that it is a rate, and what to go and
            # measure.
            needs = (f"Priced at ${described['cost']:,.2f} "
                     f"{refcosts.UNIT_LABELS.get(unit, unit)}. Needs "
                     f"{refcosts.measurement_needed(unit)} before it can be "
                     f"totalled; an instance count is not a measurement.")

        line = {
            "item_key": f.get("item_key"),
            "label": (f.get("instance_label")
                      or labels.get(f.get("item_key"))
                      or f.get("item_key")),
            "category": cat,
            "category_name": cl.CATEGORY_NAMES.get(cat, "Uncategorised"),
            "condition": f.get("condition"),
            "scope": f.get("scope"),
            "unit_cost": described["cost"],
            "unit": unit,
            "unit_label": refcosts.UNIT_LABELS.get(unit, ""),
            "is_rate": refcosts.is_rate(unit),
            "instances": instances,
            "source": described["source"],
            "source_label": SOURCE_COLUMN[described["source"]],
            "quantity": quantity,
            # Left None on purpose: line_total() below derives it. Kept on
            # the row so this line has the same shape as an Underwriting
            # capex line, where an explicit total legitimately overrides
            # quantity x unit cost.
            "total_cost": None,
            "reason": (refcosts.reason(f.get("item_key"))
                       if described["cost"] is None else needs),
        }
        # None, not 0.0, when it cannot be computed honestly. A zero would
        # sum into the total as though the work were free; None keeps the
        # line visible, keeps its rate on screen, and sends it to the
        # unpriced set where summarize() will report it.
        line["total"] = (ucx.line_total(line)
                         if described["cost"] is not None and quantity is not None
                         else None)
        out.append(line)
    return out


def summarize(lines: list[dict[str, Any]]) -> dict[str, Any]:
    """Totals, split by where the money came from.

    Deliberately three totals rather than one. A single figure would
    average a site visit and a national average into a number that
    describes neither.
    """
    by_source = {k: 0.0 for k in SOURCE_COLUMN}
    by_category: dict[str, float] = {}
    priced: list[dict[str, Any]] = []
    unmeasured: list[dict[str, Any]] = []
    unresearched: list[dict[str, Any]] = []
    for l in lines:
        # THREE BUCKETS, AND THEY MUST NOT COLLAPSE INTO TWO
        #
        # priced                 a figure and a quantity to apply it to
        # researched, unmeasured a real researched rate, nothing measured
        #                        to multiply it by. Needs a tape measure.
        # unresearched           no figure at all. Needs research that may
        #                        not exist in published form.
        #
        # Merging the last two is the same mistake as multiplying a rate
        # by a headcount, one level up: it reports "no cost data" for an
        # item whose cost we know to the cent.
        if l["total"] is not None:
            priced.append(l)
            by_source[l["source"]] += l["total"]
            if l["total"]:
                by_category[l["category_name"]] = (
                    by_category.get(l["category_name"], 0.0) + l["total"])
        elif l.get("is_rate") and l["unit_cost"] is not None:
            unmeasured.append(l)
        else:
            unresearched.append(l)

    priced_total = sum(by_source.values())
    # None, not 0.0, when there are lines and none of them could be
    # priced. "$0.00" beside an unmeasured line reads as "this costs
    # nothing" rather than "this cost is not known yet" -- the line
    # refuses to state a number and the summary must not state one on its
    # behalf.
    #
    # An EMPTY budget is different and really is zero: nothing was
    # recorded as needing work, which is a finding, not a gap.
    total = priced_total if (priced or not lines) else None
    unpriced = unmeasured + unresearched
    return {
        "total": total,
        "priced_total": priced_total,
        # True whenever the total describes only part of the work, so a
        # caller cannot render it as a finished budget by accident.
        "total_is_partial": bool(unpriced),
        "by_source": by_source,
        "by_category": dict(sorted(by_category.items(),
                                   key=lambda kv: -kv[1])),
        "priced": priced,
        "priced_count": len(priced),
        "unpriced": unpriced,
        "unpriced_count": len(unpriced),
        "unmeasured": unmeasured,
        "unmeasured_count": len(unmeasured),
        "unresearched": unresearched,
        "unresearched_count": len(unresearched),
        "line_count": len(lines),
        # Share OF THE PRICED TOTAL, and None when there is no priced
        # total. Zero would claim we hold no research, which is false when
        # every unpriced line is a researched rate waiting on a
        # measurement.
        "researched_pct": ((by_source[costs.SOURCE_REFERENCE] / priced_total * 100)
                           if priced_total else None),
        "researched_on": refcosts.RESEARCHED_ON,
        "coverage_sentence": coverage_sentence(
            len(priced), len(lines), len(unmeasured), len(unresearched)),
    }


def coverage_sentence(priced: int, total_lines: int, unmeasured: int,
                      unresearched: int) -> str:
    """One sentence naming all three buckets, in Michelle's words.

    Written once, here, because the PDF and the XLSX must not be able to
    describe the same budget differently -- and because the sentence is
    the part that stops a partial total being read as a finished one.
    """
    def lines_(n):
        return f"{n} line" if n == 1 else f"{n} lines"

    if not total_lines:
        return "No items were recorded as needing work."
    if priced == total_lines:
        return (f"All {lines_(total_lines)} priced. This total is the whole "
                f"recorded budget.")
    bits = []
    if unmeasured:
        bits.append(f"{lines_(unmeasured)} "
                    f"{'has' if unmeasured == 1 else 'have'} a researched rate "
                    f"but nothing measured to apply it to")
    if unresearched:
        bits.append(f"{lines_(unresearched)} "
                    f"{'has' if unresearched == 1 else 'have'} no researched "
                    f"figure at all")
    if not priced:
        return ("Nothing here can be priced yet, so there is NO total: "
                + " and ".join(bits) + ".")
    return (f"This total covers {priced} of {lines_(total_lines)} and is NOT "
            f"the full budget: " + " and ".join(bits) + ".")


def _money(value: Any) -> str:
    return "—" if value in (None, "") else f"${float(value):,.0f}"


def _qty(value: Any) -> str:
    """Whole counts read as counts: 40, not 40.0."""
    if value in (None, ""):
        return "—"
    number = float(value)
    return f"{number:,.0f}" if number == int(number) else f"{number:,.2f}"


def build_pdf(path, assessment: dict[str, Any], lines: list[dict[str, Any]],
              summary: dict[str, Any]) -> Path:
    path = Path(path)
    label = assessment.get("property_label") or "Property"
    pages = [lines[i:i + ROWS_PER_PAGE]
             for i in range(0, len(lines), ROWS_PER_PAGE)] or [[]]

    with PdfPages(str(path)) as pdf:
        for page_no, page in enumerate(pages, start=1):
            fig = plt.figure(figsize=PAGE_SIZE)
            fig.text(0.06, 0.94, "Capital Budget", fontsize=16,
                     fontweight="bold", color=INK)
            fig.text(0.06, 0.915, label, fontsize=11, color="#4b5563")
            fig.text(0.94, 0.94,
                     _money(summary["total"]) if summary["total"] is not None
                     else "no priced lines",
                     ha="right", fontsize=15, fontweight="bold",
                     color=INK if summary["total"] is not None else WARN)
            fig.text(0.94, 0.915,
                     (f"{summary['line_count']} items · "
                      f"{summary['priced_count']} priced · "
                      f"{summary['unmeasured_count']} need measuring · "
                      f"{summary['unresearched_count']} unresearched"),
                     ha="right", fontsize=9, color=MUTED)

            if page_no == 1:
                y = 0.865
                fig.text(0.06, y, "Where these numbers come from",
                         fontsize=10, fontweight="bold", color=INK)
                y -= 0.028
                # Three buckets, three rows. The last two used to share a
                # line reading "N item(s), not costed", which said the same
                # thing about an item priced at $5.75/sqft and an item
                # nobody has ever researched.
                have_total = summary["total"] is not None
                rows = [
                    (SOURCE_COLUMN[costs.SOURCE_MANUAL],
                     _money(summary["by_source"][costs.SOURCE_MANUAL])
                     if have_total else "—", BODY),
                    (SOURCE_COLUMN[costs.SOURCE_REFERENCE],
                     _money(summary["by_source"][costs.SOURCE_REFERENCE])
                     if have_total else "—", BODY),
                    ("Researched rate, not yet measured",
                     f"{summary['unmeasured_count']} item(s)", WARN),
                    ("No researched figure",
                     f"{summary['unresearched_count']} item(s)", MUTED),
                ]
                for name, value, colour in rows:
                    fig.text(0.06, y, name, fontsize=9, color=BODY)
                    fig.text(0.42, y, value, fontsize=9, color=colour)
                    y -= 0.024
                y -= 0.012
                for text in textwrap.wrap(summary["coverage_sentence"], 108):
                    fig.text(0.06, y, text, fontsize=8.5, color=WARN)
                    y -= 0.020
                if summary["researched_pct"] is not None:
                    fig.text(0.06, y,
                             f"{summary['researched_pct']:.0f}% of the priced "
                             f"subtotal is researched national averages "
                             f"({summary['researched_on']}), not quotes.",
                             fontsize=8.5, color=MUTED)
                else:
                    fig.text(0.06, y,
                             f"Researched national averages "
                             f"({summary['researched_on']}) are used where a "
                             f"figure exists; none could be applied here.",
                             fontsize=8.5, color=MUTED)
                y -= 0.020
                top = y - 0.035
            else:
                top = 0.865

            # Quantity earns a column now that it can be more than 1: a
            # $600 unit cost beside a $24,000 total is unreadable without
            # the 40 that connects them.
            # "Per" is the unit of measure and it is not decoration. A
            # $5.75 rate beside a blank total only makes sense once the
            # column says "per sq ft"; without it the reader sees a cheap
            # item that failed to add up.
            cols = (0.06, 0.26, 0.42, 0.54, 0.68, 0.76, 0.87, 0.94)
            heads = ("Item", "Category", "Condition", "Source", "Rate",
                     "Per", "Qty", "Total")
            for x, head, align in zip(cols, heads,
                                      ("left",) * 7 + ("right",)):
                fig.text(x, top, head, fontsize=8.5, fontweight="bold",
                         color=MUTED, ha=align)
            y = top - 0.024

            for row in page:
                fig.text(cols[0], y, textwrap.shorten(str(row["label"]), 30,
                                                      placeholder="…"),
                         fontsize=8.5, color=BODY)
                fig.text(cols[1], y, textwrap.shorten(row["category_name"], 26,
                                                      placeholder="…"),
                         fontsize=8, color=MUTED)
                fig.text(cols[2], y, (row["condition"] or "—").title(),
                         fontsize=8, color=MUTED)
                fig.text(cols[3], y, row["source_label"], fontsize=8,
                         color=WARN if row["source"] == costs.SOURCE_REFERENCE
                         else (MUTED if row["source"] == costs.SOURCE_NONE else BODY))
                fig.text(cols[4], y, _money(row["unit_cost"]), fontsize=8.5,
                         color=BODY)
                fig.text(cols[5], y, row["unit_label"] or "—", fontsize=8,
                         color=WARN if row["is_rate"] else MUTED)
                qty = row["quantity"]
                fig.text(cols[6], y, _qty(qty), fontsize=8.5,
                         color=BODY if (qty or 0) > 1 else MUTED)
                fig.text(cols[7], y,
                         _money(row["total"]) if row["total"] is not None else "—",
                         fontsize=8.5, color=BODY, ha="right")
                y -= 0.026

            fig.text(0.06, 0.05,
                     "Researched averages are national figures for budgeting, "
                     "not quotes. Items with no estimate, and rates with no "
                     "measured quantity, are listed but contribute nothing to "
                     "the total.",
                     fontsize=7.5, color=MUTED)
            fig.text(0.94, 0.05, f"Page {page_no} of {len(pages)}",
                     ha="right", fontsize=8, color=MUTED)
            pdf.savefig(fig)
            plt.close(fig)
    return path


def build_xlsx(path, assessment: dict[str, Any], lines: list[dict[str, Any]],
               summary: dict[str, Any],
               labels: dict[str, str] | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Capital Budget"
    bold = Font(bold=True)

    ws.append(["Capital Budget"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([assessment.get("property_label") or ""])
    ws.append([f"Inspected {assessment.get('assessed_on') or '—'}"])
    ws.append([])
    # The label carries the caveat, so the number can never be lifted out
    # of the cell beside it and read as a finished budget.
    if summary["total"] is None:
        ws.append(["Total", "No priced lines — see below"])
    else:
        ws.append(["Priced subtotal" if summary["total_is_partial"] else "Total",
                   summary["total"]])
    for key in (costs.SOURCE_MANUAL, costs.SOURCE_REFERENCE):
        # Dash, not 0.00, when there is no priced subtotal for them to be
        # a part of -- the same reason the total itself declines to be a
        # number here.
        ws.append([SOURCE_COLUMN[key],
                   summary["by_source"][key] if summary["total"] is not None
                   else "—"])
    ws.append(["Researched rate, not yet measured",
               f"{summary['unmeasured_count']} item(s)"])
    ws.append(["No researched figure",
               f"{summary['unresearched_count']} item(s)"])
    ws.append([summary["coverage_sentence"]])
    if summary["researched_pct"] is not None:
        ws.append([f"{summary['researched_pct']:.0f}% of the priced subtotal is "
                   f"researched national averages ({summary['researched_on']}), "
                   f"not quotes."])
    else:
        ws.append([f"Researched national averages ({summary['researched_on']}) "
                   f"are used where a figure exists; none could be applied here."])
    ws.append([])

    header = ["Item", "Category", "Scope", "Condition", "Cost source",
              "Unit cost", "Unit", "Qty", "Total", "Why no estimate"]
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.font = bold

    for l in lines:
        ws.append([l["label"], l["category_name"], l["scope"],
                   (l["condition"] or ""), l["source_label"],
                   l["unit_cost"], l["unit_label"], l["quantity"],
                   l["total"], l["reason"]])

    for col, width in zip("ABCDEFGHIJ", (30, 26, 10, 12, 20, 12, 13, 8, 12, 60)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=11, min_col=10, max_col=10):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # Rates that have a researched figure and no measurement to apply it
    # to. Its own sheet rather than a footnote: these are the lines a
    # walk can turn into real money, and each one names what to measure.
    if summary["unmeasured"]:
        nm = wb.create_sheet("Needs measurement")
        nm.append(["Item", "Where", "Rate", "Unit", "What is needed"])
        for cell in nm[1]:
            cell.font = bold
        for l in summary["unmeasured"]:
            nm.append([l["label"], l["scope"], l["unit_cost"],
                       l["unit_label"], l["reason"]])
        for col, width in zip("ABCDE", (30, 12, 12, 14, 60)):
            nm.column_dimensions[col].width = width
        for row in nm.iter_rows(min_row=2, min_col=5, max_col=5):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # The reference table itself, so a reader can audit any figure that
    # appeared above without leaving the file.
    ref = wb.create_sheet("Reference costs")
    ref.append(["Item", "Key", "Unit cost", "Unit", "Sources",
                "How it was derived"])
    for cell in ref[1]:
        cell.font = bold
    for key in sorted(refcosts.REFERENCE_COSTS):
        c = refcosts.REFERENCE_COSTS[key]
        ref.append([(labels or {}).get(c.key, c.key), c.key, c.unit_cost,
                    refcosts.UNIT_LABELS[c.unit],
                    ", ".join(c.sources), c.note])
    for col, width in zip("ABCDEF", (30, 24, 12, 14, 42, 80)):
        ref.column_dimensions[col].width = width
    for row in ref.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # And what has NO figure, with the reason. This sheet is the ask, so
    # it gets the labels a person recognises -- "ADA parking & path of
    # travel", not "ada_parking_path". A list meant to be read by
    # somebody who does not work in this codebase should not be written
    # in its identifiers.
    un = wb.create_sheet("Not priced")
    un.append(["Item", "Key", "Why it has no researched figure"])
    for cell in un[1]:
        cell.font = bold
    for row in refcosts.unpriced_report(labels or {}):
        un.append([row["label"], row["key"], row["reason"]])
    for col, width in zip("ABC", (34, 24, 90)):
        un.column_dimensions[col].width = width
    for r in un.iter_rows(min_row=2, min_col=3, max_col=3):
        r[0].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(str(path))
    return path


def suggested_filename(assessment: dict[str, Any], ext: str) -> str:
    label = "".join(ch if ch.isalnum() or ch in " -_" else ""
                    for ch in (assessment.get("property_label") or "property"))
    label = "-".join(label.split()).lower()[:48] or "property"
    return f"capex-budget-{label}.{ext}"
