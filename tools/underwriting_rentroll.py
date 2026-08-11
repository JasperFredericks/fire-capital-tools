"""
FIRE Capital Tools - Rent roll parsing for Underwriting.

Turns a ResMan rent roll export into per-unit lines: unit, type, square
footage, status, in-place rent, market rent, and lease dates.

Why this is new code rather than reuse. mmr_report.parsers.parse_rent_roll
exists, but it answers a different question -- it returns two scalars,
{"total_rental", "avg_rent"}, for an operations report, and takes an MMR
worksheet plus an occupied count from the box score. Nothing in the MMR
package captures per-unit rents, square footage or lease dates, and
parse_available_units only covers the vacant/notice/holding subset. So the
extraction is new; the *helpers* are not -- the cell utilities and the
rent-line allowlist below are imported from the MMR package rather than
reimplemented, because they already encode real knowledge about which
ledger lines count as rent (HAP and other subsidies do, renters insurance
does not, concessions offset).

ResMan only for this beta. Any other layout raises UnrecognizedRentRoll
rather than guessing -- the Scorecard Pro property-name collision came
from a parser that guessed when it did not recognize a file, and the same
mistake here would silently under- or over-state income for the whole
model.

Structure of a real export (confirmed against Eagle Rock, May 2026):

    row 8   Unit | Type | Sq. Feet | Residents | Status | Market Rent | ...
                 ... | Ledger | Description | Amount | Move In | Lease Start | Lease End
    row 9   0101 | 1/1 Upgraded | 690 | ... | C | 1065 | Resident | Rent   | 310
    row 10                                                        | HAP Rent| 665
    row 11                                                        | Total   | 975

Each unit spans several rows: one carrying the unit's attributes and its
first charge line, then further charge lines, then a Total. In-place rent
is summed from the charge lines that _is_rent_line accepts rather than
read off the Total row, so a unit whose ledger mixes rent with
non-rent charges is still counted correctly.

The Unit column is merged in the export, so its value lands one column to
the left of its own header. That is handled explicitly below rather than
by trusting the header index.
"""

from __future__ import annotations

import datetime
from typing import Any

import openpyxl

from tools.mmr_report.helpers import (
    coerce_num,
    find_col,
    find_col_contains,
    looks_like_unit_value,
    norm,
    rows_of,
    safe_get,
)
from tools.mmr_report.parsers import _is_rent_line

MAX_HEADER_SCAN_ROWS = 40

# The export ends with a charge-type summary block ("Total Charges", then
# "Rent", "HAP Rent", "Pet Rent", ... with totals). Those labels satisfy
# looks_like_unit_value(), so without an explicit stop the summary reads as
# 27 extra phantom units on a real 92-unit Eagle Rock roll -- inflating the
# unit count and dragging every per-unit average toward zero.
_SUMMARY_TERMINATORS = ("total charges", "total credits", "grand total",
                        "summary", "charge summary")


class UnrecognizedRentRoll(ValueError):
    """Raised when the file is not a rent roll layout this parser
    understands. The message is written to be shown to the user."""


# Sheets that only ever appear in a Weekly Property Summary (MMR) export.
# Used to tell the user *which* wrong file they uploaded rather than just
# that it was wrong -- an MMR and a rent roll are both ResMan exports for
# the same property, so "unrecognized" alone would be an unhelpful answer.
#
# Two dialects are in circulation and both are covered. The long-form one
# (Eagle Rock, Canyon, OXPT, High Caliber) uses names like "Cash Flow
# Statement"; the short-form one (Maple Valley) uses "Cash Flow", "Work
# Order", "Tenant Tickler". Listing both matters only for the quality of
# the error message -- either way the file is rejected -- but naming the
# actual mistake is the difference between a user fixing it in seconds and
# re-uploading the same wrong file.
_MMR_SHEET_MARKERS = (
    # long form
    "box score", "cash flow statement", "delinquency", "bank deposit register",
    "bank deposits by category", "work order summary", "expiring leases",
    "new and renewed leases", "prospect source summary", "renewal percentages",
    "available units",
    # short form
    "cash flow", "work order", "tenant tickler", "vacancy", "check register",
    "deposit register", "general ledger",
)
# Three markers, not one: a genuine rent roll is a single unnamed sheet, so
# even one marker would in practice be decisive -- but requiring three means
# a future rent-roll variant that happens to carry a "Vacancy" tab is not
# mislabelled as an MMR.
_MMR_SHEET_MATCH_THRESHOLD = 3


def _looks_like_mmr(sheetnames) -> bool:
    """True when the workbook is a Weekly Property Summary export.

    Matched on several marker sheets rather than one, so a rent roll that
    happens to carry a single similarly-named tab is not misclassified.
    """
    present = {norm(s) for s in sheetnames}
    return sum(1 for m in _MMR_SHEET_MARKERS if m in present) >= _MMR_SHEET_MATCH_THRESHOLD


def _header_index(rows) -> int | None:
    """Row index of the column header band.

    Requires Unit, Market Rent, AND the Description/Amount charge-line pair.

    The charge-line requirement is the load-bearing part, and its absence is
    exactly what let an MMR through before. An MMR's "Available Units" sheet
    does carry Unit and Market Rent, so demanding only those two matched it
    happily -- and because that sheet lists vacant units and has no charge
    lines at all, every unit came back with an in-place rent of zero. The
    result was a fully-computed model built from nothing but vacant units.

    Description and Amount are the columns in-place rent is actually summed
    from, so requiring them is not a heuristic: a sheet without them cannot
    produce a rent roll, only a plausible-looking shell.
    """
    for idx, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        has_unit = find_col(row, "unit") is not None
        has_market = (find_col(row, "market rent") is not None
                      or find_col_contains(row, "market rent") is not None)
        has_charges = (find_col(row, "description") is not None
                       and find_col(row, "amount") is not None)
        if has_unit and has_market and has_charges:
            return idx
    return None


def _as_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _unit_value(row, unit_col):
    """The unit number. Checked at the header's own column and the one to
    its left, because the export merges the Unit cell and openpyxl reports a
    merged value at the range's anchor."""
    for col in (unit_col - 1, unit_col, unit_col + 1):
        if col is None or col < 0:
            continue
        v = safe_get(row, col)
        if looks_like_unit_value(v):
            return str(v).strip()
    return None


def parse_rent_roll_workbook(path) -> dict[str, Any]:
    """Parse a ResMan rent roll .xlsx into per-unit lines.

    Raises UnrecognizedRentRoll if the layout is not recognized or yields no
    units -- never returns a partially-guessed result."""
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as exc:
        raise UnrecognizedRentRoll(f"Could not open the rent roll file: {exc}") from exc

    ws = wb[wb.sheetnames[0]]
    rows = rows_of(ws)
    header_idx = _header_index(rows)
    if header_idx is None:
        # Name the actual mistake when it is recognizable. An MMR is the file
        # most likely to be uploaded here by accident -- it is the same
        # property, the same system and a similar filename -- so it gets its
        # own message rather than a generic rejection.
        if _looks_like_mmr(wb.sheetnames):
            raise UnrecognizedRentRoll(
                "This looks like a Weekly Property Summary / MMR export, not a "
                "rent roll. Please upload the property's actual rent roll file."
            )
        raise UnrecognizedRentRoll(
            "This does not look like a ResMan rent roll — no row was found "
            "carrying a 'Unit' column, a 'Market Rent' column and the "
            "'Description'/'Amount' charge lines that in-place rent is read "
            "from. Only ResMan rent roll exports are supported in this beta; "
            "upload one of those, or enter the rent roll manually."
        )

    header = rows[header_idx]
    cols = {
        "unit": find_col(header, "unit"),
        "type": find_col(header, "type", "unit type"),
        "sqft": find_col(header, "sq. feet", "sq feet", "sqft", "square feet"),
        "status": find_col(header, "status"),
        "market": find_col(header, "market rent") or find_col_contains(header, "market rent"),
        "description": find_col(header, "description"),
        "amount": find_col(header, "amount"),
        "lease_start": find_col(header, "lease start"),
        "lease_end": find_col(header, "lease end", "lease expires"),
        "move_in": find_col(header, "move in"),
        "move_out": find_col(header, "move out"),
    }
    if cols["unit"] is None or cols["market"] is None:
        raise UnrecognizedRentRoll("Rent roll header found but its Unit/Market Rent columns could not be read.")
    # _header_index already required these, so this is a belt-and-braces
    # guard against the two ever drifting apart -- without the charge lines
    # every in-place rent would silently come back as zero.
    if cols["description"] is None or cols["amount"] is None:
        raise UnrecognizedRentRoll(
            "Rent roll header found but it has no 'Description'/'Amount' charge "
            "lines, so in-place rent cannot be read. Upload a full rent roll "
            "export rather than a summary."
        )

    units: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    warnings: list[str] = []

    def close(u):
        if u is not None:
            u["in_place_rent"] = round(u.pop("_rent_accum", 0.0), 2)
            units.append(u)

    for row in rows[header_idx + 1:]:
        first = norm(safe_get(row, 0) or "")
        if first in _SUMMARY_TERMINATORS:
            break                      # per-unit detail is over

        unit_val = _unit_value(row, cols["unit"])
        # A genuine unit row carries at least one unit attribute alongside
        # its number. Requiring corroboration keeps stray labels elsewhere in
        # the sheet from opening a phantom unit block.
        if unit_val and not any((
            safe_get(row, cols["type"]),
            safe_get(row, cols["sqft"]),
            safe_get(row, cols["status"]),
            coerce_num(safe_get(row, cols["market"]), default=None) is not None,
        )):
            unit_val = None

        if unit_val:
            close(current)
            current = {
                "unit": unit_val,
                "unit_type": (str(safe_get(row, cols["type"]) or "").strip() or None),
                "sqft": coerce_num(safe_get(row, cols["sqft"]), default=None),
                "status": (str(safe_get(row, cols["status"]) or "").strip() or None),
                "market_rent": coerce_num(safe_get(row, cols["market"]), default=None),
                "lease_start": _as_date(safe_get(row, cols["lease_start"])),
                "lease_end": _as_date(safe_get(row, cols["lease_end"])),
                "move_in": _as_date(safe_get(row, cols["move_in"])),
                "move_out": _as_date(safe_get(row, cols["move_out"])),
                "_rent_accum": 0.0,
            }

        if current is None:
            continue

        desc = safe_get(row, cols["description"])
        amt = coerce_num(safe_get(row, cols["amount"]), default=None)
        if desc is None or amt is None:
            continue
        label = str(desc).strip()
        if norm(label) in ("total", "totals"):
            continue          # summary line; the charge lines above already counted
        if _is_rent_line(label, amt):
            current["_rent_accum"] += amt

    close(current)

    if not units:
        raise UnrecognizedRentRoll(
            "The rent roll header was recognized but no unit rows could be read "
            "from it. Check the file is a full rent roll export rather than a "
            "summary."
        )

    missing_market = sum(1 for u in units if u["market_rent"] is None)
    if missing_market:
        warnings.append(f"{missing_market} unit(s) have no market rent on file; "
                        f"their in-place rent is used for gross potential rent instead.")
    missing_sqft = sum(1 for u in units if u["sqft"] is None)
    if missing_sqft:
        warnings.append(f"{missing_sqft} unit(s) have no square footage; they are "
                        f"excluded from average-sqft figures rather than counted as zero.")

    return {
        "units": units,
        "unit_count": len(units),
        "warnings": warnings,
        "source_format": "ResMan Rent Roll",
    }
